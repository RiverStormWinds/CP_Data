# -*- coding:gbk -*-
from ctypes import *
# from ctypes import wintypes
import MySQLdb
import pyodbc
import json
import time
import subprocess
import shutil
from operator import itemgetter, attrgetter
import datetime
from datetime import date
from datetime import timedelta
import sys
import random

import chardet
reload(sys)
sys.setdefaultencoding('gbk')
from gl import *
import wx
import ConfigParser
import os
import cx_Oracle
import string as Str
import gl
import csv
from decimal import Decimal
from gl import GlobalLogging as Logging
import urllib
import urllib2
import suds
import telnetlib

try:
    import xml.etree.cElementTree as ET
except ImportError:
    import xml.etree.ElementTree as ET
from xml.dom import minidom
import codecs
import struct
import socket
import wmi
import Adapter.T2.TimesvcT2SDK as t2
import O32T2SDK

#新接口
# import TimesvcT2SDKdll as t2
# import O32T2SDKdll as O32T2SDK

import message
import random
import bizmod
from lxml import etree
from collections import OrderedDict
import zipfile
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from ftpClient import XFer
import decimal
import time
import binascii

try:
    from agw import pyprogress as PP
except ImportError:  # if it's not there locally, try the wxPython lib.
    import wx.lib.agw.pyprogress as PP
import cPickle
import re
import zlib
import base64
from xmlrpclib import ServerProxy
import subprocess
import param


# import dir_process
# import secrecy as sc
def readConfigFile():
    '''读取XML系统配置文件 '''
    try:
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        elif __file__:
            application_path = os.path.dirname(__file__)
        else:
            Logging.getLog().critical(u"无法获取application_path")
            return -1
        Logging.getLog().debug("application_path = %s" % application_path)

        Logging.getLog().debug('try to switch cwd to %s' % application_path)
        os.chdir(application_path)

        if os.path.exists("settings.json"):
            _setting = json.loads(open("settings.json", "r").read())
            gl.g_testcaseDBDict["DBType"] = _setting["casedb"]["DBType"]
            gl.g_testcaseDBDict["casedb"] = _setting["casedb"][gl.g_testcaseDBDict["DBType"]]
            gl.gl_ftp_info = _setting

        if os.path.exists(".\\init\\%s\ZHLCAutoTestSetting.xml" % gl.current_system_id):
            tree = ET.parse(".\\init\\%s\\ZHLCAutoTestSetting.xml" % gl.current_system_id)  # fixme 不能使用这个固定的路径，以后更改
            if tree:
                root = tree.getroot()
                try:
                    # 解析帐号信息
                    # 帐号信息格式为以类型为键值的字典类型 {type1:value1,type2:value2 ……}
                    accountmessage = {}
                    AccountInfoElem = root.find("AccountInfo")
                    account_group_id = tree.findall('.//AccountInfo')
                    dist = {}
                    for elem in account_group_id:
                        for xx in elem.getchildren():
                            tx = xx.text.replace(',', '&comma&')
                            tx = tx.replace(':', '&colon&')
                            dist[xx.tag] = tx
                        gl.g_accountDict[elem.attrib["id"]] = dist  # fixme 账号信息以后尽量不从这个文件中读取
                        dist = {}
                        #
                        # 解析案例数据库信息

                        # 将用例数据库配置信息
                        # testcaseDBElem = root.find("TestCaseDataBase")
                        # for elem in testcaseDBElem:
                        #     gl.g_testcaseDBDict[elem.tag] = elem.text

                        # 解析BeyondCompare程序路径
                        # ByondCmpElem = root.find("BeyondComparePath/path")
                        # gl.g_bcmppath = ByondCmpElem.text

                        # 解析集中交易数据库信息
                        # SqlServerElem = root.find("ConnectSqlServerSetting")
                        # for serverElem in SqlServerElem:
                        #     gl.g_connectSqlServerSetting[serverElem.attrib["core"]] ={}
                        #     for elem in serverElem:
                        #         gl.g_connectSqlServerSetting[serverElem.attrib["core"]][elem.tag]=elem.text
                        #
                        #
                        # #解析综合理财数据库信息
                        # OracleElem = root.find("ConnectOracleSetting")
                        # for serverElem in OracleElem:
                        #     gl.g_connectOracleSetting[serverElem.attrib["core"]] ={}
                        #     for elem in serverElem:
                        #         gl.g_connectOracleSetting[serverElem.attrib["core"]][elem.tag]=elem.text
                        #
                        # 解析综合理财系统信息
                        # ZHLCSystemSettingElem = root.find("ZHLCSystemSetting")
                        # for elem in ZHLCSystemSettingElem:
                        #    gl.g_ZHLCSystemSetting[elem.tag] = elem.text

                        # #解析需要修改运行时间的服务器列表
                        # TimeSyncServerSettingElem = root.find("TimeSyncServerSetting")
                        # if TimeSyncServerSettingElem != None:
                        #     for elem in TimeSyncServerSettingElem:
                        #         gl.g_timeSyncServerSettingList.append(elem.attrib["ip"])
                        # loadSYSCases()
                except BaseException, ex:
                    exc_info = getExceptionInfo()
                    strerror = u"解析通用配置文件ZHLCAutoTestSetting.xml 出现异常。原因：%s" % (exc_info)
                    Logging.getLog().critical(strerror)
                    return -1
                Logging.getLog().info(u"解析通用配置文件ZHLCAutoTestSetting.xml 完毕")
                return 0, ""
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        strerror = u"无法解析配置文件ZHLCAutoTestSetting.xml 原因：%s" % (ex)
        # wx.MessageBox(strerror,u"提示")
        Logging.getLog().critical(strerror)
        return -1


def readJzjyAccountInfo(system_id):
    try:
        gl.g_default_account_info = {}
        system_id = system_id.upper()
        homedir = os.getcwd()
        tree = ""
        if os.path.exists("%s\\init\\%s\orginCount.xml" % (homedir, system_id)):
            tree = ET.parse("%s\\init\\%s\orginCount.xml" % (homedir, system_id))
        if not tree:
            Logging.getLog().debug(u"未读取系统%s账户组的配置文件" % system_id)
        else:
            root = tree.getroot()
            JZJYSystemSettingElem = root.find("JZJYSystemSetting")
            if JZJYSystemSettingElem != None:
                for ServerSettingElem in JZJYSystemSettingElem:
                    gl.g_JZJYSystemSetting[ServerSettingElem.attrib["id"]] = ServerSettingElem
                    AccountGroupDict = {}
                    for AccountGroupElem in ServerSettingElem:
                        paramDict = {}
                        for elem in AccountGroupElem:
                            paramDict[elem.tag] = elem.text
                        AccountGroupDict[AccountGroupElem.attrib["id"]] = paramDict
                    gl.g_default_account_info[ServerSettingElem.attrib["id"]] = AccountGroupDict
        return 0, ""
    except:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def readNewAccountInfo(system_id):
    try:
        # gl.g_paramInfoDict = {}
        gl.g_accountDict = {}
        system_id = system_id.upper()
        suffix = get_suffix(system_id)
        homedir = os.getcwd()
        tree = ""
        Logging.getLog().debug("system_id:%s" % system_id)
        if os.path.exists("%s\\init\\%s\\%sAccountGroupSetting.xml" % (homedir, system_id, suffix)):
            tree = ET.parse("%s\\init\\%s\\%sAccountGroupSetting.xml" % (homedir, system_id, suffix))
        if not tree:
            Logging.getLog().debug(u"未读取系统%s账户组的配置文件" % system_id)
        else:
            root = tree.getroot()
            JZJYSystemSettingElem = root.find("JZJYSystemSetting")
            if JZJYSystemSettingElem != None:
                for ServerSettingElem in JZJYSystemSettingElem:
                    gl.g_JZJYSystemSetting[ServerSettingElem.attrib["id"]] = ServerSettingElem
                    AccountGroupDict = {}
                    for AccountGroupElem in ServerSettingElem:
                        paramDict = {}
                        for elem in AccountGroupElem:
                            paramDict[elem.tag] = elem.text
                        AccountGroupDict[AccountGroupElem.attrib["id"]] = paramDict
                    gl.g_paramInfoDict[ServerSettingElem.attrib["id"]] = AccountGroupDict
            else:
                account_group_id = tree.findall('.//AccountInfo')
                dist = {}
                accountDict = OrderedDict()
                for elem in account_group_id:
                    for xx in elem.getchildren():
                        if xx.text:
                            tx = xx.text.replace(',', '&comma&')
                            dist[xx.tag] = tx
                        else:
                            tx = xx.text
                            dist[xx.tag] = tx
                    accountDict[elem.attrib["id"]] = dist
                    dist = {}
                gl.g_accountDict = accountDict

        if os.path.exists("settings.json"):
            _setting = json.loads(open("settings.json", "r").read())
            gl.g_testcaseDBDict["DBType"] = _setting["casedb"]["DBType"]
            gl.g_testcaseDBDict["casedb"] = _setting["casedb"][gl.g_testcaseDBDict["DBType"]]
    except:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def readSystemConfigFile(system_id):
    '''读取和系统相关的XML系统配置文件 '''
    try:
        # 对去集中交易固定账户组数据
        readJzjyAccountInfo(system_id)
        readNewAccountInfo(system_id)
        # if gl.g_server_id == "":
        #     return
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        elif __file__:
            application_path = os.path.dirname(__file__)
        Logging.getLog().debug("application_path = %s" % application_path)

        Logging.getLog().debug('try to switch cwd to %s' % application_path)
        # os.chdir(application_path)

        if system_id == "":
            Logging.getLog().critical(u"未提供系统ID信息")
            return -1
        homedir = os.getcwd()
        tree, root = "", ""
        if gl.g_server_id != "":  # FIXME 需要修改为上线时实际的system_id
            if os.path.exists("%s\\init\\%s\\ZHLCAutoTestSetting_%s.xml" % (homedir, system_id, gl.g_server_id)):
                tree = ET.parse("%s\\init\\%s\\ZHLCAutoTestSetting_%s.xml" % (homedir, system_id, gl.g_server_id))
        else:
            gl.g_connectOracleSetting = {}
            gl.g_connectSqlServerSetting = {}
            if os.path.exists("%s\\init\\%s\\ZHLCAutoTestSetting.xml" % (homedir, system_id)):
                tree = ET.parse("%s\\init\\%s\\ZHLCAutoTestSetting.xml" % (homedir, system_id))
        if tree:
            root = tree.getroot()
        if os.path.exists("%s\\init\\%s\\T2InterfaceDetails.xml" % (homedir, system_id)):
            gl.g_interfaceDetailTree = ET.parse("%s\\init\\%s\\T2InterfaceDetails.xml" % (homedir, system_id))
        else:
            Logging.getLog().debug(u'系统不存在该接口配置文件T2InterfaceDetails.xml')
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        strerror = u"无法解析配置文件AutoTestSetting.xml 原因：%s" % (ex)
        # wx.MessageBox(strerror,u"提示")
        Logging.getLog().critical(strerror)
        return -1
    try:
        # readNewAccountInfo(system_id)
        if not tree:
            return 0, ""
        # 解析集中交易数据库信息
        SqlServerElem = root.find("ConnectSqlServerSetting")
        for serverElem in SqlServerElem:
            gl.g_connectSqlServerSetting[serverElem.attrib["core"]] = {}
            for elem in serverElem:
                gl.g_connectSqlServerSetting[serverElem.attrib["core"]][elem.tag] = elem.text

        # 解析综合理财数据库信息
        OracleElem = root.find("ConnectOracleSetting")
        if OracleElem != None:
            for serverElem in OracleElem:
                gl.g_connectOracleSetting[serverElem.attrib["core"]] = {}
                for elem in serverElem:
                    gl.g_connectOracleSetting[serverElem.attrib["core"]][elem.tag] = elem.text

        # 解析综合理财系统信息
        ZHLCSystemSettingElem = root.find("ZHLCSystemSetting")

        if ZHLCSystemSettingElem != None:
            for elem in ZHLCSystemSettingElem:
                gl.g_ZHLCSystemSetting[elem.tag] = elem.text
                if str(elem.tag) == "KCBP_IP":
                    if not gl.g_core_ip_port_info.has_key(str(gl.g_server_id)):
                        gl.g_core_ip_port_info[str(gl.g_server_id)] = {}
                        gl.g_core_ip_port_info[str(gl.g_server_id)]["ip"] = elem.text
                    else:
                        gl.g_core_ip_port_info[str(gl.g_server_id)]["ip"] = elem.text
                if str(elem.tag) == "KCBP_PORT":
                    if not gl.g_core_ip_port_info.has_key(str(gl.g_server_id)):
                        gl.g_core_ip_port_info[str(gl.g_server_id)] = {}
                        gl.g_core_ip_port_info[str(gl.g_server_id)]["KCBP_PORT"] = elem.text
                    else:
                        gl.g_core_ip_port_info[str(gl.g_server_id)]["KCBP_PORT"] = elem.text
                if str(elem.tag) == "KcbpConfigure":
                    KcbpConfigureList = elem.text.split(",")
                    for key_value in KcbpConfigureList:
                        key, value = key_value.split(":")
                        value = value.replace("|", ":")
                        gl.g_KcxpDataElemDict[str(key).strip()] = str(value).strip()
                if str(elem.tag) == "deviation":
                    gl.g_deviationPCT = str(elem.text).strip()

                if str(elem.tag) == "SendQName":
                    gl.SendQName = str(elem.text).strip()

                if str(elem.tag) == "ReceiveQName":
                    gl.ReceiveQName = str(elem.text).strip()
        # 解析需要修改运行时间的服务器列表
        TimeSyncServerSettingElem = root.find("TimeSyncServerSetting")
        if TimeSyncServerSettingElem != None:
            gl.g_timeSyncServerSettingList = []
            for elem in TimeSyncServerSettingElem:
                gl.g_timeSyncServerSettingList.append(elem.attrib["ip"])
                # cF.loadSYSCases()
        CasesSleepTimeElem = root.find("OtherSetting/CasesSleepTime")
        if CasesSleepTimeElem != None:
            gl.g_cases_sleep_time = float(CasesSleepTimeElem.text)
        # gl.g_allows_error_dict = {}
        AllowErrorsElem = root.find("OtherSetting/AllowErrors")
        if AllowErrorsElem != None:
            for FuncidErrorElem in AllowErrorsElem:
                gl.g_allows_error_dict.setdefault(FuncidErrorElem.attrib["funcid"], [])
                for errorElem in FuncidErrorElem:
                    text = errorElem.text
                    texts = text.split(",")
                    gl.g_allows_error_dict[FuncidErrorElem.attrib["funcid"]].extend(texts)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        strerror = u"解析配置文件AutoTestSetting.xml 出现异常。原因：%s" % (exc_info)
        Logging.getLog().critical(strerror)
        return -1
    # bizmod.traversalDir_XMLFile(system_id)
    Logging.getLog().info(u"解析配置文件AutoTestSetting.xml 完毕")
    return 0


def add_tnsnames_info():
    '''添加ORACLE信息 '''
    try:
        all_oracle_info = []
        for i in gl.g_connectOracleSetting:
            DB_ALIAS = gl.g_connectOracleSetting[i]["DB_ALIAS"]
            ServiceName = gl.g_connectOracleSetting[i]["ServiceName"]
            IP = gl.g_connectOracleSetting[i]["IP"]
            DB_PORT = gl.g_connectOracleSetting[i]["DB_PORT"]
            oracle_info = """%s =
(DESCRIPTION =
  (ADDRESS_LIST =
    (ADDRESS = (PROTOCOL = TCP)(HOST = %s)(PORT = %s))
  )
  (CONNECT_DATA =
    (SERVICE_NAME = %s)
  )
)""" % (DB_ALIAS, IP, DB_PORT, ServiceName)
            all_oracle_info.append(oracle_info)
        all_oracle_info = list(set(all_oracle_info))
        if all_oracle_info:
            environ_info = os.environ
            oracle_path = ""
            if "TNS_ADMIN" in environ_info.keys():
                oracle_path = environ_info['TNS_ADMIN']
            f = open(r'%s\tnsnames.ora' % oracle_path, 'w+')
            for i in all_oracle_info:
                f.write(i)
                f.write("\n")
            f.close()
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1


def readConfigFileForTools():
    '''读取XML系统配置文件 '''
    try:
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        elif __file__:
            application_path = os.path.dirname(__file__)
        else:
            Logging.getLog().critical(u"无法获取application_path")
            return -1
        Logging.getLog().debug("application_path = %s" % application_path)

        Logging.getLog().debug('try to switch cwd to %s' % application_path)
        os.chdir(application_path)

        tree = ET.parse(".\\init\\14774643386617P4VW\\ZHLCAutoTestSetting.xml")  # fixme 不能使用这个固定的路径，以后更改
        root = tree.getroot()
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        strerror = u"无法解析配置文件ZHLCAutoTestSetting.xml 原因：%s" % (ex)
        # wx.MessageBox(strerror,u"提示")
        Logging.getLog().critical(strerror)
        return -1
    try:
        # 解析帐号信息
        # 帐号信息格式为以类型为键值的字典类型 {type1:value1,type2:value2 ……}
        accountmessage = {}
        AccountInfoElem = root.find("AccountInfo")
        account_group_id = tree.findall('.//AccountInfo')
        dist = {}
        for elem in account_group_id:
            for xx in elem.getchildren():
                tx = xx.text.replace(',', '&comma&')
                tx = tx.replace(':', '&colon&')
                dist[xx.tag] = tx
            gl.g_accountDict[elem.attrib["id"]] = dist  # fixme 账号信息以后尽量不从这个文件中读取
            dist = {}
        #
        # 解析案例数据库信息

        # 将用例数据库配置信息
        # testcaseDBElem = root.find("TestCaseDataBase")
        # for elem in testcaseDBElem:
        #     gl.g_testcaseDBDict[elem.tag] = elem.text

        if os.path.exists("settings.json"):
            _setting = json.loads(open("settings.json", "r").read())
            gl.g_testcaseDBDict["DBType"] = _setting["casedb"]["DBType"]
            gl.g_testcaseDBDict["casedb"] = _setting["casedb"][gl.g_testcaseDBDict["DBType"]]

        # 解析BeyondCompare程序路径
        # ByondCmpElem = root.find("BeyondComparePath/path")
        # gl.g_bcmppath = ByondCmpElem.text

        # 解析集中交易数据库信息
        SqlServerElem = root.find("ConnectSqlServerSetting")
        for serverElem in SqlServerElem:
            gl.g_connectSqlServerSetting[serverElem.attrib["core"]] = {}
            for elem in serverElem:
                gl.g_connectSqlServerSetting[serverElem.attrib["core"]][elem.tag] = elem.text
        # SqlServerElem = root.find("ConnectSqlServerSetting")
        # for serverElem in SqlServerElem:
        #     gl.g_connectSqlServerSetting[serverElem.attrib["servertype"]] = {}
        #     for elem in serverElem:
        #         gl.g_connectSqlServerSetting[serverElem.attrib["servertype"]][elem.tag] = elem.text

        # 解析综合理财数据库信息
        OracleElem = root.find("ConnectOracleSetting")
        for serverElem in OracleElem:
            gl.g_connectOracleSetting[serverElem.attrib["core"]] = {}
            for elem in serverElem:
                gl.g_connectOracleSetting[serverElem.attrib["core"]][elem.tag] = elem.text

        # 解析综合理财系统信息
        ZHLCSystemSettingElem = root.find("ZHLCSystemSetting")
        for elem in ZHLCSystemSettingElem:
            gl.g_ZHLCSystemSetting[elem.tag] = elem.text

            # #解析需要修改运行时间的服务器列表
            # TimeSyncServerSettingElem = root.find("TimeSyncServerSetting")
            # if TimeSyncServerSettingElem != None:
            #     for elem in TimeSyncServerSettingElem:
            #         gl.g_timeSyncServerSettingList.append(elem.attrib["ip"])
            # loadSYSCases()
    except BaseException, ex:
        exc_info = getExceptionInfo()
        strerror = u"解析通用配置文件ZHLCAutoTestSetting.xml 出现异常。原因：%s" % (exc_info)
        Logging.getLog().critical(strerror)
        return -1
    Logging.getLog().info(u"解析通用配置文件ZHLCAutoTestSetting.xml 完毕")
    return 0


def loadSYSCases():
    '''读取XML系统配置文件 '''
    try:
        if gl.g_interfaceDetailTree == None:
            gl.g_interfaceDetailTree = ET.parse(".\\init\\142891090059622AV2\\T2InterfaceDetails.xml")
        tree = ET.parse(".\\init\\142891090059622AV2\\T2InterfaceDetails.xml")
        root = tree.getroot()
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        strerror = u"无法解析配置文件T2InterfaceDetails.xml 原因：%s" % (ex)
        wx.MessageBox(strerror, u"提示")
        Logging.getLog().critical(strerror)
        return -1
    try:
        # 解析案例库
        for elem in root:
            # gl.g_SysCasesList.append(elem.attrib['name'])
            gl.g_SysCasesList[elem.attrib['funcid']] = {}
            try:
                gl.g_SysCasesList[elem.attrib['funcid']]['method'] = elem.attrib['funcname']
            except BaseException, ex:
                gl.g_SysCasesList[elem.attrib['funcid']]['method'] = ""

                # 去除重复案例
                # gl.g_SysCasesList = set(gl.g_SysCasesList)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        strerror = u"解析配置文件T2InterfaceDetails.xml 出现异常。原因：%s" % (exc_info)
        Logging.getLog().critical(strerror)
        return -1
        # Logging.getLog().info(u"解析配置文件T2InterfaceDetails.xml 完毕")


def GetTimeStamp():  # 获取本地时间
    return time.strftime('%Y-%m-%d %X', time.localtime(time.time()))


def ConnectOracleDB(server, cmd='-1'):  # 连接Oracle数据库
    # Logging.getLog().info(u"cbs---g_connectOracleSetting %s" % str(gl.g_connectOracleSetting))
    AUTOName = 'AUTO_' + gl.g_connectOracleSetting[server]['UserName']
    try:
        # Logging.getLog().debug("gl.g_connectOracleSetting:%s" % str(gl.g_connectOracleSetting))
        if len(gl.g_connectOracleSetting.keys()) > 0 and cmd == '-1':  # 已经读取了配置文件
            # conn = cx_Oracle.connect(gl.g_ZHLCSystemSetting["OracleUserName"],gl.g_ZHLCSystemSetting["OraclePassword"],gl.g_ZHLCSystemSetting["OracleServiceName"])
            conn = cx_Oracle.connect(gl.g_connectOracleSetting[server]['UserName'],
                                     gl.g_connectOracleSetting[server]['Password'],
                                     gl.g_connectOracleSetting[server]['DB_ALIAS'])
            print gl.g_connectOracleSetting[server]['UserName'], gl.g_connectOracleSetting[server]['Password'], \
                gl.g_connectOracleSetting[server]['ServiceName']
            print type(gl.g_connectOracleSetting[server]['UserName']), type(
                gl.g_connectOracleSetting[server]['Password']), type(gl.g_connectOracleSetting[server]['DB_ALIAS'])
            # else: #仅作调试之用,如果测试环境发生变化，需要修改
            #   conn = cx_Oracle.connect('hs_acct','hundsun','ACCT200_gtja')
        if cmd == '0':
            conn = cx_Oracle.connect(AUTOName.upper(), gl.g_connectOracleSetting[server]['Password'],
                                     gl.g_connectOracleSetting[server]['DB_ALIAS'])
    except Exception, ex:
        exc_info = getExceptionInfo()
        strerror = u"CreateOracleConn 连接  Oracle数据库 出现问题，原因：%s" % (exc_info)
        Logging.getLog().critical(strerror)
        return -1
    return conn


def ConnectCaseDB(type='MYSQL'):  # 连接数据库
    """ using the mysqldb to connect mysql database"""
    dbstring = ""
    try:
        db_dict = gl.g_testcaseDBDict
        if not db_dict:
            readSystemConfigFile(gl.current_system_id)
        conn = None
        if type == 'MYSQL':
            conn = MySQLdb.connect(host=db_dict["casedb"]["casedbip"],
                                   user=db_dict["casedb"]["casedbuser"],
                                   passwd=db_dict["casedb"]["casedbpwd"],
                                   port=int(db_dict["casedb"]["casedbport"]),
                                   db=db_dict["casedb"]["casedbname"],
                                   charset='gbk')
        elif type == 'DB2':
            dbstring = "driver={IBM DB2 ODBC DRIVER};database=%s;hostname=%s;port=%s;protocol=tcpip;uid=%s;pwd=%s;LONGDATACOMPAT=1;LOBMAXCOLUMNSIZE=10485875;" % (
                db_dict["casedb"]["casedbname"], db_dict["casedb"]["casedbip"], int(db_dict["casedb"]["casedbport"]),
                db_dict["casedb"]["casedbuser"], db_dict["casedb"]["casedbpwd"])
            conn = pyodbc.connect(dbstring)
        return conn
    except:
        Logging.getLog().info(dbstring)
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1


# def ConnectToRunDB(servertype='p'):
#     # servertype = 'p1'
#     dbstring = "DRIVER={SQL Server};SERVER=%s;DATABASE=%s;UID=%s;PWD=%s;autocommit=True" % (
#         gl.g_connectSqlServerSetting[servertype]["ip"], gl.g_connectSqlServerSetting[servertype]["database"],
#         gl.g_connectSqlServerSetting[servertype]["username"], gl.g_connectSqlServerSetting[servertype]["password"])
#     return pyodbc.connect(dbstring)
def ConnectToRunDB(server=0):
    dbstring = "DRIVER={SQL Server};SERVER=%s;DATABASE=%s;UID=%s;PWD=%s;autocommit=True" % (
        gl.g_connectSqlServerSetting[str(server)]["ip"], gl.g_connectSqlServerSetting[str(server)]["database"],
        gl.g_connectSqlServerSetting[str(server)]["username"], gl.g_connectSqlServerSetting[str(server)]["password"])
    try:
        conn = pyodbc.connect(dbstring)
        return 0, conn
    except BaseException, ex:
        return -1, "SQL Server connect error, ip:%s" % gl.g_connectSqlServerSetting[str(server)]["ip"]


def executeUpdateSQL(sqlstr, servertype="p"):
    ret, conn = ConnectToRunDB(servertype)
    crs = conn.cursor()
    crs.execute(sqlstr)
    crs.commit()
    crs.close()
    conn.close()


def get_core(servertype):
    core = -1
    if gl.g_connectSqlServerSetting:
        for server in gl.g_connectSqlServerSetting.keys():
            if str(gl.g_connectSqlServerSetting[str(server)]["servertype"]) == str(servertype):
                core = str(server)
                break
    return core


def executeCaseSQL(sqlstr, cmd='-1'):
    """ execute the sql command in mysql database, and return the dataset in dict structure"""
    # print(sqlstr)
    '''
    needReConnect = True
    ReConnectCount = 0
    ConnectCount = 0
    while (needReConnect == True and ReConnectCount <= 3):
        
        try:
            conn = ConnectCaseDB()
        except Exception,ex:
            strerror = u"连接MySQL数据库出现问题，原因：%s" %(ex)
            #wx.MessageBox(strerror,u"提示")
            Logging.getLog().critical(strerror)
            needReConnect = True
            ReConnectCount = ReConnectCount + 1
            ConnectCount = ReConnectCount
            #print 'ConnectCount:',ConnectCount
        
        else:
            #print 'ReConnectCount:',ReConnectCount
            #print 'ConnectCount Final:',ConnectCount
            needReConnect = False
            ReConnectCount = 0
        
    if (ConnectCount > 2):
        strerror = u"连接MySQL数据库出现问题，原因：连接失败次数超过三次!  %s" %(ex)
        #wx.MessageBox(strerror,u"提示")
        Logging.getLog().critical(strerror)
        return []
            
    '''

    # if gl.g_ZHLCSystemSetting['TestCaseDBType'] == 'DB2':
    if gl.g_testcaseDBDict["DBType"] == "DB2":
        return executeCaseSQL_DB2(sqlstr)

    flag = False
    if g_SysRunFlag == False:
        # readConfigFile()
        g_MySQLConn = ConnectCaseDB()
        flag = True
    ds = []
    columns = []
    if sqlstr.__contains__("update") or sqlstr.__contains__("insert") or sqlstr.__contains__(
            "delete") or sqlstr.__contains__("replace"):
        crs = g_MySQLConn.cursor()
        if cmd == '-1':
            crs.execute(sqlstr)
        else:
            crs.execute(sqlstr, cmd)
        g_MySQLConn.commit()
    else:
        crs = g_MySQLConn.cursor(MySQLdb.cursors.DictCursor)
        # rowNumber = crs.execute(sqlstr)
        if cmd == '-1':
            rowNumber = crs.execute(sqlstr)
        else:
            rowNumber = crs.execute(sqlstr, cmd)
        if rowNumber != 0:
            for rec in crs.fetchall():
                ds.append(rec)
    crs.close()
    if flag:
        g_MySQLConn.close()
    return ds


def executeCaseSinGelSQL(sql, Data):
    try:
        conn = ConnectCaseDB()
        crs = conn.cursor()
        crs.execute(sql, Data)
        conn.commit()
        crs.close()
        conn.close()
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info
    return 0, ""


def executeCaseManySQL(strModel, listData):
    try:
        conn = ConnectCaseDB()
        crs = conn.cursor(MySQLdb.cursors.DictCursor)
        crs.executemany(strModel, listData)
        conn.commit()
        crs.close()
        conn.close()
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info
    return 0, ""


def executeCaseManySQL_DB2(strModel, listData):  # Fixme 修改为MYSQL兼容的
    try:
        if not listData:
            return 0, ""
        conn = ConnectCaseDB('DB2')
        crs = conn.cursor()
        listData = [listData[i:i + 1000] for i in range(0, len(listData), 1000)]
        for child_listData in listData:
            crs.executemany(strModel, child_listData)
            conn.commit()
        crs.close()
        conn.close()
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().info(str(listData))
        Logging.getLog().error(exc_info)
        crs.close()
        conn.rollback()
        conn.close()
        return -1, exc_info
    return 0, ""


def executeCaseSQL_DB2(sqlstr):
    '''
    执行DB2 SQL
    '''
    try:
        ds = []
        conn = ConnectCaseDB('DB2')
        crs = conn.cursor()
        crs.execute(sqlstr)
        if sqlstr.lower().__contains__("update") or sqlstr.__contains__("insert") or sqlstr.__contains__(
                "delete") or sqlstr.__contains__("replace"):
            conn.commit()
        else:
            rows = crs.fetchall()
            ds = []

            for row in rows:
                rec = {}
                for i, desc in enumerate(row.cursor_description):
                    if isinstance(row[i], str):
                        rec[desc[0]] = row[i].decode('gbk')
                    else:
                        rec[desc[0]] = row[i]
                ds.append(rec)
        # conn.commit()
        crs.close()
        conn.close()
        return ds
    except BaseException, ex:
        Logging.getLog().info(sqlstr)
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return []


def exeRunSqlserver(sqlstr, login_info):
    try:
        dbstring = "DRIVER={SQL Server};SERVER=%s;DATABASE=%s;UID=%s;PWD=%s;autocommit=True" % (
            login_info["ip"], login_info["database"],
            login_info["username"],
            login_info["password"])
        conn = pyodbc.connect(dbstring, timeout=0)
        crs = conn.cursor()
        crs.execute(sqlstr)
        ds = []
        if sqlstr.lower().__contains__("select") and "select * into" not in sqlstr.lower():
            rows = crs.fetchall()  # row objects are a tuple-like objects which returned by Cursor fetch function,as specified in the DB API.

            for row in rows:
                rec = {}
                for i, desc in enumerate(row.cursor_description):
                    if isinstance(row[i], decimal.Decimal):
                        rec[desc[0]] = str(row[i])
                    elif isinstance(row[i], str):
                        rec[desc[0]] = str(row[i]).strip().decode("gbk")
                    else:
                        rec[desc[0]] = row[i]
                ds.append(rec)
        conn.commit()
        crs.close()
        conn.close()
        return 0, ds
    except Exception as e:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1, exc_info


def executeRunSQL(sqlstr, server_id):
    try:
        try:
            server_id = int(server_id)
        except Exception as e:
            server_id = get_core(server_id)
        ret, conn = ConnectToRunDB(server_id)
        crs = conn.cursor()
        crs.execute(sqlstr)
        ds = []
        if sqlstr.startswith("insert into"):
            pass
        elif sqlstr.lower().__contains__("select") and "select * into" not in sqlstr.lower():
            rows = crs.fetchall()  # row objects are a tuple-like objects which returned by Cursor fetch function,as specified in the DB API.
            for row in rows:
                rec = {}
                for i, desc in enumerate(row.cursor_description):
                    n_desx = desc[0]
                    if isinstance(n_desx, str):
                        n_desx = n_desx.strip().decode("gbk")
                    if row[i] == None:
                        rec[n_desx] = ""
                        continue
                    if isinstance(row[i], decimal.Decimal):
                        rec[n_desx] = str(row[i])
                    elif isinstance(row[i], str):
                        rec[n_desx] = str(row[i]).strip().decode("gbk")
                    else:
                        rec[n_desx] = row[i]
                ds.append(rec)
        else:
            pass
        conn.commit()
        crs.close()
        conn.close()
        return ds
    except Exception as e:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1


# def executeRunSQL_COMPARE(sqlstr, servertype="p"):
#     conn = ConnectToRunDB_COMPARE(server)
#     crs = conn.cursor()
#     crs.execute(sqlstr)
#     rows = crs.fetchall()  # row objects are a tuple-like objects which returned by Cursor fetch function,as specified in the DB API.
#     ds = []

#     for row in rows:
#         rec = {}
#         for i, desc in enumerate(row.cursor_description):
#             rec[desc[0]] = row[i]
#         ds.append(rec)
#     crs.close()
#     conn.close()
#     return ds


def getLastWorkDay(nowday):
    d = date(int(nowday[0:4]), int(nowday[4:6]), int(nowday[6:8]))
    d = d + timedelta(days=-1)
    while d.weekday() == 5 or d.weekday() == 6:
        d = d + timedelta(days=-1)
    return d.strftime('%Y%m%d')


def getNextWorkDay(nowday):
    d = date(int(nowday[0:4]), int(nowday[4:6]), int(nowday[6:8]))
    d = d + timedelta(days=1)
    while d.weekday() == 5 or d.weekday() == 6:
        d = d + timedelta(days=1)
    return d.strftime('%Y%m%d')


def getNextyearOfsysday():
    date1 = int(param.getSysDate()) + 10000
    return str(date1)


def getValueByKey(cmd, key):
    '''
    Args:
        cmd: 命令行
        key: 关键字
    Returns:
        关键字对应的值
        如果没找到，返回None。如果对应值为空，返回''
    '''

    lpos = cmd.find(key)
    if lpos == -1:
        return None
    if lpos + len(key) == len(cmd) or cmd[lpos + len(key)] == '#':
        return ''
    rpos = cmd.find('#', lpos + 1)
    if rpos == -1:
        value = cmd[lpos + len(key):]
        if value[:2] != '$$':
            return None
        return value
    else:
        value = cmd[lpos + len(key):rpos]
        if value[:2] != '$$':
            return None
        return value


def getValueFromCmd(cmd, key):
    keylist = ['price:', 'begindate:', 'enddate:', 'regdate:']
    if key in keylist:
        lpos = cmd.find(',' + key)
        if lpos != -1:
            if lpos + len(key) + 1 == len(cmd) or cmd[lpos + len(key) + 1] == ',':
                return ''
            rpos = cmd.find(',', lpos + 1)
            if rpos == -1:
                value = cmd[lpos + len(key) + 1:]
                return value
            else:
                value = cmd[lpos + len(key) + 1:rpos]
                return value
        else:
            return None
    else:
        lpos = cmd.find(key)
        if lpos == -1:
            return None
        if lpos + len(key) == len(cmd) or cmd[lpos + len(key)] == '#':
            return ''
        rpos = cmd.find('#', lpos + 1)
        if rpos == -1:
            value = cmd[lpos + len(key):]
            return value
        else:
            value = cmd[lpos + len(key):rpos]
            return value


def getDateStringByKey(datetype):
    '''
    Args:
        datetype: 参数值类型，可能是参数值($$开头），也可能就是普通日期值
    Return:
        转化后的参数值
    '''
    if datetype == '$$date_1':
        return str(param.getSysDate())
    elif datetype == '$$date_2':
        return getLastWorkDay(str(param.getSysDate()))
    elif datetype == '$$date_3':
        return getNextWorkDay(str(param.getSysDate()))
    else:
        return datetype


# def FindKeyFromCmdFix(cmd):  # 案例录制时候不对委托号做参数处理
#     for item in g_keylistprice:
#         if cmd.find(item) == -1:
#             continue
#         else:
#             return 0
#     for item in g_keylistdate:
#         if cmd.find(item) == -1:
#             continue
#         else:
#             return 0
#     return 1


def queryFileByName(describe):
    title = "Autotest_report"
    new_path = os.path.join(os.getcwd(), title)
    if not os.path.isdir(new_path):
        os.mkdir(new_path)
    rootname = os.getcwd() + '\\Autotest_report\\'
    temp_file_list = os.listdir(rootname)
    filelist = []
    for item in temp_file_list:
        if item.find(describe) != -1:
            filelist.append(item[-18:-4])
    filelist.sort(reverse=True)
    if len(filelist) == 0:
        return "None_Report.txt"
    filename = rootname + describe[0:-1] + filelist[0] + '.txt'
    return filename


def formatspecialcharacter(cmd):
    if cmd.find('\\') != -1 and cmd.find('\\\\') == -1:
        cmd = (r'%s' % (cmd)).replace('\\', '\\\\')
        pass
    return cmd


def queryMenupos(funcid):
    sql = "select menupos from tbl_sysmenu where funcid = '%s'" % (funcid)
    ds = executeCaseSQL(sql)
    return ds[0]["menupos"]


def testOracleConn():
    conn = cx_Oracle.connect("system", "hbzq", "ABOSS2")  # system/密码/数据库名

    cursor = conn.cursor()
    r = cursor.execute("SELECT * FROM aboss.tqs_qsrwzt_20130910")
    result = r.fetchall()
    testdict = {}
    FINE_csv = open("_oracle.csv", "w")
    number = 0
    sum = 0
    for i in result:
        sum = 0
        number = number + 1
        for app in i:
            if sum > 0:
                FINE_csv.write(",")
            FINE_csv.write("%s" % app)
            sum = sum + 1

        FINE_csv.write("\n")
    FINE_csv.close()


def AutoSettingSql(sql):
    if sql.find('$$date_') == -1:
        return sql
    # find datetype
    num = sql.count('$$date_')
    for i in range(num):
        left = sql.find('$$date_')
        datetype = sql[left:left + 8]
        datestring = getDateStringByKey(datetype)  # 根据数据类型，算出日期值  比如"$$date_1" --> "20131108"
        lpos = sql.find('$$date_')
        rpos = lpos + 8
        if rpos == -1:
            sql = sql[:lpos] + str(datestring)
        else:
            sql = sql[:lpos] + str(datestring) + sql[rpos:]
    return sql
    pass


def ReadConf():
    cf = ConfigParser.ConfigParser()

    cf.read("TimesvcTrader.conf")

    # return all section
    secs = cf.sections()
    print 'sections:', secs

    opts = cf.options("portal")
    print 'options:', opts

    kvs = cf.items("portal")
    print 'db:', kvs

    print "use ConfigParser() write"
    cf.set("portal", "url", "%(host)s:%(port)s")
    print cf.get("portal", "sql2")
    cf.write(open("TimesvcTrader.conf", "w"))
    pass


def setReqParam(req, cmdstring):
    cmdlist = cmdstring.strip('#').split('#')
    for item in cmdlist:
        b = item.split('=')
        if (hasattr(req, b[0])):
            if (isinstance(getattr(req, b[0]), int) == True):
                if (b[1] == ""):
                    setattr(req, b[0], 0)
                elif (b[1] == "FALSE"):
                    setattr(req, b[0], 0)
                elif (b[1] == "TRUE"):
                    setattr(req, b[0], 1)
                else:
                    setattr(req, b[0], int(b[1]))
            elif (isinstance(getattr(req, b[0]), float) == True):
                if (b[1] == ""):
                    setattr(req, b[0], 0.0)
                else:
                    setattr(req, b[0], float(b[1]))
            else:
                setattr(req, b[0], str(b[1]))
    return req


def ReadCsvfile():
    RootName = os.getcwd()
    reqrespfile = RootName + '\\reqresp.csv'
    execreportfile = RootName + '\\execreport.csv'
    csvfile = file(reqrespfile, 'rb')
    reader = csv.reader((line.replace('\0', '') for line in csvfile), delimiter=',')
    execcsvfile = file(execreportfile, 'rb')
    execreader = csv.reader((line.replace('\0', '') for line in execcsvfile), delimiter=',')

    reqresplist = []  # 从CSV文件读取的报单发送、响应数据
    execreportlist = []  # 从CSV文件读取成交响应数据
    execreportDict = {}  # 以报单编号为主键的成交响应数据字典

    # ReqrespKeyList = ['reqnum','reqid','reff','pbu','securityid','recordtimestamp','ordstatus','remark','reqtext','resptext','localrespnum']
    # ExecreportKeyList = ['execnum','bcasttype','setid','seqnum','recordtimestamp','execreporttext']
    try:
        for line in reader:
            reqresplist.append(line)
        reqresplist = reqresplist[:-1]
        reqresplist[0][0] = (str(reqresplist[0][0])).strip('\xff\xfe')
        gl.g_reqrespList = reqresplist

        for line in execreader:
            execreportlist.append(line)
        execreportlist = execreportlist[:-1]
        execreportlist[0][0] = (str(execreportlist[0][0])).strip('\xff\xfe')

        for line in execreportlist:
            rep = ((str(line[5])).strip()).split('\x01')[:-1]
            line[5] = str(line[5]).strip()
            for item in rep:
                element = item.split('=')
                if element[0] == '11':
                    execreportDict[element[1]] = line
                    break
        gl.g_execreportDict = execreportDict

    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
    csvfile.close()
    execcsvfile.close()
    pass


def CaseSendOutToXML():
    '''
    MOON将测试案例库中的案例导出    ZHLCAutoTestCases.xml
    '''
    sql = "select * from tbl_cases where disableflag = '0' order by itempos,groupname,case_index"
    ds = executeCaseSQL(sql)
    # 开始创建XML文件
    sdkpos_id = ""
    group_name = ""
    case_name = ""
    case_index = 0
    case_sum_sdkpos = 0
    case_sum_groupname = 0
    false_flag_one = "FALSE"
    false_flag_two = "FALSE"
    false_flag_three = "FALSE"

    functionList = ET.ElementTree()
    root = ET.Element("ZHLCAutoTestCases")  # 创建根目录·······第一层
    functionList._setroot(root)

    sdkposesElem = ET.Element("sdkpos")  # 创建交易根目录·······第二层
    root.append(sdkposesElem)

    for sql_result in ds:
        if sql_result["sdkpos"] != sdkpos_id:  # 创建业务类别目录：创建一个目录写一个目录，本目录写完后创建新目录·······第三层
            if false_flag_one == "FALSE":
                false_flag_one = "SUCCESS"
            else:
                sdkposElem.set("case_sum", str(case_sum_sdkpos))
            case_sum_sdkpos = 0
            sdkpos_id = sql_result["businessid"]
            sdkposElem = ET.Element("businessid", {"business_id": sdkpos_id})
            sdkposesElem.append(sdkposElem)

        if sql_result["groupname"] != group_name:
            if false_flag_three == "FALSE":
                false_flag_three = "SUCCESS"
            else:
                groupname.set("case_sum", str(case_sum_groupname))
            case_sum_groupname = 0
            group_name = sql_result["groupname"]
            groupname = ET.Element("groupname", {"group_name": group_name})
            sdkposElem.append(groupname)

        if sql_result["casename"] != case_name:
            case_index = sql_result["case_index"]
            case_name = sql_result["casename"]
            casename = ET.Element("casename", {"case_name": case_name, "case_index": str(case_index)})
            groupname.append(casename)

        cmdstring = []

        case_sum_sdkpos = case_sum_sdkpos + 1
        case_sum_groupname = case_sum_groupname + 1

    sdkposElem.set("case_sum", str(case_sum_sdkpos))
    groupname.set("case_sum", str(case_sum_groupname))

    rough_string = ET.tostring(root, 'UTF-8')
    reparsed = minidom.parseString(rough_string)
    a = reparsed.toprettyxml(indent='    ', newl='\r\n')  # 采用windows的换行符
    codecs.open("ZHLCAutoTestCases.xml", "w", "utf-8-sig").write(a)  # 使用UTF-8方式写入


def createRspTable(filename):
    f = open(filename, "r")
    w = open("create_rsp_table.sql", "w")
    firstColumnFlag = 1
    create_table_string = ""
    for line in f:
        line = line.strip('\n')

        if line.find('struct ') >= 0:
            if (len(create_table_string) != 0):
                # 处理上一条记录
                create_table_string = create_table_string + "\n)ENGINE=InnoDB DEFAULT CHARSET=utf8;\n\n"
                w.write(create_table_string)
                create_table_string = ""
                firstColumnFlag = 1

            line = line.strip(";").strip()
            colTemp = line.split()
            tablename = colTemp[1]
            create_table_string = "DROP TABLE IF EXISTS tbl_lts_%s;\nCREATE TABLE tbl_lts_%s(\n" % (
                tablename, tablename)
            continue
        if line.find('///') >= 0:
            n = line.find('///')
            comment = line[n + 3:]
            continue
        if line.find('TSecurity') >= 0:
            line = line.strip(';').strip()
            colTemp = line.split()
            column = colTemp[1]
            if firstColumnFlag != 1:
                create_table_string = create_table_string + ",\n"
            else:
                firstColumnFlag = 0
            create_table_string = create_table_string + "\t%s char(128) CHARACTER SET utf8 COLLATE utf8_unicode_ci DEFAULT NULL COMMENT '%s'" % (
                column, comment)
    create_table_string = create_table_string + "\n)ENGINE=InnoDB DEFAULT CHARSET=utf8;\n\n"
    w.write(create_table_string)
    w.close()
    f.close()


def ExecuteOneTestCase(sdk, funcid, cmdstring):
    '''
    命令行的发送,此时的cmdstring为已经替换过参数了

    返回值存在以下几种可能性：

    返回 小于0，判定案例失败
    返回 0，案例执行完毕，需要进一步判定（通过获取返回结果与期望结果来判定）
    返回 1，业务操作失败，需要进一步判定 （通过获取返回结果与期望结果判定）
    返回 2，获取错误信息，判定失败
    返回 3，业务解包失败，判定失败
    '''
    try:
        # readSystemConfigFile("1428908551835S0NXP")
        DefaultTimeOut = gl.g_ZHLCSystemSetting["DefaultTimeOut"]
        print funcid
        ret, msg = sendT2Biz(sdk, funcid, cmdstring)
        try:
            msg = msg.decode("utf-8")
        except Exception as e:
            try:
                msg = msg.decode("gbk")
            except Exception as e:
                msg = msg
        if ret < 0:
            return ret, msg
        timeout = getFuncidTimeout(funcid)
        if timeout == 0:
            ret = sdk.RecvBiz(ret, int(DefaultTimeOut))  # 如果没设置，缺省值使用6000
        else:
            ret = sdk.RecvBiz(ret, timeout)
        print "Recvbiz return %d" % int(ret)
        if ret < 0:
            return -1, sdk.GetErrorMsg(ret)

        if ret == 1:
            return 1, u"业务操作失败"
        elif ret == 2:
            return 2, sdk.GetLastExtraErrorMsg().decode("gbk")
        elif ret == 3:
            return 3, u"业务解包失败"
        else:
            return 0, "OK"
    except:
        exc_info = getExceptionInfo()
        Logging.getLog.error(exc_info)
        return -1, exc_info


def GetReturnRows(sdk):
    '''
    读取案例执行后的返回结果
    '''
    count = sdk.GetDatasetCount()
    dataset_list = []
    for i in range(count):
        row_lists = getRecvBizPacket(sdk, i)
        dataset_list.append(row_lists)
    return dataset_list


def CreateNewCaseRunRootDir():
    root_dir = os.getcwd()
    root_dir = root_dir + '\\CaseRunData'
    new_path = os.path.join(root_dir, str(time.strftime('%Y_%m_%d %H-%M-%S', time.localtime(time.time()))))
    os.makedirs(new_path)
    return new_path


def CreateOneCaseRunDir(rootpath, sno, sdkpos, funcid, groupname, case_index):
    dir_name = "%s_%s_%s_%s" % (sno, sdkpos, groupname, case_index)
    path = os.path.join(rootpath, dir_name)
    os.makedirs(path)
    return path


def SaveReturnRows(funcid, path, dataset_list, flag=0):
    """
    将功能号调用的返回值进行保存

    Args:
        funcid: 功能号
        path: 案例运行时创建的目录路径
        dataset_list: 返回值数据列表
        flag: 返回值标识
            0 代表为正常的功能调用返回（中间件返回为0）
            1 代表为数据操作失败的功能调用返回（中间件返回为1）
    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息
    """
    try:
        orig_return_file_path = os.path.join(path, "orig")
        reduce_return_file_path = os.path.join(path, "reduce")

        if os.path.exists(orig_return_file_path) == False:
            os.makedirs(orig_return_file_path)

        if os.path.exists(reduce_return_file_path) == False:
            os.makedirs(reduce_return_file_path)

        # return_file_path = os.path.join(orig_return_file_path,"return.csv")
        for i, row_lists in enumerate(dataset_list):
            return_file_path = os.path.join(orig_return_file_path, "%s_return_%d.csv" % (funcid, i))
            writer = csv.writer(file(return_file_path, 'wb'), quoting=csv.QUOTE_ALL)
            writer.writerows(row_lists)

        # return_file_path = os.path.join(reduce_return_file_path,"return.csv")
        if flag == 1:
            for i, row_lists in enumerate(dataset_list):
                return_file_path = os.path.join(reduce_return_file_path, "%s_return_%d.csv" % (funcid, i))
                writer = csv.writer(file(return_file_path, 'wb'), quoting=csv.QUOTE_ALL)
                writer.writerows(row_lists)
        else:
            ret, field_dict = GetReturnEnableInfo(funcid)
            if ret < 0:
                print field_dict
                return ret, field_dict
            for i, row_lists in enumerate(dataset_list):
                return_file_path = os.path.join(reduce_return_file_path, "%s_return_%d.csv" % (funcid, i))
                writer = csv.writer(file(return_file_path, 'wb'), quoting=csv.QUOTE_ALL)

                orig_field_dict = OrderedDict()
                for j, row in enumerate(row_lists):
                    if j == 0:
                        for k, field in enumerate(row):
                            orig_field_dict[k] = field

                    reduce_row = []

                    for k, field in enumerate(row):
                        if field_dict.get(orig_field_dict[k], '0') == '0':
                            reduce_row.append(row[k])
                    writer.writerow(reduce_row)
        return 0, ""
    except Exception, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1, "%s" % (ex)


def SaveAllRunDbTables(path):
    for servertype in g_connectSqlServerSetting.keys():
        SaveRunDbTables(path, servertype)


def SaveRunDbTables(path, servertype="p"):
    '''
    保存变化的表信息
    '''
    try:
        ret, conn = ConnectToRunDB(servertype)
        if ret == -1:
            err = '核心%s,数据库连接异常原因：%s' % (str(servertype), conn)
            Logging.getLog().critical(err)
            return -1, err
        cursor = conn.cursor()
        sql = "select username, tablename,operation,row_id from run..TMP_AUTOTEST_TRIGGER"
        cursor.execute(sql)
        ds = cursor.fetchall()
        for rec in ds:
            # 获取需要备份字段名
            if gl.g_JZJY_column_Tree == None:
                gl.g_JZJY_column_Tree = ET.parse("JZJY_column_report.xml")

            root = gl.g_JZJY_column_Tree.getroot()
            xpath = "schema[@core='%s']/table[@tablename='%s']" % (servertype, rec[1])
            table_elem = root.find(xpath)
            if table_elem == None or table_elem.attrib["disableflag"] == '1':
                continue

            field_list = []
            field_list_removed = []  # 被disable的索引

            for i, item in enumerate(table_elem):
                field_list.append(item.attrib["column_name"])
                if item.attrib["disableflag"] == '1':
                    field_list_removed.append(i)
            if len(field_list) == 0:
                continue

            if rec[2] == 1 or rec[2] == 2:  # 更新或新增
                sql = "select %s from %s..AUTO_I_%s" % (','.join(field_list), rec[0], rec[1])
            else:  # 删除
                sql = "select %s from %s..AUTO_D_%s" % (','.join(field_list), rec[0], rec[1])
            cursor.execute(sql)
            ds1 = cursor.fetchall()

            # 部分删除触发器出发时，deleted表内容为空，此时可以不用处理。fixme
            if len(ds1) <= 0:
                continue

            orig_path = os.path.join(path, "orig")
            if os.path.exists(orig_path) == False:
                os.makedirs(orig_path)

            reduce_path = os.path.join(path, "reduce")
            if os.path.exists(reduce_path) == False:
                os.makedirs(reduce_path)

            if rec[2] == 1 or rec[2] == 2:  # 更新或新增
                orig_file_path = os.path.join(orig_path, "%s.%s.%s.csv" % (rec[0], str(servertype), rec[1]))
                reduce_file_path = os.path.join(reduce_path, "%s.%s.%s.csv" % (rec[0], str(servertype), rec[1]))
            else:  # 删除  删除的数据文件采用文件名前加一个"_" 已做区分
                orig_file_path = os.path.join(orig_path, "_%s.%s.%s.csv" % (rec[0], str(servertype), rec[1]))
                reduce_file_path = os.path.join(reduce_path, "_%s.%s.%s.csv" % (rec[0], str(servertype), rec[1]))

            # 处理orig 数据
            if os.path.exists(orig_file_path):
                orig_writer = csv.writer(file(orig_file_path, "ab+"), quoting=csv.QUOTE_ALL)
            else:
                orig_writer = csv.writer(file(orig_file_path, "wb"), quoting=csv.QUOTE_ALL)
                orig_writer.writerow(field_list)
            # 处理reduce数据
            field_list_removed.sort(reverse=True)
            for i in field_list_removed:
                field_list.pop(i)

            if len(field_list) > 0:  # 还有数据
                if os.path.exists(reduce_file_path):
                    reduce_writer = csv.writer(file(reduce_file_path, "ab+"), quoting=csv.QUOTE_ALL)
                else:
                    reduce_writer = csv.writer(file(reduce_file_path, "wb"), quoting=csv.QUOTE_ALL)
                    reduce_writer.writerow(field_list)

            for rec1 in ds1:
                orig_writer.writerow(rec1)
                if len(field_list) > 0:  # 还有数据
                    list1 = list(rec1)
                    for i in field_list_removed:
                        list1.pop(i)
                    reduce_writer.writerow(list1)

            if rec[2] == 3:  # 删除
                sql = "truncate table %s..AUTO_D_%s" % (rec[0], rec[1])
                Logging.getLog().debug(sql)
                print sql
                cursor.execute(sql)
            else:
                sql = "truncate table %s..AUTO_I_%s" % (rec[0], rec[1])
                Logging.getLog().debug(sql)
                print sql
                cursor.execute(sql)

        sql = "truncate table run..TMP_AUTOTEST_TRIGGER"
        cursor.execute(sql)
        cursor.close()
        conn.commit()
        conn.close()
    except Exception, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
    pass


def SaveDbTables(path, asy_flag=0):
    '''
    保存变化的表信息
    '''
    for server in gl.g_connectOracleSetting.keys():
        UserName = gl.g_connectOracleSetting[server]['UserName']
        try:
            conn = ConnectOracleDB(server)
            cursor = conn.cursor()
            sql = "select username, tablename,operation,row_id from auto_%s.tmp_autotest_trigger" % (UserName)
            cursor.execute(sql)
            ds = cursor.fetchall()
            if ds == []:
                continue
            for rec in ds:
                # 获取需要备份字段名
                root = ET.parse("ZHLC_column_report.xml").getroot()
                xpath = "schema[@schemaname='%s']/table[@tablename='%s']" % (rec[0], rec[1])
                table_elem = root.find(xpath)
                if table_elem == None:
                    continue

                field_list = []
                field_list_removed = []  # 被disable的索引

                for i, item in enumerate(table_elem):
                    field_list.append(item.attrib["column_name"])
                    if item.attrib["disableflag"] == '1':
                        field_list_removed.append(i)
                if len(field_list) == 0:
                    continue

                if rec[2] == 1:  # 更新或新增
                    sql = "select %s from %s.%s where rowid = '%s'" % (','.join(field_list), rec[0], rec[1], rec[3])
                else:  # 删除
                    sql = "select %s from auto_%s.c_%s" % (','.join(field_list), rec[0], rec[1])
                cursor.execute(sql)
                ds1 = cursor.fetchall()

                orig_path = os.path.join(path, "orig")
                if os.path.exists(orig_path) == False:
                    os.makedirs(orig_path)

                reduce_path = os.path.join(path, "reduce")
                if os.path.exists(reduce_path) == False:
                    os.makedirs(reduce_path)

                if rec[2] == 1:  # 更新或新增
                    if asy_flag == 0:
                        orig_file_path = os.path.join(orig_path, "%s.%s.csv" % (rec[0], rec[1]))
                        reduce_file_path = os.path.join(reduce_path, "%s.%s.csv" % (rec[0], rec[1]))
                    else:
                        orig_file_path = os.path.join(orig_path, "a_%s.%s.csv" % (rec[0], rec[1]))
                        reduce_file_path = os.path.join(reduce_path, "a_%s.%s.csv" % (rec[0], rec[1]))
                else:  # 删除  删除的数据文件采用文件名前加一个"_" 已做区分
                    if asy_flag == 0:
                        orig_file_path = os.path.join(orig_path, "_%s.%s.csv" % (rec[0], rec[1]))
                        reduce_file_path = os.path.join(reduce_path, "_%s.%s.csv" % (rec[0], rec[1]))
                    else:
                        orig_file_path = os.path.join(orig_path, "_a_%s.%s.csv" % (rec[0], rec[1]))
                        reduce_file_path = os.path.join(reduce_path, "_a_%s.%s.csv" % (rec[0], rec[1]))

                # 处理orig 数据
                if os.path.exists(orig_file_path):
                    orig_writer = csv.writer(file(orig_file_path, "ab+"), quoting=csv.QUOTE_ALL)
                else:
                    orig_writer = csv.writer(file(orig_file_path, "wb"), quoting=csv.QUOTE_ALL)
                    orig_writer.writerow(field_list)
                # 处理reduce数据
                field_list_removed.sort(reverse=True)
                for i in field_list_removed:
                    field_list.pop(i)

                if len(field_list) > 0:  # 还有数据
                    if os.path.exists(reduce_file_path):
                        reduce_writer = csv.writer(file(reduce_file_path, "ab+"), quoting=csv.QUOTE_ALL)
                    else:
                        reduce_writer = csv.writer(file(reduce_file_path, "wb"), quoting=csv.QUOTE_ALL)
                        reduce_writer.writerow(field_list)

                for rec1 in ds1:
                    orig_writer.writerow(rec1)
                    if len(field_list) > 0:  # 还有数据
                        list1 = list(rec1)
                        for i in field_list_removed:
                            list1.pop(i)
                        reduce_writer.writerow(list1)

                if rec[2] == 2:  # 删除
                    sql = "truncate table auto_%s.c_%s" % (rec[0], rec[1])
                    cursor.execute(sql)

            sql = "truncate table auto_%s.tmp_autotest_trigger" % (UserName)
            cursor.execute(sql)
            cursor.close()
            conn.commit()
            conn.close()
            print "end of SaveDbTables"
            return 0, ""

        except cx_Oracle.DatabaseError, ex:
            err = u"数据库执行%s出现异常,%s" % (sql, ex[0].message.decode('gbk'))
            Logging.getLog().critical(err)
            cursor.close()
            conn.rollback()
            conn.close()
            return -1, err
        except BaseException, ex:
            exc_info = getExceptionInfo()
            Logging.getLog().critical(exc_info)
            return -1, "%s" % (exc_info)
    pass


def GetRunEnableColumnInfo(schema, server, table):
    """
    获取集中交易数据库的字段enable信息

    Args:
        schema: 用户名
        server: 核心号
        table:表名
    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息。若成功，则返回为{字段名:disableflag}的有序字典
    """
    try:
        if gl.g_JZJY_column_Tree == None:
            gl.g_JZJY_column_Tree = ET.parse("JZJY_column_report.xml")
        root = gl.g_JZJY_column_Tree.getroot()

        xpath = "schema[@core='%s']/table[@tablename='%s']" % (server, table)
        table_elem = root.find(xpath)
        if table_elem == None:
            return -1, u"未找到指定的表"

        field_dict = OrderedDict()

        for i, item in enumerate(table_elem):
            field_dict[item.attrib["column_name"]] = item.attrib["disableflag"]
        return 0, field_dict

    except Exception, ex:
        exc_info = getExceptionInfo()
        msg = u"解析xml文件出现异常 提示信息为:%s" % (exc_info)
        Logging.getLog().critical(msg)
        return -1, msg


def GetEnableColumnInfo(schema, table):
    """
    获取综合理财数据库的字段enable信息

    Args:
        schema: 用户名
        table:表名
    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息。若成功，则返回为{字段名:disableflag}的有序字典
    """
    try:
        if gl.g_ZHLC_column_Tree == None:
            gl.g_ZHLC_column_Tree = ET.parse("ZHLC_column_report.xml")

        root = gl.g_ZHLC_column_Tree.getroot()

        xpath = "schema[@schemaname='%s']/table[@tablename='%s']" % (schema, table)
        table_elem = root.find(xpath)
        if table_elem == None:
            return -1, u"未找到指定的表"

        field_dict = OrderedDict()

        for i, item in enumerate(table_elem):
            field_dict[item.attrib["column_name"]] = item.attrib["disableflag"]
        return 0, field_dict

    except Exception, ex:
        exc_info = getExceptionInfo()
        msg = u"解析xml文件出现异常 提示信息为:%s" % (exc_info)
        Logging.getLog().critical(msg)
        return -1, msg


def GetReturnEnableInfo(funcid):
    """
    获取综合理财功能号返回字段enable信息
    Args:
        funcid: 外围功能号
    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息。若成功，则返回为{字段名:disableflag}的字典
    """
    try:
        if gl.g_interfaceDetailTree == None:
            gl.g_interfaceDetailTree = ET.parse("T2InterfaceDetails.xml")

        root = gl.g_interfaceDetailTree.getroot()

        # xpath = "Interface[@funcid='%s']/OutParams/out" %funcid

        # 考虑到文档中有重复的问题，直接考虑获取第一个
        xpath = "Interface[@funcid='%s']" % funcid
        interface_elem = root.find(xpath)
        if interface_elem == None:
            print u"未找到%s功能输出信息" % str(funcid)
            return -1, u"未找到%s功能输出信息" % str(funcid)
        xpath = "OutParams/out"
        out_elem_list = interface_elem.findall(xpath)

        # out_elem_list = root.findall(xpath)
        if len(out_elem_list) <= 0:
            # 没有输出信息
            return 0, OrderedDict()

        field_dict = OrderedDict()
        for item in out_elem_list:
            field_dict[item.attrib["name"]] = item.attrib["disableflag"]
            # field_list.append((item.attrib["name"],item.attrib["disableflag"]))
        return 0, field_dict

    except Exception, ex:
        exc_info = getExceptionInfo()
        msg = u"解析xml文件出现异常 提示信息为:%s" % (exc_info)
        Logging.getLog().critical(msg)
        return -1, msg


def GetComparedDir(left_path, right_path):
    '''
    根据比较的2个路径，返回比较结果的存放路径
    '''
    left_running_dir = os.path.basename(left_path)
    right_running_dir = os.path.basename(right_path)

    root = os.getcwd()
    compared_result_dir = os.path.join(root, "Compared\\%s VS %s" % (left_running_dir, right_running_dir))

    return compared_result_dir, left_running_dir, right_running_dir


def GetCaseRunDataDir(left_path, right_path):
    '''
    根据比较的2个路径，返回比较结果的存放路径
    '''
    left_running_dir = os.path.basename(left_path)
    right_running_dir = os.path.basename(right_path)

    root = os.getcwd()
    compared_result_dir = os.path.join(root, "CaseRunData")

    return compared_result_dir, left_running_dir, right_running_dir


def ProcessOrigCSVToReduce(srcPath, dstPath, file, field_dict, type):
    """
    fixme 考虑升级后字段不一致的问题 xiaorz
    Args:
        srcPath: 源文件路径
        dstPath： 目标文件路径
        file：需要处理的文件名
        field_dict： 字段disable信息字典
        type:
            0   综合理财数据库表
            1   综合理财外围功能返回值
            2   集中交易数据库表
        funcid: 外围功能号
    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息。
    """
    try:
        src = os.path.join(srcPath, file)
        reader = csv.reader(open(src, 'rb'))
        dst = os.path.join(dstPath, file)
        writer = csv.writer(open(dst, 'wb'), quoting=csv.QUOTE_ALL)

        copy_direct_flag = False

        orig_field_dict = OrderedDict()
        for j, line in enumerate(reader):
            if j == 0:
                if len(line) != len(field_dict):
                    print u"%s\\%s 字段与设置不匹配" % (srcPath, file)
                if len(line) > 1 and line[1] == 'error_no':  # 以这种方法来判断是否是数据操作失败的状态
                    copy_direct_flag = True  # 数据操作失败状态，直接拷贝过去即可
                else:
                    # 记录实际的输出字段情况（因为可能和文档中描述的不一致）
                    for k, field in enumerate(line):
                        orig_field_dict[k] = field

            if copy_direct_flag == True:
                writer.writerow(line)
            else:
                row = []
                for i, field in enumerate(line):
                    if field_dict.get(orig_field_dict[i], '0') == '0':
                        row.append(field)
                writer.writerow(row)
    except Exception, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1, "异常发生,提示：%s" % ex
    return 0, ""
    # reader = csdv


def FlushReduceData(left_dir, right_dir):
    '''
    从orig目录重新按照ZHLC_column_report.xml 和 JZJY_column_report.xml的配置导出reduce
    '''
    try:
        # 考虑到配置文件可能在程序启动后进行修改，此处强制重新读取一次
        gl.g_interfaceDetailTree = ET.parse("T2InterfaceDetails.xml")
        gl.g_JZJY_column_Tree = ET.parse("JZJY_column_report.xml")
        gl.g_ZHLC_column_Tree = ET.parse("ZHLC_column_report.xml")

        for dir in [left_dir, right_dir]:
            root = os.getcwd()

            # 列出案例目录
            case_dir_list = os.listdir(dir)

            # 处理每一个案例
            for case_dir in case_dir_list:
                # 如果存在reduce目录就删除
                case_dir = case_dir.decode('gbk')
                reduce_dir = os.path.join(dir, "%s\\reduce" % case_dir)
                if os.path.exists(reduce_dir):
                    shutil.rmtree(reduce_dir)
                os.mkdir(reduce_dir)

                orig_dir = os.path.join(dir, "%s\\orig" % case_dir)
                print orig_dir
                if os.path.exists(orig_dir):
                    table_list = os.listdir(orig_dir)
                    for table in table_list:
                        if table.find("return") >= 0:  # 处理返回结果集


                            # 获取功能号

                            # workaround 考虑到部分sdkpos带有"_"
                            return_file_name = os.path.basename(table)
                            funcid = return_file_name.split('_')[0]

                            ret, field_dict = GetReturnEnableInfo(funcid)
                            if ret < 0:
                                shutil.copyfile(os.path.join(orig_dir, table), os.path.join(reduce_dir, table))
                                continue
                                # return ret,field_dict
                            ProcessOrigCSVToReduce(orig_dir, reduce_dir, table, field_dict, 1)
                            # shutil.copyfile(os.path.join(orig_dir,table),os.path.join(reduce_dir,table))
                            continue
                        # 解析schema和tablename
                        schema = table[:table.find(".")]
                        if schema.startswith("_"):
                            schema = schema[1:]

                        if schema == 'run':
                            server = table.split('.')[1]
                            tablename = table.split('.')[2]
                            ret, field_dict = GetRunEnableColumnInfo(schema, server, tablename)
                        else:
                            tablename = table[table.find(".") + 1:table.rfind(".")]
                            ret, field_dict = GetEnableColumnInfo(schema, tablename)

                        if ret < 0:
                            return ret, field_dict
                        ProcessOrigCSVToReduce(orig_dir, reduce_dir, table, field_dict, 0)

    except Exception, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1, exc_info
    return 0, ""


def CopyDiffFilesToDir(compared_dir, line):
    """
    将有差异的文件拷贝到Compared的目录下，便于查看

    Args:
        compared_dir: 保存对比结果的绝对路径，比如 %CURRENT_DIRECTORY%\Compared\YYYY_MM_DD HH-MM-SS VS YYYY_MM_DD HH-MM-SS\
        line: diff 返回的一行表示差异的字符串

    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息
    """

    compared_case_dir = os.path.basename(compared_dir)
    left_case_dir = compared_case_dir[:compared_case_dir.find(" VS")]
    right_case_dir = compared_case_dir[compared_case_dir.find("VS") + 3:]

    left_path = os.path.join(compared_dir, left_case_dir)
    right_path = os.path.join(compared_dir, right_case_dir)

    if os.path.exists(left_path) == False:
        os.mkdir(left_path)
    if os.path.exists(right_path) == False:
        os.mkdir(right_path)

    if line[:5] == "Files":

        left_full_path = line[6:line.find(" and")]
        right_full_path = line[line.find(" and ") + 5:line.find(" diff")]

        left_compared_dir = os.path.join(left_path, os.path.basename(os.path.dirname(os.path.dirname(left_full_path))))
        right_compared_dir = os.path.join(right_path,
                                          os.path.basename(os.path.dirname(os.path.dirname(right_full_path))))

        if os.path.exists(left_compared_dir) == False:
            os.mkdir(left_compared_dir)
        if os.path.exists(right_compared_dir) == False:
            os.mkdir(right_compared_dir)

        shutil.copy(left_full_path, left_compared_dir)
        shutil.copy(right_full_path, right_compared_dir)

    if line[:4] == "Only":
        full_path = line[8:]
        if full_path != '':
            full_path = full_path.replace(': ', '/')

        if line.find(left_case_dir) > 0:
            # 仅左边有，负责到左边文件夹下
            compare_dir = os.path.join(left_path, os.path.basename(os.path.dirname(os.path.dirname(full_path))))
        else:
            # 仅右边有，负责到右边文件夹下
            compare_dir = os.path.join(right_path, os.path.basename(os.path.dirname(os.path.dirname(full_path))))

        if os.path.exists(compare_dir) == False:
            os.mkdir(compare_dir)

        shutil.copy(full_path, compare_dir)
    return 0, ""


# def CasesDataCompare(left_path, right_path, force=False):
#     """
#     比较两个目录文件内容，本函数用于整个案例运行结果之间的比较
#     Args:
#         left_path: 左侧目录 D:\01_projects\ZHLCAutoTest\CaseRunData\2015_10_29 18-52-02
#         right_path: 右侧目录 D:\01_projects\ZHLCAutoTest\CaseRunData\2015_10_29 20-26-08
#         force: 是否强制重新比较。考虑到JZJY_column_report.xml 和 ZHLC_column_report.xml可能会经常发生变化，设置为True时。系统将从orig目录从重新按照新的配置导出reduce，再进行比较。
#     Returns:
#         一个元组（只有左侧有文件数，不同文件文件数，只有右侧有文件数）：
#     """
#     left_cases_dir = os.path.basename(left_path)
#     right_cases_dir = os.path.basename(right_path)

#     compared_cases_dir = ""
#     if left_cases_dir > right_cases_dir:
#         compared_cases_dir = "%s VS %s" % (left_cases_dir, right_cases_dir)
#     else:
#         compared_cases_dir = "%s VS %s" % (right_cases_dir, left_cases_dir)
#         left_cases_dir, right_cases_dir = right_cases_dir, left_cases_dir
#         left_path, right_path = right_path, left_path

# 比较结果的保存路径
#     compare_dir_root = os.path.join(os.getcwd(), "Compared")
#     if not os.path.exists(compare_dir_root):
#         os.mkdir(compare_dir_root)

#     compared_dir = os.path.join(os.getcwd(), "Compared\\%s" % (compared_cases_dir))

# 清空比较结果的保存路径
#     if os.path.exists(compared_dir):
#         shutil.rmtree(compared_dir)
#     os.mkdir(compared_dir)

#     if force == True:
#         ret, msg = FlushReduceData(left_path, right_path)
#         if ret < 0:
#             gl.threadQueue.put((-1, msg, "CasesDataCompare"))
#             return ret, msg, -1

#     cmd_str = "diff -arq -x 'orig' '%s' '%s' " % (left_path, right_path)
#     resultList = []
#     try:
#         si = subprocess.STARTUPINFO
#         si.wShowWindow = subprocess.SW_HIDE
#         proc = subprocess.Popen(cmd_str, shell=True, startupinfo=si, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
#                                 stdin=subprocess.PIPE)
#     except BaseException, ex:
#         exc_info = getExceptionInfo()
#         Logging.getLog().critical("启动diff.exe出现异常，请检查后重新启动程序,提示信息为:%s" % (exc_info))
#         gl.threadQueue.put((-1, exc_info, "CasesDataCompare"))
#         return -1, -1, -1

# lines = proc.stdout.readlines()
# print lines
# only_left = 0
# diff_count = 0
# only_right = 0
# for line in lines:
#    if line[:4] == 'File':
#        diff_count = diff_count + 1
#        CopyDiffFilesToDir(compared_dir,line)
#    elif line[:4] == 'Only':
#        if line.decode('utf8').find(left_cases_dir)>0:
#            only_left = only_left+1
#        else:
#            only_right = only_right + 1
# print "only_left=%d,diff_count=%d,only_right=%d" %(only_left,diff_count,only_right)
# return only_left, diff_count, only_right

#     result = []

#     lines = proc.stdout.readlines()
# print lines
#     only_left = 0
#     diff_count = 0
#     only_right = 0
#     for line in lines:
#         line = line.decode('utf8').strip('\n')

#         if line[:4] == 'File':
#             diff_count = diff_count + 1
#             CopyDiffFilesToDir(compared_dir, line)
#             ret, left_group_dir, right_group_dir, left_filename, right_filename = dir_process.GetLeftandRightDirectoryInfo(
#                 line)
#             if ret == 0:
#                 result.append(
#                     (left_cases_dir, right_cases_dir, left_group_dir, right_group_dir, left_filename, right_filename))
#         elif line[:4] == 'Only':
#             if line.find('.csv') >= 0:  # 只有一侧有不同的文件
#                 if line.find(left_cases_dir) > 0:
#                     only_left = only_left + 1
#                     CopyDiffFilesToDir(compared_dir, line)

# 解析相应目录和文件信息
#                     Logging.getLog().debug(line)
#                     print line
#                     left_group_dir = line[line.find('/') + 1:line.find('/reduce')]
#                     print left_group_dir
#                     Logging.getLog().debug(left_group_dir)

#                     left_filename = line[line.find('/reduce: ') + len('/reduce: '):]
#                     result.append((left_cases_dir, right_cases_dir, left_group_dir, "", left_filename, ""))
#                 else:
#                     only_right = only_right + 1
#                     CopyDiffFilesToDir(compared_dir, line)

# 解析相应目录和文件信息
#                     Logging.getLog().debug(line)
#                     right_group_dir = line[line.find('/') + 1:line.find('/reduce')]
#                     Logging.getLog().debug(right_group_dir)
#                     right_filename = line[line.find('/reduce: ') + len('/reduce: '):]
#                     result.append((left_cases_dir, right_cases_dir, "", right_group_dir, "", right_filename))
#             else:  # 只有一侧有不同的目录
#                 '''
#                 需要将该目录下的所有csv文件全部进行记录
#                 '''
#                 if line.find(left_cases_dir) > 0:
#                     left_group_dir = line[line.rfind(":") + 1:]
#                     left_group_dir = left_group_dir.strip()

#                     full_path = r"%s\%s\reduce" % (left_path, left_group_dir)
#                     ret, csv_list = dir_process.GetCSVFilesListInPath(full_path)
#                     if ret == 0:
#                         only_left = only_left + len(csv_list)
#                         for file in csv_list:
#                             result.append((left_cases_dir, right_cases_dir, left_group_dir, "", file, ""))
#                 else:
#                     right_group_dir = line[line.rfind(":") + 1:]
#                     right_group_dir = right_group_dir.strip()

#                     full_path = r"%s\%s\reduce" % (right_path, right_group_dir)
#                     ret, csv_list = dir_process.GetCSVFilesListInPath(full_path)
#                     if ret == 0:
#                         only_right = only_right + len(csv_list)
#                         for file in csv_list:
#                             result.append((left_cases_dir, right_cases_dir, "", right_group_dir, "", file))
#                 pass

#     print "only_left=%d,diff_count=%d,only_right=%d" % (only_left, diff_count, only_right)

#     try:
#         conn = ConnectCaseDB()
#         crs = conn.cursor()
#         querySql = "select left_dir,right_dir,left_group_dir,right_group_dir,left_filename,right_filename,ignoreflag from tbl_compare_result_setting where left_dir='%s' and right_dir='%s' and ignoreflag = '1'" % (
#             left_cases_dir, right_cases_dir)
#         queryDs = executeCaseSQL(querySql)

#         snoArray = []
#         for item in queryDs:
#             snoMap = {}
#             snoMap['left_dir'] = item['left_dir']
#             snoMap['right_dir'] = item['right_dir']
#             snoMap['left_group_dir'] = item['left_group_dir']
#             snoMap['right_group_dir'] = item['right_group_dir']
#             snoMap['left_filename'] = item['left_filename']
#             snoMap['right_filename'] = item['right_filename']
#             snoMap['ignoreflag'] = item['ignoreflag']

#             snoArray.append(snoMap)

#         sql = "delete from tbl_compare_result_setting where left_dir=%s and right_dir=%s"
#         crs.execute(sql, (left_cases_dir, right_cases_dir))
# 数据对比结果保存进入数据库
#         insert_sql = "insert into tbl_compare_result_setting(left_dir,right_dir,left_group_dir,right_group_dir,left_filename,right_filename) values(%s,%s,%s,%s,%s,%s)"
#         crs.executemany(insert_sql, tuple(result))
#         crs.close()
#         conn.commit()
#         conn.close()

#         for item in snoArray:
#             updateSql = "update tbl_compare_result_setting set ignoreflag = '%s' where left_dir = '%s' and right_dir = '%s' and left_group_dir = '%s' and right_group_dir = '%s' and left_filename = '%s' and right_filename = '%s'" % (
#                 item['ignoreflag'], item['left_dir'], item['right_dir'], item['left_group_dir'],
#                 item['right_group_dir'],
#                 item['left_filename'], item['right_filename'])
#             executeCaseSQL(updateSql)
#     except:
#         exc_info = getExceptionInfo()
#         Logging.getLog().error(exc_info)
#         crs.close()
#         conn.rollback()
#         conn.close()
#         gl.threadQueue.put((-1, exc_info, "CasesDataCompare"))
#         return -1, only_left, diff_count, only_right
#     gl.threadQueue.put((0, only_left, diff_count, only_right, "CasesDataCompare"))
#     sql = "delete from tbl_compare_result where left_dir='%s' and right_dir='%s' " % (left_cases_dir, right_cases_dir)
#     executeCaseSQL(sql)
#     sql = "insert into tbl_compare_result (left_dir,right_dir,only_left_count,differ_count,only_right_count,need_fresh) value('%s','%s',%d,%d,%d,0)" % (
#         left_cases_dir, right_cases_dir, only_left, diff_count, only_right)
#     executeCaseSQL(sql)
#     return 0, only_left, diff_count, only_right


def CaseDataCompare(caes_path, left_dir):
    """
    单案例差异数据结果对比

    Args:
        caes_path: 当前运行案例新创建的目录路径，%CURRENT_DIRECTORY%\CaseRunData\YYYY_MM_DD HH-MM-SS\caseid_...
        left_dir: 需要对比的运行目录名，YYYY_MM_DD HH-MM-SS

    Returns:
        一个元组（only_left,diff_count,only_right）：
        only_left:只有左侧有文件数,如果异常，返回-1
        diff_count:不同文件文件数，如果异常，返回-1
        only_right:只有右侧有文件数，如果异常，返回-1
    """
    time1 = time.time()
    basename = os.path.basename(caes_path)
    r = caes_path.find(basename) - 1
    l = caes_path[:r].rfind("\\") + 1

    right_dir = caes_path[l:r]

    left_base_path = caes_path.replace(right_dir, left_dir)

    if left_dir < right_dir:
        left_dir, right_dir = right_dir, left_dir

    compared_dir = os.path.join(os.getcwd(), "Compared\\%s VS %s" % (left_dir, right_dir))
    if os.path.exists(compared_dir) == False:
        os.mkdir(compared_dir)

    # 比较reduce 目录
    right_path = os.path.join(caes_path, "reduce")
    left_path = os.path.join(left_base_path, "reduce")

    if left_path < right_path:
        left_path, right_path = right_path, left_path

    if not os.path.exists(left_path):
        return -1, -1, -1

    cmd_str = "diff -aq '%s' '%s' " % (left_path, right_path)
    resultList = []
    try:
        si = subprocess.STARTUPINFO
        si.wShowWindow = subprocess.SW_HIDE
        proc = subprocess.Popen(cmd_str, shell=True, startupinfo=si, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                stdin=subprocess.PIPE)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical("启动diff.exe出现异常，请检查后重新启动程序,提示信息为:%s" % (exc_info))
        return -1, -1, -1

    lines = proc.stdout.readlines()
    print lines
    only_left = 0
    diff_count = 0
    only_right = 0
    for line in lines:
        if line[:4] == 'File':
            diff_count = diff_count + 1
            CopyDiffFilesToDir(compared_dir, line)
        elif line[:4] == 'Only':
            if line.decode('utf8').find(left_path) > 0:
                only_left = only_left + 1
            else:
                only_right = only_right + 1
    print "only_left=%d,diff_count=%d,only_right=%d" % (only_left, diff_count, only_right)
    print (time.time() - time1)
    return only_left, diff_count, only_right


def GetErrorInfoFromReturnRows(dataset_list):
    try:
        Logging.getLog().debug("dataset_list:%s" % str(dataset_list))
        pathinfo = ""
        err_no = ""
        error_info = ""
        msg = ""
        try:
            pathinfo = dataset_list[0][1][0].decode('gbk')
            err_no = dataset_list[0][1][1]
            error_info = dataset_list[0][1][2].decode('gbk')
            msg = "pathinfo:%s, err_no:%s, error_info:%s" % (pathinfo, err_no, error_info)
        except Exception as e:
            try:
                pathinfo = dataset_list[0][1][0].decode('gbk')
                err_no = dataset_list[0][1][1].decode('gbk')
                error_info = dataset_list[0][1][2].decode('gbk')
                msg = "pathinfo:%s, err_no:%s, error_info:%s" % (pathinfo, err_no, error_info)
            except Exception as e:
                pass
        return msg
    except BaseException, ex:
        exc_info = getExceptionInfo()
        err = u"获取返回信息异常！"
        Logging.getLog().critical(err)
        return err


def SaveRunningResultToDB(rec, result, dataset_list, msg, cmdstring=""):
    # cmdstring = MySQLdb.escape_string(r'%s'%cmdstring)
    if dataset_list == None:
        msg = msg
    else:
        if result == 'fail':
            msg = GetErrorInfoFromReturnRows(dataset_list)
        else:
            msg = 'OK'

    return_detail = ""
    if dataset_list != None:
        return_detail = "\r\n".join(",".join(r) for r in dataset_list[0])
        # .decode('gbk')
        print "return_detail = %s" % return_detail

    if cmdstring == "":
        cmd = (rec["itempos"], rec["sdkposdesc"], rec["groupname"], rec["case_index"], rec["casename"], rec["funcid"],
               rec["funcidname"], rec["cmdstring"], result, msg.decode('gbk'), return_detail)
        sql = "insert into tbl_cases_run_result (itempos,sdkposdesc,groupname,case_index,casename,funcid,funcname,cmdstring,result,message,detail) values \
            ('%s','%s','%s','%d','%s','%s','%s','%s','%s','%s','%s')" % cmd
        # cmd=(rec["itempos"],rec["sdkposdesc"],rec["groupname"],rec["case_index"],rec["casename"],rec["funcid"],rec["funcidname"],rec["cmdstring"],result,msg.decode('gbk'),return_detail)
    else:
        cmd = (rec["itempos"], rec["sdkposdesc"], rec["groupname"], rec["case_index"], rec["casename"], rec["funcid"],
               rec["funcidname"], cmdstring, result, msg.decode('gbk'), return_detail)
        sql = "insert into tbl_cases_run_result (itempos,sdkposdesc,groupname,case_index,casename,funcid,funcname,cmdstring,result,message,detail) values \
            ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % cmd
        # cmd=(rec["itempos"],rec["sdkposdesc"],rec["groupname"],rec["case_index"],rec["casename"],rec["funcid"],rec["funcidname"],cmdstring,result,msg.decode('gbk'),return_detail)
    # print sql
    executeCaseSQL(sql)


def ClearZHLCDB():
    Logging.getLog().info(u"准备清理综合理财运行环境...")
    for server in gl.g_connectOracleSetting.keys():
        try:
            conn = ConnectOracleDB(server)
        except BaseException, ex:
            exc_info = getExceptionInfo()
            Logging.getLog().critical(exc_info)
            gl.threadQueue.put((-1, exc_info, "ClearZHLCDB"))
            return -1, exc_info
        cursor = conn.cursor()
        UserName = gl.g_connectOracleSetting[server]['UserName']
        ServiceName = gl.g_connectOracleSetting[server]['ServiceName']
        try:
            sql = "truncate table auto_%s.tmp_autotest_trigger" % (UserName)
            cursor.execute(sql)

            if gl.g_ZHLC_column_Tree == None:
                gl.g_ZHLC_column_Tree = ET.parse("ZHLC_column_report.xml")

            root = gl.g_ZHLC_column_Tree.getroot()

            for schemaElem in root:
                for tableElem in schemaElem:
                    # if self.m_threadExitFlag == True:
                    #    return -1,"threadExit";
                    if tableElem.attrib["disableflag"] == "1":
                        continue
                    SerName = schemaElem.attrib["ServiceName"]
                    if SerName == ServiceName.upper():
                        sql = "truncate table auto_%s.c_%s" % (
                            schemaElem.attrib["schemaname"], tableElem.attrib["tablename"])
                        cursor.execute(sql)
        except cx_Oracle.DatabaseError, ex:
            err = u"数据库执行%s出现异常,%s" % (sql, ex[0].message.decode('gbk'))
            Logging.getLog().critical(err)
            cursor.close()
            conn.rollback()
            conn.close()
            gl.threadQueue.put((-1, err, "ClearZHLCDB"))
            return -1, err
        except BaseException, ex:
            exc_info = getExceptionInfo()
            Logging.getLog().critical(exc_info)
            cursor.close()
            conn.rollback()
            conn.close()
            gl.threadQueue.put((-1, exc_info, "ClearZHLCDB"))
            return -1, exc_info
    cursor.close()
    conn.commit()
    conn.close()
    gl.threadQueue.put((0, "", "ClearZHLCDB"))
    Logging.getLog().info(u"清理综合理财运行环境完毕")
    return 0, ""


def ClearJZJYDB(server=0):
    Logging.getLog().info(u"准备清理集中交易运行环境...")
    # 清理SQL Server 临时表
    try:
        ret, conn = ConnectToRunDB(server)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        gl.threadQueue.put((-1, exc_info, "ClearJZJYDB", server))
        return -1, exc_info
    cursor = conn.cursor()

    try:
        sql = "truncate table run..TMP_AUTOTEST_TRIGGER"
        cursor.execute(sql)

        if gl.g_JZJY_column_Tree == None:
            gl.g_JZJY_column_Tree = ET.parse("JZJY_column_report.xml")

        root = gl.g_JZJY_column_Tree.getroot()

        for schemaElem in root:
            for tableElem in schemaElem:
                # if self.m_threadExitFlag == True:
                #    return -1,"threadExit";
                if tableElem.attrib["disableflag"] == '1':
                    continue
                sql = "truncate table run..AUTO_I_%s" % tableElem.attrib["tablename"]
                cursor.execute(sql)
                sql = "truncate table run..AUTO_D_%s" % tableElem.attrib["tablename"]
                cursor.execute(sql)

    except BaseException, ex:
        exc_info = getExceptionInfo()
        err = u"数据库执行%s出现异常,%s" % (sql, exc_info)
        Logging.getLog().critical(err)
        cursor.close()
        conn.rollback()
        conn.close()
        gl.threadQueue.put((-1, err, "ClearJZJYDB", server))
        return -1, err
    cursor.close()
    conn.commit()
    conn.close()
    Logging.getLog().info(u"清理集中交易server = %s运行环境完毕" % str(server))
    gl.threadQueue.put((0, "", "ClearJZJYDB", server))
    return 0, ""


def InitRunningEnv(self):
    # 清理Oracle 临时表
    Logging.getLog().info(u"准备清理综合理财运行环境...")
    try:
        conn = ConnectOracleDB()
    except BaseException, ex:
        exc_info = getExceptionInfo()
        err = u"连接数据库出现异常,%s" % exc_info
        Logging.getLog().critical(err)
        return -1, err
    cursor = conn.cursor()
    try:
        sql = "truncate table auto_hs_acct.tmp_autotest_trigger"
        cursor.execute(sql)

        if gl.g_ZHLC_column_Tree == None:
            gl.g_ZHLC_column_Tree = ET.parse("ZHLC_column_report.xml")

        root = gl.g_ZHLC_column_Tree.getroot()

        for schemaElem in root:
            for tableElem in schemaElem:
                # if self.m_threadExitFlag == True:
                #    return -1,"threadExit";
                if tableElem.attrib["disableflag"] == "1":
                    continue
                sql = "truncate table %s.c_%s" % (schemaElem.attrib["schemaname"], tableElem.attrib["tablename"])
                cursor.execute(sql)
    except cx_Oracle.DatabaseError, ex:
        err = u"数据库执行%s出现异常,%s" % (sql, ex[0].message.decode('gbk'))
        Logging.getLog().critical(err)
        cursor.close()
        conn.rollback()
        conn.close()
        return -1, err
    except BaseException, ex:
        exc_info = getExceptionInfo()
        err = u"数据库执行%s出现异常,%s" % (sql, exc_info)
        Logging.getLog().critical(err)
        cursor.close()
        conn.rollback()
        conn.close()
        return -1, err
    cursor.close()
    conn.commit()
    conn.close()

    Logging.getLog().info(u"准备清理集中交易运行环境...")
    # 清理SQL Server 临时表
    for server in gl.g_connectSqlServerSetting.keys():
        try:
            ret, conn = ConnectToRunDB(server)
        except BaseException, ex:
            exc_info = getExceptionInfo()
            err = u"连接SQL Server数据库出现异常，%s" % exc_info
            Logging.getLog().critical(err)
            return -1, err
        cursor = conn.cursor()

        try:
            sql = "truncate table run..TMP_AUTOTEST_TRIGGER"
            cursor.execute(sql)

            if gl.g_JZJY_column_Tree == None:
                gl.g_JZJY_column_Tree = ET.parse("JZJY_column_report.xml")

            root = gl.g_JZJY_column_Tree.getroot()

            for schemaElem in root:
                for tableElem in schemaElem:
                    # if self.m_threadExitFlag == True:
                    #    return -1,"threadExit";
                    if tableElem.attrib["disableflag"] == '1':
                        continue
                    sql = "truncate table run..I_%s" % tableElem.attrib["tablename"]
                    cursor.execute(sql)
                    sql = "truncate table run..D_%s" % tableElem.attrib["tablename"]
                    cursor.execute(sql)

        except BaseException, ex:
            exc_info = getExceptionInfo()
            err = u"数据库执行%s出现异常,%s" % (sql, exc_info)
            Logging.getLog().critical(err)
            cursor.close()
            conn.rollback()
            conn.close()
            return -1, err
        cursor.close()
        conn.commit()
        conn.close()

    return 0, "OK"

    # 创建触发器
    # ret,msg = createTrigger()
    # if ret < 0:
    #    return ret,msg

    # 创建临时表
    # ret, msg = createTempTable()
    # return ret,msg


# def RunTest(self):
# createTrigger()

#     sql = "select sno,itempos,sdkposdesc,funcidname,groupname,case_index,casename,funcid,cmdstring,expect_ret,pre_action,pro_action from tbl_cases where checkflag = 1 and disableflag = 0 order by itempos,groupname,case_index"
#     ds = executeCaseSQL(sql)

#     if len(ds) <= 0:
#         Logging.getLog().info(u"未选中案例，退出执行")
#         wx.MessageBox(u"未选中案例，退出执行", u"测试案例运行")
#         return -1, u"未选中案例，退出执行";
#     TestRuning(self)


# def TestRuning(self=0):
#     dataCompareList = []  # 记录数据有差异的数据 [案例名，left_only_count, diff_count,right_only_count]

#     left_dir = ""  # 左侧对比目录名
#     right_dir = ""  # 右侧对比目录名

#     Logging.getLog().info(u"案例执行开始")
#     thread_list = []

#     t = threading.Thread(target=ClearZHLCDB)
#     thread_list.append(t)

#     for server in gl.g_connectSqlServerSetting.keys():
#         t = threading.Thread(target=ClearJZJYDB, args=(server,))
#         thread_list.append(t)

#     while not gl.threadQueue.empty():
#         gl.threadQueue.get()

#     for t in thread_list:
#         t.setDaemon(True)
#         t.start()

#     is_threads_exit = False
#     while (is_threads_exit != True):
#         for t in thread_list:
#             t.join(0.03)
#             if t.is_alive() == False:
#                 thread_list.remove(t)
# dlg.UpdatePulse()

#             if len(thread_list) == 0:
#                 is_threads_exit = True

# dlg.Destroy()

#     result = []
#     while not gl.threadQueue.empty():
#         result.append(gl.threadQueue.get())

#     for item in result:
#         if str(item[0]) != "0":
#             return -1, u"初始化环境异常";
#     if self != 0:
#         if self.m_threadExitFlag == True:
#             return -1, "threadExit";

#     gl.g_caseCount = 0
#     gl.g_succCount = 0
#     gl.g_failCount = 0

# 启用sdk长连接
#     if (gl.g_sdk == None):
#         sdk, msg = connectT2(gl.g_ZHLCSystemSetting["T2Server"])
#         if sdk == None:
#             gl.g_sdk = None
#             wx.MessageBox(u"连接T2服务器失败，原因:%s" % (msg), u"测试案例运行")
#             Logging.getLog().critical(u"连接T2服务器失败，原因:%s" % (msg))
#             return
#         else:
#             gl.g_sdk = sdk
#     else:
#         sdk = gl.g_sdk
#     try:
#         sql = "delete from tbl_cases_run_result"
#         executeCaseSQL(sql)

#         if gl.g_dataCompareFlag == True:
#             sql = "select running_dir from tbl_running_record where compare_flag = 1 LIMIT 0,1"
#             ds = executeCaseSQL(sql)
#             if len(ds) <= 0:
#                 return -1, u"找不到对比目录"
#             left_dir = ds[0]["running_dir"]

#         sql = "select sno,itempos,sdkposdesc,funcidname,groupname,case_index,casename,funcid,cmdstring,expect_ret,pre_action,pro_action,account_group_id from tbl_cases where checkflag = 1 and disableflag = 0 order by itempos,groupname,case_index"
#         ds = executeCaseSQL(sql)
# Logging.getLog().debug(u"选中%d个案例" %len(ds))
#         run_path = CreateNewCaseRunRootDir()

#         right_dir = os.path.basename(run_path)

# 逐一分组执行
#         group_variable_dict = {}
#         group_skip_cases_flag = False
#         start_time = datetime.datetime.now()
#         for i, rec in enumerate(ds):
#             if self != 0:
#                 if self.m_threadExitFlag == True:  # 用户取消
#                     Logging.getLog().info(u"用户取消案例运行")
#                     break;
#             sno = rec["sno"]
#             itempos = rec["itempos"]
#             funcidname = rec["funcidname"]
#             groupname = rec["groupname"]
#             case_index = rec["case_index"]
#             funcid = rec["funcid"]
#             cmdstring = rec["cmdstring"]
#             cstring = {}
#             cmd_list = cmdstring.split(',')
#             for i, inpar in enumerate(cmd_list):
#                 cmdStr = cmd_list[i].split(":")
#                 field, cmdValue = cmdStr
#                 cstring[field] = cmdValue
#             xpath = "Interface[@funcid='%s']" % (funcid)
#             elem = gl.g_interfaceDetailTree.find(xpath)
#             if elem == None:
#                 Logging.getLog().critical("功能号：%s,的入参信息不存在案例未被执行" % funcid)
#                 gl.g_failCount = gl.g_failCount + 1
#                 gl.g_caseCount = gl.g_caseCount + 1
#                 SaveRunningResultToDB(rec, "fail", None, u"功能号：%s,的入参信息不存在案例未被执行" % funcid, cmdstring)
#                 group_skip_cases_flag = True
#                 continue
# in_elem_list=elem.findall("InputParams/in")
# cmdstring=""
# for i,inpar in enumerate(in_elem_list):
#    if cstring.has_key(inpar.attrib["name"]):
#        cmdstring+=str(inpar.attrib["name"]+":"+cstring[inpar.attrib["name"]])+","
#    else:
#        cmdstring+=str(inpar.attrib["name"]+":"+"")+","
# cmdstring=cmdstring[:-1]
#             casename = rec["casename"]
#             expect_ret = rec["expect_ret"]
#             pre_action = rec["pre_action"]
#             pro_action = rec["pro_action"]
#             account_group_id = rec["account_group_id"]

#             cmdstring = GlobalParameterReplace(cmdstring, str(account_group_id))

# 案例名串
#             dir_name = "%s_%s_%s_%s_%s" % (sno, itempos, funcidname, groupname, case_index)
#             Logging.getLog().info("执行案例:%s" % dir_name)

# 设置组级变量


#             if case_index == 1:
# 清空组级变量
#                 group_variable_dict = {}
#                 group_skip_cases_flag = False
#             elif group_skip_cases_flag == True:
# 跳过剩下的组案例
#                 gl.g_failCount = gl.g_failCount + 1
#                 gl.g_caseCount = gl.g_caseCount + 1
#                 SaveRunningResultToDB(rec, "fail", None, u"案例被跳过", cmdstring)
#                 message.pub(gl.FINISH_ONE_TEST_CASES)
#                 continue

#             if pre_action != "" and pre_action != None:
#                 ret, msg = action_process(pre_action, group_variable_dict, [], funcid, cmdstring, str(account_group_id))
#                 if ret != 0:
#                     gl.g_failCount = gl.g_failCount + 1
#                     gl.g_caseCount = gl.g_caseCount + 1
# err_info = u"测试用例组名：%s, 步骤编号：%s, 案例名称：%s, 功能号:%s,执行结果:  失败(--前置动作失败 %s--)\n入参: %s\n" %(groupname,case_index,casename,funcname,msg,cmdstring)
# Logging.getLog().debug(err_info)
#                     SaveRunningResultToDB(rec, "fail", None, u"前置任务失败", cmdstring)
#                     group_skip_cases_flag = True
#                     continue

#             cmdstring = GroupParameterReplace(cmdstring, group_variable_dict)
# 执行案例
#             ret, msg = ExecuteOneTestCase(sdk, funcid, cmdstring)

#             case_run_path = CreateOneCaseRunDir(run_path, sno, itempos, funcid, groupname, case_index)

# 处理异步功能号，如果是异步功能，需要等待指定时间后，再次轮询一次变化值
#             rtn, timeout = zhlc.GetASYCfunctionTimeout(funcid)
#             if timeout == "":
#                 timeout = 0
# 获取返回值
#             if ret == 0 or ret == 1:
#                 dataset_list = GetReturnRows(sdk)
# 保存返回值
#                 print u"begin SaveReturnRows"
#                 SaveReturnRows(funcid, case_run_path, dataset_list, ret)
#                 if timeout == 0:
# if funcid=='10203201':
# pass
#                     print u"begin SaveDbTables"
# 保存综合理财表变化值
#                     SaveDbTables(case_run_path)
#                     print u"begin SaveAllRunDbTables"
# 保存集中交易表变化值
#                     SaveAllRunDbTables(case_run_path)
#                     print "end SaveAllRunDbTables"
#                 if rtn == 0:
# 保存综合理财表变化值
#                     Logging.getLog().debug("等待异步返回结果，timeout = %s 秒" % timeout)
#                     time.sleep(int(timeout))
# SaveDbTables(case_run_path,1)
#                     SaveDbTables(case_run_path)
#                     print u"begin SaveAllRunDbTables"
# 保存集中交易表变化值
# SaveAllRunDbTables(case_run_path,1)
#                     SaveAllRunDbTables(case_run_path)

#                 if ret == 0:
# 这里有一个假设，当返回0的时候，不会出现error_no的字段，那么如果expect_id为非零值时，就判定案例失败
#                     if int(expect_ret) != 0:
#                         gl.g_failCount = gl.g_failCount + 1
#                         gl.g_caseCount = gl.g_caseCount + 1
#                         SaveRunningResultToDB(rec, "fail", dataset_list, u"返回和期望不一致，期望%d 返回 0" % int(expect_ret),
#                                               cmdstring)
#                         group_skip_cases_flag = True
#                     else:
#                         if pro_action != "" and pro_action != None:
#                             ret, msg = action_process(pro_action, group_variable_dict, dataset_list[0], funcid,
#                                                       cmdstring, str(account_group_id))  # fixme rzxiao
#                             if ret < 0:
#                                 gl.g_failCount = gl.g_failCount + 1
#                                 gl.g_caseCount = gl.g_caseCount + 1
#                                 SaveRunningResultToDB(rec, "fail", dataset_list, u"失败(--后续动作失败 %s--)" % msg, cmdstring)
#                                 group_skip_cases_flag = True
#                             else:
#                                 if gl.g_dataCompareFlag == True:
#                                     leftonly, diff_count, rightonly = CaseDataCompare(case_run_path, left_dir)
#                                     if leftonly == 0 and diff_count == 0 and rightonly == 0:  # fixme xiaorz
#                                         gl.g_succCount = gl.g_succCount + 1
#                                         gl.g_caseCount = gl.g_caseCount + 1
#                                         SaveRunningResultToDB(rec, "pass", dataset_list, u"成功", cmdstring)
#                                     else:
#                                         gl.g_failCount = gl.g_failCount + 1
#                                         gl.g_caseCount = gl.g_caseCount + 1
#                                         SaveRunningResultToDB(rec, "fail", dataset_list, u"数据对比失败", cmdstring)
#                                         dataCompareList.append((dir_name, leftonly, diff_count, rightonly))
# 数据对比失败，不需要跳过后续案例
#                                 else:
#                                     gl.g_succCount = gl.g_succCount + 1
#                                     gl.g_caseCount = gl.g_caseCount + 1
#                                     SaveRunningResultToDB(rec, "pass", dataset_list, u"成功", cmdstring)
#                         else:
#                             if gl.g_dataCompareFlag == True:
#                                 leftonly, diff_count, rightonly = CaseDataCompare(case_run_path, left_dir)
#                                 if leftonly == 0 and diff_count == 0 and rightonly == 0:  # fixme xiaorz
#                                     gl.g_succCount = gl.g_succCount + 1
#                                     gl.g_caseCount = gl.g_caseCount + 1
#                                     SaveRunningResultToDB(rec, "pass", dataset_list, u"成功", cmdstring)
#                                 else:
#                                     gl.g_failCount = gl.g_failCount + 1
#                                     gl.g_caseCount = gl.g_caseCount + 1
#                                     SaveRunningResultToDB(rec, "fail", dataset_list, u"数据对比失败", cmdstring)
# 数据对比失败，不需要跳过后续案例
#                                     dataCompareList.append((dir_name, leftonly, diff_count, rightonly))
#                             else:
#                                 gl.g_succCount = gl.g_succCount + 1
#                                 gl.g_caseCount = gl.g_caseCount + 1
#                                 SaveRunningResultToDB(rec, "pass", dataset_list, u"成功", cmdstring)

#                 else:  # ret == 1
#                     rows_list = dataset_list[0]
#                     error_no = ''
#                     if rows_list[0][1] == "error_no":  # 确认是error_no 字段
#                         error_no = rows_list[1][1]
#                     if error_no == '':  # 无法正确提取error_no
# 判定案例失败
#                         gl.g_failCount = gl.g_failCount + 1
#                         gl.g_caseCount = gl.g_caseCount + 1
#                         SaveRunningResultToDB(rec, "fail", dataset_list, u"业务操作失败", cmdstring)
#                         group_skip_cases_flag = True
#                         continue
# err_info = u"测试用例组名：%s, 步骤编号：%s, 案例名称：%s, 功能号:%s,执行结果:  失败(--业务操作失败--)\n入参: %s\n" %(groupname,case_index,casename,funcname,cmdstring)
# Logging.getLog().debug(err_info)

#                     else:
#                         if int(error_no) == int(expect_ret):  # 符合预期，判定成功
#                             if gl.g_dataCompareFlag == True:
#                                 leftonly, diff_count, rightonly = CaseDataCompare(case_run_path, left_dir)
#                                 if leftonly == 0 and diff_count == 0 and rightonly == 0:  # fixme xiaorz
#                                     gl.g_succCount = gl.g_succCount + 1
#                                     gl.g_caseCount = gl.g_caseCount + 1
#                                     SaveRunningResultToDB(rec, "pass", dataset_list, u"成功", cmdstring)
#                                 else:
#                                     gl.g_failCount = gl.g_failCount + 1
#                                     gl.g_caseCount = gl.g_caseCount + 1
#                                     SaveRunningResultToDB(rec, "fail", dataset_list, u"数据对比失败", cmdstring)
# 数据对比失败，不需要跳过后续案例
#                                     dataCompareList.append((dir_name, leftonly, diff_count, rightonly))
#                             else:
#                                 gl.g_succCount = gl.g_succCount + 1
#                                 gl.g_caseCount = gl.g_caseCount + 1
#                                 SaveRunningResultToDB(rec, "pass", dataset_list, u"成功", cmdstring)
#                         else:
#                             gl.g_failCount = gl.g_failCount + 1
#                             gl.g_caseCount = gl.g_caseCount + 1

# err_msg = "error_pathinfo:%s, error_no:%s, error_info:%s" %(rows_list[1][0],rows_list[1][1],rows_list[1][2].decode('gbk'))
#                             SaveRunningResultToDB(rec, "fail", dataset_list, u"业务操作失败", cmdstring)
#                             group_skip_cases_flag = True

#             else:
#                 gl.g_failCount = gl.g_failCount + 1
#                 gl.g_caseCount = gl.g_caseCount + 1

#                 SaveRunningResultToDB(rec, "fail", None, msg, cmdstring)
#                 print msg
#             message.pub(gl.FINISH_ONE_TEST_CASES)
#         end_time = datetime.datetime.now()
#         during = end_time - start_time

# 写入执行记录
#         if self != 0:
#             if self.m_threadExitFlag == True:  # 用户取消
#                 sql = "insert into tbl_running_record (running_dir,start_date,end_date,during,state,casenum,succCount,failCount) values ('%s','%s','%s','%s','%s',%d,%d,%d)" % (
#                     os.path.basename(run_path), start_time.strftime("%Y-%m-%d %H:%M:%S"),
#                     end_time.strftime("%Y-%m-%d %H:%M:%S"), str(during), u'用户取消', len(ds), gl.g_succCount,
#                     gl.g_failCount)
#             else:
#                 sql = "insert into tbl_running_record (running_dir,start_date,end_date,during,state,casenum,succCount,failCount) values ('%s','%s','%s','%s','%s',%d,%d,%d)" % (
#                     os.path.basename(run_path), start_time.strftime("%Y-%m-%d %H:%M:%S"),
#                     end_time.strftime("%Y-%m-%d %H:%M:%S"), str(during), u'全部结束', len(ds), gl.g_succCount,
#                     gl.g_failCount)
# print sql
#             executeCaseSQL(sql)
#         else:
#             sql = "insert into tbl_running_record (running_dir,start_date,end_date,during,state,casenum,succCount,failCount) values ('%s','%s','%s','%s','%s',%d,%d,%d)" % (
#                 os.path.basename(run_path), start_time.strftime("%Y-%m-%d %H:%M:%S"),
#                 end_time.strftime("%Y-%m-%d %H:%M:%S"), str(during), u'全部结束', len(ds), gl.g_succCount, gl.g_failCount)
#             executeCaseSQL(sql)
# 写入数据对比记录
#         WriteDataCompareResult(left_dir, right_dir, dataCompareList)

#     except BaseException, ex:
#         exc_info = getExceptionInfo()
#         Logging.getLog().critical(exc_info)
#         end_time = datetime.datetime.now()
#         during = end_time - start_time
#         sql = "insert into tbl_running_record (running_dir,start_date,end_date,during,state,casenum,succCount,failCount) values ('%s','%s','%s','%s','%s',%d,%d,%d)" % (
#             os.path.basename(run_path), start_time.strftime("%Y-%m-%d %H:%M:%S"),
#             end_time.strftime("%Y-%m-%d %H:%M:%S"),
#             str(during), u'异常结束', len(ds), gl.g_succCount, gl.g_failCount)
# print sql
#         executeCaseSQL(sql)

#         WriteDataCompareResult(left_dir, right_dir, dataCompareList)

#         return -1

#     ExportRunningResult(run_path)


# def WriteDataCompareResult(left_dir, right_dir, dataCompareList):
#     if len(dataCompareList) <= 0:
#         return -1, ""

#     leftonly_sum = 0
#     diff_count_sum = 0
#     rightonly_sum = 0

#     root = etree.Element("CompareResult")
#     for dir_name, leftonly, diff_count, rightonly in dataCompareList:
#         etree.SubElement(root, "result", case=dir_name, only_left_count=str(leftonly), differ_count=str(diff_count),
#                          right_only_count=str(rightonly))
#         leftonly_sum = leftonly_sum + leftonly
#         diff_count_sum = diff_count_sum + diff_count
#         rightonly_sum = rightonly_sum + rightonly

#     doc = etree.ElementTree(root)  # root 为根元素
#     xml_path = os.path.join(os.getcwd(), "Compared\\%s VS %s.xml" % (left_dir, right_dir))
#     doc.write(xml_path, pretty_print=True, xml_declaration=True, encoding='utf-8')

#     sql = "insert into tbl_compare_result(left_dir,right_dir,only_left_count,differ_count,only_right_count,need_fresh) value ('%s','%s',%d,%d,%d,0)" % (
#         left_dir, right_dir, leftonly_sum, diff_count_sum, rightonly_sum)
#     executeCaseSQL(sql)


# def ExportRunningResult(path):
#     basename = os.path.basename(path)
#     root_dir = os.getcwd() + '\\Autotest_report\\Autotest_Total_report'
#     root_dir1 = os.getcwd() + '\\Autotest_report\\Autotest_Error_report'
#     root_dir2 = os.getcwd() + '\\Autotest_report\\Autotest_Correct_report'
#     root_dir3 = os.getcwd() + '\\Autotest_report\\Autotest_Coverage_html_report'
#     JuageDirectoryExist(root_dir)
#     JuageDirectoryExist(root_dir1)
#     JuageDirectoryExist(root_dir2)
#     JuageDirectoryExist(root_dir3)

#     new_path = os.path.join(root_dir, basename + '.xml')
#     return_flag = gl.g_ZHLCSystemSetting["OutParamPrint"]
#     runningResult_to_xml(basename, return_flag)
#     runingResultBy_need_log(basename)
#     runningErrorResult_to_xml(basename)
#     runningCorrectResult_to_xml(basename)
# SqlTranslateToXML(new_path)
# exportHtml(root_dir,basename+'.html',return_flag)
#     exportCoverageHtml(root_dir3 + "\\" + basename + ".html")


def JuageDirectoryExist(path):
    '''
    判断目录是否存在，不存在就创建
    '''
    if os.path.exists(path) == False:
        os.makedirs(path)


def convBinaryToStr(array):
    '''
        将二进制数据转为为字符串
        比如: array[] = '0x11,0x22,0x33,0x44" 会转化为 "0x11223344"
        入参:
            array 应该是二进制数组对象
    '''
    strList = []
    for i in array:
        strList.append("%x" % (i))
    str = "0x%s" % (''.join(strList))
    return str


# def SelectSQLserverToXML(database_name, table_name, whereclause, primary_key_list, column_list, path):
#     if table_name == "sjshb" or table_name == "sjswt":
#         try:
#             Sd.readDbf(database_name, table_name, whereclause, primary_key_list, column_list, path)
#         except BaseException, ex:
#             exc_info = getExceptionInfo()
#             Logging.getLog().critical(exc_info)
#             return -1
#     else:
#         table_name_str = database_name + "." + table_name

#         sql_column = ",".join(column_list)
#         sql_primary_key = ",".join(primary_key_list)

#         sql = "SELECT %s FROM %s %s order by %s" % (sql_column, table_name_str, whereclause, sql_primary_key)
#         print "sql=", sql
#         try:
#             ds = executeRunSQL(sql)
#         except Exception, ex:
#             exc_info = getExceptionInfo()
#             Logging.getLog().critical(exc_info)

#             return -1
#         file_name = path + "\\" + table_name_str + ".csv"

#         FINE_csv = open(file_name, "w")
#         FINE_csv.write(sql_column)
#         FINE_csv.write("\n")
#         number = 0
#         sum = 0
#         for i in ds:
#             sum = 0
#             number = number + 1
#             for app in column_list:
#                 if sum > 0:
#                     FINE_csv.write(",")
# if app == "checkord":
#                     column_value = i["%s"%(app)]
#                     try:
#                         colmn_value = convBinaryToStr(column_value)
#                     except BaseException,ex:
#                         exc_type, exc_value,exc_traceback = sys.exc_info()
#                         traceback_details = {
#                               'filename': exc_traceback.tb_frame.f_code.co_filename,
#                               'lineno' : exc_traceback.tb_lineno,
#                               'name'  : exc_traceback.tb_frame.f_code.co_name,
#                               'type'  : exc_type.__name__,
#                               'message' : exc_value.message, # or see traceback._some_str()
#                               }
#                         Logging.getLog().critical("异常发生,提示信息为:%s" %(gl.g_traceback_template % traceback_details))
#                 else:
#                 column_value = i["%s" % (app)]
#                 column_value = str(column_value)
#                 if column_value == "None":
#                     FINE_csv.write("")
#                 else:
#                     FINE_csv.write('"')
#                     FINE_csv.write(column_value)
#                     FINE_csv.write('"')
#                 sum = sum + 1
#             FINE_csv.write("\n")
#         FINE_csv.close()
#         pass


def read_xml(in_path):
    '''读取并解析xml文件
       in_path: xml路径
       return: ElementTree'''
    tree = ET.ElementTree()
    tree.parse(in_path)
    return tree


def find_nodes(tree, path):
    '''''查找某个路径匹配的所有节点
       tree: xml树
       path: 节点路径'''
    return tree.findall(path)


def get_node_by_keyvalue(nodelist):
    '''''根据属性及属性值定位符合的节点，返回节点
       nodelist: 节点列表
       kv_map: 匹配属性及属性值map'''
    result_nodes = []

    for node in nodelist:
        result_nodes.append(node)
    return result_nodes


def create_node(tag):
    '''''新造一个节点
       tag:节点标签
       property_map:属性及属性值map
       content: 节点闭合标签里的文本内容
       return 新节点'''
    element = ET.Element(tag)
    return element


def add_child_node(nodelist, element):
    '''''给一个节点添加子节点
       nodelist: 节点列表
       element: 子节点'''
    for node in nodelist:
        node.append(element)


def write_xml(root, out_path):
    '''''将xml文件写出
       tree: xml树
       out_path: 写出路径'''
    root.write(out_path)


def create_column_node(tag, property_map):
    '''''新造一个节点
       tag:节点标签
       property_map:属性及属性值map
       content: 节点闭合标签里的文本内容
       return 新节点'''
    element = ET.Element(tag, property_map)
    return element


# def InsertPrimaryKey(user_name, database_name, table_name):
#     sql = "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE WHERE TABLE_CATALOG = '%s' and TABLE_schema = '%s' and TABLE_NAME='%s'" % (
#         user_name, database_name, table_name)
#     ds = executeRunSQL_COMPARE(sql)

# 1. 读取xml文件
#     tree = read_xml("LTS_column_report.xml")
#     path = "schemas/" + database_name + "/" + table_name
# 2. 找到父节点
#     nodes = find_nodes(tree, path)
# 3. 通过属性准确定位子节点
#     result_nodes = get_node_by_keyvalue(nodes)
# 4.新建节点
#     a = create_node("PrimaryKeys")
# 5.插入到父节点之下
#     add_child_node(result_nodes, a)

#     for PrimaryKeys in ds:
#         path = "schemas/" + database_name + "/" + table_name + "/PrimaryKeys"
# 2. 找到父节点
#         nodes = find_nodes(tree, path)
# 3. 通过属性准确定位子节点
#         result_nodes = get_node_by_keyvalue(nodes)
# 4.新建节点
#         a = create_column_node("primaryKey", {"column_name": PrimaryKeys["COLUMN_NAME"]})
# 5.插入到父节点之下
#         add_child_node(result_nodes, a)

# 6. 输出到结果文件
#     write_xml(tree, "LTS_column_report.xml")


# def InsertColumns(user_name, database_name, table_name):
#     sql = "SELECT COLUMN_NAME,DATA_TYPE FROM INFORMATION_SCHEMA.columns WHERE TABLE_CATALOG = '%s' and TABLE_schema = '%s' and TABLE_NAME='%s' ORDER BY COLUMN_NAME;" % (
#         user_name, database_name, table_name)
#     ds = executeRunSQL_COMPARE(sql)

# 1. 读取xml文件
#     tree = read_xml("LTS_column_report.xml")
# 2. 找到父节点
#     path = "schemas/" + database_name + "/" + table_name
#     nodes = find_nodes(tree, path)
# 3. 通过属性准确定位子节点
#     result_nodes = get_node_by_keyvalue(nodes)
# 4.新建节点
#     a = create_node("Columns")
# 5.插入到父节点之下
#     add_child_node(result_nodes, a)

#     for columns in ds:
#         path = "schemas/" + database_name + "/" + table_name + "/Columns"
# 2. 找到父节点
#         nodes = find_nodes(tree, path)
# 3. 通过属性准确定位子节点
#         result_nodes = get_node_by_keyvalue(nodes)
# 4.新建节点
#         a = create_column_node("Column", {"column_name": columns["COLUMN_NAME"]})
# 5.插入到父节点之下
#         add_child_node(result_nodes, a)
# 6. 输出到结果文件
#     write_xml(tree, "LTS_column_report.xml")


def Done():
    tree = ET.ElementTree()
    tree.parse("LTS_column_report.xml")
    root = tree.getroot()

    rough_string = ET.tostring(root, 'UTF-8')
    reparsed = minidom.parseString(rough_string)
    a = reparsed.toprettyxml(indent='    ', newl='\r\n')  # 采用windows的换行符
    codecs.open("LTS_column_report.xml", "w", "utf-8-sig").write(a)  # 使用UTF-8方式写入


def convBinaryToStr(array):
    '''
        将二进制数据转为为字符串
        比如: array[] = '0x11,0x22,0x33,0x44" 会转化为 "0x11223344"
        入参:
            array 应该是二进制数组对象
    '''
    strList = []
    for i in array:
        strList.append("%x" % (i))
    str = "0x%s" % (''.join(strList))
    return str


def extractInstrumentIDfromCmdstring(cmdstringList):
    '''
    从命令行列表中，解析出合约号和交易所信息

    *入参：
        cmdstringList: 命令行列表

    *返回值：
        [(instrumentID,ExchangeID),……] 合约信息列表
    '''
    Exchange = {}
    instruList = []

    for item in cmdstringList:
        left = item.find("InstrumentID=")
        if left == -1:
            continue
        right = item.find("#", left)
        instru = item[left + len("InstrumentID="):right]  # 截取合约号

        left = item.find("ExchangeID=")
        if left == -1:
            continue
        right = item.find("#", left)
        exchangeID = item[left + len("ExchangeID="):right]  # 截取交易所信息
        if instru != '' and exchangeID != '' and (instru, exchangeID) not in instruList:
            # print item
            print instru
            instruList.append((instru, exchangeID))
    return instruList


# def scanInstrumentIDInCasesDB():
#     '''
#     扫描案例数据库获取合约信息

#     *返回值：
#         instrumentList[],内容为(InstrumentID, ExchangeID)
#     '''
#     cmdstringList = []
#     sql = "select cmdstring from tbl_cases"
#     ds = executeCaseSQL(sql)
#     for rec in ds:
#         if not rec["cmdstring"].__contains__(':'):
#             rec["cmdstring"] = sc.decrypt_mode_cbc(rec["cmdstring"].encode('gbk'), 'abcdefghijklmnop')

#         cmdstringList.append(rec["cmdstring"])
#     return extractInstrumentIDfromCmdstring(cmdstringList)


def get_process_name_by_id(pid):
    c = wmi.WMI()
    msg = u"后台进程扫描"
    Logging.getLog().info(msg)

    for proc in c.Win32_Process():
        if proc.ProcessId == pid:
            return proc.Name
    return ""


def running_environment_check():
    """
     运行环境检查。

     * 程序启动时，检查tbl_running_control表。
        无记录：继续执行，无提示
        有记录：
            status为0或1，判断记录中进程是否存在。
                如果存在 则提示，已经有程序在运行，是否需要强制结束，还是继续等待。
                如果不存在，说明之前的进程异常退出，未能正确设置状态。删除该记录，继续执行。
            status 为2。删除该记录，继续执行

     * 返回值为(ret,msg) 值对
        ret:
            -1: 有异常抛出
            0: 可以继续执行后续代码
            1: 需要提示客户做出选择(命令行执行情况下，直接退出)
        msg:
            提示信息
    """
    try:
        sql = "select process_id, process_name,hostname,hostIP,running_status from tbl_running_control"
        ds = executeCaseSQL(sql)
        if ds == []:
            return 0, ""
        if ds[0]["running_status"] == 2:
            sql = "delete from tbl_running_control"
            executeCaseSQL(sql)
            return 0, ""
        if ds[0]["running_status"] == 0 or ds[0]["running_status"] == 1:
            pid = ds[0]["process_id"]
            pname = ds[0]["process_name"]
            pname_local = get_process_name_by_id(pid)
            if pname_local == "" or pname_local != pname:  # 进程不存在,或进程存在但是和数据库记录的进程名不一致
                sql = "delete from tbl_running_control"
                executeCaseSQL(sql)
                return 0, ""
            else:
                msg = u"自动化测试软件%s正在运行中!" % (pname)
                return 1, msg
    except Exception, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1, str(exc_info)
    return 0


def kill_console_process():
    try:
        sql = "select process_id,process_name from tbl_running_control"
        ds = executeCaseSQL(sql)
        process_id = -1
        if ds != []:
            process_id = ds[0]["process_id"]
            process_name = ds[0]["process_name"]
        if process_id == -1:
            return -1
        Logging.getLog().info("准备终止后台进程")
        # kill process
        c = wmi.WMI()

        for proc in c.Win32_Process():
            if proc.ProcessId == process_id:
                try:
                    proc.Terminate()
                except BaseException, ex:
                    exc_info = getExceptionInfo()
                    Logging.getLog().critical(exc_info)
                    # 由于时间延迟，有可能进程已经自动终止了，这时再做一次判断，进程是否还在
                    for proc in c.Win32_Process():
                        if proc.ProcessId == process_id:
                            # 进程还在
                            return -1
                    Logging.getLog().info("后台进程已经终止")
                    return 0
                Logging.getLog().info("后台进程已经终止")
                break;
    except:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1
    return 0


def set_running_state(state):
    '''
    设置程序运行状态
    state: 运行状态
        0：程序已经启动，尚未进入案例执行阶段
        1：案例运行状态
        2：案例运行结束状态

    * 返回值
        0：成功
        -1：失败
    '''
    msg = u"设置运行状态为%d" % (state)
    Logging.getLog().debug(msg)
    try:
        pid = os.getpid()
        if (state == 0):
            pname = get_process_name_by_id(pid)
            hostname = socket.getfqdn(socket.gethostname())
            hostIP = socket.gethostbyname(hostname)

            sql = "delete from tbl_running_control"
            executeCaseSQL(sql)
            sql = "insert into tbl_running_control(process_name,process_id,hostname,hostIP,running_status) values('%s',%d,'%s','%s',0)" % (
                pname, pid, hostname, hostIP)
            executeCaseSQL(sql)
        elif (state == 1):
            start_time = time.strftime("%Y-%m-%d %H:%M:%S")
            sql = "update tbl_running_control set start_time = '%s', running_status = 1" % (start_time)
            # print sql
            executeCaseSQL(sql)
        elif (state == 2):
            end_time = time.strftime("%Y-%m-%d %H:%M:%S")
            sql = "update tbl_running_control set end_time = '%s', running_status = 2" % (end_time)
            executeCaseSQL(sql)

        else:
            return -1
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1
    return 0


def TestRunSql():
    sql = "select reqtext from reqresp_20141120  order by reqnum "
    ds = executeRunSQL(sql)
    for item in ds:
        print item['reqtext']
        reqlist = (((str(item['reqtext'])).strip()).split('\x01'))[:-1]
        for req in reqlist:
            element = req.split('=')
            if element[0] == '38':
                Volume = element[1]
                print 'Volume:', Volume

        print reqlist


def UseBeyondCompare(src, dst):
    if os.path.exists(src) == False:
        ErrMsg = src + ' not exists !'
        Logging.getLog().error(ErrMsg)
        # print ErrMsg
        return ErrMsg
    if os.path.exists(dst) == False:
        ErrMsg = dst + ' not exists !'
        Logging.getLog().error(ErrMsg)
        return ErrMsg
    if os.path.exists(gl.g_bcmppath) == False:
        ErrMsg = u'应用程序BCompare.exe的路径: ' + gl.g_bcmppath + u'配置不正确 !'
        Logging.getLog().error(ErrMsg)
        return ErrMsg

    exe = gl.g_bcmppath

    cmd_str = '"%s" -aru "%s" "%s"' % (exe, src, dst)

    try:
        si = subprocess.STARTUPINFO
        si.wShowWindow = subprocess.SW_HIDE
        proc = subprocess.Popen(cmd_str, shell=False, startupinfo=si, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                stdin=subprocess.PIPE)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        ErrMsg = u'Open Beyond Compare Exception: %s' % (exc_info)
        return ErrMsg
    return None


def SortSourceFile(srcrootname=None):  # 排序需要做数据对比的文件,按案例执行的顺序进行数据对比
    srcfilelist = os.listdir(srcrootname)
    tempdict = {}
    xmlfile = None
    for item in srcfilelist:
        path = srcrootname + '\\' + item
        if (os.path.isfile(path)):
            xmlfile = item
            continue
        key = item[:item.find('_')]
        num = int(key)
        tempdict[num] = item
    if xmlfile != None:
        count = len(tempdict)
        tempdict[count + 1] = xmlfile
    srcfilelist = sorted(tempdict.iteritems(), key=itemgetter(0))
    dstlist = []
    for item in srcfilelist:
        dstlist.append(item[1])
    return dstlist


def autoSettingUser(cmdstring, UserId):  # 自动替换用户代码和投资者代码
    if cmdstring.find("InvestorID=") >= 0:  # 根据配置文件中的个股期权帐号自动替换命令行中的投资者代码
        # 查找投资者代码
        left = cmdstring.find("InvestorID=")
        if left == -1:
            # 不应该发生
            pass
        else:
            right = cmdstring.find("#", left)
            investorid = cmdstring[left + len("InvestorID="):right]  # 投资者代码
            oldstr = "InvestorID=" + investorid
            newstr = "InvestorID=" + str(UserId)
            cmdstring = cmdstring.replace(oldstr, newstr)
    if cmdstring.find("UserID=") >= 0:  # 根据配置文件中的个股期权账户自动替换命令行中的用户代码
        # 查找用户代码
        left = cmdstring.find("UserID=")
        if left == -1:
            # 不应该发生
            pass
        else:
            right = cmdstring.find("#", left)
            investorid = cmdstring[left + len("UserID="):right]  # 用户代码
            oldstr = "UserID=" + investorid
            newstr = "UserID=" + str(UserId)
            cmdstring = cmdstring.replace(oldstr, newstr)
    return cmdstring


def InitETFElementPosition():
    try:
        InsertPositonFromCsv()  # 从CSV导入持仓明细
        rootname = os.getcwd()
        temp_file_list = os.listdir(rootname)
        fileDict = {}
        for item in temp_file_list:
            file = (str(item).strip())
            if file.__contains__(".PCF"):
                fileDict[file] = 'SZE'
            if file.__contains__(".ETF"):
                fileDict[file] = 'SSE'
        print fileDict
        if len(fileDict) == 0:
            log = u'无ETF成份股文件 , 请拷贝需要初始化持仓的ETF成份股文件  !'
            wx.MessageBox(log, u"提示")
            return
        sql = "SELECT max(TradeID)  as TrdId from tbl_etfelementposition where TradeID < 106394"
        ds = executeCaseSQL(sql)
        tradeid = ds[0]['TrdId']  # 查询当前最大的成交编号,成交编号不能重复,插入成交编号递增
        print tradeid
        # tradeid = 1000
        for file in fileDict.keys():
            # file = '15990120141104.PCF'
            ETFName = file[:6]
            log = u'正在初始化 %s 的成份股持仓.......' % (ETFName)
            # print log
            Logging.getLog().info(log)
            ExchangeId = fileDict[file]
            f = open(file, "r")
            lines = f.readlines()
            f.close()
            InstrumentDict = {}
            for line in lines:
                if line.find('|') != -1:
                    # print line
                    instrumentList = line.split('|')
                    instrumentId = (str(instrumentList[0])).strip()
                    volume = (str(instrumentList[2])).strip()
                    InstrumentDict[instrumentId] = volume

            # print len(InstrumentDict)
            # print InstrumentDict
            for key in InstrumentDict.keys():
                InstrumentID = key
                volume = InstrumentDict[key]
                if (volume == None) or (len(volume) == 0):
                    log = u'现金替代证券 : %s' % (InstrumentID)
                    # print log
                    Logging.getLog().info(log)
                    continue
                Volume = int(volume) * 100
                # print 'InstrumentID: %s  Volume:%d'%(InstrumentID,Volume)
                sql = "SELECT Volume from tbl_etfelementposition   where InvestorID = '010000035178' and AccountID = '01000003517801' and InstrumentID = '%s'" % (
                    InstrumentID)
                ds = executeCaseSQL(sql)
                if len(ds) > 0:
                    OriginalVolume = ds[0]['Volume']
                    # print OriginalVolume
                    NewVolume = OriginalVolume + Volume
                    sql = "update tbl_etfelementposition set Volume = %d where InvestorID = '010000035178' and AccountID = '01000003517801' and InstrumentID = '%s'" % (
                        NewVolume, InstrumentID)
                    print sql
                    ds = executeCaseSQL(sql)
                else:
                    tradeid = tradeid + 1
                    sql = "insert into tbl_etfelementposition (InstrumentID, BrokerID, InvestorID, HedgeFlag, Direction, OpenDate, TradeID, Volume, OpenPrice, TradingDay, TradeType, ExchangeID, Margin, ExchMargin, LastSettlementPrice, SettlementPrice, CloseVolume, CloseAmount, TransferFee, StampTax, Commission, AccountID, PledgeInPosition, PledgeInFrozenPosition, RepurchasePosition, Amount, UnderlyingInstrumentID)\
                            values('%s', '2011', '010000035178', '1', '0', '20141009', '%d', '%d', 10.5, '20141017', '0', '%s', 0, 0, 0, 0, 0, 0, 0, 0, 0, '01000003517801', 0, 0, 0, 0, ' ')" \
                          % (InstrumentID, tradeid, Volume, ExchangeId)
                    # print sql
                    executeCaseSQL(sql)
            log = u'初始化 %s 的成份股持仓结束  !' % (ETFName)
            # print log
            Logging.getLog().info(log)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        strerror = u"ETF成份股初始化抛出异常，原因：%s" % (exc_info)
        wx.MessageBox(strerror, u"提示")
        Logging.getLog().critical(strerror)
        return -1


def InsertPositonFromCsv():
    try:
        log = u'正准备从CSV导入持仓明细........'
        # readConfigFile()
        print log
        Logging.getLog().info(log)
        file = gl.g_ETFCsvFilePath + '\\t_InvestorPositionDetail.csv'
        print file
        f = open(file, "r")
        lines = f.readlines()
        f.close()
        sql = "delete from tbl_etfelementposition"  # 清空原持仓表记录
        executeCaseSQL(sql)
        i = 0
        for line in lines:
            line = line.strip('\n')
            i = i + 1
            if i == 1:  # 首行为字段名,不需要插入
                continue
            # print line
            ValueList = line.split(',')
            InstrumentID = ValueList[0][1:-1]  # %s
            BrokerID = ValueList[1][1:-1]  # %s
            InvestorID = ValueList[2][1:-1]  # %s
            HedgeFlag = ValueList[3][1:-1]  # %s
            Direction = ValueList[4][1:-1]  # %s
            OpenDate = ValueList[5][1:-1]  # %s
            TradeID = int(ValueList[6][1:-1])  # %d
            Volume = int(ValueList[7][1:-1])  # %d
            OpenPrice = float(ValueList[8][1:-1])  # %f
            TradingDay = ValueList[9][1:-1]  # %s
            TradeType = ValueList[10][1:-1]  # %s
            ExchangeID = ValueList[11][1:-1]  # %s
            Margin = float(ValueList[12][1:-1])  # %f
            ExchMargin = float(ValueList[13][1:-1])  # %f
            LastSettlementPrice = float(ValueList[14][1:-1])  # %f
            SettlementPrice = float(ValueList[15][1:-1])  # %f
            CloseVolume = int(ValueList[16][1:-1])  # %d
            CloseAmount = float(ValueList[17][1:-1])  # %f
            TransferFee = float(ValueList[18][1:-1])  # %f
            StampTax = float(ValueList[19][1:-1])  # %f
            Commission = float(ValueList[20][1:-1])  # %f
            AccountID = ValueList[21][1:-1]  # %s
            PledgeInPosition = int(ValueList[22][1:-1])  # %d
            PledgeInFrozenPosition = int(ValueList[23][1:-1])  # %d
            RepurchasePosition = int(ValueList[24][1:-1])  # %d
            Amount = float(ValueList[25][1:-1])  # %f
            UnderlyingInstrumentID = ValueList[26][1:-1]  # %s
            sql = "insert into tbl_etfelementposition (InstrumentID,BrokerID,InvestorID,HedgeFlag,Direction,OpenDate,TradeID,Volume,OpenPrice,TradingDay,TradeType,ExchangeID,Margin,ExchMargin,LastSettlementPrice,SettlementPrice,CloseVolume,CloseAmount,TransferFee,StampTax,Commission,AccountID,PledgeInPosition,PledgeInFrozenPosition,RepurchasePosition,Amount,UnderlyingInstrumentID)\
                    values('%s', '%s', '%s', '%s', '%s', '%s', '%d', '%d', %f, '%s', '%s', '%s', %f, %f, %f, %f, %d, %f, %f, %f, %f, '%s', %d, %d, %d, %f, '%s')" \
                  % (InstrumentID, BrokerID, InvestorID, HedgeFlag, Direction, OpenDate, TradeID, Volume, OpenPrice,
                     TradingDay, TradeType, ExchangeID, Margin, ExchMargin, LastSettlementPrice, SettlementPrice,
                     CloseVolume, CloseAmount, TransferFee, StampTax, Commission, AccountID, PledgeInPosition,
                     PledgeInFrozenPosition, RepurchasePosition, Amount, UnderlyingInstrumentID)
            # print sql
            try:
                executeCaseSQL(sql)
            except Exception, ex:
                exc_info = getExceptionInfo()
                err = 'Line: %d InstrumentID: %s  Volume: %d   AccountID: %s' % (i, InstrumentID, Volume, AccountID)
                print err
                log = u'从CSV导入持仓表 :tbl_etfelementposition出现异常: %s' % (exc_info)
                Logging.getLog().critical(log)
        log = u'从CSV导入持仓明细结束  !'
        print log
        Logging.getLog().info(log)
        print 'InsertPositonFromCsv Finished !'
    except BaseException, ex:
        exc_info = getExceptionInfo()
        strerror = u"从CSV导入持仓明细抛出异常，原因：%s" % (exc_info)
        wx.MessageBox(strerror, u"提示")
        Logging.getLog().critical(strerror)
        return -1


def TranslateSqlToCSV():
    try:
        log = u'正准备从数据库导出CSV文件........'
        # readConfigFile()
        print log
        Logging.getLog().info(log)
        #     file = 'D:\\t_InvestorPositionDetail.csv'
        file = gl.g_ETFCsvFilePath + '\\t_InvestorPositionDetail.csv'
        print file
        f = open(file, "r")
        lines = f.readlines()
        f.close()
        line = lines[0]
        # print line
        Columnline = line.strip('\n')
        templist = Columnline.split(',')
        # print templist
        Columns = ''
        columnlist = []
        for item in templist:
            columnlist.append(item[1:-1])
            Columns = Columns + item[1:-1] + ','
        # print Columns[:-1]
        # print columnlist
        rootname = os.getcwd()
        file_name = rootname + '\\t_InvestorPositionDetail.csv'
        fileIn_csv = open(file_name, "w")
        fileIn_csv.write(lines[0])

        sql = "SELECT %s from tbl_etfelementposition order by InvestorID,InstrumentID,AccountID" % (Columns[:-1])
        ds = executeCaseSQL(sql)
        for Record in ds:
            line = ''
            # print Record
            for column in columnlist:
                line = line + '"' + str(Record[column]) + '",'
            line = line[:-1] + '\n'
            # print line
            fileIn_csv.write(line)
        fileIn_csv.close()
        log = u'从数据库导出CSV文件结束  !'
        print log
        Logging.getLog().info(log)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        strerror = u"从数据库导出CSV文件抛出异常，原因：%s" % (exc_info)
        wx.MessageBox(strerror, u"提示")
        Logging.getLog().critical(strerror)
        return -1


def QueryCmpResultByChoice(choiceText):  # 根据对比文件的选择,查询符合条件的对比的结果
    ChoiceDict = {}
    Root = os.getcwd() + '\\Compared'
    VSFileList = os.listdir(Root)
    if choiceText == None:  # 对比数据文件一个都没有选择
        for Item in VSFileList:
            filepath = Root + '\\' + Item
            cmplist = os.listdir(filepath)
            if len(cmplist) == 0:  # 判断VS文件下是否为空,如果为空,遍历下个VS文件夹;如果不为空,取出该文件夹下所有对比结果
                continue
            else:
                for data in cmplist:
                    ChoiceDict[data] = Item
    elif choiceText.find("VS") != -1:  # 对比数据文件选择了两个
        if choiceText not in VSFileList:  # 选择的对比的数据文件,没有匹配的数据对比结果
            return ChoiceDict
        else:
            filepath = Root + '\\' + choiceText
            cmplist = os.listdir(filepath)
            if len(cmplist) == 0:  # 判断VS文件下是否为空,如果为空,遍历下个VS文件夹;如果不为空,取出该文件夹下所有对比结果
                return ChoiceDict
            else:
                for data in cmplist:
                    ChoiceDict[data] = choiceText
    else:  # 对比数据文件选择了一个
        for Item in VSFileList:
            if Item.find(choiceText) != -1:
                filepath = Root + '\\' + Item
                cmplist = os.listdir(filepath)
                if len(cmplist) == 0:  # 判断VS文件下是否为空,如果为空,遍历下个VS文件夹;如果不为空,取出该文件夹下所有对比结果
                    continue
                else:
                    for data in cmplist:
                        ChoiceDict[data] = Item
    return ChoiceDict


# ZHLC Auto Test
def connectT2(addr, license_path):
    '''
    获取一个
    '''
    try:
        sdk = t2.TimesvcT2SDK()
        ret = sdk.InitInterface(addr, license_path)
        if ret != 0:
            return None, "InitInterface fail"
        ret = sdk.Connect(10000)
        if ret != 0:
            return None, sdk.GetErrorMsg(ret)
        return sdk, "OK"
    except Exception as e:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(u"connectT2:%s" % (exc_info))
        return -1, "T2 error"


def connectUFX(addr, license_path):
    '''
    连接UFX
    '''
    Logging.getLog().debug(gl.g_ZHLCSystemSetting["T2Server"])
    sdk = O32T2SDK.O32T2SDK()
    ret = sdk.InitInterface(addr, license_path)
    if ret != 0:
        return None, "InitInterface fail"
    ret = sdk.Connect(1000)
    if ret != 0:
        return None, sdk.GetErrorMsg(ret)
    return sdk, "OK"


def sendT2Biz(sdk, funcid, cmdstring):
    Logging.getLog().critical("funcid %s" % funcid)
    # gl.HsBranchNo_char_to_int = False
    # if str(funcid) in gl.funcids_list:
    #     gl.HsBranchNo_char_to_int = True
    try:
        if not funcid:
            Logging.getLog().critical("sendT2Biz 传入功能号funcid为空，不能从T2InterfaceDetails.xml 获取信息")
            return -1, "sendT2Biz 传入功能号funcid为空，不能从T2InterfaceDetails.xml 获取信息"
        if gl.g_interfaceDetailTree == None:
            # gl.g_interfaceDetailTree = ET.parse(".\\Adapter\\T2\\T2InterfaceDetails.xml")
            gl.g_interfaceDetailTree = ET.parse(".\\init\\%s\\T2InterfaceDetails.xml" % gl.current_system_id)

        func_root = gl.g_interfaceDetailTree.getroot()
        xpath = "Interface[@funcid='%s']" % (funcid)
        elem = ""
        try:
            elem = func_root.find(xpath)
        except Exception, ex:
            exc_info_zzp = getExceptionInfo()
            Logging.getLog().critical("方法sendT2Biz 该功能号 不存在")
            Logging.getLog().critical("%s" % exc_info_zzp)
            return -1, "%s" % exc_info_zzp
        if not elem:
            return -1, "fail to find interface detail"

        # in_elem_list = elem.findall("InputParams/in")
        ##cmd_list = cmdstring.split('|')
        cmdList = cmdstring.split('#')
        ## print "beginpack"
        sdk.GenNewPacker(2)
        sdk.BeginPack()
        cmd_list = [i.replace("&bfb&", "#") for i in cmdList]
        print "cmd_list:%s" % (cmd_list)
        errMessage = ''
        for i, inpar in enumerate(cmd_list):
            # if cmd_list[i].split(':')[1] == "":   //TODO
            #    continue
            cmdValue = ""

            # if i == len(cmd_list):
            #    break
            cmdStr = cmd_list[i].split("=")
            if len(cmdStr) == 2:
                field, cmdValue = cmdStr
                # if field != inpar.attrib["name"]:#fixme
                #    continue
                # 如果值中有&comma&，予以替换
                cmdValue = cmdValue.replace('&comma&', ',')
                cmdValue = cmdValue.replace('&colon&', ':')
            else:
                field = cmdStr[0]
                cmdValue = ""

                for cindex, cmditem in enumerate(cmdStr):
                    if cindex > 0:
                        if cindex == 1:
                            cmdValue = cmditem
                        else:
                            cmdValue = cmdValue + ":" + cmditem
                    # 如果值中有&comma&，予以替换
                    cmdValue = cmdValue.replace('&comma&', ',')
                    cmdValue = cmdValue.replace('&colon&', ':')
            type = ""
            for test in elem[0].getchildren():
                if test.attrib["name"] == field:
                    if test.attrib["type"] == "":  # fixme
                        # Logging.getLog().error("interface param type is not set, using String instead")
                        type = 'S'
                    else:
                        if test.attrib["name"] == field:
                            if not test.attrib["type"]:
                                return -1, u"funcid:%s,接口字段 field:%s 数据类型不能为空!" % (funcid, field)
                        type = judge_data_type(test.attrib["type"])
                    break
            try:
                sdk.AddField(test.attrib["name"], type, 1000)
            except BaseException, ex:
                exc_info = getExceptionInfo()
                Logging.getLog().error(exc_info)
                Logging.getLog().critical(u"%s案例中的%s字段在配置文件中不存在" % (funcid, field))
                return -1, u"%s案例中的%s字段在配置文件中不存在" % (funcid, field)
        for i, inpar in enumerate(cmd_list):
            # cmdValue = cmd_list[i].split(':')[1]
            cmdValue = ""

            if i == len(cmd_list):
                break

            cmdStr = cmd_list[i].split("=")
            if len(cmdStr) == 2:
                field, cmdValue = cmdStr
                # if field != inpar.attrib["name"]:#fixme
                #    continue
                # 如果值中有&comma&，予以替换
                cmdValue = cmdValue.replace('&comma&', ',')
                cmdValue = cmdValue.replace('&colon&', ':')

            else:
                field = cmdStr[0]
                cmdValue = ""

                for cindex, cmditem in enumerate(cmdStr):
                    if cindex > 0:
                        if cindex == 1:
                            cmdValue = cmditem
                        else:
                            cmdValue = cmdValue + ":" + cmditem
                    # 如果值中有&comma&，予以替换
                    cmdValue = cmdValue.replace('&comma&', ',')
                    cmdValue = cmdValue.replace('&colon&', ':')
            # if cmdValue == "":
            #    continue
            type = ""
            for test in elem[0].getchildren():
                if test.attrib["name"] == field:
                    if not test.attrib["type"]:
                        return -1, u"funcid:%s,接口字段 field:%s 数据类型不能为空!" % (funcid, field)
                    type = judge_data_type(test.attrib["type"])
                    break
            # type = judge_data_type(elem[0][i].attrib["type"])
            # sdk.AddField(inpar.attrib["name"],type)
            # if field == 'USER_CODE':
            #     pass
            # print type, field, cmdValue, type
            try:
                if type == 'I':
                    #print "I:%s" %(str(cmd_list[i]))
                    if cmdValue == "":  # 如果为空，默认输入空字符 fixme
                        sdk.AddStr('')
                        # sdk.AddChar('')
                    else:
                        # print "addInt"
                        sdk.AddInt(int(cmdValue))
                elif type == 'F':
                    # print "F:%s" %(cmd_list[i])
                    if cmdValue == "":  # 如果为空，默认输入空字符 fixme
                        sdk.AddChar('')
                    else:
                        sdk.AddDouble(float(cmdValue))
                elif type == 'C':
                    try:
                        sdk.AddChar(cmdValue.encode('ascii'))
                    except Exception, ex:
                        sdk.AddChar(cmdValue)
                        # if str(field) in ['GENERATOR_NO', 'NICKNAME']:
                        #     sdk.AddChar(cmdValue)
                        # else:
                        #     sdk.AddChar(cmdValue.encode('ascii'))
                        # try:
                        #     sdk.AddChar(cmdValue.encode('ascii'))
                        # except Exception, ex:
                        #     try:
                        #         if cmdValue == "":  # 如果为空，默认输入空字符 fixme
                        #             sdk.AddChar('')
                        #         else:
                        #             sdk.AddInt(int(cmdValue))
                        #     except Exception, ex:
                        #         sdk.AddStr(cmdValue.encode('ascii'))
                elif type == 'S':
                    # print "S:%s" %(cmd_list[i].encode('gbk'))
                    sdk.AddStr(cmdValue.encode("gbk"))
                else:
                    print "Unknow data type"
                    # return -1,"Unknow data type"
                    errMessage = errMessage + u"“" + field + u"”列字段类型获取失败； "
            except Exception, ex:
                exc_info_zzp = getExceptionInfo()
                Logging.getLog().critical(exc_info_zzp)
                columnName = field
                errMessage = str(errMessage) + "“" + str(columnName) + "”列输入值“" + str(cmdValue) + "”类型有误，"
                if type == 'I':
                    errMessage = errMessage + "必须为整数类型； ";
                elif type == 'F':
                    errMessage = errMessage + "必须为数值类型； ";
                elif type == 'C':
                    errMessage = errMessage + "必须为字符类型； ";
                elif type == 'S':
                    errMessage = errMessage + "必须为字符串类型； ";
                else:
                    errMessage = errMessage + "请重新输入； ";

                Logging.getLog().info(errMessage)
        if errMessage != '':
            return -1, errMessage
        sdk.EndPack()

        ret = sdk.SendBiz(int(funcid))
        if ret < 0:
            return -1, sdk.GetErrorMsg(ret)
        return ret, "OK"
    except:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1, u"请检查案例对应的接口字段类型是否缺少或者填写有误,谢谢!.%s" % exc_info


def judge_data_type(type):
    # if gl.HsBranchNo_char_to_int == True and type == 'HsBranchNo':
    #     return "I"
    homedir = os.getcwd()
    # gl.current_system_id = '1428908551835S0NXP'
    path = homedir + '\init\%s\datatype.xml' % gl.current_system_id
    # Logging.getLog().info(path)
    root = ET.parse(path).getroot()
    xpath = "test[@name='%s']" % type
    # if type=='HsIDNo':
    #    pass
    try:
        elem = root.find(xpath).attrib['type']
    except Exception, ex:  # fixme
        if type[0] == 'N':
            if type.find(".") < 0:
                return "I"
            else:
                return "F"
        elif type[0] == 'C':
            if type == 'C1':
                return 'C'
            else:
                return 'S'
        else:
            return 'S'
    if elem == 'int':
        return "I"
    elif elem == 'double':
        return "F"
    elif elem == 'char':
        return 'C'
    else:
        return 'S'


def getRecvBizPacket(sdk, index):
    row_lists = []
    sdk.SetCurrentDatasetByIndex(index)

    col_count = sdk.GetColCount()
    col_list = []
    for j in xrange(col_count):
        # print "%s" %(sdk.GetColName(j))
        col_list.append(sdk.GetColName(j))
    row_lists.append(col_list)
    row_count = sdk.GetRowCount()
    for k in xrange(row_count):
        row_list = []
        for l in xrange(col_count):
            type = sdk.GetColType(l)
            val = None
            if type == 'I':
                # print "%d" %(sdk.GetIntByIndex(l))
                val = sdk.GetIntByIndex(l)

            elif type == 'C':
                val = sdk.GetCharByIndex(l)
            elif type == 'S':
                val = sdk.GetStrByIndex(l)
            elif type == 'F':
                val = sdk.GetDoubleByIndex(l)
            elif type == 'R':
                lpLen = t2.new_intp()
                raw = sdk.GetRawByIndex(l, lpLen)
                len = t2.intp_value(lpLen)
                str1 = t2.cdata(raw, len)
                val = binascii.b2a_hex(str1)
            else:
                val = "None"
            # print val
            if isinstance(val, float):
                row_list.append("%.2f" % val)
            else:
                row_list.append(str(val))
        row_lists.append(row_list)
        sdk.Next()
    return row_lists


def findAndDeleteCaps(toRePlaceDict, filePath):
    # toPlaceDict specifies which items to replace and with what
    # e.g. Aname will be replaced with name, Btype is replaced with type
    # toRePlaceDict = {"Aname":"name", "Benabled":"enabled", "Btype":"type", "Cshadow_caster":"shadow_caster", "Ctemplate":"template", "Cpadding":"padding", "Cmaterial":"material"}

    # Read in file as a string and open it to be written to
    textToReplace = open(filePath).read()
    writeFile = open(filePath, "w")

    # run through the dictionary using the original key and the new text to use
    for original, new in toRePlaceDict.iteritems():
        textToReplace = textToReplace.replace(original, new)  # Replace the specified text
    writeFile.write(textToReplace)
    writeFile.close()


def ExcelProcessParam(system_id, funcid, param_index, group_dict, account_group_id, param_data_id):
    """
    处理Excel参数
    :param system_id:
    :param funcid:
    :param param_index:
    :param group_dict:
    :param account_group_id:
    :return:
    """
    ret, value_list = read_funcid_param_mang_step(system_id, funcid, param_index, param_data_id)
    if ret < 0:
        Logging.getLog().error("read_funcid_param fail")
    value_list = ExcelParameterReplace(value_list, group_dict, account_group_id)
    write_funcid_param_many_step(system_id, funcid, param_index, value_list, param_data_id)


def export_data(colname, FLAG, data):
    """导出excel"""
    wb = Workbook()
    # ws = wb.active

    # all_data是字典类型， key值是sheet的名称， values是对应的行数据
    # 判断sheet的名称是不是以文件名开头的数据，如果是是同一个文件的数据，保存在同一个文件，否则反正
    ws = wb.create_sheet('sheet')
    # 创建sheet
    for ch in xrange(0x41, 0x5A):
        # 设置列宽
        ws.column_dimensions[str(unichr(ch))].width = 25
    # 数据可以直接分配到单元格中
    ws.append(colname)
    for j in data:
        # ws.append(j.values())
        ws.append(j)
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.save("%s.xlsx" % FLAG)


def ExcelProcessParamTools(system_id, funcid, group_dict, account_group_id, old_param_new_params):
    """
    处理Excel参数
    :param system_id:
    :param funcid:
    :param group_dict:
    :param account_group_id:
    :return:
    """
    PARAM_DATA_IDS = []
    ret, value_list, col_names = read_funcid_param_tools(system_id, funcid, old_param_new_params)
    if value_list:
        for i, valuelist in enumerate(value_list):
            if valuelist:
                # PARAM_DATA_IDS.append(valuelist[1])
                param_index = i + 1
                write_funcid_param_tool(system_id, funcid, param_index, valuelist)
    if ret < 0:
        Logging.getLog().info(u"JZYY_EXCEL 替换参数完毕")
    mycopyfile(funcid)
    Logging.getLog().info(u"JZYY_EXCEL\%s copy 到init\\14774643386617P4VW下" % funcid[0:3])
    value_list = ExcelParameterReplaceTools(value_list, group_dict, account_group_id)
    if value_list:
        for i, valuelist in enumerate(value_list):
            if valuelist:
                PARAM_DATA_IDS.append(valuelist[1])
                param_index = i + 1
                write_funcid_param(system_id, funcid, param_index, valuelist)
    return PARAM_DATA_IDS


def mycopyfile(funcid):
    try:
        home_dir = os.getcwd()
        srcfile = r'%s\JZYY_EXCEL\%s.xlsx' % (home_dir, funcid[0:3])
        dstfile = r'%s\init\14774643386617P4VW\%s.xlsx' % (home_dir, funcid[0:3])
        if not os.path.isfile(srcfile):
            print "%s not exist!" % (srcfile)
        else:
            fpath, fname = os.path.split(dstfile)  # 分离文件名和路径
            if not os.path.exists(fpath):
                os.makedirs(fpath)  # 创建路径
            shutil.copyfile(srcfile, dstfile)  # 复制文件
            print "copy %s -> %s" % (srcfile, dstfile)
    except:
        exc_info = getExceptionInfo()
        print exc_info
        return -1, exc_info


def ExcelParameterReplace(value_list, group_dict, account_group_id=-1, ):
    """

    :param value_list: 待处理的值列表
    :param account_group_id: 账号组ID
    :param group_dict: 组变量
    :return: 值列表
    """
    return_value_list = []
    for value in value_list:
        if isinstance(value, basestring):
            value = GlobalParameterReplace(str(value), account_group_id)
            value = GroupParameterReplace(str(value), group_dict)
        return_value_list.append(value)

    return return_value_list


def ExcelParameterReplaceTools(value_list, group_dict, account_group_id=-1, ):
    """

    :param value_list: 待处理的值列表
    :param account_group_id: 账号组ID
    :param group_dict: 组变量
    :return: 值列表
    """
    all_data = []
    for values in value_list:
        if values:
            return_value_list = []
            for value in values:
                if isinstance(value, basestring):
                    value = GlobalParameterReplace(str(value), account_group_id)
                    value = GroupParameterReplace(str(value), group_dict)
                return_value_list.append(value)
            all_data.append(return_value_list)
    return all_data


def GroupParameterReplace(string, dict):
    '''
     案例组级别的参数替换
    '''
    if string == "" or string == None:
        return ""
    if "compare_list" in string:
        return string
    if gl.g_case_adapter_type != 'SPLX_SELENIUM':
        string = string.replace("$bfh$", '"')
    # if string.find('@random_number@') >= 0:
    print string
    string = string.replace("@random_number@", str(int(time.time())))
    for key in dict.keys():
        field = "@%s@" % key
        string = string.replace(field, str(dict[key]))
        # if str(dict[key]):
        #     string = string.replace(field, str(dict[key]))
        # else:
        #     string = string.replace(field, '""')
    return string


def getCardID():
    # 区域码
    Region = [110101, 110102, 110105, 110106, 110107, 110108, 110109, 110111, 110112, 110113, 110114, 110115, 110116,
              110117, 110228, 110229, 120101, 120102, 120103, 120104, 120105, 120106, 120110, 120111, 120112, 120113,
              120114, 120115, 120116, 120221, 120223, 120225, 130101, 130102, 130103, 130104, 130105, 130107, 130108,
              130121, 130123, 130124, 130125, 130126, 130127, 130128, 130129, 130130, 130131, 130132, 130133, 130181,
              130182, 130183, 130184, 130185, 130201, 130202, 130203, 130204, 130205, 130207, 130208, 130209, 130223,
              130224, 130225, 130227, 130229, 130281, 130283, 130301, 130302, 130303, 130304, 130321, 130322, 130323,
              130324, 130401, 130402, 130403, 130404, 130406, 130421, 130423, 130424, 130425, 130426, 130427, 130428,
              130429, 130430, 130431, 130432, 130433, 130434, 130435, 130481, 130501, 130502, 130503, 130521, 130522,
              130523, 130524, 130525, 130526, 130527, 130528, 130529, 130530, 130531, 130532, 130533, 130534, 130535,
              130581, 130582, 130601, 130602, 130603, 130604, 130621, 130622, 130623, 130624, 130625, 130626, 130627,
              130628, 130629, 130630, 130631, 130632, 130633, 130634, 130635, 130636, 130637, 130638, 130681, 130682,
              130683, 130684, 130701, 130702, 130703, 130705, 130706, 130721, 130722, 130723, 130724, 130725, 130726,
              130727, 130728, 130729, 130730, 130731, 130732, 130733, 130801, 130802, 130803, 130804, 130821, 130822,
              130823, 130824, 130825, 130826, 130827, 130828, 130901, 130902, 130903, 130921, 130922, 130923, 130924,
              130925, 130926, 130927, 130928, 130929, 130930, 130981, 130982, 130983, 130984, 131001, 131002, 131003,
              131022, 131023, 131024, 131025, 131026, 131028, 131081, 131082, 131101, 131102, 131121, 131122, 131123,
              131124, 131125, 131126, 131127, 131128, 131181, 131182, 140101, 140105, 140106, 140107, 140108, 140109,
              140110, 140121, 140122, 140123, 140181, 140201, 140202, 140203, 140211, 140212, 140221, 140222, 140223,
              140224, 140225, 140226, 140227, 140301, 140302, 140303, 140311, 140321, 140322, 140401, 140402, 140411,
              140421, 140423, 140424, 140425, 140426, 140427, 140428, 140429, 140430, 140431, 140481, 140501, 140502,
              140521, 140522, 140524, 140525, 140581, 140601, 140602, 140603, 140621, 140622, 140623, 140624, 140701,
              140702, 140721, 140722, 140723, 140724, 140725, 140726, 140727, 140728, 140729, 140781, 140801, 140802,
              140821, 140822, 140823, 140824, 140825, 140826, 140827, 140828, 140829, 140830, 140881, 140882, 140901,
              140902, 140921, 140922, 140923, 140924, 140925, 140926, 140927, 140928, 140929, 140930, 140931, 140932,
              140981, 141001, 141002, 141021, 141022, 141023, 141024, 141025, 141026, 141027, 141028, 141029, 141030,
              141031, 141032, 141033, 141034, 141081, 141082, 141101, 141102, 141121, 141122, 141123, 141124, 141125,
              141126, 141127, 141128, 141129, 141130, 141181, 141182, 150101, 150102, 150103, 150104, 150105, 150121,
              150122, 150123, 150124, 150125, 150201, 150202, 150203, 150204, 150205, 150206, 150207, 150221, 150222,
              150223, 150301, 150302, 150303, 150304, 150401, 150402, 150403, 150404, 150421, 150422, 150423, 150424,
              150425, 150426, 150428, 150429, 150430, 150501, 150502, 150521, 150522, 150523, 150524, 150525, 150526,
              150581, 150601, 150602, 150621, 150622, 150623, 150624, 150625, 150626, 150627, 150701, 150702, 150721,
              150722, 150723, 150724, 150725, 150726, 150727, 150781, 150782, 150783, 150784, 150785, 150801, 150802,
              150821, 150822, 150823, 150824, 150825, 150826, 150901, 150902, 150921, 150922, 150923, 150924, 150925,
              150926, 150927, 150928, 150929, 150981, 152201, 152202, 152221, 152222, 152223, 152224, 152501, 152502,
              152522, 152523, 152524, 152525, 152526, 152527, 152528, 152529, 152530, 152531, 152921, 152922, 152923,
              210101, 210102, 210103, 210104, 210105, 210106, 210111, 210112, 210113, 210114, 210122, 210123, 210124,
              210181, 210201, 210202, 210203, 210204, 210211, 210212, 210213, 210224, 210281, 210282, 210283, 210301,
              210302, 210303, 210304, 210311, 210321, 210323, 210381, 210401, 210402, 210403, 210404, 210411, 210421,
              210422, 210423, 210501, 210502, 210503, 210504, 210505, 210521, 210522, 210601, 210602, 210603, 210604,
              210624, 210681, 210682, 210701, 210702, 210703, 210711, 210726, 210727, 210781, 210782, 210801, 210802,
              210803, 210804, 210811, 210881, 210882, 210901, 210902, 210903, 210904, 210905, 210911, 210921, 210922,
              211001, 211002, 211003, 211004, 211005, 211011, 211021, 211081, 211101, 211102, 211103, 211121, 211122,
              211201, 211202, 211204, 211221, 211223, 211224, 211281, 211282, 211301, 211302, 211303, 211321, 211322,
              211324, 211381, 211382, 211401, 211402, 211403, 211404, 211421, 211422, 211481, 220101, 220102, 220103,
              220104, 220105, 220106, 220112, 220122, 220181, 220182, 220183, 220201, 220202, 220203, 220204, 220211,
              220221, 220281, 220282, 220283, 220284, 220301, 220302, 220303, 220322, 220323, 220381, 220382, 220401,
              220402, 220403, 220421, 220422, 220501, 220502, 220503, 220521, 220523, 220524, 220581, 220582, 220601,
              220602, 220605, 220621, 220622, 220623, 220681, 220701, 220702, 220721, 220722, 220723, 220724, 220801,
              220802, 220821, 220822, 220881, 220882, 222401, 222402, 222403, 222404, 222405, 222406, 222424, 222426,
              230101, 230102, 230103, 230104, 230108, 230109, 230110, 230111, 230112, 230123, 230124, 230125, 230126,
              230127, 230128, 230129, 230182, 230183, 230184, 230201, 230202, 230203, 230204, 230205, 230206, 230207,
              230208, 230221, 230223, 230224, 230225, 230227, 230229, 230230, 230231, 230281, 230301, 230302, 230303,
              230304, 230305, 230306, 230307, 230321, 230381, 230382, 230401, 230402, 230403, 230404, 230405, 230406,
              230407, 230421, 230422, 230501, 230502, 230503, 230505, 230506, 230521, 230522, 230523, 230524, 230601,
              230602, 230603, 230604, 230605, 230606, 230621, 230622, 230623, 230624, 230701, 230702, 230703, 230704,
              230705, 230706, 230707, 230708, 230709, 230710, 230711, 230712, 230713, 230714, 230715, 230716, 230722,
              230781, 230801, 230803, 230804, 230805, 230811, 230822, 230826, 230828, 230833, 230881, 230882, 230901,
              230902, 230903, 230904, 230921, 231001, 231002, 231003, 231004, 231005, 231024, 231025, 231081, 231083,
              231084, 231085, 231101, 231102, 231121, 231123, 231124, 231181, 231182, 231201, 231202, 231221, 231222,
              231223, 231224, 231225, 231226, 231281, 231282, 231283, 232721, 232722, 232723, 310101, 310104, 310105,
              310106, 310107, 310108, 310109, 310110, 310112, 310113, 310114, 310115, 310116, 310117, 310118, 310120,
              310230, 320101, 320102, 320103, 320104, 320105, 320106, 320107, 320111, 320113, 320114, 320115, 320116,
              320124, 320125, 320201, 320202, 320203, 320204, 320205, 320206, 320211, 320281, 320282, 320301, 320302,
              320303, 320305, 320311, 320312, 320321, 320322, 320324, 320381, 320382, 320401, 320402, 320404, 320405,
              320411, 320412, 320481, 320482, 320501, 320505, 320506, 320507, 320508, 320509, 320581, 320582, 320583,
              320585, 320601, 320602, 320611, 320612, 320621, 320623, 320681, 320682, 320684, 320701, 320703, 320705,
              320706, 320721, 320722, 320723, 320724, 320801, 320802, 320803, 320804, 320811, 320826, 320829, 320830,
              320831, 320901, 320902, 320903, 320921, 320922, 320923, 320924, 320925, 320981, 320982, 321001, 321002,
              321003, 321012, 321023, 321081, 321084, 321101, 321102, 321111, 321112, 321181, 321182, 321183, 321201,
              321202, 321203, 321281, 321282, 321283, 321284, 321301, 321302, 321311, 321322, 321323, 321324, 330101,
              330102, 330103, 330104, 330105, 330106, 330108, 330109, 330110, 330122, 330127, 330182, 330183, 330185,
              330201, 330203, 330204, 330205, 330206, 330211, 330212, 330225, 330226, 330281, 330282, 330283, 330301,
              330302, 330303, 330304, 330322, 330324, 330326, 330327, 330328, 330329, 330381, 330382, 330401, 330402,
              330411, 330421, 330424, 330481, 330482, 330483, 330501, 330502, 330503, 330521, 330522, 330523, 330601,
              330602, 330621, 330624, 330681, 330682, 330683, 330701, 330702, 330703, 330723, 330726, 330727, 330781,
              330782, 330783, 330784, 330801, 330802, 330803, 330822, 330824, 330825, 330881, 330901, 330902, 330903,
              330921, 330922, 331001, 331002, 331003, 331004, 331021, 331022, 331023, 331024, 331081, 331082, 331101,
              331102, 331121, 331122, 331123, 331124, 331125, 331126, 331127, 331181, 340101, 340102, 340103, 340104,
              340111, 340121, 340122, 340123, 340124, 340181, 340201, 340202, 340203, 340207, 340208, 340221, 340222,
              340223, 340225, 340301, 340302, 340303, 340304, 340311, 340321, 340322, 340323, 340401, 340402, 340403,
              340404, 340405, 340406, 340421, 340501, 340503, 340504, 340506, 340521, 340522, 340523, 340601, 340602,
              340603, 340604, 340621, 340701, 340702, 340703, 340711, 340721, 340801, 340802, 340803, 340811, 340822,
              340823, 340824, 340825, 340826, 340827, 340828, 340881, 341001, 341002, 341003, 341004, 341021, 341022,
              341023, 341024, 341101, 341102, 341103, 341122, 341124, 341125, 341126, 341181, 341182, 341201, 341202,
              341203, 341204, 341221, 341222, 341225, 341226, 341282, 341301, 341302, 341321, 341322, 341323, 341324,
              341501, 341502, 341503, 341521, 341522, 341523, 341524, 341525, 341601, 341602, 341621, 341622, 341623,
              341701, 341702, 341721, 341722, 341723, 341801, 341802, 341821, 341822, 341823, 341824, 341825, 341881,
              350101, 350102, 350103, 350104, 350105, 350111, 350121, 350122, 350123, 350124, 350125, 350128, 350181,
              350182, 350201, 350203, 350205, 350206, 350211, 350212, 350213, 350301, 350302, 350303, 350304, 350305,
              350322, 350401, 350402, 350403, 350421, 350423, 350424, 350425, 350426, 350427, 350428, 350429, 350430,
              350481, 350501, 350502, 350503, 350504, 350505, 350521, 350524, 350525, 350526, 350527, 350581, 350582,
              350583, 350601, 350602, 350603, 350622, 350623, 350624, 350625, 350626, 350627, 350628, 350629, 350681,
              350701, 350702, 350721, 350722, 350723, 350724, 350725, 350781, 350782, 350783, 350784, 350801, 350802,
              350821, 350822, 350823, 350824, 350825, 350881, 350901, 350902, 350921, 350922, 350923, 350924, 350925,
              350926, 350981, 350982, 360101, 360102, 360103, 360104, 360105, 360111, 360121, 360122, 360123, 360124,
              360201, 360202, 360203, 360222, 360281, 360301, 360302, 360313, 360321, 360322, 360323, 360401, 360402,
              360403, 360421, 360423, 360424, 360425, 360426, 360427, 360428, 360429, 360430, 360481, 360482, 360501,
              360502, 360521, 360601, 360602, 360622, 360681, 360701, 360702, 360721, 360722, 360723, 360724, 360725,
              360726, 360727, 360728, 360729, 360730, 360731, 360732, 360733, 360734, 360735, 360781, 360782, 360801,
              360802, 360803, 360821, 360822, 360823, 360824, 360825, 360826, 360827, 360828, 360829, 360830, 360881,
              360901, 360902, 360921, 360922, 360923, 360924, 360925, 360926, 360981, 360982, 360983, 361001, 361002,
              361021, 361022, 361023, 361024, 361025, 361026, 361027, 361028, 361029, 361030, 361101, 361102, 361121,
              361122, 361123, 361124, 361125, 361126, 361127, 361128, 361129, 361130, 361181, 370101, 370102, 370103,
              370104, 370105, 370112, 370113, 370124, 370125, 370126, 370181, 370201, 370202, 370203, 370205, 370211,
              370212, 370213, 370214, 370281, 370282, 370283, 370284, 370285, 370301, 370302, 370303, 370304, 370305,
              370306, 370321, 370322, 370323, 370401, 370402, 370403, 370404, 370405, 370406, 370481, 370501, 370502,
              370503, 370521, 370522, 370523, 370601, 370602, 370611, 370612, 370613, 370634, 370681, 370682, 370683,
              370684, 370685, 370686, 370687, 370701, 370702, 370703, 370704, 370705, 370724, 370725, 370781, 370782,
              370783, 370784, 370785, 370786, 370801, 370802, 370811, 370826, 370827, 370828, 370829, 370830, 370831,
              370832, 370881, 370882, 370883, 370901, 370902, 370911, 370921, 370923, 370982, 370983, 371001, 371002,
              371081, 371082, 371083, 371101, 371102, 371103, 371121, 371122, 371201, 371202, 371203, 371301, 371302,
              371311, 371312, 371321, 371322, 371323, 371324, 371325, 371326, 371327, 371328, 371329, 371401, 371402,
              371421, 371422, 371423, 371424, 371425, 371426, 371427, 371428, 371481, 371482, 371501, 371502, 371521,
              371522, 371523, 371524, 371525, 371526, 371581, 371601, 371602, 371621, 371622, 371623, 371624, 371625,
              371626, 371701, 371702, 371721, 371722, 371723, 371724, 371725, 371726, 371727, 371728, 410101, 410102,
              410103, 410104, 410105, 410106, 410108, 410122, 410181, 410182, 410183, 410184, 410185, 410201, 410202,
              410203, 410204, 410205, 410211, 410221, 410222, 410223, 410224, 410225, 410301, 410302, 410303, 410304,
              410305, 410306, 410311, 410322, 410323, 410324, 410325, 410326, 410327, 410328, 410329, 410381, 410401,
              410402, 410403, 410404, 410411, 410421, 410422, 410423, 410425, 410481, 410482, 410501, 410502, 410503,
              410505, 410506, 410522, 410523, 410526, 410527, 410581, 410601, 410602, 410603, 410611, 410621, 410622,
              410701, 410702, 410703, 410704, 410711, 410721, 410724, 410725, 410726, 410727, 410728, 410781, 410782,
              410801, 410802, 410803, 410804, 410811, 410821, 410822, 410823, 410825, 410882, 410883, 410901, 410902,
              410922, 410923, 410926, 410927, 410928, 411001, 411002, 411023, 411024, 411025, 411081, 411082, 411101,
              411102, 411103, 411104, 411121, 411122, 411201, 411202, 411221, 411222, 411224, 411281, 411282, 411301,
              411302, 411303, 411321, 411322, 411323, 411324, 411325, 411326, 411327, 411328, 411329, 411330, 411381,
              411401, 411402, 411403, 411421, 411422, 411423, 411424, 411425, 411426, 411481, 411501, 411502, 411503,
              411521, 411522, 411523, 411524, 411525, 411526, 411527, 411528, 411601, 411602, 411621, 411622, 411623,
              411624, 411625, 411626, 411627, 411628, 411681, 411701, 411702, 411721, 411722, 411723, 411724, 411725,
              411726, 411727, 411728, 411729, 419001, 420101, 420102, 420103, 420104, 420105, 420106, 420107, 420111,
              420112, 420113, 420114, 420115, 420116, 420117, 420201, 420202, 420203, 420204, 420205, 420222, 420281,
              420301, 420302, 420303, 420321, 420322, 420323, 420324, 420325, 420381, 420501, 420502, 420503, 420504,
              420505, 420506, 420525, 420526, 420527, 420528, 420529, 420581, 420582, 420583, 420601, 420602, 420606,
              420607, 420624, 420625, 420626, 420682, 420683, 420684, 420701, 420702, 420703, 420704, 420801, 420802,
              420804, 420821, 420822, 420881, 420901, 420902, 420921, 420922, 420923, 420981, 420982, 420984, 421001,
              421002, 421003, 421022, 421023, 421024, 421081, 421083, 421087, 421101, 421102, 421121, 421122, 421123,
              421124, 421125, 421126, 421127, 421181, 421182, 421201, 421202, 421221, 421222, 421223, 421224, 421281,
              421301, 421303, 421321, 421381, 422801, 422802, 422822, 422823, 422825, 422826, 422827, 422828, 429004,
              429005, 429006, 429021, 430101, 430102, 430103, 430104, 430105, 430111, 430112, 430121, 430124, 430181,
              430201, 430202, 430203, 430204, 430211, 430221, 430223, 430224, 430225, 430281, 430301, 430302, 430304,
              430321, 430381, 430382, 430401, 430405, 430406, 430407, 430408, 430412, 430421, 430422, 430423, 430424,
              430426, 430481, 430482, 430501, 430502, 430503, 430511, 430521, 430522, 430523, 430524, 430525, 430527,
              430528, 430529, 430581, 430601, 430602, 430603, 430611, 430621, 430623, 430624, 430626, 430681, 430682,
              430701, 430702, 430703, 430721, 430722, 430723, 430724, 430725, 430726, 430781, 430801, 430802, 430811,
              430821, 430822, 430901, 430902, 430903, 430921, 430922, 430923, 430981, 431001, 431002, 431003, 431021,
              431022, 431023, 431024, 431025, 431026, 431027, 431028, 431081, 431101, 431102, 431103, 431121, 431122,
              431123, 431124, 431125, 431126, 431127, 431128, 431129, 431201, 431202, 431221, 431222, 431223, 431224,
              431225, 431226, 431227, 431228, 431229, 431230, 431281, 431301, 431302, 431321, 431322, 431381, 431382,
              433101, 433122, 433123, 433124, 433125, 433126, 433127, 433130, 440101, 440103, 440104, 440105, 440106,
              440111, 440112, 440113, 440114, 440115, 440116, 440183, 440184, 440201, 440203, 440204, 440205, 440222,
              440224, 440229, 440232, 440233, 440281, 440282, 440301, 440303, 440304, 440305, 440306, 440307, 440308,
              440401, 440402, 440403, 440404, 440501, 440507, 440511, 440512, 440513, 440514, 440515, 440523, 440601,
              440604, 440605, 440606, 440607, 440608, 440701, 440703, 440704, 440705, 440781, 440783, 440784, 440785,
              440801, 440802, 440803, 440804, 440811, 440823, 440825, 440881, 440882, 440883, 440901, 440902, 440903,
              440923, 440981, 440982, 440983, 441201, 441202, 441203, 441223, 441224, 441225, 441226, 441283, 441284,
              441301, 441302, 441303, 441322, 441323, 441324, 441401, 441402, 441421, 441422, 441423, 441424, 441426,
              441427, 441481, 441501, 441502, 441521, 441523, 441581, 441601, 441602, 441621, 441622, 441623, 441624,
              441625, 441701, 441702, 441721, 441723, 441781, 441801, 441802, 441821, 441823, 441825, 441826, 441827,
              441881, 441882, 445101, 445102, 445121, 445122, 445201, 445202, 445221, 445222, 445224, 445281, 445301,
              445302, 445321, 445322, 445323, 445381, 450101, 450102, 450103, 450105, 450107, 450108, 450109, 450122,
              450123, 450124, 450125, 450126, 450127, 450201, 450202, 450203, 450204, 450205, 450221, 450222, 450223,
              450224, 450225, 450226, 450301, 450302, 450303, 450304, 450305, 450311, 450321, 450322, 450323, 450324,
              450325, 450326, 450327, 450328, 450329, 450330, 450331, 450332, 450401, 450403, 450404, 450405, 450421,
              450422, 450423, 450481, 450501, 450502, 450503, 450512, 450521, 450601, 450602, 450603, 450621, 450681,
              450701, 450702, 450703, 450721, 450722, 450801, 450802, 450803, 450804, 450821, 450881, 450901, 450902,
              450921, 450922, 450923, 450924, 450981, 451001, 451002, 451021, 451022, 451023, 451024, 451025, 451026,
              451027, 451028, 451029, 451030, 451031, 451101, 451102, 451121, 451122, 451123, 451201, 451202, 451221,
              451222, 451223, 451224, 451225, 451226, 451227, 451228, 451229, 451281, 451301, 451302, 451321, 451322,
              451323, 451324, 451381, 451401, 451402, 451421, 451422, 451423, 451424, 451425, 451481, 460101, 460105,
              460106, 460107, 460108, 460201, 460321, 460322, 460323, 469001, 469002, 469003, 469005, 469006, 469007,
              469021, 469022, 469023, 469024, 469025, 469026, 469027, 469028, 469029, 469030, 500101, 500102, 500103,
              500104, 500105, 500106, 500107, 500108, 500109, 500110, 500111, 500112, 500113, 500114, 500115, 500116,
              500117, 500118, 500119, 500223, 500224, 500226, 500227, 500228, 500229, 500230, 500231, 500232, 500233,
              500234, 500235, 500236, 500237, 500238, 500240, 500241, 500242, 500243, 510101, 510104, 510105, 510106,
              510107, 510108, 510112, 510113, 510114, 510115, 510121, 510122, 510124, 510129, 510131, 510132, 510181,
              510182, 510183, 510184, 510301, 510302, 510303, 510304, 510311, 510321, 510322, 510401, 510402, 510403,
              510411, 510421, 510422, 510501, 510502, 510503, 510504, 510521, 510522, 510524, 510525, 510601, 510603,
              510623, 510626, 510681, 510682, 510683, 510701, 510703, 510704, 510722, 510723, 510724, 510725, 510726,
              510727, 510781, 510801, 510802, 510811, 510812, 510821, 510822, 510823, 510824, 510901, 510903, 510904,
              510921, 510922, 510923, 511001, 511002, 511011, 511024, 511025, 511028, 511101, 511102, 511111, 511112,
              511113, 511123, 511124, 511126, 511129, 511132, 511133, 511181, 511301, 511302, 511303, 511304, 511321,
              511322, 511323, 511324, 511325, 511381, 511401, 511402, 511421, 511422, 511423, 511424, 511425, 511501,
              511502, 511503, 511521, 511523, 511524, 511525, 511526, 511527, 511528, 511529, 511601, 511602, 511621,
              511622, 511623, 511681, 511701, 511702, 511721, 511722, 511723, 511724, 511725, 511781, 511801, 511802,
              511803, 511822, 511823, 511824, 511825, 511826, 511827, 511901, 511902, 511921, 511922, 511923, 512001,
              512002, 512021, 512022, 512081, 513221, 513222, 513223, 513224, 513225, 513226, 513227, 513228, 513229,
              513230, 513231, 513232, 513233, 513321, 513322, 513323, 513324, 513325, 513326, 513327, 513328, 513329,
              513330, 513331, 513332, 513333, 513334, 513335, 513336, 513337, 513338, 513401, 513422, 513423, 513424,
              513425, 513426, 513427, 513428, 513429, 513430, 513431, 513432, 513433, 513434, 513435, 513436, 513437,
              520101, 520102, 520103, 520111, 520112, 520113, 520114, 520121, 520122, 520123, 520181, 520201, 520203,
              520221, 520222, 520301, 520302, 520303, 520321, 520322, 520323, 520324, 520325, 520326, 520327, 520328,
              520329, 520330, 520381, 520382, 520401, 520402, 520421, 520422, 520423, 520424, 520425, 520502, 520521,
              520522, 520523, 520524, 520525, 520526, 520527, 520602, 520603, 520621, 520622, 520623, 520624, 520625,
              520626, 520627, 520628, 522301, 522322, 522323, 522324, 522325, 522326, 522327, 522328, 522601, 522622,
              522623, 522624, 522625, 522626, 522627, 522628, 522629, 522630, 522631, 522632, 522633, 522634, 522635,
              522636, 522701, 522702, 522722, 522723, 522725, 522726, 522727, 522728, 522729, 522730, 522731, 522732,
              530101, 530102, 530103, 530111, 530112, 530113, 530114, 530122, 530124, 530125, 530126, 530127, 530128,
              530129, 530181, 530301, 530302, 530321, 530322, 530323, 530324, 530325, 530326, 530328, 530381, 530402,
              530421, 530422, 530423, 530424, 530425, 530426, 530427, 530428, 530501, 530502, 530521, 530522, 530523,
              530524, 530601, 530602, 530621, 530622, 530623, 530624, 530625, 530626, 530627, 530628, 530629, 530630,
              530701, 530702, 530721, 530722, 530723, 530724, 530801, 530802, 530821, 530822, 530823, 530824, 530825,
              530826, 530827, 530828, 530829, 530901, 530902, 530921, 530922, 530923, 530924, 530925, 530926, 530927,
              532301, 532322, 532323, 532324, 532325, 532326, 532327, 532328, 532329, 532331, 532501, 532502, 532503,
              532523, 532524, 532525, 532526, 532527, 532528, 532529, 532530, 532531, 532532, 532601, 532622, 532623,
              532624, 532625, 532626, 532627, 532628, 532801, 532822, 532823, 532901, 532922, 532923, 532924, 532925,
              532926, 532927, 532928, 532929, 532930, 532931, 532932, 533102, 533103, 533122, 533123, 533124, 533321,
              533323, 533324, 533325, 533421, 533422, 533423, 540101, 540102, 540121, 540122, 540123, 540124, 540125,
              540126, 540127, 542121, 542122, 542123, 542124, 542125, 542126, 542127, 542128, 542129, 542132, 542133,
              542221, 542222, 542223, 542224, 542225, 542226, 542227, 542228, 542229, 542231, 542232, 542233, 542301,
              542322, 542323, 542324, 542325, 542326, 542327, 542328, 542329, 542330, 542331, 542332, 542333, 542334,
              542335, 542336, 542337, 542338, 542421, 542422, 542423, 542424, 542425, 542426, 542427, 542428, 542429,
              542430, 542521, 542522, 542523, 542524, 542525, 542526, 542527, 542621, 542622, 542623, 542624, 542625,
              542626, 542627, 610101, 610102, 610103, 610104, 610111, 610112, 610113, 610114, 610115, 610116, 610122,
              610124, 610125, 610126, 610201, 610202, 610203, 610204, 610222, 610301, 610302, 610303, 610304, 610322,
              610323, 610324, 610326, 610327, 610328, 610329, 610330, 610331, 610401, 610402, 610403, 610404, 610422,
              610423, 610424, 610425, 610426, 610427, 610428, 610429, 610430, 610431, 610481, 610501, 610502, 610521,
              610522, 610523, 610524, 610525, 610526, 610527, 610528, 610581, 610582, 610601, 610602, 610621, 610622,
              610623, 610624, 610625, 610626, 610627, 610628, 610629, 610630, 610631, 610632, 610701, 610702, 610721,
              610722, 610723, 610724, 610725, 610726, 610727, 610728, 610729, 610730, 610801, 610802, 610821, 610822,
              610823, 610824, 610825, 610826, 610827, 610828, 610829, 610830, 610831, 610901, 610902, 610921, 610922,
              610923, 610924, 610925, 610926, 610927, 610928, 610929, 611001, 611002, 611021, 611022, 611023, 611024,
              611025, 611026, 620101, 620102, 620103, 620104, 620105, 620111, 620121, 620122, 620123, 620201, 620301,
              620302, 620321, 620401, 620402, 620403, 620421, 620422, 620423, 620501, 620502, 620503, 620521, 620522,
              620523, 620524, 620525, 620601, 620602, 620621, 620622, 620623, 620701, 620702, 620721, 620722, 620723,
              620724, 620725, 620801, 620802, 620821, 620822, 620823, 620824, 620825, 620826, 620901, 620902, 620921,
              620922, 620923, 620924, 620981, 620982, 621001, 621002, 621021, 621022, 621023, 621024, 621025, 621026,
              621027, 621101, 621102, 621121, 621122, 621123, 621124, 621125, 621126, 621201, 621202, 621221, 621222,
              621223, 621224, 621225, 621226, 621227, 621228, 622901, 622921, 622922, 622923, 622924, 622925, 622926,
              622927, 623001, 623021, 623022, 623023, 623024, 623025, 623026, 623027, 630101, 630102, 630103, 630104,
              630105, 630121, 630122, 630123, 632121, 632122, 632123, 632126, 632127, 632128, 632221, 632222, 632223,
              632224, 632321, 632322, 632323, 632324, 632521, 632522, 632523, 632524, 632525, 632621, 632622, 632623,
              632624, 632625, 632626, 632721, 632722, 632723, 632724, 632725, 632726, 632801, 632802, 632821, 632822,
              632823, 640101, 640104, 640105, 640106, 640121, 640122, 640181, 640201, 640202, 640205, 640221, 640301,
              640302, 640303, 640323, 640324, 640381, 640401, 640402, 640422, 640423, 640424, 640425, 640501, 640502,
              640521, 640522, 650101, 650102, 650103, 650104, 650105, 650106, 650107, 650109, 650121, 650201, 650202,
              650203, 650204, 650205, 652101, 652122, 652123, 652201, 652222, 652223, 652301, 652302, 652323, 652324,
              652325, 652327, 652328, 652701, 652722, 652723, 652801, 652822, 652823, 652824, 652825, 652826, 652827,
              652828, 652829, 652901, 652922, 652923, 652924, 652925, 652926, 652927, 652928, 652929, 653001, 653022,
              653023, 653024, 653101, 653121, 653122, 653123, 653124, 653125, 653126, 653127, 653128, 653129, 653130,
              653131, 653201, 653221, 653222, 653223, 653224, 653225, 653226, 653227, 654002, 654003, 654021, 654022,
              654023, 654024, 654025, 654026, 654027, 654028, 654201, 654202, 654221, 654223, 654224, 654225, 654226,
              654301, 654321, 654322, 654323, 654324, 654325, 654326, 659001, 659002, 659003, 659004]
    # 生成随机年份
    Birth_Y = random.randint(1950, 1996)
    # 生成随机月份
    Birth_M = random.randint(01, 12)
    # 生成随机区域码
    R_id = random.choice(Region)
    # 3位随机数
    T_id = random.randint(1, 999)
    # 生成随机日期
    if Birth_M in [1, 3, 5, 7, 8, 10, 12]:
        Birth_D = random.randint(01, 31)
    elif Birth_M in [4, 6, 9, 11]:
        Birth_D = random.randint(01, 30)
    else:
        if Birth_Y % 4 != 0:
            Birth_D = random.randint(01, 28)
        else:
            Birth_D = random.randint(01, 29)
    # 拼接8位出生日期
    DirthDate = str(Birth_Y) + "%02d" % (Birth_M) + "%02d" % (Birth_D)
    # 拼接身份证钱17位
    base = str(R_id) + DirthDate + str("%03d" % (T_id))

    d = {0: 1, 1: 0, 2: 'X', 3: 9, 4: 8, 5: 7, 6: 6, 7: 5, 8: 4, 9: 3, 10: 2}
    f = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
    sums = 0
    # 根据前17位数字算出最后一个字符
    for i in range(17):
        sums = sums + f[i] * int(base[i])
    mods = sums % 11
    # 拼接并返回18位身份证号
    CardID = base + str(d[mods])
    # print CardID
    return CardID


def runInit_Script():
    bizmod.RunInitScript()
    return 0, ""


def replace_account_parameter(cif_account):
    bizmod.replaceparameter(cif_account)
    return 0, ""


def get_kcbp_core(system_id, TEST_SYSTEM_ARR):  # system_id系统id
    try:
        if os.path.exists("settings.json"):
            _setting = json.loads(open("settings.json", "r").read())
            gl.g_testcaseDBDict["DBType"] = _setting["casedb"]["DBType"]
            gl.g_testcaseDBDict["casedb"] = _setting["casedb"][gl.g_testcaseDBDict["DBType"]]
        sql = ""
        if TEST_SYSTEM_ARR:
            TEST_SYSTEM_ARR_STR = ",".join(TEST_SYSTEM_ARR)
            # 表中的SYSTEM_ID  代表主键id   sys_id 代表系统id
            sql = "select system_id from sx_test_system where sys_id = '%s' and SYSTEM_ID in (%s)" % (
                system_id, TEST_SYSTEM_ARR_STR)
        else:
            sql = "select system_id from sx_test_system where sys_id = '%s'" % system_id
        ds = executeCaseSQL_DB2(sql)
        core_data = {}
        for i in ds:
            sys_test_id = i["SYSTEM_ID"]
            sql_info = "select SETTING_TYPE, SETTING_VALUE from sx_system_setting where sys_test_id =%s and " \
                       "setting_type in ('KCBP_IP', 'KCBP_CORE', 'KCBP_TYPE', 'KCBP_DESC','dbip')" % sys_test_id
            ds_2 = executeCaseSQL_DB2(sql_info)
            if ds_2:
                children_dict = {}
                for tmp in ds_2:
                    SETTING_TYPE = tmp['SETTING_TYPE']
                    SETTING_VALUE = tmp['SETTING_VALUE']
                    children_dict[SETTING_TYPE] = SETTING_VALUE
                if children_dict['KCBP_CORE'] == "" or children_dict['KCBP_CORE'] == "null" or children_dict[
                    'KCBP_CORE'] == None:
                    continue
                KCBP_CORE = str(children_dict['KCBP_CORE']).strip()
                core_data[str(KCBP_CORE)] = children_dict
        return 0, core_data
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(u"配置文件更新失败%s" % (exc_info))
        return -1, exc_info


def updateZHLCAutoTestSetting(system_id, runtask_id, TEST_SYSTEM_ID):
    try:
        root = ""
        core = 0
        core_flag = False
        sql_setting = "select SETTING_VALUE from sx_system_setting where sys_test_id = %d and setting_type = 'KCBP_CORE'" % (
            TEST_SYSTEM_ID)
        rs = executeCaseSQL(sql_setting)
        if rs:
            core = rs[0]["SETTING_VALUE"]
            if core == "" or core == "null" or core == None:
                Logging.getLog().debug(u"未找到 核心号")
                return 0, ''
            else:
                core_flag = True
                core = core.strip()
            if os.path.exists((".\\init\\%s\ZHLCAutoTestSetting_%s.xml") % (system_id, core)):
                os.remove((".\\init\\%s\ZHLCAutoTestSetting_%s.xml") % (system_id, core))
            if not os.path.exists(".\\init\\%s\ZHLCAutoTestSetting_%s.xml" % (system_id, core)):
                if not os.path.exists(r'.\init\%s' % system_id):
                    os.mkdir('init\%s' % system_id)
            root = etree.Element("ZHLCAutoTestSetting")
            etree.SubElement(root, "ConnectOracleSetting")
            etree.SubElement(root, "ZHLCSystemSetting")
            etree.SubElement(root, "ConnectSqlServerSetting")
            etree.SubElement(root, "TimeSyncServerSetting")
            etree.SubElement(root, "OtherSetting")
            doc = etree.ElementTree(root)
            bizmod.indent(root)
            path = ".\\init\\%s\\ZHLCAutoTestSetting_%s.xml" % (system_id, core)
            doc.write(path, pretty_print=True, xml_declaration=True, encoding='utf-8')
            doc = etree.ElementTree(file=".\\init\\%s\\ZHLCAutoTestSetting_%s.xml" % (system_id, core))
            root = doc.getroot()
        else:
            if os.path.exists(".\\init\\%s\ZHLCAutoTestSetting.xml" % system_id):
                os.remove(".\\init\\%s\ZHLCAutoTestSetting.xml" % system_id)
                # 读取文件
            if not os.path.exists(".\\init\\%s\ZHLCAutoTestSetting.xml" % system_id):
                if not os.path.exists(r'.\init\%s' % system_id):
                    os.mkdir('init\%s' % system_id)
            root = etree.Element("ZHLCAutoTestSetting")
            etree.SubElement(root, "ConnectOracleSetting")
            etree.SubElement(root, "ZHLCSystemSetting")
            etree.SubElement(root, "ConnectSqlServerSetting")
            etree.SubElement(root, "TimeSyncServerSetting")
            etree.SubElement(root, "OtherSetting")
            doc = etree.ElementTree(root)
            bizmod.indent(root)
            path = ".\\init\\%s\\ZHLCAutoTestSetting.xml" % (system_id)
            doc.write(path, pretty_print=True, xml_declaration=True, encoding='utf-8')
            doc = etree.ElementTree(file=".\\init\\%s\\ZHLCAutoTestSetting.xml" % (system_id))
            root = doc.getroot()
        sql = "select SETTING_TYPE, SETTING_VALUE from SX_SYSTEM_SETTING where SYSTEM_ID='%s' and SYS_TEST_ID=%s" % (
            system_id, TEST_SYSTEM_ID)
        rs = executeCaseSQL(sql)
        # print rs
        if rs:
            for i in rs:
                tagName = i["SETTING_TYPE"].strip()
                tagValue = i["SETTING_VALUE"].strip()
                # Logging.getLog().info(system_id)
                # Logging.getLog().info(tagName)
                # Logging.getLog().info(tagValue)
                # 获取节点信息
                if str(tagName) == "DBFDatabasePath":
                    xpath = "DBFDatabasePath"
                    DBFDatabasePath = root.find(xpath)
                    if DBFDatabasePath != None:
                        DBFDatabasePath.text = str(tagValue).replace('amp;amp;', 'amp;')
                    else:
                        dasd = etree.SubElement(root, str(tagName))
                        dasd.text = str(tagValue).replace('amp;amp;', 'amp;')
                    continue
                elif str(tagName) == "ETFCsvFilePath":
                    xpath = "ETFCsvFilePath"
                    ETFCsvFilePath = root.find(xpath)
                    if ETFCsvFilePath != None:
                        ETFCsvFilePath.text = str(tagValue).replace('amp;amp;', 'amp;')
                    else:
                        dasd = etree.SubElement(root, str(tagName))
                        dasd.text = str(tagValue).replace('amp;amp;', 'amp;')
                    continue
                else:
                    pass
                if str(tagName) == "TimeSyncServerSetting":
                    xpath = "TimeSyncServerSetting/server[@ip='%s']" % tagValue
                    TimeSyncServerSetting = root.find(xpath)
                    if TimeSyncServerSetting == None:
                        xpath = "TimeSyncServerSetting"
                        TimeSyncServerSetting = root.find(xpath)
                        dasd = etree.SubElement(TimeSyncServerSetting, "server", ip=str(tagValue), desc="")
                    continue

                xpath = "ZHLCSystemSetting"
                ZHLCSystemSettingInfo = root.find(xpath)
                dictlist = []
                for n in ZHLCSystemSettingInfo:
                    dictlist.append(n.tag)
                if tagName in dictlist:
                    for n in ZHLCSystemSettingInfo:
                        if n.tag == tagName:
                            n.text = tagValue.replace('amp;amp;', 'amp;')
                            break
                else:
                    dasd = etree.SubElement(ZHLCSystemSettingInfo, str(tagName))
                    dasd.text = tagValue.replace('amp;amp;', 'amp;')
        sql = "select PARAMETER from sx_sleep_time where SYSTEM_ID = '%s'" % (system_id)
        rs = executeCaseSQL(sql)
        CasesSleepTime_num = 0
        if rs:
            print u'案例间休眠时间xml正在进行 --> ', rs
            CasesSleepTime_num = rs[0]['PARAMETER']
        xpath = "OtherSetting"
        OtherSetting = root.find(xpath)
        CasesSleepTime = etree.SubElement(OtherSetting, "CasesSleepTime")
        CasesSleepTime.text = str(CasesSleepTime_num)
        sql = "select * from SX_INTERFACE_FAULT_TOLERANCE where SYSTEM_ID = '%s'" % (system_id)
        rs = executeCaseSQL(sql)
        print sql
        opr_data = []
        if rs:
            print u"正在进行xml容错处理生成 --> ", rs
            for i in rs:
                print i
                ele_dict = {}
                ele_dict['ERROR_CODE'] = i['ERROR_CODE']
                ele_dict['FUNCTION_NUMBER'] = i['FUNCTION_NUMBER']
                opr_data.append(ele_dict)
        xpath = "OtherSetting"
        ConnectSqlServerSetting = root.find(xpath)
        AllowErrors = etree.SubElement(ConnectSqlServerSetting, 'AllowErrors')
        for i in opr_data:
            fundid_error = etree.SubElement(AllowErrors, "fundid_error", funcid=i['FUNCTION_NUMBER'])
            error = etree.SubElement(fundid_error, "error")
            error.text = i['ERROR_CODE']
        # 更新oracle或sqlserver
        sql = "select * from SX_DATABASE_INFO " \
              "where SYSTEM_ID = '%s' and SYS_TEST_ID=%s" % (system_id, TEST_SYSTEM_ID)
        rs = executeCaseSQL(sql)
        print sql
        # print rs
        if rs:
            # 因为sqlserver核心号,只有集中交易才有，其他系统不涉及核心问题。防止核心号有冲突, 没有核心的，随机从100起。
            for i in rs:
                DB_TYPE = i["DB_TYPE"]
                if str(DB_TYPE) == "SQLServer":
                    tags = {}
                    if core_flag == False:
                        core = i['ID']
                    # if i['DB_ID'] + 1 == 0:
                    #     core = sql_srver_core
                    ip = i['DB_IP']
                    xpip = '0.0.0.0'
                    if "," in ip:
                        ips = ip.split(",")
                        ip = ips[0]
                        xpip = ips[1]
                    DB_PORT = i['DB_PORT']
                    # if system_id == "1428908551835S0NXP":
                    ip = ip + "," + DB_PORT
                    username = i['DB_USERNAME']
                    password = i['DB_PASSWORD']
                    database = i['DB_SCHEMA']
                    servertype = i['DB_SERVICETYPE']
                    xpath = "ConnectSqlServerSetting/server[@servertype='%s'][@core='%s']" % (servertype, core)
                    ConnectSqlServerSetting = root.find(xpath)
                    tags = {"xpip": xpip, "ip": ip, "username": username, "password": password, "database": database,
                            "servertype": servertype, "serverid": core}
                    if ConnectSqlServerSetting != None:
                        for tagName in tags:
                            dictlist = []
                            for n in ConnectSqlServerSetting:
                                dictlist.append(n.tag)
                            if tagName in dictlist:
                                for n in ConnectSqlServerSetting:
                                    if n.tag == tagName:
                                        n.text = str(tags[tagName]).replace('amp;amp;', 'amp;')
                                        break
                            else:
                                dasd = etree.SubElement(ConnectSqlServerSetting, tagName)
                                dasd.text = str(tags[tagName]).replace('amp;amp;', 'amp;')
                    else:
                        xpath = "ConnectSqlServerSetting"
                        ConnectSqlServerSetting = root.find(xpath)
                        server = etree.SubElement(ConnectSqlServerSetting, 'server', core=str(core),
                                                  servertype=str(servertype))
                        for tagName in tags:
                            dasd = etree.SubElement(server, tagName)
                            dasd.text = str(tags[tagName]).replace('amp;amp;', 'amp;')

                if str(DB_TYPE) == "Oracle":
                    # core = i['DB_ID']
                    if core_flag == False:
                        core = i['ID']
                    ip = i['DB_IP']
                    DB_PORT = i['DB_PORT']
                    # if i['DB_ID'] + 1 == 0:
                    #     core = core_count
                    UserName = i['DB_USERNAME']
                    Password = i['DB_PASSWORD']
                    UserList = i['DB_SCHEMA']
                    DB_ALIAS = i['DB_ALIAS']
                    ServiceName = i['DB_SERVICENAME']
                    servertype = i['DB_SERVICETYPE']
                    tags = {"UserName": UserName, "Password": Password, "UserList": UserList,
                            "ServiceName": ServiceName, "servertype": servertype, 'IP': ip, 'DB_PORT': DB_PORT,
                            "DB_ALIAS": DB_ALIAS}
                    xpath = "ConnectOracleSetting/server[@core='%s']" % (core)
                    ConnectOracleSetting = root.find(xpath)
                    if ConnectOracleSetting != None:
                        for tagName in tags:
                            dictlist = []
                            for n in ConnectOracleSetting:
                                dictlist.append(n.tag)
                            if tagName in dictlist:
                                for n in ConnectOracleSetting:
                                    if n.tag == tagName:
                                        n.text = str(tags[tagName]).replace('amp;amp;', 'amp;')
                                        break
                            else:
                                dasd = etree.SubElement(ConnectOracleSetting, tagName)
                                dasd.text = str(tags[tagName]).replace('amp;amp;', 'amp;')
                    else:
                        xpath = "ConnectOracleSetting"
                        ConnectOracleSetting = root.find(xpath)
                        server = etree.SubElement(ConnectOracleSetting, 'server', core=str(core))
                        for tagName in tags:
                            dasd = etree.SubElement(server, tagName)
                            dasd.text = str(tags[tagName]).replace('amp;amp;', 'amp;')

        homedir = os.getcwd()
        path = "%s/init/%s/" % (homedir, system_id)
        f = XFer(gl.gl_ftp_info['FTPServer']['IP'], gl.gl_ftp_info['FTPServer']['USER'], gl.gl_ftp_info['FTPServer']['PASSWORD'], '.', int(gl.gl_ftp_info['FTPServer']['PORT']))
        f.login()
        doc = etree.ElementTree(root)
        bizmod.indent(root)
        if core_flag == True:
            doc.write(".\\init\\%s\\ZHLCAutoTestSetting_%s.xml" % (system_id, core), pretty_print=True,
                      xml_declaration=True,
                      encoding='utf-8')
            strik = '%s/%s' % (path, "ZHLCAutoTestSetting_%s.xml" % (core))
            f.upload_file_ex(strik, './init/%s' % system_id)
        else:
            doc.write(".\\init\\%s\\ZHLCAutoTestSetting.xml" % (system_id), pretty_print=True, xml_declaration=True,
                      encoding='utf-8')

            strik = '%s/%s' % (path, "ZHLCAutoTestSetting.xml")
            f.upload_file_ex(strik, './init/%s' % system_id)
        print 'ftp ok'
        return 0, ""
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(u"配置文件更新失败%s" % (exc_info))
        return -1, exc_info


def updateZHLCAutoTestSettingXml(system_id, runtask_id,
                                 TEST_SYSTEM_ARR):  # '1428908551835S0NXP', 20277，'' 如果是清算则会是(1, 2)这种形式
    try:
        sql_setting = "select SYS_TEST_ID from sx_system_setting where SYSTEM_ID = '%s' and setting_type = 'KCBP_CORE'" % (
            system_id)
        rs = executeCaseSQL(sql_setting)
        if rs:
            for i in rs:
                TEST_SYSTEM_ID = int(i["SYS_TEST_ID"])
                updateZHLCAutoTestSetting(system_id, runtask_id, TEST_SYSTEM_ID)
        else:
            if runtask_id:
                sql = "select TEST_SYSTEM_ARR, AGENT_ID from SX_CASES_COLLECTION_RUN_TASK where COLLECTION_RUN_TASK_ID=%s" % runtask_id
                rs = executeCaseSQL(sql)
                if rs:
                    TEST_SYSTEM_ID_Str = str(rs[0]["TEST_SYSTEM_ARR"])
                    if TEST_SYSTEM_ID_Str == "" or TEST_SYSTEM_ID_Str == None or TEST_SYSTEM_ID_Str == 'null':
                        Logging.getLog().critical(u"未找到 运行环境标识 TEST_SYSTEM_ID 无法更新配置文件")
                        return 0, ''
                    TEST_SYSTEM_ID_LIST = TEST_SYSTEM_ID_Str.split(",")
                    for TEST_SYSTEM_ID in TEST_SYSTEM_ID_LIST:
                        TEST_SYSTEM_ID = int(TEST_SYSTEM_ID)
                        updateZHLCAutoTestSetting(system_id, runtask_id, TEST_SYSTEM_ID)
                else:
                    Logging.getLog().critical(u"未找到 运行环境标识 TEST_SYSTEM_ID 无法更新配置文件")
                    return 0, ''
            else:
                TEST_SYSTEM_ID_LIST = TEST_SYSTEM_ARR.split(",")
                for TEST_SYSTEM_ID in TEST_SYSTEM_ID_LIST:
                    TEST_SYSTEM_ID = int(TEST_SYSTEM_ID)
                    updateZHLCAutoTestSetting(system_id, runtask_id, TEST_SYSTEM_ID)
    except Exception as e:
        Logging.getLog().critical(u"未找到 运行环境标识 TEST_SYSTEM_ID 无法更新配置文件")
        return 0, ''


def GlobalParameterReplace(cmdstring, account_group_id=-1):
    '''
    全局参数替换
    '''
    # Logging.getLog().debug("gl.g_account_id:%s" % gl.g_account_id)
    # Logging.getLog().debug("gl.g_accountDict:%s" % gl.g_accountDict)
    print cmdstring, account_group_id
    if gl.g_case_adapter_type != 'SPLX_SELENIUM':
        cmdstring = cmdstring.replace("$bfh$", '"')
    try:
        if cmdstring.find('@random_number@') >= 0:
            cmdstring = cmdstring.replace("@random_number@", str(int(time.time())))
        if gl.random_args:
            for key in gl.random_args:
                key1 = "@" + str(key) + "@"
                cmdstring = cmdstring.replace(key1, gl.random_args[key])
        if cmdstring.find('@openprice@') >= 0:
            stkcode = getStkcodeByString(cmdstring)
            price = getPriceByStkcode(stkcode)
            cmdstring = cmdstring.replace('@openprice@', str(price))
        # cmdstring = cmdstring.replace("@generate_idno@", getCardID())
        if len(gl.g_accountDict.keys()) <= 0:
            return cmdstring
        if account_group_id != -1:
            if str(account_group_id) in gl.lh_account_info.keys():
                for key in gl.lh_account_info[str(account_group_id)].keys():
                    cmdstring = cmdstring.replace(("@%s@" % key).upper(),
                                                  gl.lh_account_info[str(account_group_id)][key])  # fixme
                    cmdstring = cmdstring.replace(("@%s@" % key).lower(),
                                                  gl.lh_account_info[str(account_group_id)][key])  # fixme
        if account_group_id != -1:
            if str(account_group_id) in gl.g_accountDict.keys():
                for key in gl.g_accountDict[str(account_group_id)].keys():
                    if str(key).lower() == "datatime":
                        continue
                    args_key = "@%s@" % key
                    args_key = args_key.lower()
                    # print"cmdstring"
                    # print args_key
                    # print cmdstring
                    if args_key in cmdstring.lower():
                        cmdstring = cmdstring.replace(("@%s@" % key).upper(),
                                                      gl.g_accountDict[str(account_group_id)][key])  # fixme
                        cmdstring = cmdstring.replace(("@%s@" % key).lower(),
                                                      gl.g_accountDict[str(account_group_id)][key])  # fixme
        if cmdstring.find('@datetime@') >= 0:
            datetime = time.strftime('%Y%m%d', time.localtime())
            cmdstring = cmdstring.replace('@datetime@', str(datetime))
        Blicence = bizmod.getBlicence('0')[1][0]
        Orgcode = bizmod.getOrgcode('0')[1][0]
        Id_no = getCardID()
        Birthday = Id_no[6:14]
        return cmdstring.replace("@generate_idno@", Id_no).replace("@birthday@", Birthday).replace(
            "@generate_business_licence@", Blicence).replace("@generate_organ_code@", Orgcode)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1
        # Blicence=bizmod.getBlicence('0')[1][0]
        # Orgcode=bizmod.getOrgcode('0')[1][0]
        # return cmdstring.replace("@generate_idno@",getCardID()).replace("@generate_business_licence@",Blicence).replace("@generate_organ_code@",Orgcode)

        # return cmdstring    \
        #    .replace("@id_no@",gl.g_accountDict["id_no"]) \
        #    .replace("@cif_account@",gl.g_accountDict["cif_account"])\
        #    .replace("@password@",gl.g_accountDict["password"]) \
        #    .replace("@busin_account@",gl.g_accountDict["busin_account"]) \
        #    .replace("@client_id@",gl.g_accountDict["client_id"]) \
        #    .replace("@branch_no@",gl.g_accountDict["branch_no"]) \
        #    .replace("@mobile_tel@",gl.g_accountDict["mobile_tel"]) \
        #    .replace("@fund_password@",gl.g_accountDict["fund_password"]) \
        #    .replace("@client_name@",gl.g_accountDict["client_name"]) \
        #    .replace("@Cpassword@",gl.g_accountDict["Cpassword"]) \
        #    .replace("@paypwd@",gl.g_accountDict["paypwd"]) \
        #    .replace("@sh_stock_account@",gl.g_accountDict["sh_stock_account"]) \
        #    .replace("@sz_stock_account@",gl.g_accountDict["sz_stock_account"]) \
        #    .replace("@generate_idno@",getCardID())


unique_id = 1


def walkData(root_node, level, result_list, result_dict):
    global unique_id

    if unique_id != 1:
        result_dict[root_node.attrib["itempos"]] = root_node.attrib["desc"]
    children_node = root_node.getchildren()
    if len(children_node) == 0:
        temp_list = [unique_id, level, root_node.tag, root_node.attrib, 1]
        result_list.append(temp_list)
        unique_id += 1
        return
    else:
        temp_list = [unique_id, level, root_node.tag, root_node.attrib, 0]
        result_list.append(temp_list)
        unique_id += 1
    for child in children_node:
        walkData(child, level + 1, result_list, result_dict)
    return


def getXmlData(file_name):
    level = 0
    result_list = []
    result_dict = {}
    root = ET.parse(file_name).getroot()
    walkData(root, level, result_list, result_dict)
    return result_list, result_dict


def getCardID():
    # 区域码
    Region = [110101, 110102, 110105, 110106, 110107, 110108, 110109, 110111, 110112, 110113, 110114, 110115, 110116,
              110117, 110228, 110229, 120101, 120102, 120103, 120104, 120105, 120106, 120110, 120111, 120112, 120113,
              120114, 120115, 120116, 120221, 120223, 120225, 130101, 130102, 130103, 130104, 130105, 130107, 130108,
              130121, 130123, 130124, 130125, 130126, 130127, 130128, 130129, 130130, 130131, 130132, 130133, 130181,
              130182, 130183, 130184, 130185, 130201, 130202, 130203, 130204, 130205, 130207, 130208, 130209, 130223,
              130224, 130225, 130227, 130229, 130281, 130283, 130301, 130302, 130303, 130304, 130321, 130322, 130323,
              130324, 130401, 130402, 130403, 130404, 130406, 130421, 130423, 130424, 130425, 130426, 130427, 130428,
              130429, 130430, 130431, 130432, 130433, 130434, 130435, 130481, 130501, 130502, 130503, 130521, 130522,
              130523, 130524, 130525, 130526, 130527, 130528, 130529, 130530, 130531, 130532, 130533, 130534, 130535,
              130581, 130582, 130601, 130602, 130603, 130604, 130621, 130622, 130623, 130624, 130625, 130626, 130627,
              130628, 130629, 130630, 130631, 130632, 130633, 130634, 130635, 130636, 130637, 130638, 130681, 130682,
              130683, 130684, 130701, 130702, 130703, 130705, 130706, 130721, 130722, 130723, 130724, 130725, 130726,
              130727, 130728, 130729, 130730, 130731, 130732, 130733, 130801, 130802, 130803, 130804, 130821, 130822,
              130823, 130824, 130825, 130826, 130827, 130828, 130901, 130902, 130903, 130921, 130922, 130923, 130924,
              130925, 130926, 130927, 130928, 130929, 130930, 130981, 130982, 130983, 130984, 131001, 131002, 131003,
              131022, 131023, 131024, 131025, 131026, 131028, 131081, 131082, 131101, 131102, 131121, 131122, 131123,
              131124, 131125, 131126, 131127, 131128, 131181, 131182, 140101, 140105, 140106, 140107, 140108, 140109,
              140110, 140121, 140122, 140123, 140181, 140201, 140202, 140203, 140211, 140212, 140221, 140222, 140223,
              140224, 140225, 140226, 140227, 140301, 140302, 140303, 140311, 140321, 140322, 140401, 140402, 140411,
              140421, 140423, 140424, 140425, 140426, 140427, 140428, 140429, 140430, 140431, 140481, 140501, 140502,
              140521, 140522, 140524, 140525, 140581, 140601, 140602, 140603, 140621, 140622, 140623, 140624, 140701,
              140702, 140721, 140722, 140723, 140724, 140725, 140726, 140727, 140728, 140729, 140781, 140801, 140802,
              140821, 140822, 140823, 140824, 140825, 140826, 140827, 140828, 140829, 140830, 140881, 140882, 140901,
              140902, 140921, 140922, 140923, 140924, 140925, 140926, 140927, 140928, 140929, 140930, 140931, 140932,
              140981, 141001, 141002, 141021, 141022, 141023, 141024, 141025, 141026, 141027, 141028, 141029, 141030,
              141031, 141032, 141033, 141034, 141081, 141082, 141101, 141102, 141121, 141122, 141123, 141124, 141125,
              141126, 141127, 141128, 141129, 141130, 141181, 141182, 150101, 150102, 150103, 150104, 150105, 150121,
              150122, 150123, 150124, 150125, 150201, 150202, 150203, 150204, 150205, 150206, 150207, 150221, 150222,
              150223, 150301, 150302, 150303, 150304, 150401, 150402, 150403, 150404, 150421, 150422, 150423, 150424,
              150425, 150426, 150428, 150429, 150430, 150501, 150502, 150521, 150522, 150523, 150524, 150525, 150526,
              150581, 150601, 150602, 150621, 150622, 150623, 150624, 150625, 150626, 150627, 150701, 150702, 150721,
              150722, 150723, 150724, 150725, 150726, 150727, 150781, 150782, 150783, 150784, 150785, 150801, 150802,
              150821, 150822, 150823, 150824, 150825, 150826, 150901, 150902, 150921, 150922, 150923, 150924, 150925,
              150926, 150927, 150928, 150929, 150981, 152201, 152202, 152221, 152222, 152223, 152224, 152501, 152502,
              152522, 152523, 152524, 152525, 152526, 152527, 152528, 152529, 152530, 152531, 152921, 152922, 152923,
              210101, 210102, 210103, 210104, 210105, 210106, 210111, 210112, 210113, 210114, 210122, 210123, 210124,
              210181, 210201, 210202, 210203, 210204, 210211, 210212, 210213, 210224, 210281, 210282, 210283, 210301,
              210302, 210303, 210304, 210311, 210321, 210323, 210381, 210401, 210402, 210403, 210404, 210411, 210421,
              210422, 210423, 210501, 210502, 210503, 210504, 210505, 210521, 210522, 210601, 210602, 210603, 210604,
              210624, 210681, 210682, 210701, 210702, 210703, 210711, 210726, 210727, 210781, 210782, 210801, 210802,
              210803, 210804, 210811, 210881, 210882, 210901, 210902, 210903, 210904, 210905, 210911, 210921, 210922,
              211001, 211002, 211003, 211004, 211005, 211011, 211021, 211081, 211101, 211102, 211103, 211121, 211122,
              211201, 211202, 211204, 211221, 211223, 211224, 211281, 211282, 211301, 211302, 211303, 211321, 211322,
              211324, 211381, 211382, 211401, 211402, 211403, 211404, 211421, 211422, 211481, 220101, 220102, 220103,
              220104, 220105, 220106, 220112, 220122, 220181, 220182, 220183, 220201, 220202, 220203, 220204, 220211,
              220221, 220281, 220282, 220283, 220284, 220301, 220302, 220303, 220322, 220323, 220381, 220382, 220401,
              220402, 220403, 220421, 220422, 220501, 220502, 220503, 220521, 220523, 220524, 220581, 220582, 220601,
              220602, 220605, 220621, 220622, 220623, 220681, 220701, 220702, 220721, 220722, 220723, 220724, 220801,
              220802, 220821, 220822, 220881, 220882, 222401, 222402, 222403, 222404, 222405, 222406, 222424, 222426,
              230101, 230102, 230103, 230104, 230108, 230109, 230110, 230111, 230112, 230123, 230124, 230125, 230126,
              230127, 230128, 230129, 230182, 230183, 230184, 230201, 230202, 230203, 230204, 230205, 230206, 230207,
              230208, 230221, 230223, 230224, 230225, 230227, 230229, 230230, 230231, 230281, 230301, 230302, 230303,
              230304, 230305, 230306, 230307, 230321, 230381, 230382, 230401, 230402, 230403, 230404, 230405, 230406,
              230407, 230421, 230422, 230501, 230502, 230503, 230505, 230506, 230521, 230522, 230523, 230524, 230601,
              230602, 230603, 230604, 230605, 230606, 230621, 230622, 230623, 230624, 230701, 230702, 230703, 230704,
              230705, 230706, 230707, 230708, 230709, 230710, 230711, 230712, 230713, 230714, 230715, 230716, 230722,
              230781, 230801, 230803, 230804, 230805, 230811, 230822, 230826, 230828, 230833, 230881, 230882, 230901,
              230902, 230903, 230904, 230921, 231001, 231002, 231003, 231004, 231005, 231024, 231025, 231081, 231083,
              231084, 231085, 231101, 231102, 231121, 231123, 231124, 231181, 231182, 231201, 231202, 231221, 231222,
              231223, 231224, 231225, 231226, 231281, 231282, 231283, 232721, 232722, 232723, 310101, 310104, 310105,
              310106, 310107, 310108, 310109, 310110, 310112, 310113, 310114, 310115, 310116, 310117, 310118, 310120,
              310230, 320101, 320102, 320103, 320104, 320105, 320106, 320107, 320111, 320113, 320114, 320115, 320116,
              320124, 320125, 320201, 320202, 320203, 320204, 320205, 320206, 320211, 320281, 320282, 320301, 320302,
              320303, 320305, 320311, 320312, 320321, 320322, 320324, 320381, 320382, 320401, 320402, 320404, 320405,
              320411, 320412, 320481, 320482, 320501, 320505, 320506, 320507, 320508, 320509, 320581, 320582, 320583,
              320585, 320601, 320602, 320611, 320612, 320621, 320623, 320681, 320682, 320684, 320701, 320703, 320705,
              320706, 320721, 320722, 320723, 320724, 320801, 320802, 320803, 320804, 320811, 320826, 320829, 320830,
              320831, 320901, 320902, 320903, 320921, 320922, 320923, 320924, 320925, 320981, 320982, 321001, 321002,
              321003, 321012, 321023, 321081, 321084, 321101, 321102, 321111, 321112, 321181, 321182, 321183, 321201,
              321202, 321203, 321281, 321282, 321283, 321284, 321301, 321302, 321311, 321322, 321323, 321324, 330101,
              330102, 330103, 330104, 330105, 330106, 330108, 330109, 330110, 330122, 330127, 330182, 330183, 330185,
              330201, 330203, 330204, 330205, 330206, 330211, 330212, 330225, 330226, 330281, 330282, 330283, 330301,
              330302, 330303, 330304, 330322, 330324, 330326, 330327, 330328, 330329, 330381, 330382, 330401, 330402,
              330411, 330421, 330424, 330481, 330482, 330483, 330501, 330502, 330503, 330521, 330522, 330523, 330601,
              330602, 330621, 330624, 330681, 330682, 330683, 330701, 330702, 330703, 330723, 330726, 330727, 330781,
              330782, 330783, 330784, 330801, 330802, 330803, 330822, 330824, 330825, 330881, 330901, 330902, 330903,
              330921, 330922, 331001, 331002, 331003, 331004, 331021, 331022, 331023, 331024, 331081, 331082, 331101,
              331102, 331121, 331122, 331123, 331124, 331125, 331126, 331127, 331181, 340101, 340102, 340103, 340104,
              340111, 340121, 340122, 340123, 340124, 340181, 340201, 340202, 340203, 340207, 340208, 340221, 340222,
              340223, 340225, 340301, 340302, 340303, 340304, 340311, 340321, 340322, 340323, 340401, 340402, 340403,
              340404, 340405, 340406, 340421, 340501, 340503, 340504, 340506, 340521, 340522, 340523, 340601, 340602,
              340603, 340604, 340621, 340701, 340702, 340703, 340711, 340721, 340801, 340802, 340803, 340811, 340822,
              340823, 340824, 340825, 340826, 340827, 340828, 340881, 341001, 341002, 341003, 341004, 341021, 341022,
              341023, 341024, 341101, 341102, 341103, 341122, 341124, 341125, 341126, 341181, 341182, 341201, 341202,
              341203, 341204, 341221, 341222, 341225, 341226, 341282, 341301, 341302, 341321, 341322, 341323, 341324,
              341501, 341502, 341503, 341521, 341522, 341523, 341524, 341525, 341601, 341602, 341621, 341622, 341623,
              341701, 341702, 341721, 341722, 341723, 341801, 341802, 341821, 341822, 341823, 341824, 341825, 341881,
              350101, 350102, 350103, 350104, 350105, 350111, 350121, 350122, 350123, 350124, 350125, 350128, 350181,
              350182, 350201, 350203, 350205, 350206, 350211, 350212, 350213, 350301, 350302, 350303, 350304, 350305,
              350322, 350401, 350402, 350403, 350421, 350423, 350424, 350425, 350426, 350427, 350428, 350429, 350430,
              350481, 350501, 350502, 350503, 350504, 350505, 350521, 350524, 350525, 350526, 350527, 350581, 350582,
              350583, 350601, 350602, 350603, 350622, 350623, 350624, 350625, 350626, 350627, 350628, 350629, 350681,
              350701, 350702, 350721, 350722, 350723, 350724, 350725, 350781, 350782, 350783, 350784, 350801, 350802,
              350821, 350822, 350823, 350824, 350825, 350881, 350901, 350902, 350921, 350922, 350923, 350924, 350925,
              350926, 350981, 350982, 360101, 360102, 360103, 360104, 360105, 360111, 360121, 360122, 360123, 360124,
              360201, 360202, 360203, 360222, 360281, 360301, 360302, 360313, 360321, 360322, 360323, 360401, 360402,
              360403, 360421, 360423, 360424, 360425, 360426, 360427, 360428, 360429, 360430, 360481, 360482, 360501,
              360502, 360521, 360601, 360602, 360622, 360681, 360701, 360702, 360721, 360722, 360723, 360724, 360725,
              360726, 360727, 360728, 360729, 360730, 360731, 360732, 360733, 360734, 360735, 360781, 360782, 360801,
              360802, 360803, 360821, 360822, 360823, 360824, 360825, 360826, 360827, 360828, 360829, 360830, 360881,
              360901, 360902, 360921, 360922, 360923, 360924, 360925, 360926, 360981, 360982, 360983, 361001, 361002,
              361021, 361022, 361023, 361024, 361025, 361026, 361027, 361028, 361029, 361030, 361101, 361102, 361121,
              361122, 361123, 361124, 361125, 361126, 361127, 361128, 361129, 361130, 361181, 370101, 370102, 370103,
              370104, 370105, 370112, 370113, 370124, 370125, 370126, 370181, 370201, 370202, 370203, 370205, 370211,
              370212, 370213, 370214, 370281, 370282, 370283, 370284, 370285, 370301, 370302, 370303, 370304, 370305,
              370306, 370321, 370322, 370323, 370401, 370402, 370403, 370404, 370405, 370406, 370481, 370501, 370502,
              370503, 370521, 370522, 370523, 370601, 370602, 370611, 370612, 370613, 370634, 370681, 370682, 370683,
              370684, 370685, 370686, 370687, 370701, 370702, 370703, 370704, 370705, 370724, 370725, 370781, 370782,
              370783, 370784, 370785, 370786, 370801, 370802, 370811, 370826, 370827, 370828, 370829, 370830, 370831,
              370832, 370881, 370882, 370883, 370901, 370902, 370911, 370921, 370923, 370982, 370983, 371001, 371002,
              371081, 371082, 371083, 371101, 371102, 371103, 371121, 371122, 371201, 371202, 371203, 371301, 371302,
              371311, 371312, 371321, 371322, 371323, 371324, 371325, 371326, 371327, 371328, 371329, 371401, 371402,
              371421, 371422, 371423, 371424, 371425, 371426, 371427, 371428, 371481, 371482, 371501, 371502, 371521,
              371522, 371523, 371524, 371525, 371526, 371581, 371601, 371602, 371621, 371622, 371623, 371624, 371625,
              371626, 371701, 371702, 371721, 371722, 371723, 371724, 371725, 371726, 371727, 371728, 410101, 410102,
              410103, 410104, 410105, 410106, 410108, 410122, 410181, 410182, 410183, 410184, 410185, 410201, 410202,
              410203, 410204, 410205, 410211, 410221, 410222, 410223, 410224, 410225, 410301, 410302, 410303, 410304,
              410305, 410306, 410311, 410322, 410323, 410324, 410325, 410326, 410327, 410328, 410329, 410381, 410401,
              410402, 410403, 410404, 410411, 410421, 410422, 410423, 410425, 410481, 410482, 410501, 410502, 410503,
              410505, 410506, 410522, 410523, 410526, 410527, 410581, 410601, 410602, 410603, 410611, 410621, 410622,
              410701, 410702, 410703, 410704, 410711, 410721, 410724, 410725, 410726, 410727, 410728, 410781, 410782,
              410801, 410802, 410803, 410804, 410811, 410821, 410822, 410823, 410825, 410882, 410883, 410901, 410902,
              410922, 410923, 410926, 410927, 410928, 411001, 411002, 411023, 411024, 411025, 411081, 411082, 411101,
              411102, 411103, 411104, 411121, 411122, 411201, 411202, 411221, 411222, 411224, 411281, 411282, 411301,
              411302, 411303, 411321, 411322, 411323, 411324, 411325, 411326, 411327, 411328, 411329, 411330, 411381,
              411401, 411402, 411403, 411421, 411422, 411423, 411424, 411425, 411426, 411481, 411501, 411502, 411503,
              411521, 411522, 411523, 411524, 411525, 411526, 411527, 411528, 411601, 411602, 411621, 411622, 411623,
              411624, 411625, 411626, 411627, 411628, 411681, 411701, 411702, 411721, 411722, 411723, 411724, 411725,
              411726, 411727, 411728, 411729, 419001, 420101, 420102, 420103, 420104, 420105, 420106, 420107, 420111,
              420112, 420113, 420114, 420115, 420116, 420117, 420201, 420202, 420203, 420204, 420205, 420222, 420281,
              420301, 420302, 420303, 420321, 420322, 420323, 420324, 420325, 420381, 420501, 420502, 420503, 420504,
              420505, 420506, 420525, 420526, 420527, 420528, 420529, 420581, 420582, 420583, 420601, 420602, 420606,
              420607, 420624, 420625, 420626, 420682, 420683, 420684, 420701, 420702, 420703, 420704, 420801, 420802,
              420804, 420821, 420822, 420881, 420901, 420902, 420921, 420922, 420923, 420981, 420982, 420984, 421001,
              421002, 421003, 421022, 421023, 421024, 421081, 421083, 421087, 421101, 421102, 421121, 421122, 421123,
              421124, 421125, 421126, 421127, 421181, 421182, 421201, 421202, 421221, 421222, 421223, 421224, 421281,
              421301, 421303, 421321, 421381, 422801, 422802, 422822, 422823, 422825, 422826, 422827, 422828, 429004,
              429005, 429006, 429021, 430101, 430102, 430103, 430104, 430105, 430111, 430112, 430121, 430124, 430181,
              430201, 430202, 430203, 430204, 430211, 430221, 430223, 430224, 430225, 430281, 430301, 430302, 430304,
              430321, 430381, 430382, 430401, 430405, 430406, 430407, 430408, 430412, 430421, 430422, 430423, 430424,
              430426, 430481, 430482, 430501, 430502, 430503, 430511, 430521, 430522, 430523, 430524, 430525, 430527,
              430528, 430529, 430581, 430601, 430602, 430603, 430611, 430621, 430623, 430624, 430626, 430681, 430682,
              430701, 430702, 430703, 430721, 430722, 430723, 430724, 430725, 430726, 430781, 430801, 430802, 430811,
              430821, 430822, 430901, 430902, 430903, 430921, 430922, 430923, 430981, 431001, 431002, 431003, 431021,
              431022, 431023, 431024, 431025, 431026, 431027, 431028, 431081, 431101, 431102, 431103, 431121, 431122,
              431123, 431124, 431125, 431126, 431127, 431128, 431129, 431201, 431202, 431221, 431222, 431223, 431224,
              431225, 431226, 431227, 431228, 431229, 431230, 431281, 431301, 431302, 431321, 431322, 431381, 431382,
              433101, 433122, 433123, 433124, 433125, 433126, 433127, 433130, 440101, 440103, 440104, 440105, 440106,
              440111, 440112, 440113, 440114, 440115, 440116, 440183, 440184, 440201, 440203, 440204, 440205, 440222,
              440224, 440229, 440232, 440233, 440281, 440282, 440301, 440303, 440304, 440305, 440306, 440307, 440308,
              440401, 440402, 440403, 440404, 440501, 440507, 440511, 440512, 440513, 440514, 440515, 440523, 440601,
              440604, 440605, 440606, 440607, 440608, 440701, 440703, 440704, 440705, 440781, 440783, 440784, 440785,
              440801, 440802, 440803, 440804, 440811, 440823, 440825, 440881, 440882, 440883, 440901, 440902, 440903,
              440923, 440981, 440982, 440983, 441201, 441202, 441203, 441223, 441224, 441225, 441226, 441283, 441284,
              441301, 441302, 441303, 441322, 441323, 441324, 441401, 441402, 441421, 441422, 441423, 441424, 441426,
              441427, 441481, 441501, 441502, 441521, 441523, 441581, 441601, 441602, 441621, 441622, 441623, 441624,
              441625, 441701, 441702, 441721, 441723, 441781, 441801, 441802, 441821, 441823, 441825, 441826, 441827,
              441881, 441882, 445101, 445102, 445121, 445122, 445201, 445202, 445221, 445222, 445224, 445281, 445301,
              445302, 445321, 445322, 445323, 445381, 450101, 450102, 450103, 450105, 450107, 450108, 450109, 450122,
              450123, 450124, 450125, 450126, 450127, 450201, 450202, 450203, 450204, 450205, 450221, 450222, 450223,
              450224, 450225, 450226, 450301, 450302, 450303, 450304, 450305, 450311, 450321, 450322, 450323, 450324,
              450325, 450326, 450327, 450328, 450329, 450330, 450331, 450332, 450401, 450403, 450404, 450405, 450421,
              450422, 450423, 450481, 450501, 450502, 450503, 450512, 450521, 450601, 450602, 450603, 450621, 450681,
              450701, 450702, 450703, 450721, 450722, 450801, 450802, 450803, 450804, 450821, 450881, 450901, 450902,
              450921, 450922, 450923, 450924, 450981, 451001, 451002, 451021, 451022, 451023, 451024, 451025, 451026,
              451027, 451028, 451029, 451030, 451031, 451101, 451102, 451121, 451122, 451123, 451201, 451202, 451221,
              451222, 451223, 451224, 451225, 451226, 451227, 451228, 451229, 451281, 451301, 451302, 451321, 451322,
              451323, 451324, 451381, 451401, 451402, 451421, 451422, 451423, 451424, 451425, 451481, 460101, 460105,
              460106, 460107, 460108, 460201, 460321, 460322, 460323, 469001, 469002, 469003, 469005, 469006, 469007,
              469021, 469022, 469023, 469024, 469025, 469026, 469027, 469028, 469029, 469030, 500101, 500102, 500103,
              500104, 500105, 500106, 500107, 500108, 500109, 500110, 500111, 500112, 500113, 500114, 500115, 500116,
              500117, 500118, 500119, 500223, 500224, 500226, 500227, 500228, 500229, 500230, 500231, 500232, 500233,
              500234, 500235, 500236, 500237, 500238, 500240, 500241, 500242, 500243, 510101, 510104, 510105, 510106,
              510107, 510108, 510112, 510113, 510114, 510115, 510121, 510122, 510124, 510129, 510131, 510132, 510181,
              510182, 510183, 510184, 510301, 510302, 510303, 510304, 510311, 510321, 510322, 510401, 510402, 510403,
              510411, 510421, 510422, 510501, 510502, 510503, 510504, 510521, 510522, 510524, 510525, 510601, 510603,
              510623, 510626, 510681, 510682, 510683, 510701, 510703, 510704, 510722, 510723, 510724, 510725, 510726,
              510727, 510781, 510801, 510802, 510811, 510812, 510821, 510822, 510823, 510824, 510901, 510903, 510904,
              510921, 510922, 510923, 511001, 511002, 511011, 511024, 511025, 511028, 511101, 511102, 511111, 511112,
              511113, 511123, 511124, 511126, 511129, 511132, 511133, 511181, 511301, 511302, 511303, 511304, 511321,
              511322, 511323, 511324, 511325, 511381, 511401, 511402, 511421, 511422, 511423, 511424, 511425, 511501,
              511502, 511503, 511521, 511523, 511524, 511525, 511526, 511527, 511528, 511529, 511601, 511602, 511621,
              511622, 511623, 511681, 511701, 511702, 511721, 511722, 511723, 511724, 511725, 511781, 511801, 511802,
              511803, 511822, 511823, 511824, 511825, 511826, 511827, 511901, 511902, 511921, 511922, 511923, 512001,
              512002, 512021, 512022, 512081, 513221, 513222, 513223, 513224, 513225, 513226, 513227, 513228, 513229,
              513230, 513231, 513232, 513233, 513321, 513322, 513323, 513324, 513325, 513326, 513327, 513328, 513329,
              513330, 513331, 513332, 513333, 513334, 513335, 513336, 513337, 513338, 513401, 513422, 513423, 513424,
              513425, 513426, 513427, 513428, 513429, 513430, 513431, 513432, 513433, 513434, 513435, 513436, 513437,
              520101, 520102, 520103, 520111, 520112, 520113, 520114, 520121, 520122, 520123, 520181, 520201, 520203,
              520221, 520222, 520301, 520302, 520303, 520321, 520322, 520323, 520324, 520325, 520326, 520327, 520328,
              520329, 520330, 520381, 520382, 520401, 520402, 520421, 520422, 520423, 520424, 520425, 520502, 520521,
              520522, 520523, 520524, 520525, 520526, 520527, 520602, 520603, 520621, 520622, 520623, 520624, 520625,
              520626, 520627, 520628, 522301, 522322, 522323, 522324, 522325, 522326, 522327, 522328, 522601, 522622,
              522623, 522624, 522625, 522626, 522627, 522628, 522629, 522630, 522631, 522632, 522633, 522634, 522635,
              522636, 522701, 522702, 522722, 522723, 522725, 522726, 522727, 522728, 522729, 522730, 522731, 522732,
              530101, 530102, 530103, 530111, 530112, 530113, 530114, 530122, 530124, 530125, 530126, 530127, 530128,
              530129, 530181, 530301, 530302, 530321, 530322, 530323, 530324, 530325, 530326, 530328, 530381, 530402,
              530421, 530422, 530423, 530424, 530425, 530426, 530427, 530428, 530501, 530502, 530521, 530522, 530523,
              530524, 530601, 530602, 530621, 530622, 530623, 530624, 530625, 530626, 530627, 530628, 530629, 530630,
              530701, 530702, 530721, 530722, 530723, 530724, 530801, 530802, 530821, 530822, 530823, 530824, 530825,
              530826, 530827, 530828, 530829, 530901, 530902, 530921, 530922, 530923, 530924, 530925, 530926, 530927,
              532301, 532322, 532323, 532324, 532325, 532326, 532327, 532328, 532329, 532331, 532501, 532502, 532503,
              532523, 532524, 532525, 532526, 532527, 532528, 532529, 532530, 532531, 532532, 532601, 532622, 532623,
              532624, 532625, 532626, 532627, 532628, 532801, 532822, 532823, 532901, 532922, 532923, 532924, 532925,
              532926, 532927, 532928, 532929, 532930, 532931, 532932, 533102, 533103, 533122, 533123, 533124, 533321,
              533323, 533324, 533325, 533421, 533422, 533423, 540101, 540102, 540121, 540122, 540123, 540124, 540125,
              540126, 540127, 542121, 542122, 542123, 542124, 542125, 542126, 542127, 542128, 542129, 542132, 542133,
              542221, 542222, 542223, 542224, 542225, 542226, 542227, 542228, 542229, 542231, 542232, 542233, 542301,
              542322, 542323, 542324, 542325, 542326, 542327, 542328, 542329, 542330, 542331, 542332, 542333, 542334,
              542335, 542336, 542337, 542338, 542421, 542422, 542423, 542424, 542425, 542426, 542427, 542428, 542429,
              542430, 542521, 542522, 542523, 542524, 542525, 542526, 542527, 542621, 542622, 542623, 542624, 542625,
              542626, 542627, 610101, 610102, 610103, 610104, 610111, 610112, 610113, 610114, 610115, 610116, 610122,
              610124, 610125, 610126, 610201, 610202, 610203, 610204, 610222, 610301, 610302, 610303, 610304, 610322,
              610323, 610324, 610326, 610327, 610328, 610329, 610330, 610331, 610401, 610402, 610403, 610404, 610422,
              610423, 610424, 610425, 610426, 610427, 610428, 610429, 610430, 610431, 610481, 610501, 610502, 610521,
              610522, 610523, 610524, 610525, 610526, 610527, 610528, 610581, 610582, 610601, 610602, 610621, 610622,
              610623, 610624, 610625, 610626, 610627, 610628, 610629, 610630, 610631, 610632, 610701, 610702, 610721,
              610722, 610723, 610724, 610725, 610726, 610727, 610728, 610729, 610730, 610801, 610802, 610821, 610822,
              610823, 610824, 610825, 610826, 610827, 610828, 610829, 610830, 610831, 610901, 610902, 610921, 610922,
              610923, 610924, 610925, 610926, 610927, 610928, 610929, 611001, 611002, 611021, 611022, 611023, 611024,
              611025, 611026, 620101, 620102, 620103, 620104, 620105, 620111, 620121, 620122, 620123, 620201, 620301,
              620302, 620321, 620401, 620402, 620403, 620421, 620422, 620423, 620501, 620502, 620503, 620521, 620522,
              620523, 620524, 620525, 620601, 620602, 620621, 620622, 620623, 620701, 620702, 620721, 620722, 620723,
              620724, 620725, 620801, 620802, 620821, 620822, 620823, 620824, 620825, 620826, 620901, 620902, 620921,
              620922, 620923, 620924, 620981, 620982, 621001, 621002, 621021, 621022, 621023, 621024, 621025, 621026,
              621027, 621101, 621102, 621121, 621122, 621123, 621124, 621125, 621126, 621201, 621202, 621221, 621222,
              621223, 621224, 621225, 621226, 621227, 621228, 622901, 622921, 622922, 622923, 622924, 622925, 622926,
              622927, 623001, 623021, 623022, 623023, 623024, 623025, 623026, 623027, 630101, 630102, 630103, 630104,
              630105, 630121, 630122, 630123, 632121, 632122, 632123, 632126, 632127, 632128, 632221, 632222, 632223,
              632224, 632321, 632322, 632323, 632324, 632521, 632522, 632523, 632524, 632525, 632621, 632622, 632623,
              632624, 632625, 632626, 632721, 632722, 632723, 632724, 632725, 632726, 632801, 632802, 632821, 632822,
              632823, 640101, 640104, 640105, 640106, 640121, 640122, 640181, 640201, 640202, 640205, 640221, 640301,
              640302, 640303, 640323, 640324, 640381, 640401, 640402, 640422, 640423, 640424, 640425, 640501, 640502,
              640521, 640522, 650101, 650102, 650103, 650104, 650105, 650106, 650107, 650109, 650121, 650201, 650202,
              650203, 650204, 650205, 652101, 652122, 652123, 652201, 652222, 652223, 652301, 652302, 652323, 652324,
              652325, 652327, 652328, 652701, 652722, 652723, 652801, 652822, 652823, 652824, 652825, 652826, 652827,
              652828, 652829, 652901, 652922, 652923, 652924, 652925, 652926, 652927, 652928, 652929, 653001, 653022,
              653023, 653024, 653101, 653121, 653122, 653123, 653124, 653125, 653126, 653127, 653128, 653129, 653130,
              653131, 653201, 653221, 653222, 653223, 653224, 653225, 653226, 653227, 654002, 654003, 654021, 654022,
              654023, 654024, 654025, 654026, 654027, 654028, 654201, 654202, 654221, 654223, 654224, 654225, 654226,
              654301, 654321, 654322, 654323, 654324, 654325, 654326, 659001, 659002, 659003, 659004]
    # 生成随机年份
    Birth_Y = random.randint(1950, 1996)
    # 生成随机月份
    Birth_M = random.randint(01, 12)
    # 生成随机区域码
    R_id = random.choice(Region)
    # 3位随机数
    T_id = random.randint(1, 999)
    # 生成随机日期
    if Birth_M in [1, 3, 5, 7, 8, 10, 12]:
        Birth_D = random.randint(01, 31)
    elif Birth_M in [4, 6, 9, 11]:
        Birth_D = random.randint(01, 30)
    else:
        if Birth_Y % 4 != 0:
            Birth_D = random.randint(01, 28)
        else:
            Birth_D = random.randint(01, 29)
    # 拼接8位出生日期
    DirthDate = str(Birth_Y) + "%02d" % (Birth_M) + "%02d" % (Birth_D)
    # 拼接身份证钱17位
    base = str(R_id) + DirthDate + str("%03d" % (T_id))

    d = {0: 1, 1: 0, 2: 'X', 3: 9, 4: 8, 5: 7, 6: 6, 7: 5, 8: 4, 9: 3, 10: 2}
    f = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
    sums = 0
    # 根据前17位数字算出最后一个字符
    for i in range(17):
        sums = sums + f[i] * int(base[i])
    mods = sums % 11
    # 拼接并返回18位身份证号
    CardID = base + str(d[mods])
    # print CardID
    return CardID


def generate_cmd_column_dict(funcid, cmdstring):
    '''
    基于功能号和命令串，获取每个字段对应值的字典
    return:

    '''
    try:
        column_dict = OrderedDict()
        gl.g_interfaceDetailTree = ET.parse(r".\init\%s\T2InterfaceDetails.xml" % gl.current_system_id)
        detail_root = gl.g_interfaceDetailTree.getroot()

        # xpath = "Interface[@funcid='%s']/InputParams/in" %(fundid)

        # 如果存在重复的功能号，挑选第一个
        xpath = "Interface[@funcid='%s']/InputParams" % (funcid)
        input_elem = detail_root.find(xpath)
        if input_elem == None:
            return -1, u"未能找到接口细节信息"

        elem_list = input_elem.findall("./in")

        if elem_list == []:
            return -1, u"未能找到接口细节信息"

        cmd_list = cmdstring.split("#")
        # if len(elem_list) != len(cmd_list):
        #    return -1,u"接口参数和命令参数不匹配"

        for i in xrange(len(cmd_list)):
            cmdStr = cmd_list[i].split("=")
            if len(cmdStr) == 2:
                field, value = cmdStr
                # 如果值中有&comma&，予以替换
                value = value.replace('&comma&', ',')
                value = value.replace('&colon&', ':')
            else:
                field = cmdStr[0]
                value = ""

                for cindex, cmditem in enumerate(cmdStr):
                    if cindex > 0:
                        if cindex == 1:
                            value = cmditem
                        else:
                            value = value + "=" + cmditem
                    # 如果值中有&comma&，予以替换
                    value = value.replace('&comma&', ',')
                    value = value.replace('&colon&', ':')

            # column_dict[elem_list[i].attrib["name"]] = cmd_list[i]
            column_dict[field] = value.replace("&bfb&", "#").replace("&bft&", "=")

        return 0, column_dict
    except:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1, exc_info


def createDeleteTrigger():
    root = ET.parse("ZHLC_column_report.xml").getroot()
    for schemaElem in root:
        for tableElem in schemaElem:
            trigger = ""


def GetSdkPosDescBySdkPos(sdkpos):
    sdkposdesc = ""
    desc_list = []
    sdkposStr = sdkpos.split('.')
    sdkcode = ''
    for i in xrange(len(sdkposStr)):
        if i == 0:
            sdkcode = sdkposStr[i]
        else:
            sdkcode = sdkcode + '.' + sdkposStr[i]
        desc_list.append(gl.g_interfaceResultDict[sdkcode])
    sdkposdesc = "-->".join(desc_list)
    return sdkposdesc


# def GetSdkPosDescBySdkPos(sdkpos):
#    sdkposdesc = ""
#    desc_list = []
#    for i in xrange(len(sdkpos)):
#        desc_list.append(gl.g_interfaceResultDict[sdkpos[:i+1]])
#    sdkposdesc = "-->".join(desc_list)
#    return sdkposdesc

def deal_with_action_str(action_str):
    try:
        if action_str == "" or action_str == None or action_str == "null":
            return 0, ""
        action_list = []
        action_list_flag = False
        if not isinstance(action_str, list):
            action_str = action_str.replace("\n", '')
            try:
                action_str_new = ""
                if action_str.count('$bfh$') > 2:
                    action_str_new = action_str.replace("$bfh$", '"')
                else:
                    action_str_new = action_str
                action_list = json.loads(action_str_new)
                action_list_flag = True
            except Exception as e:
                pass
            if action_list_flag == False:
                try:
                    action_list = json.loads(action_str)
                    action_list_flag = True
                except Exception as e:
                    pass
            if action_list_flag == False:
                try:
                    action_str_new = ""
                    if action_str.count('$bfh$') > 2:
                        action_str_new = action_str.replace("$bfh$", '"')
                    else:
                        action_str_new = action_str
                    action_list = eval(action_str_new)
                    action_list_flag = True
                except Exception as e:
                    pass
            if action_list_flag == False:
                try:
                    action_list = eval(action_str)
                    action_list_flag = True
                except Exception as e:
                    pass
        else:
            action_list = action_str
        return action_list
    except Exception, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1, exc_info


# actions, group_variable_dict, [], None, None, account_group_id
def action_process(action_str="", group_dict={}, dataset=[], funcid=None, cmdstring=None, account_group_id=-1):
    '''
    前置条件和后续动作处理

    入参：
        action_str:命令串
        group_dict:组参数
        dataset : 数据集
        funcid: 功能号
        cmdstring:已经参数替换过的命令串
        account_group_id:全局参数组
    返回：
        n： 0 成功， -1 失败
        msg: 相关信息
    '''
    try:
        gl.g_account_id = account_group_id
        if group_dict:
            gl.all_group_dict = group_dict
        # 032保存多个变量
        action_flag = True
        error_list = []

        if action_str == "" or action_str == None or action_str == "null":
            return 0, ""
        error_action = ""
        error_param_info = ""
        action_list = []
        action_list_flag = False
        if not isinstance(action_str, list):
            action_str = action_str.replace("\n", '')
            try:
                action_str_new = ""
                if action_str.count('$bfh$') > 2:
                    action_str_new = action_str.replace("$bfh$", '"')
                else:
                    action_str_new = action_str
                action_list = json.loads(action_str_new)
                action_list_flag = True
            except Exception as e:
                pass
            if action_list_flag == False:
                try:
                    action_list = json.loads(action_str)
                    action_list_flag = True
                except Exception as e:
                    pass
            if action_list_flag == False:
                try:
                    action_str_new = ""
                    if action_str.count('$bfh$') > 2:
                        action_str_new = action_str.replace("$bfh$", '"')
                    else:
                        action_str_new = action_str
                    action_list = eval(action_str_new)
                    action_list_flag = True
                except Exception as e:
                    pass
            if action_list_flag == False:
                try:
                    action_list = eval(action_str)
                    action_list_flag = True
                except Exception as e:
                    pass
        else:
            action_list = action_str
        for action_dict in action_list:
            try:
                call_when = action_dict['call_when']
                type = action_dict['type']
                action = action_dict['content']
                if "clinch_a_deal_the_pair_and_the_data_into" in action:
                    continue
                action = action.replace("$bfh$", '"')
                if str(action.strip()) == "":
                    continue
                desc = action_dict['desc']
                gl.check_data = OrderedDict()
                application_type = ""
                Logging.getLog().debug(str(action_dict))
                Logging.getLog().debug(str(group_dict))
                if action:
                    try:
                        action = action.strip().replace(u"（", "(").replace(u"）", ")")
                    except Exception as e:
                        action = action.strip().replace("（", "(").replace("）", ")").decode("gbk")
                if type != 'General_Operation':
                    # action = action.replace("count(*)", 'count')
                    actions = action.split("&dot&")
                    if len(actions) > 1:
                        for single_func_info in actions:
                            if "exec_sql" in single_func_info:
                                action = single_func_info
                                continue
                            single_func_info = single_func_info.replace("count(*)", 'count')
                            func_name = single_func_info.split("(")[0]

                            func_param = ""
                            if "max(" in single_func_info or "MAX(" in single_func_info:
                                func_param = single_func_info.split("%s(" % func_name)[1][:-1].replace("max(",
                                                                                                       "").replace(
                                    "MAX(", "").replace(")", "")
                            else:
                                func_param = single_func_info.split("(")[1].replace(")", "")
                            func_param = func_param.replace("count", "count(*)")
                            gl.check_data[str(func_name)] = func_param
                    exec_sql_index = action.find("exec_sql")
                    variable_name_or_application_type = action[:exec_sql_index].replace(".", "")
                    if "&equ&" in variable_name_or_application_type:
                        variable_names = variable_name_or_application_type.split("&equ&")[0]
                        application_type = variable_name_or_application_type.split("&equ&")[1]
                        action = variable_names + "=" + action[exec_sql_index:]
                    else:
                        application_type = variable_name_or_application_type
                        action = action[exec_sql_index:]
                action = call_when + ":" + action.replace("&equ&", "=")
                error_action = action
                if action.startswith('#') or action.startswith('--') or not action:
                    continue
                if call_when == 'normal_call' and action_flag == False:  # except_call final_call normal_call
                    pass
                elif call_when == 'except_call' and action_flag == False:  # except_call final_call normal_call
                    pass
                else:  # example: "call:ret1,ret2,ret3 = set_cert_state(@cif_account@,5)"
                    if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
                        return_list_str = action[action.find(":") + 1:action.find("=")].strip()
                        param_list_str = action[action.find("=") + 1:].strip()
                    else:
                        return_list_str = ""
                        param_list_str = action[action.find(":") + 1:].strip()
                    func = param_list_str[:param_list_str.find("(")]
                    param_info = param_list_str[param_list_str.find("(") + 1:param_list_str.rfind(")")]
                    if gl.g_case_adapter_type in ["SPLX_HUARUI", "SPLX_UFX"]:
                        param_info = GroupParameterReplace(param_info, group_dict)
                    param_info = GlobalParameterReplace(param_info, account_group_id)
                    param_info = GroupParameterReplace(param_info, group_dict)
                    param_info = param_info.replace("&bfb&", "#").replace("&bft&", "=").replace("&bft&", "=")
                    error_param_info = param_info
                    if return_list_str != "":
                        return_list = return_list_str.split(',')
                    else:
                        return_list = []
                    try:
                        Logging.getLog().debug("replace action:%s" % param_info)
                        if func == 'mobile_data_by_normalOfTable_to_dormancyOfTable':
                            func = "mobile_data_by_normalOfTable_to_dormancyOfTable_pre"
                        if "exec_sql" == func:
                            param_info = param_info + "," + application_type
                        if "save_param" == func:
                            ret, ret_param = bizmod.save_param(param_info, dataset)
                        elif "save_in_param" == func:
                            ret, ret_param = bizmod.save_in_param(param_info, funcid, cmdstring)
                        elif "getparam" == func and gl.current_system_id in ["14288901937052FXJA",
                                                                             "142891263425277DZ1"]:  # fixmecbs
                            ret, ret_param = getparam(param_info)
                        else:
                            if param_info == '':  # fixme
                                param_info = '1'
                                ret, ret_param = eval("bizmod.%s" % func)(param_info)
                            else:
                                ret, ret_param = eval("bizmod.%s" % func)(param_info)
                    except NameError, ex:
                        exc_info = getExceptionInfo()
                        Logging.getLog().critical(exc_info)
                        action_flag = False
                        error_list.append(exc_info)
                    group_dict.update(gl.random_args)
                    if ret == 0:
                        Logging.getLog().debug(u"执行成功")
                    if ret == 0 and "expression_evaluation" in action:
                        group_dict.update(ret_param)
                    if ret != 0:
                        if isinstance(ret_param, tuple):
                            try:
                                errmsg = u"执行 %s 函数失败,原因是：%s, %s, %s" % (
                                    func, ret_param[0].decode('utf-8'), action, param_info)
                            except:
                                errmsg = u"执行 %s 函数失败,原因是：%s, %s, %s" % (
                                    func, ret_param, action, param_info)
                                # errmsg = u"执行 %s 函数失败,原因是：%s, %s, %s" % (func, ret_param[0].decode('utf-8'), action, param_info)
                        else:
                            errmsg = u"执行 %s 函数失败,原因是：%s, %s, %s" % (
                                func, ret_param, action, param_info)
                        Logging.getLog().debug(errmsg)
                        action_flag = False
                        error_list.append(errmsg)
                    else:
                        try:
                            if (len(return_list) > 0):
                                if gl.o32_for_flag == True:
                                    gl.o32_for_field[return_list[0]] = ret_param[0]
                                    continue
                                for i, r in enumerate(return_list):
                                    uni_value = ""
                                    try:
                                        if not ret_param:
                                            Logging.getLog().critical(u"查询数值为空, 请查看原语输入参数是否有误. %s" % str(action))
                                            return -1, u"查询数值为空, 请查看原语输入参数是否有误.%s" % str(action)
                                        if len(return_list) != len(ret_param):
                                            Logging.getLog().critical(
                                                "获取数据有误, return_list:%s, ret_param:%s" % (
                                                    str(return_list), str(ret_param)))
                                            return -1, "获取数据有误, return_list:%s, ret_param:%s" % (
                                                str(return_list), str(ret_param))
                                        if isinstance(ret_param[i], unicode):
                                            uni_value = ret_param[i]
                                        else:
                                            uni_value = ret_param[i].decode('gbk')
                                    except BaseException, ex:
                                        uni_value = ret_param[i]
                                    print uni_value
                                    group_dict[r.strip()] = uni_value
                                    Logging.getLog().debug("return %s" % str(uni_value))
                                    Logging.getLog().debug("group_dict %s" % str(group_dict))
                            gl.compare_list_group_list.update(group_dict)
                        except Exception, ex:
                            exc_info = getExceptionInfo()
                            errmsg = u"执行 %s 函数后处理返回值失败 提示%s" % (func, exc_info)
                            Logging.getLog().critical(errmsg)
                            action_flag = False
                            error_list.append(errmsg)
                            # return -1, errmsg
            except Exception, ex:
                exc_info = getExceptionInfo()
                Logging.getLog().critical(exc_info)
                action_flag = False
                error_list.append(exc_info)
                continue
        cmdstring = GroupParameterReplace(cmdstring, group_dict)
        if 'openprice' in group_dict.keys():
            gl.openprice_zhlc = group_dict['openprice']

        if action_flag == False:
            return -1, ",".join(error_list).replace("&comma&", ',')
        Logging.getLog().debug(u"check_point:%s" % (json.dumps(gl.return_check_data, sort_keys=True,
                                                               indent=4, ensure_ascii=False, encoding="gbk",
                                                               separators=(',', ':'))))
        return 0, ""
    except Exception, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1, "%s  %s  %s" % (exc_info, error_action, error_param_info)


def getparam(param):
    fundid, test, serverid, secuid, orgid = param.split(',')
    if len(test) > 1:
        sql = "select stkavl,stksale from run..stkasset where fundid ='%s' and secuid = '%s' and orgid = '%s' and stkcode='%s'" % (
            fundid, secuid, orgid, test)
    else:
        if test == 'D' or test == 'd':
            moneytype = '1'
        elif test == 'H' or test == 'h':
            moneytype = '2'
        else:
            moneytype = '0'
        sql = "select fundavl,fundbuy from run..fundasset where fundid ='%s' and orgid = '%s' and moneytype='%s'" % (
            fundid, orgid, moneytype)
    ret, r = bizmod.change_db(sql, serverid)

    if ret < 0:
        return ret, (u"没有在run..fundasset表查到该客户%s持仓信息！" % (test),)

    return ret, r[0]


def action_process_selenium(action_str="", group_dict={}, succ_flag=0, dataset=[], funcid=None, cmdstring=None,
                            account_group_id=-1):
    '''
    前置条件和后续动作处理

    入参：
        action_str:命令串
        group_dict:组参数
        succ_flag:案例成功标志
        dataset : 数据集
        funcid: 功能号
        cmdstring:已经参数替换过的命令串
        account_group_id:全局参数组
    返回：
        n： 0 成功， -1 失败
        msg: 相关信息
    '''
    try:
        gl.g_account_id = account_group_id
        action_flag = True
        error_list = []
        # reture_flag = 0
        Logging.getLog().info("后置---账户组信息为---%s" % account_group_id)
        if action_str == "" or action_str == None:
            return 0, ""
        # Logging.getLog().debug("执行action %s" %action_str)
        action_list = []
        # try:
        #     action_list = json.loads(action_str)
        # except Exception as e:
        #     action_list = eval(action_str)
        action_list_flag = False
        # if not isinstance(action_str, list):
        #     try:
        #         action_str_new = action_str.replace("$bfh$", '"')
        #         action_list = json.loads(action_str_new)
        #         action_list_flag = True
        #     except Exception as e:
        #         pass
        #     if action_list_flag == False:
        #         try:
        #             action_list = json.loads(action_str)
        #             action_list_flag = True
        #         except Exception as e:
        #             pass
        #     if action_list_flag == False:
        #         try:
        #             action_str_new = action_str.replace("$bfh$", '"')
        #             action_list = eval(action_str_new)
        #             action_list_flag = True
        #         except Exception as e:
        #             pass
        #     if action_list_flag == False:
        #         try:
        #             action_list = eval(action_str)
        #             action_list_flag = True
        #         except Exception as e:
        #             pass
        if not isinstance(action_str, list):
            action_str = action_str.replace("\n", '')
            try:
                action_str_new = ""
                if action_str.count('$bfh$') > 2:
                    action_str_new = action_str.replace("$bfh$", '"')
                else:
                    action_str_new = action_str
                action_list = json.loads(action_str_new)
                if action_list is None:
                    action_list = []
                action_list_flag = True
            except Exception as e:
                pass
            if action_list_flag == False:
                try:
                    action_list = json.loads(action_str)
                    action_list_flag = True
                except Exception as e:
                    pass
            if action_list_flag == False:
                try:
                    action_str_new = ""
                    if action_str.count('$bfh$') > 2:
                        action_str_new = action_str.replace("$bfh$", '"')
                    else:
                        action_str_new = action_str
                    action_list = eval(action_str_new)
                    action_list_flag = True
                except Exception as e:
                    pass
            if action_list_flag == False:
                try:
                    action_list = eval(action_str)
                    action_list_flag = True
                except Exception as e:
                    pass
        else:
            action_list = action_str
        for action_dict in action_list:
            try:
                flag = False
                call_when = action_dict['call_when']
                type = action_dict['type']
                action = action_dict['content']
                if "clinch_a_deal_the_pair_and_the_data_into" in action:
                    continue
                if str(action.strip()) == "":
                    continue
                desc = action_dict['desc']
                gl.check_data = OrderedDict()
                application_type = ""
                Logging.getLog().debug(str(group_dict))
                Logging.getLog().debug(str(action_dict))
                if action:
                    action = action.strip().replace(u"（", "(").replace(u"）", ")")
                if type == 'DB_Operation':
                    actions = action.split("&dot&")
                    if len(actions) > 1:
                        for single_func_info in actions:
                            if "exec_sql" in single_func_info:
                                action = single_func_info
                                continue
                            single_func_info = single_func_info.replace("count(*)", 'count')
                            func_name = single_func_info.split("(")[0]

                            func_param = ""
                            if "max(" in single_func_info or "MAX(" in single_func_info:
                                func_param = single_func_info.split("%s(" % func_name)[1][:-1].replace("max(",
                                                                                                       "").replace(
                                    "MAX(", "").replace(")", "")
                            else:
                                func_param = single_func_info.split("(")[1].replace(")", "")
                            func_param = func_param.replace("count", "count(*)")
                            gl.check_data[str(func_name)] = func_param
                    exec_sql_index = action.find("exec_sql")
                    variable_name_or_application_type = action[:exec_sql_index].replace(".", "")
                    if "&equ&" in variable_name_or_application_type:
                        variable_names = variable_name_or_application_type.split("&equ&")[0]
                        application_type = variable_name_or_application_type.split("&equ&")[1]
                        action = variable_names + "=" + action[exec_sql_index:]
                    else:
                        application_type = variable_name_or_application_type
                        action = action[exec_sql_index:]
                action = call_when + ":" + action.replace("&equ&", "=")
                Logging.getLog().info("group_dict---%s" % str(group_dict))
                Logging.getLog().debug(action)
                if action.startswith('#') or action.startswith('--'):
                    continue
                if succ_flag == 0 and call_when == 'normal_call' and action_flag == True:
                    flag = True
                elif succ_flag < 0 and call_when == 'except_call':
                    flag = True
                if call_when == 'final_call':
                    flag = True
                if flag == True:
                    if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
                        return_list_str = action[action.find(":") + 1:action.find("=")].strip()
                        param_list_str = action[action.find("=") + 1:].strip()
                    else:
                        return_list_str = ""
                        param_list_str = action[action.find(":") + 1:].strip()
                    func = param_list_str[:param_list_str.find("(")]
                    param_info = param_list_str[param_list_str.find("(") + 1:param_list_str.rfind(")")]
                    param_info = GlobalParameterReplace(param_info, account_group_id)
                    param_info = GroupParameterReplace(param_info, group_dict)
                    param_info = param_info.replace("&bfb&", "#").replace("&bft&", "=")
                    Logging.getLog().debug(param_info)
                    if return_list_str != "":
                        return_list = return_list_str.split(',')
                    else:
                        return_list = []
                    try:
                        Logging.getLog().debug("replace action:%s" % action)
                        if func == 'mobile_data_by_normalOfTable_to_dormancyOfTable':
                            func = "mobile_data_by_normalOfTable_to_dormancyOfTable_pro"
                        if "exec_sql" == func:
                            param_info = param_info + "," + application_type
                        if "save_param" == func:
                            ret, ret_param = bizmod.save_param(param_info, dataset)
                        elif "save_in_param" == func:
                            ret, ret_param = bizmod.save_in_param(param_info, funcid, cmdstring)
                        elif "getparam" == func and gl.current_system_id in ["14288901937052FXJA",
                                                                             "142891263425277DZ1"]:  # fixmecbs
                            ret, ret_param = getparam(param_info)
                        else:
                            if param_info == '':  # fixme
                                param_info = '1'
                                ret, ret_param = eval("bizmod.%s" % func)(param_info)
                            else:
                                ret, ret_param = eval("bizmod.%s" % func)(param_info)
                    except NameError, ex:
                        exc_info = getExceptionInfo()
                        Logging.getLog().critical(exc_info)
                        error_list.append(exc_info)
                        action_flag = False
                    group_dict.update(gl.random_args)
                    if ret == 0:
                        Logging.getLog().debug(u"执行成功")
                    if ret == 0 and "expression_evaluation" in action:
                        group_dict.update(ret_param)
                    if ret != 0:
                        errmsg = ""
                        if isinstance(ret_param, tuple):
                            try:
                                errmsg = u"执行 %s 函数失败,原因是：%s" % (func, ret_param[0].decode("utf-8"))
                            except:
                                errmsg = u"执行 %s 函数失败,原因是：%s" % (func, ret_param[0])
                                # errmsg = u"执行 %s 函数失败,原因是：%s" % (func, ret_param[0].decode("utf-8"))
                        else:
                            errmsg = u"执行 %s 函数失败,原因是：%s" % (func, ret_param)
                        Logging.getLog().debug(errmsg)
                        # reture_flag = -1
                        error_list.append(errmsg)
                        action_flag = False
                        # return -1, errmsg
                    else:
                        try:
                            if (len(return_list) > 0):
                                for i, r in enumerate(return_list):
                                    try:
                                        if isinstance(ret_param[i], list):
                                            param_i = json.dumps(ret_param[i], encoding='gbk')
                                            uni_value = json.loads(param_i)
                                        elif isinstance(ret_param[i], unicode):
                                            uni_value = ret_param[i]
                                        else:
                                            uni_value = ret_param[i].decode('gbk')
                                    except BaseException, ex:
                                        uni_value = ret_param[i]
                                    group_dict[r.strip()] = uni_value
                                    Logging.getLog().debug("return %s" % uni_value)
                        except Exception, ex:
                            exc_info = getExceptionInfo()
                            errmsg = u"执行 %s 函数后处理返回值失败 提示%s" % (func, exc_info)
                            Logging.getLog().critical(errmsg)
                            error_list.append(errmsg)
                            action_flag = False
                gl.compare_list_group_list.update(group_dict)
            except Exception, ex:
                exc_info = getExceptionInfo()
                errmsg = u"执行 %s 函数后处理返回值失败 提示%s" % (func, exc_info)
                Logging.getLog().critical(errmsg)
                error_list.append(errmsg)
                action_flag = False
                continue
        if action_flag == False:
            return -1, ",".join(error_list).replace("&comma&", ',')
        return 0, ""
    except Exception, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().critical(exc_info)
        return -1, "%s" % (exc_info)


# def action_process_kcbp(serverid, action_str="", group_dict={}, dataset=[], funcid=None, cmdstring=None):
#     '''
#     前置条件和后续动作处理,使用与KCBP，和综合理财不同的点，在于KCBP需要根据serverid来发送命令行到不同的KCXP。
#
#     入参：
#         serverid: 核心号
#         action_str:命令串
#         group_dict:组参数
#         dataset : 数据集
#         funcid: 功能号
#         cmdstring:已经参数替换过的命令串
#     返回：
#         n： 0 成功， -1 失败
#         msg: 相关信息
#     '''
#     try:
#         if action_str == "":
#             return 0, ""
#         Logging.getLog().debug("执行action %s" % action_str)
#         action_list = action_str.strip().split('\n')
#         for action in action_list:
#             if action.startswith('#'):
#                 continue
#             if action.startswith('call:'):  # example: "call:ret1,ret2,ret3 = set_cert_state(@cif_account@,5)"
#                 if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
#                     return_list_str = action[action.find(":") + 1:action.find("=")].strip()
#                     param_list_str = action[action.find("=") + 1:].strip()
#                 else:
#                     return_list_str = ""
#                     param_list_str = action[action.find(":") + 1:].strip()
#                 func = param_list_str[:param_list_str.find("(")]
#                 param_info = param_list_str[param_list_str.find("(") + 1:param_list_str.rfind(")")]
#                 param_info = GlobalParameterReplace(param_info, serverid)
#                 param_info = GroupParameterReplace(param_info, group_dict)
#
#                 if return_list_str != "":
#                     return_list = return_list_str.split(',')
#                 else:
#                     return_list = []
#                 try:
#                     ret, ret_param = eval("bizmod.%s" % func)(param_info)
#                 except NameError, ex:
#                     exc_info = getExceptionInfo()
#                     Logging.getLog().critical(exc_info)
#                     return -1, exc_info
#                 if ret != 0:
#                     errmsg = u"执行 %s 函数失败,原因:%s" % (func, ret_param[0])
#                     Logging.getLog().error(errmsg)
#                     return -1, errmsg
#                 else:
#                     try:
#                         if (len(return_list) > 0):
#                             for i, r in enumerate(return_list):
#                                 group_dict[r.strip()] = ret_param[i]
#                     except Exception, ex:
#                         exc_info = getExceptionInfo()
#                         errmsg = u"执行 %s 函数后处理返回值失败 提示%s" % (func, exc_info)
#                         Logging.getLog().critical(errmsg)
#                         return -1, errmsg
#
#             if action.startswith('save_param:'):  # example: "save_param:cif_account = [1,2],account = [2,1]
#                 param_list_str = action[action.find(":") + 1:]
#                 param_list = param_list_str.split(",")
#                 for p in param_list:
#                     par, value = p.strip().split('=')
#
#                     # 移除前后的空格
#                     par = par.strip()
#                     value = value.strip()
#
#                     # row,col = value.strip()[1:-1].split(',')
#                     # 解析[1][2]
#                     value = value.replace(' ', '')  # 去除所有空格
#                     row, col = value[1:-1].split('][')
#
#                     row = int(row) + 1  # 数据集中第0行是字段名，所以需要从第1行开始
#                     col = int(col)
#                     try:
#                         group_dict[par] = dataset[int(row)][int(col)]
#                     except IndexError as e:
#                         errmsg = "dataset IndexError"
#                         Logging.getLog().debug("dataset IndexError")
#                         return -1, errmsg
#             if action.startswith(
#                     'save_in_param:'):  # example: "save_in_param:cif_account = [cif_account],account = [account]"
#                 param_list_str = action[action.find(":") + 1:]
#                 param_list = param_list_str.split(",")
#
#                 ret, column_dict = generate_cmd_column_dict(funcid, cmdstring)
#                 if ret != 0:
#                     # return -1,column_dict # 如果出错，直接从命令行中解析
#                     column_dict = param.getInParamDictFromCmdstring(cmdstring)
#
#                 for p in param_list:
#                     par, value = p.strip().split('=')
#
#                     # 移除前后的空格
#                     par = par.strip()
#                     value = value.strip()
#
#                     # row,col = value.strip()[1:-1].split(',')
#                     # 解析[1][2]
#                     value = value.replace(' ', '')  # 去除所有空格
#                     # row,col = value[1:-1].split('][')
#                     value = value[1:-1]
#
#                     group_dict[par] = column_dict[value]
#         return 0, ""
#     except Exception, ex:
#         exc_info = getExceptionInfo()
#         Logging.getLog().critical(exc_info)
#         return -1, exc_info


def testtest(param_string):
    print param_string


def delete_file_folder(src):
    '''删除文件和文件夹'''
    if os.path.isfile(src):
        try:
            os.remove(src)
        except:
            exc_info = getExceptionInfo()
            Logging.getLog().critical(exc_info)
            pass
    elif os.path.isdir(src):
        for item in os.listdir(src):
            itemsrc = os.path.join(src, item)
            delete_file_folder(itemsrc)
        try:
            os.rmdir(src)
        except:
            exc_info = getExceptionInfo()
            Logging.getLog().critical(exc_info)
            pass


def runningResult_to_xml(running_dir, return_flag):
    """
    将当前运行记录导出为xml格式

    Args:
        running_dir: 运行目录

    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息
    """
    try:
        keydict = {'pass': u'通过', 'fail': u'失败'}
        resultdict = {'0': u'相等', '1': u'不相等', '-1': u'未比较'}
        TOTAL = 0
        SUCC = 1
        FAIL = 2

        time1 = time.time()
        sql = "select itempos,sdkposdesc, groupname,case_index,casename,funcid,funcname,cmdstring,result,message,detail,cmpresult from tbl_cases_run_result order by itempos,groupname,case_index"
        ds = executeCaseSQL(sql)
        print (time.time() - time1)
        time1 = time.time()
        current_itempos = None
        current_groupname = None
        current_case_index = None

        current_itempos_item = None
        current_groupname_item = None
        current_case_index_item = None

        current_itempos_case_list = [0, 0, 0]  # totalcase, succCase,failCase
        current_groupname_case_list = [0, 0, 0]  # totalcase, succCase,failCase

        total_case_list = [0, 0, 0]  # totalcase, succCase,failCase
        root = etree.Element(u"案例运行结果")
        for rec in ds:
            if (current_itempos != rec["itempos"]):
                # 新的业务节点，增加一个新的树节点
                sdkpos_item = etree.SubElement(root, u"接口")
                sdkpos_item.attrib[u"业务编号"] = rec["itempos"]
                sdkpos_item.attrib[u"业务描述"] = rec["sdkposdesc"]
                # sdkpos_item.attrib["funcid"] = rec["funcid"]
                groupname_item = etree.SubElement(sdkpos_item, u"案例组")
                groupname_item.attrib[u"案例组名"] = rec["groupname"]

                # 更新上一个业务节点的统计信息
                if current_itempos_case_list[TOTAL] != 0:
                    if current_itempos_case_list[TOTAL] == current_itempos_case_list[SUCC]:
                        # 全部成功
                        tips = u"全部成功(%d/%d)" % (current_itempos_case_list[SUCC], current_itempos_case_list[TOTAL])
                    elif current_itempos_case_list[TOTAL] == current_itempos_case_list[FAIL]:
                        # 全部失败
                        tips = u"全部失败(%d/%d)" % (current_itempos_case_list[SUCC], current_itempos_case_list[TOTAL])
                    else:
                        # 部分成功
                        tips = u"部分成功(%d/%d)" % (current_itempos_case_list[SUCC], current_itempos_case_list[TOTAL])
                    current_itempos_item.attrib[u"运行状态"] = tips

                    # 更新上一个组节点的统计信息
                    if current_groupname_case_list[TOTAL] != 0:
                        tips = ""
                        if current_groupname_case_list[TOTAL] == current_groupname_case_list[SUCC]:
                            # 全部成功
                            tips = u"全部成功(%d/%d)" % (
                                current_groupname_case_list[SUCC], current_groupname_case_list[TOTAL])
                        elif current_groupname_case_list[TOTAL] == current_groupname_case_list[FAIL]:
                            # 全部失败
                            tips = u"全部失败(%d/%d)" % (
                                current_groupname_case_list[SUCC], current_groupname_case_list[TOTAL])
                        else:
                            # 部分成功
                            tips = u"部分成功(%d/%d)" % (
                                current_groupname_case_list[SUCC], current_groupname_case_list[TOTAL])
                        current_groupname_item.attrib[u"运行状态"] = tips

                # 预备处理新的一行案例结果
                case_index_item = etree.SubElement(groupname_item, u"案例")
                # case_index_item.attrib["sdkpos"] = "%s" %rec["sdkpos"]
                # case_index_item.attrib["sdkposdesc"] ="%s" %rec["sdkposdesc"]
                # case_index_item.attrib["groupname"] = "%s" %rec["groupname"]
                case_index_item.attrib["case_index"] = "%d" % rec["case_index"]
                case_index_item.attrib["funcid"] = "%s" % rec["funcid"]
                case_index_item.attrib["casename"] = "%s" % rec["casename"]
                case_index_item.attrib["funcname"] = "%s" % rec["funcname"]
                case_index_item.attrib["result"] = "%s" % (keydict[rec["result"]])
                case_index_item.attrib["message"] = "%s " % rec["message"]
                case_index_item.attrib["cmdstring"] = "%s" % re.sub(r'[\x01-\x1f]', '',
                                                                    rec["cmdstring"])  # 去除可能存在的不可见字符
                ret, inparam = getInparamListFromCmdstring(rec["funcid"], rec["cmdstring"])
                if ret == 0:
                    case_index_item.attrib["InputParam"] = inparam
                # case_index_item.attrib["cmpresult"]= "%s" %resultdict[rec["cmpresult"]]

                if return_flag == '1':
                    return_item = etree.SubElement(case_index_item, u"返回值")

                    if rec["detail"] != None or rec["detail"] != "":
                        return_list = rec["detail"].split('\r\n')
                        for k, line in enumerate(return_list):
                            if k == 0:
                                column_name_list = line.split(',')
                            else:
                                value_list = line.split(',')
                                out = ",".join("%s=%s" % (column_name_list[m], value_list[m]) for m in
                                               xrange(min(len(value_list), len(column_name_list))))
                                outElem = etree.SubElement(return_item, "OutParam")
                                outElem.attrib["lineno"] = str(k)
                                outElem.attrib["value"] = re.sub(r'[\x00-x1f]', '', out)  # 去除可能存在的不可见字符

                current_itempos = rec["itempos"]
                current_groupname = rec["groupname"]

                current_itempos_item = sdkpos_item
                current_groupname_item = groupname_item

                if (rec["result"] == "pass"):
                    current_itempos_case_list = [1, 1, 0]
                    current_groupname_case_list = [1, 1, 0]
                    total_case_list[TOTAL] = total_case_list[TOTAL] + 1
                    total_case_list[SUCC] = total_case_list[SUCC] + 1
                else:
                    current_itempos_case_list = [1, 0, 1]
                    current_groupname_case_list = [1, 0, 1]

                    total_case_list[TOTAL] = total_case_list[TOTAL] + 1
                    total_case_list[FAIL] = total_case_list[FAIL] + 1
                continue
            elif (current_groupname != rec["groupname"]):
                groupname_item = etree.SubElement(current_itempos_item, u"案例组")
                groupname_item.attrib[u"案例组名"] = rec["groupname"]
                if current_groupname_case_list[TOTAL] != 0:
                    tips = ""
                    if current_groupname_case_list[TOTAL] == current_groupname_case_list[SUCC]:
                        # 全部成功
                        tips = u"全部成功(%d/%d)" % (current_groupname_case_list[SUCC], current_groupname_case_list[TOTAL])
                    elif current_groupname_case_list[TOTAL] == current_groupname_case_list[FAIL]:
                        # 全部失败
                        tips = u"全部失败(%d/%d)" % (current_groupname_case_list[SUCC], current_groupname_case_list[TOTAL])
                    else:
                        # 部分成功
                        tips = u"部分成功(%d/%d)" % (current_groupname_case_list[SUCC], current_groupname_case_list[TOTAL])
                    current_groupname_item.attrib[u"运行状态"] = tips

                    # 预备处理新的一行案例结果
                    # case_index_item = dialog.tree.AppendItem(groupname_item, "%s"  %(str(rec["caserunsequence"])))
                    case_index_item = etree.SubElement(groupname_item, u"案例")
                    # case_index_item.attrib["sdkpos"] = "%s" %rec["sdkpos"]
                    # case_index_item.attrib["sdkposdesc"] ="%s" %rec["sdkposdesc"]
                    # case_index_item.attrib["groupname"] = "%s" %rec["groupname"]
                    case_index_item.attrib["case_index"] = "%d" % rec["case_index"]
                    case_index_item.attrib["funcid"] = "%s" % rec["funcid"]
                    case_index_item.attrib["casename"] = "%s" % rec["casename"]
                    case_index_item.attrib["funcname"] = "%s" % rec["funcname"]
                    case_index_item.attrib["result"] = "%s" % (keydict[rec["result"]])
                    case_index_item.attrib["message"] = "%s " % rec["message"]
                    case_index_item.attrib["cmdstring"] = "%s" % re.sub(r'[\x01-\x1f]', '',
                                                                        rec["cmdstring"])  # 去除可能存在的不可见字符
                    # case_index_item.attrib["cmpresult"]= "%s" %resultdict[rec["cmpresult"]]

                    ret, inparam = getInparamListFromCmdstring(rec["funcid"], rec["cmdstring"])
                    if ret == 0:
                        case_index_item.attrib["InputParam"] = inparam

                    if return_flag == '1':
                        return_item = etree.SubElement(case_index_item, u"返回值")

                        if rec["detail"] != None or rec["detail"] != "":
                            return_list = rec["detail"].split('\r\n')
                            for k, line in enumerate(return_list):
                                if k == 0:
                                    column_name_list = line.split(',')
                                else:
                                    value_list = line.split(',')
                                    out = ",".join("%s=%s" % (column_name_list[m], value_list[m]) for m in
                                                   xrange(min(len(value_list), len(column_name_list))))
                                    print out
                                    outElem = etree.SubElement(return_item, "OutParam")
                                    outElem.attrib["lineno"] = str(k)
                                    outElem.attrib["value"] = re.sub(r'[\x00-x1f]', '', out)

                    current_groupname = rec["groupname"]

                    current_groupname_item = groupname_item

                    if (rec["result"] == "pass"):
                        current_itempos_case_list[TOTAL] = current_itempos_case_list[TOTAL] + 1
                        current_itempos_case_list[SUCC] = current_itempos_case_list[SUCC] + 1
                        current_groupname_case_list = [1, 1, 0]
                        total_case_list[TOTAL] = total_case_list[TOTAL] + 1
                        total_case_list[SUCC] = total_case_list[SUCC] + 1
                    else:
                        current_itempos_case_list[TOTAL] = current_itempos_case_list[TOTAL] + 1
                        current_itempos_case_list[FAIL] = current_itempos_case_list[FAIL] + 1
                        current_groupname_case_list = [1, 0, 1]
                        total_case_list[TOTAL] = total_case_list[TOTAL] + 1
                        total_case_list[FAIL] = total_case_list[FAIL] + 1
                continue
            else:
                case_index_item = etree.SubElement(groupname_item, u"案例")
                # case_index_item.attrib["sdkpos"] = "%s" %rec["sdkpos"]
                # case_index_item.attrib["sdkposdesc"] ="%s" %rec["sdkposdesc"]
                # case_index_item.attrib["groupname"] = "%s" %rec["groupname"]
                case_index_item.attrib["case_index"] = "%d" % rec["case_index"]
                case_index_item.attrib["funcid"] = "%s" % rec["funcid"]
                case_index_item.attrib["casename"] = "%s" % rec["casename"]
                case_index_item.attrib["funcname"] = "%s" % rec["funcname"]
                case_index_item.attrib["result"] = "%s" % (keydict[rec["result"]])
                case_index_item.attrib["message"] = "%s " % rec["message"]
                case_index_item.attrib["cmdstring"] = "%s" % re.sub(r'[\x01-\x1f]', '',
                                                                    rec["cmdstring"])  # 去除可能存在的不可见字符
                # case_index_item.attrib["cmpresult"]= "%s" %resultdict[rec["cmpresult"]]

                ret, inparam = getInparamListFromCmdstring(rec["funcid"], rec["cmdstring"])
                if ret == 0:
                    case_index_item.attrib["InputParam"] = inparam

                if return_flag == '1':
                    return_item = etree.SubElement(case_index_item, u"返回值")

                    if rec["detail"] != None or rec["detail"] != "":
                        return_list = rec["detail"].split('\r\n')
                        for k, line in enumerate(return_list):
                            if k == 0:
                                column_name_list = line.split(',')
                            else:
                                value_list = line.split(',')
                                out = ",".join("%s=%s" % (column_name_list[m], value_list[m]) for m in
                                               xrange(min(len(value_list), len(column_name_list))))
                                outElem = etree.SubElement(return_item, "OutParam")
                                outElem.attrib["lineno"] = str(k)
                                outElem.attrib["value"] = re.sub(r'[\x00-x1f]', '', out)

                current_groupname_item = groupname_item

                if (rec["result"] == "pass"):
                    current_itempos_case_list[TOTAL] = current_itempos_case_list[TOTAL] + 1
                    current_itempos_case_list[SUCC] = current_itempos_case_list[SUCC] + 1
                    current_groupname_case_list[TOTAL] = current_groupname_case_list[TOTAL] + 1
                    current_groupname_case_list[SUCC] = current_groupname_case_list[SUCC] + 1
                    total_case_list[TOTAL] = total_case_list[TOTAL] + 1
                    total_case_list[SUCC] = total_case_list[SUCC] + 1
                else:
                    current_itempos_case_list[TOTAL] = current_itempos_case_list[TOTAL] + 1
                    current_itempos_case_list[FAIL] = current_itempos_case_list[FAIL] + 1
                    current_groupname_case_list[TOTAL] = current_groupname_case_list[TOTAL] + 1
                    current_groupname_case_list[FAIL] = current_groupname_case_list[FAIL] + 1
                    total_case_list[TOTAL] = total_case_list[TOTAL] + 1
                    total_case_list[FAIL] = total_case_list[FAIL] + 1
                continue
        if current_itempos_case_list[TOTAL] != 0:
            tips = ""
            if current_itempos_case_list[TOTAL] == current_itempos_case_list[SUCC]:
                # 全部成功
                tips = u"全部成功(%d/%d)" % (current_itempos_case_list[SUCC], current_itempos_case_list[TOTAL])
            elif current_itempos_case_list[TOTAL] == current_itempos_case_list[FAIL]:
                # 全部失败
                tips = u"全部失败(%d/%d)" % (current_itempos_case_list[SUCC], current_itempos_case_list[TOTAL])
            else:
                # 部分成功
                tips = u"部分成功(%d/%d)" % (current_itempos_case_list[SUCC], current_itempos_case_list[TOTAL])
            current_itempos_item.attrib[u"运行状态"] = tips

            # 更新上一个组节点的统计信息
            if current_groupname_case_list[TOTAL] != 0:
                tips = ""
                if current_groupname_case_list[TOTAL] == current_groupname_case_list[SUCC]:
                    # 全部成功
                    tips = u"全部成功(%d/%d)" % (current_groupname_case_list[SUCC], current_groupname_case_list[TOTAL])
                elif current_groupname_case_list[TOTAL] == current_groupname_case_list[FAIL]:
                    # 全部失败
                    tips = u"全部失败(%d/%d)" % (current_groupname_case_list[SUCC], current_groupname_case_list[TOTAL])
                else:
                    # 部分成功
                    tips = u"部分成功(%d/%d)" % (current_groupname_case_list[SUCC], current_groupname_case_list[TOTAL])
                current_groupname_item.attrib[u"运行状态"] = tips

            # 更新根节点的统计信息
            if total_case_list[TOTAL] == total_case_list[SUCC]:
                # 全部成功
                tips = u"全部成功(%d/%d)" % (total_case_list[SUCC], total_case_list[TOTAL])
            elif total_case_list[TOTAL] == total_case_list[FAIL]:
                # 全部失败
                tips = u"全部失败(%d/%d)" % (total_case_list[SUCC], total_case_list[TOTAL])
            else:
                # 部分成功
                tips = u"部分成功(%d/%d)" % (total_case_list[SUCC], total_case_list[TOTAL])
            root.attrib[u"运行状态"] = tips

        # 写入xml文件
        doc = etree.ElementTree(root)  # root 为根元素
        root = os.getcwd()
        xml_path = os.path.join(root, "Autotest_report\\Autotest_Total_report\\%s.xml" % running_dir)
        doc.write(xml_path, pretty_print=True, xml_declaration=True, encoding='utf-8')

        # 生成执行html报告
        cover_path = os.path.join(root, "Autotest_report\\Autotest_Coverage_html_report\\%s.html" % running_dir)

        html_path = os.path.join(root, "Autotest_report\\Autotest_Total_report\\%s.html" % running_dir)
        exportHtml(html_path, 'all', cover_path)

        print (time.time() - time1)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)


def exportHtml(html_path, flag, cover_path):
    # filePath = rootdir + "/" + htmlName
    # html = open(filePath, 'w')

    html = open(html_path, 'w')

    html.write("""<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">
<html xmlns="http://www.w3.org/1999/xhtml">
  <head>
  <meta http-equiv="Content-Type" content="text/html; charset=gbk"/>
  <title>测试报告</title>
  <style type="text/css">
  <!--
body {font-family:Arial, Helvetica, sans-serif; font-size:80%; line-height:140%; color:#020202;}
      #casesRunDate {text-align:center;}
      #casesRunDetail {text-align:center;}
      #topAnchor {position:absolute; top:0; left:0; overflow:hidden; height:0; width:0;}
      #dataSource {margin:10px auto;}
        .dbImage {padding:0 5px 0 0;}
        .leftSchemas {font-size:150%;color:#666; background:#FFF; text-align:right;}
        .rightSchemas {font-size:150%;color:#666; background:#FFF; text-align:left;}
        .schemaSplitter {text-align:center; padding:10px;}
        .leftIdentifier  {font-size:140%;line-height:100%; text-align:right;}
        .rightIdentifier {font-size:140%; line-height:100%; text-align:left;}
        
      #runDetail {margin:0 left; text-align:left;}
      thead {background:#E6E6F2}
      td, th {padding:2px 10px;}
      tr:nth-child(2n) {background:#f8f8f8;}
    
      #runDetail a, .backToTop {text-decoration:none; color:#2f80cc; text-indent:10px; display:block;}
      #runDetail a:hover, .backToTop:hover {text-decoration:underline;} 
    
      .backToTop {text-indent:0px;}
    
      hr {background:#CCC; border:0 none; margin:20px 0;}
    
      h1 {font-size:140%;}
      h2 {font-size:130%; color:#666;}
    
      .data {text-align:left;}
      .data td:nth-child(2n), .data th:nth-child(2n) {border-right:4px solid #f1f1f1; margin-right:15px;}
    
      .different {background:#2f80cc; color:#FFF}
      .caseRunRecord {background:#2f80cc; color:#FFF}
      .caseRunRecordDetail {background:#2f80cc; color:#FFF}
  
      .leftDbIconSmall {background:transparent url(data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABAAAAAQCAYAAAAf8/9hAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAgY0hSTQAAeiYAAICEAAD6AAAAgOgAAHUwAADqYAAAOpgAABdwnLpRPAAAAAlwSFlzAAALEQAACxEBf2RfkQAAAwJJREFUOE91kn1Ik1EYxd+/wggig1AQRTEqRFAURVk4jERRFMUwrMRQLIdiTJTEsZhMDGXgmAiTQFFmmDJBEsdEURyKouzN6dpwTbe21Hdzn36kmZ3utjSGdeH575zffc65l0q9n0tlsItCCgsrn5aV17+vb2xf5bd0M018MVPJ4a3lFT//kJiWWR4bn3At8vZdKiwiigJwMVRyWtY9Lle4rFKtQ6tlQK9aoZo3QKHUYGR0Cb0DMxC09SOvuOJTROydhEuAiqoGlcWyD6v1EDqdDcsrZszM6TGuoDE8uugHdEkVaBPJkZlTog6LjLoSBBBL+j0McwyTyQ2tjrkMkM1AIp1AW4ccJWX1nvCYmJtBAL6g0+lwnPo38AGWlk2XAF09CrSKRpBV8MwZFh0dDKjiNDEbhj14vb9gMruxqtmGaoF0MKnBkHwB0l4lBG8HwOGKkP4gn7kVGXkjCFBdy7etqM3Q6/f8xrEJtf9WUdcY3rTJwG3qxou6dlTUCMHOeWQPj44JDQLwSAStnmRXb0I5rcHA0BzaxaNoFg6gnteD2kaxH/C/oTh1fJtx0wWL1Yu5eT1GxhYh7VOiQzJKcg+R9WV+oe+c/TzDyckpDo+O4fYcBgBVNc27atoCo9GFpRUTxpU0+gZn0dn9kQCG0ch/5xeeEvOTyuaLse95AgBfBFpjgZo2Y3J6HbIhFUSSMfCEMjQQ83mE4+Mf2D/4DpdrHza7G9ZtRwDAbWz1+No3mhyYml3H4IgKkp5xCMn6fPIDea19fqH34AgOYj7fYMvEBAD5xeXL9KoF29uHpEgzFFMaApknbz9BypSjuSUAcDi92GVc+Grdw+bWLvSGbwFAXGJ6IqeWp52apjFLPtCwfBHi7nESYRBNLf149ecZd3adQR2sfTYHAPFJLCop9eH1jMyil6VldYrqOoGhskZgL614bS94zPmSkV086ROaLTYYjDvQbVihIWZas/UXkMbKpVgZ+VQqK4dKYWWHpLJzQ1PY2aEJ6eyrccnplE/472nHbwVfrTjUnrWTAAAAAElFTkSuQmCC) right center no-repeat; padding-right:25px;}
      .rightDbIconSmall {background:transparent url(data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABAAAAAQCAYAAAAf8/9hAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAgY0hSTQAAeiYAAICEAAD6AAAAgOgAAHUwAADqYAAAOpgAABdwnLpRPAAAAAlwSFlzAAALEQAACxEBf2RfkQAAAsRJREFUOE91k0tME0EYx5ftc1taaKHlUSgUEFAQA4oe1AiJGkEPPhL1IF6MXkw8eDGcPWiC0QhGeQSMIjWKCkZQQ+QlCEHRgIAoKQgqobTIFmJJ7Xb7d7rYmqY4yZfJJjO/+f3nm6UABFUlRVFDJXuVr7dl17cYoucbVXK3WUbjnpSUjHY3MCKrOVx8524MpfStDQH052bubzbGufoz4jGRyGAyicGXZAbDRgZ98XK81EvxQCVGrYR2VTOi4hDAU4POOVe0BT93roNtgwY/UhSwEMAoAQ0SYI9BjrY4OZp0MlRLaGcI4KFOg+VjBVjcnQ37phjMpSowZVLgEwF8IIC+BDnaiUlrrBzVYjp4swDQa+E8XwK2OA8LW42YJ4AZAvDHGEhg0EUsnhOLqrUA9zUquGsu4be5Eo6TxbBvNmA2UwNLmhqjqSq8M6nQkahEY4wMN0VhoYB6tQJLV0rBXS3Fr7NHwB7Ihy1Hj6/pkfhoCkcPMfDp10VLccMHqFBL8L+aOFyAb4VZsGWoMWli8IaoP9ZJcY104XqEZPUOfAsDg+cBtxte14oAGN21Ed2pUXiilcIcKUZjlBTNRL0jQYlHxOKW3wAeLuhkr2NRmJ3njoM9tB22fCOms/QYSYtEb6ICz8lms172L4J3xQnvsgNedgFeuxW89bsAYM8cBFuUG4gwRLrQHkf6r5HgMokhdEEALNrB2+YCBvzUZ2Huyk1BS5IWTX8j3I4Qo45sfhbLoIlYVPoN+NkZ8DMW8JZxeMaH4RkZXI1w4RQcRwth25GO6TwDxtZHoy85HC/iGTSQ+ygXhS0JAI9lLHC68D3QLczWE/tg25MTiPCevMJX5BLriX45Tbku0tTpVcDwW3CDveD6O8H1tIHrbBUAtSrGUSeXcLXiMNQQ3SoRxVWIKLaMpjrLKEor/I3+U9cqMkIeWnCB+gOP0TjDKUQRBgAAAABJRU5ErkJggg==) left center no-repeat; padding-left:25px;}
      .keyIcon {background:transparent url(data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABAAAAAQCAYAAAAf8/9hAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAgY0hSTQAAeiYAAICEAAD6AAAAgOgAAHUwAADqYAAAOpgAABdwnLpRPAAAAAlwSFlzAAALEQAACxEBf2RfkQAAAeNJREFUOE+VkktIW0EUhieJlFizsujC92Nh6a64cVUXgqBSN1W68IGCMTVBqwuX3bgSFd1koTE3gl0outGqpWSlUCuKlipNGzB6G6jREFHB58293t+Z60SUK23ywWGGmfP/Z+ZwSAyXlfROfkg//dRXoM715yljnebASCsp4df/ZrzT8tO3VIdraQDHYhsu/zbhPFCNz31pGLYRK097HFrF+ftrI4LL5aB7bcWJQ1tlfzFme56q9Dybp+uZ7smKKlGnJqYwQ82AxYW/FPveZ6Dfc/N0Pd7hlwhvdzx8wVErcNgEhF4jumaG4DBt8XQ984NFOBJvK2rBxJFG4KAWCL6CtJIEt924wdP1fOy2nEphO6T9Fi5uoOI3wF4l4MtAaMEMt42s83Q9oy2kftFVhKvdKkhilfZsBEupOBPyqglysAIu221/WDwKbdJb71C6Gll+DvzKhrJuQfiLGdJmIbankyH/icNksIuksm57Oky7nnajT3AYznyTqfBPWfBDMEIWy/5vch86QC8Eu+FccJCD8ffE/82ZpCp7NYmZ0P7ke5pJ2lQtecKEgblcKIdtiZnEYCIW3ycKVCVivTNJCDpoOaPvDGpgNhOx7/Cr+GHTqu6UQJxJ0SaXH8cPq8qELFw2ghvl51sKtlRwUgAAAABJRU5ErkJggg==) 5px center no-repeat; padding-left:25px;}
      -->
      </style>
</head>
<body> 
    """)

    casesRunSql = "select sno,running_dir,start_date,end_date,during,state,casenum,succCount,failCount from tbl_running_record order by sno desc limit 1"
    runDs = executeCaseSQL(casesRunSql)

    for rec in runDs:
        html.write("""<p id="casesRunDate"><H1 align="center">案例执行情况汇总  """ + str(rec["start_date"]).strip() + """</H1></p>
    <table width="1475px">
        <thead>
            <tr>
                <th>序号</th>
                <th>结果目录</th>
                <th>开始时间</th>
                <th>结束时间</th>
                <th>运行时长</th>
                <th>运行状态</th>
                <th>案例总数</th>
                <th>成功案例个数</th>
                <th>失败案例个数</th>
            </tr>
        </thead>
            <tr align="center">
                <td>""" + str(rec["sno"]).strip() + """</td>
                <td>""" + str(rec["running_dir"]).strip() + """</td>
                <td>""" + str(rec["start_date"]).strip() + """</td>
                <td>""" + str(rec["end_date"]).strip() + """</td>
                <td>""" + str(rec["during"]).strip() + """</td>
                <td>""" + str(rec["state"]).strip() + """</td>
                <td>""" + str(rec["casenum"]).strip() + """</td>
                <td>""" + str(rec["succCount"]).strip() + """</td>
                <td>""" + str(rec["failCount"]).strip() + """</td>
                <td><a target="_blank"  href='""" + ('file:///' + cover_path) + """'>案例执行覆盖情况</a></td>
            </tr>
    </table>
    <p/><p/><p/>""")

        ##功能号覆盖情况
        # runCaseList = {}
        # runCasesSql="select funcid,                              "\
        #            "       funcname,                            "\
        #            "       count(funcid) count,                 "\
        #            "       (select count(funcid)                "\
        #            "          from tbl_cases_run_result         "\
        #            "         where result = 'pass'              "\
        #            "           and funcid = r.funcid) succount, "\
        #            "       (select count(funcid)                "\
        #            "          from tbl_cases_run_result         "\
        #            "         where result = 'fail'              "\
        #            "           and funcid = r.funcid) failcount "\
        #            "  from tbl_cases_run_result r               "\
        #            " group by funcid                            "\
        #            " order by funcid                            "
        # ds = executeCaseSQL(runCasesSql)

        # for rec in ds:
        #    runCaseList[str(rec["funcid"])]={}
        #    runCaseList[str(rec["funcid"])]["funcname"]=str(rec["funcname"])
        #    runCaseList[str(rec["funcid"])]["totalcount"]=str(rec["count"])
        #    runCaseList[str(rec["funcid"])]["successcount"]=str(rec["succount"])
        #    runCaseList[str(rec["funcid"])]["failcount"]=str(rec["failcount"])

        # if len(gl.g_SysCasesList) > 0:
        #    html.write("""
        # <p id="casesCoverage"><H1 align="center">功能号覆盖情况汇总</H1></p>
        # <table width="1300px" id="casesCoverageDetail">
        # <thead>
        #        <tr>
        #           <th>系统支持功能号</th>
        #      <th>功能号描述</th>
        #      <th>覆盖情况</th>
        #      <th>案例个数</th>
        #      <th>成功个数</th>
        #      <th>失败个数</th>
        #            <th>系统支持功能号</th>
        #      <th>功能号描述</th>
        #      <th>覆盖情况</th>
        #      <th>案例个数</th>
        #      <th>成功个数</th>
        #      <th>失败个数</th>
        #     </tr>
        # </thead>""")
    # funcindex = 1
    # isCover = ""
    # funcname = ""
    # totalcount = "0"
    # successcount = "0"
    # failcount = "0"
    ##for funcid in gl.g_SysCasesList:
    # for index,funcid in enumerate(gl.g_SysCasesList):
    #    if runCaseList.has_key(funcid):
    #        isCover = "覆盖"
    #        funcname = runCaseList[funcid]['funcname']
    #        totalcount = runCaseList[funcid]['totalcount']
    #        successcount = runCaseList[funcid]['successcount']
    #        failcount = runCaseList[funcid]['failcount']
    #    else:
    #        isCover = "未覆盖"
    #        funcname = ""
    #        totalcount = "0"
    #        successcount = "0"
    #        failcount = "0"
    #    #if index == len(gl.g_SysCasesList) - 1:
    #    if funcindex == 1:
    #        html.write("""
    #            <tr align="center">
    #                <td>"""+funcid+"""</td>
    #                <td>"""+funcname+"""</td>
    #                <td>"""+isCover+"""</td>
    #                <td>"""+totalcount+"""</td>
    #                <td>"""+successcount+"""</td>""")
    #        if failcount != "0":
    #            html.write("""
    #                <td bgcolor="#FF0000">"""+failcount+"""</td>""")
    #        else:
    #            html.write("""
    #                <td>"""+failcount+"""</td>""")

    #        funcindex = funcindex + 1
    #    else:
    #        html.write("""
    #                <td>"""+funcid+"""</td>
    #                <td>"""+funcname+"""</td>
    #                <td>"""+isCover+"""</td>
    #                <td>"""+totalcount+"""</td>
    #                <td>"""+successcount+"""</td>
    #                <td>"""+failcount+"""</td>""")
    #        if failcount != "0":
    #            html.write("""
    #                <td bgcolor="#FF0000">"""+failcount+"""</td>
    #            </tr>""")
    #        else:
    #            html.write("""
    #                <td>"""+failcount+"""</td>
    #            </tr>""")

    #        funcindex = 1

    # if len(gl.g_SysCasesList) > 0:
    #    html.write("""
    # </table>""")

    casesRunDelSql = " select serialno,                                  " \
                     "        itempos,                                   " \
                     "        sdkposdesc,                                " \
                     "        groupname,                                 " \
                     "        count(itempos) count,                      " \
                     "        (select count(funcid)                      " \
                     "           from tbl_cases_run_result               " \
                     "          where result = 'pass'                    " \
                     "            and itempos = r.itempos                " \
                     "            and groupname = r.groupname) succount, " \
                     "        (select count(funcid)                      " \
                     "           from tbl_cases_run_result               " \
                     "          where result = 'fail'                    " \
                     "            and itempos = r.itempos                " \
                     "            and groupname = r.groupname) failcount " \
                     "   from tbl_cases_run_result r                     " \
                     "  group by itempos, groupname                      "
    # "  order by itempos, groupname, case_index          "
    if flag == 'fail':
        casesRunDelSql = casesRunDelSql + " HAVING  failcount > 0 order by itempos, groupname, case_index"
    elif flag == 'succ':
        casesRunDelSql = casesRunDelSql + " HAVING  failcount = 0 order by itempos, groupname, case_index"
    else:
        casesRunDelSql = casesRunDelSql + "order by itempos, groupname, case_index"

    ds = executeCaseSQL(casesRunDelSql)
    index = 1
    for rec in ds:
        if index == 1:
            html.write("""
    <p id="casesRunDetail"><H1 align="center">案例执行情况详情汇总</H1></p>
    <table width="1300px" id="runDetail">
        <thead>
            <tr>
                <th>序号</th>
                <th>结果目录</th>
                <th>功能号描述</th>
                <th>案例组名称</th>
                <th>案例总数</th>
                <th>成功个数</th>
                <th>失败个数</th>
            </tr>
        </thead>""")

        html.write("""
            <tr align="center">
                <td>""" + str(index) + """</td>
                <td>""" + str(rec["itempos"]) + """</td>
                <td>""" + str(rec["sdkposdesc"]) + """</td>
                <td><a href="#""" + str(rec["itempos"]) + """_""" + str(rec["groupname"]) + """">""" + str(
            rec["groupname"]) + """</a></td>
                <td>""" + str(rec["count"]) + """</td>
                <td>""" + str(rec["succount"]) + """</td>""")
        if rec["failcount"] != 0:
            html.write("""
                <td bgcolor="#FF0000">""" + str(rec["failcount"]) + """</td>
            </tr>""")
        else:
            html.write("""
                <td>""" + str(rec["failcount"]) + """</td>
            </tr>""")

        if index == len(ds):
            html.write("""    
    </table>""")
        index = index + 1

    casesDetailSql = ""
    if flag == 'fail':
        casesDetailSql = "select serialno,                                               " \
                         "       itempos,                                                " \
                         "       sdkposdesc,                                             " \
                         "       groupname,                                              " \
                         "       case_index,                                             " \
                         "       casename,                                               " \
                         "       funcid,                                                 " \
                         "       funcname,                                               " \
                         "       cmdstring,                                              " \
                         "       result,                                                 " \
                         "       message,                                                " \
                         "       detail                                                  " \
                         "  from tbl_cases_run_result                                    " \
                         " where itempos in (select itempos                              " \
                         "                     from tbl_cases_run_result r               " \
                         "                    group by itempos, groupname                " \
                         "                   HAVING (select count(funcid)                " \
                         "                            from tbl_cases_run_result          " \
                         "                           where result = 'fail'               " \
                         "                             and itempos = r.itempos           " \
                         "                             and groupname = r.groupname) > 0) " \
                         "   and groupname in                                            " \
                         "       (select groupname                                       " \
                         "          from tbl_cases_run_result r                          " \
                         "         group by itempos, groupname                           " \
                         "        HAVING (select count(funcid)                           " \
                         "                 from tbl_cases_run_result                     " \
                         "                where result = 'fail'                          " \
                         "                  and itempos = r.itempos                      " \
                         "                  and groupname = r.groupname) > 0)            " \
                         " order by itempos, groupname, case_index                       "
    elif flag == 'succ':
        casesDetailSql = "select serialno,                                               " \
                         "       itempos,                                                " \
                         "       sdkposdesc,                                             " \
                         "       groupname,                                              " \
                         "       case_index,                                             " \
                         "       casename,                                               " \
                         "       funcid,                                                 " \
                         "       funcname,                                               " \
                         "       cmdstring,                                              " \
                         "       result,                                                 " \
                         "       message,                                                " \
                         "       detail                                                  " \
                         "  from tbl_cases_run_result                                    " \
                         " where itempos in (select itempos                              " \
                         "                     from tbl_cases_run_result r               " \
                         "                    group by itempos, groupname                " \
                         "                   HAVING (select count(funcid)                " \
                         "                            from tbl_cases_run_result          " \
                         "                           where result = 'fail'               " \
                         "                             and itempos = r.itempos           " \
                         "                             and groupname = r.groupname) = 0) " \
                         "   and groupname in                                            " \
                         "       (select groupname                                       " \
                         "          from tbl_cases_run_result r                          " \
                         "         group by itempos, groupname                           " \
                         "        HAVING (select count(funcid)                           " \
                         "                 from tbl_cases_run_result                     " \
                         "                where result = 'fail'                          " \
                         "                  and itempos = r.itempos                      " \
                         "                  and groupname = r.groupname) = 0)            " \
                         " order by itempos, groupname, case_index                       "
    else:
        casesDetailSql = "select serialno,itempos,sdkposdesc,groupname,case_index,casename,funcid,funcname,cmdstring,result,message,detail from tbl_cases_run_result order by itempos,groupname,case_index"
    # casesDetailSql = "select serialno,itempos,sdkposdesc,groupname,case_index,casename,funcid,funcname,cmdstring,result,message,detail from tbl_cases_run_result order by itempos,groupname,case_index"
    detailDS = executeCaseSQL(casesDetailSql)
    lastItempos = ''
    lastGroupName = ''
    i = 0
    casesindex = 0
    for detailRec in detailDS:
        i = i + 1
        casesindex = casesindex + 1
        html.write("""    
    <hr />""")
        if str(detailRec["case_index"]) == '1':
            casesindex = 1
            html.write("""
    <h1><a name='""" + str(detailRec["itempos"]) + """_""" + str(detailRec["groupname"]) + """'>""" + str(
                detailRec["itempos"]) + """_""" + str(detailRec["groupname"]) + """</a></h1>""")
        html.write("""
    <a href="#top" class="backToTop">Back to top</a>
    <h2></h2>
    <table class="data" width="1800px">
        <thead>
            <tr>
                <th class="keyIcon" width="50px">序号</th>
                <th class="keyIcon" width="50px">itempos</th>
                <th class="keyIcon" width="150px">案例描述</th>
                <th class="leftDbIconSmall" width="120px">案例组名称</th>
                <th class="leftDbIconSmall" width="50px">步骤</th>
                <th class="leftDbIconSmall" width="100px">案例名称</th>
                <th class="leftDbIconSmall" width="50px">功能号</th>
                <th class="leftDbIconSmall" width="100px">功能号名称</th>
                <!--<th class="leftDbIconSmall" width="330px">请求参数</th>-->
                <th class="rightDbIconSmall" width="50px">执行结果</th>
                <th class="rightDbIconSmall" width="50px">执行结果描述</th>
                <!--<th class="rightDbIconSmall" width="700px">返回值</th>-->
            </tr>
        </thead>
        <tbody>""")

        html.write("""      
            <tr>
                <td>""" + str(casesindex) + """</td>
                <td>""" + str(detailRec["itempos"]) + """</td>
                <td style="word-break:break-all">""" + str(detailRec["sdkposdesc"]) + """</td>
                <td style="word-break:break-all">""" + str(detailRec["groupname"]) + """</td>
                <td>""" + str(detailRec["case_index"]) + """</td>
                <td>""" + str(detailRec["casename"]) + """</td>
                <td>""" + str(detailRec["funcid"]) + """</td>
                <td>""" + str(detailRec["funcname"]) + """</td>
                <!--<td style="word-break:break-all">""" + str(detailRec["cmdstring"]) + """</td>-->""")
        if str(detailRec["result"]) == 'fail':
            html.write("""
                 <td bgcolor="#FF0000">""" + str(detailRec["result"]) + """</td>""")
        else:
            html.write("""
                 <td>""" + str(detailRec["result"]) + """</td>""")
        html.write("""
                <td>""" + str(detailRec["message"]) + """</td>
                <!--<td style="word-break:break-all">""" + str(detailRec["detail"]) + """</td>-->
            </tr> """)

        # if i == len(detailDS)-1:
        html.write("""      
        </tbody>
    </table><p /><p />""")

        # 请求参数处理  ------------------>  begin
        cmdlist = str(detailRec["cmdstring"]).strip(" ").split(',')
        column_name_list = []
        value_list = []
        for j, item in enumerate(cmdlist):
            cmdStr = cmdlist[j].split(":")
            if len(cmdStr) == 2:
                field, value = cmdStr
                # 如果值中有&comma&，予以替换
                value = value.replace('&comma&', ',')
            else:
                field = cmdStr[0]
                value = ""

                for cindex, cmditem in enumerate(cmdStr):
                    if cindex > 0:
                        if cindex == 1:
                            value = cmditem
                        else:
                            value = value + ":" + cmditem
                    # 如果值中有&comma&，予以替换
                    value = value.replace('&comma&', ',')

            column_name_list.append(field)
            value_list.append(value)

        if len(column_name_list) > 0:
            valueHtml = """"""

            html.write(""" 
    <table class="data">""")
            for index, col in enumerate(column_name_list):
                if index == 0:
                    html.write(""" 
        <tr><th colspan='""" + str(
                        len(column_name_list)) + """' bgcolor="#E6E6F2" align="left">请求参数&nbsp;&nbsp;""" + str(
                        detailRec["cmdstring"]) + """</th></tr>
        <tr> """)

                html.write(""" 
            <th>""" + col + """</th>""")

                if index == len(column_name_list) - 1:
                    html.write(""" 
        </tr> """)

            for index, col in enumerate(value_list):
                if index == 0:
                    html.write(""" 
        <tr> """)

                html.write(""" 
            <td>""" + col + """</td>""")

                if index == len(column_name_list) - 1:
                    html.write(""" 
        </tr> """)
            html.write(""" 
    </table><p /><p />""")
        # 请求参数处理  ------------------>  end

        # 返回值处理  ------------------>  begin
        detaillist = str(detailRec["detail"]).strip(" ").split('\r\n')
        # print detaillist

        return_column_list = []
        return_value_list = []
        for j, item in enumerate(detaillist):
            if j == 0:
                columnlist = item.split(",")
                for column in columnlist:
                    return_column_list.append(column)
            else:
                return_value_list.append(item)

        html.write(""" 
    <table class="data">""")
        for index, col in enumerate(return_column_list):
            if index == 0:
                html.write(""" 
        <tr><th colspan='""" + str(len(return_column_list)) + """' bgcolor="#E6E6F2" align="left">返回值</th></tr>
        <tr> """)

            html.write(""" 
            <th>""" + col + """</th>""")

            if index == len(return_column_list) - 1:
                html.write(""" 
        </tr> """)

        for index, col in enumerate(return_value_list):
            html.write(""" 
        <tr> """)
            valuelist = col.split(",")
            for value in valuelist:
                html.write(""" 
            <td>""" + value + """</td>""")

            html.write(""" 
        </tr> """)
        html.write(""" 
    </table><p /><p />""")


def exportCoverageHtml(html_path):
    html = open(html_path, 'w')

    html.write("""<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">
<html xmlns="http://www.w3.org/1999/xhtml">
  <head>
  <meta http-equiv="Content-Type" content="text/html; charset=gbk"/>
  <title>功能号覆盖情况报告</title>
  <style type="text/css">
  <!--
body {font-family:Arial, Helvetica, sans-serif; font-size:80%; line-height:140%; color:#020202;}
      #casesRunDate {text-align:center;}
      #casesRunDetail {text-align:center;}
      #topAnchor {position:absolute; top:0; left:0; overflow:hidden; height:0; width:0;}
      #dataSource {margin:10px auto;}
        .dbImage {padding:0 5px 0 0;}
        .leftSchemas {font-size:150%;color:#666; background:#FFF; text-align:right;}
        .rightSchemas {font-size:150%;color:#666; background:#FFF; text-align:left;}
        .schemaSplitter {text-align:center; padding:10px;}
        .leftIdentifier  {font-size:140%;line-height:100%; text-align:right;}
        .rightIdentifier {font-size:140%; line-height:100%; text-align:left;}
        
      #runDetail {margin:0 left; text-align:left;}
      thead {background:#E6E6F2}
      td, th {padding:2px 10px;}
      tr:nth-child(2n) {background:#f8f8f8;}
    
      #runDetail a, .backToTop {text-decoration:none; color:#2f80cc; text-indent:10px; display:block;}
      #runDetail a:hover, .backToTop:hover {text-decoration:underline;} 
    
      .backToTop {text-indent:0px;}
    
      hr {background:#CCC; border:0 none; margin:20px 0;}
    
      h1 {font-size:140%;}
      h2 {font-size:130%; color:#666;}
    
      .data {text-align:left;}
      .data td:nth-child(2n), .data th:nth-child(2n) {border-right:4px solid #f1f1f1; margin-right:15px;}
    
      .different {background:#2f80cc; color:#FFF}
      .caseRunRecord {background:#2f80cc; color:#FFF}
      .caseRunRecordDetail {background:#2f80cc; color:#FFF}
  
      .leftDbIconSmall {background:transparent url(data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABAAAAAQCAYAAAAf8/9hAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAgY0hSTQAAeiYAAICEAAD6AAAAgOgAAHUwAADqYAAAOpgAABdwnLpRPAAAAAlwSFlzAAALEQAACxEBf2RfkQAAAwJJREFUOE91kn1Ik1EYxd+/wggig1AQRTEqRFAURVk4jERRFMUwrMRQLIdiTJTEsZhMDGXgmAiTQFFmmDJBEsdEURyKouzN6dpwTbe21Hdzn36kmZ3utjSGdeH575zffc65l0q9n0tlsItCCgsrn5aV17+vb2xf5bd0M018MVPJ4a3lFT//kJiWWR4bn3At8vZdKiwiigJwMVRyWtY9Lle4rFKtQ6tlQK9aoZo3QKHUYGR0Cb0DMxC09SOvuOJTROydhEuAiqoGlcWyD6v1EDqdDcsrZszM6TGuoDE8uugHdEkVaBPJkZlTog6LjLoSBBBL+j0McwyTyQ2tjrkMkM1AIp1AW4ccJWX1nvCYmJtBAL6g0+lwnPo38AGWlk2XAF09CrSKRpBV8MwZFh0dDKjiNDEbhj14vb9gMruxqtmGaoF0MKnBkHwB0l4lBG8HwOGKkP4gn7kVGXkjCFBdy7etqM3Q6/f8xrEJtf9WUdcY3rTJwG3qxou6dlTUCMHOeWQPj44JDQLwSAStnmRXb0I5rcHA0BzaxaNoFg6gnteD2kaxH/C/oTh1fJtx0wWL1Yu5eT1GxhYh7VOiQzJKcg+R9WV+oe+c/TzDyckpDo+O4fYcBgBVNc27atoCo9GFpRUTxpU0+gZn0dn9kQCG0ch/5xeeEvOTyuaLse95AgBfBFpjgZo2Y3J6HbIhFUSSMfCEMjQQ83mE4+Mf2D/4DpdrHza7G9ZtRwDAbWz1+No3mhyYml3H4IgKkp5xCMn6fPIDea19fqH34AgOYj7fYMvEBAD5xeXL9KoF29uHpEgzFFMaApknbz9BypSjuSUAcDi92GVc+Grdw+bWLvSGbwFAXGJ6IqeWp52apjFLPtCwfBHi7nESYRBNLf149ecZd3adQR2sfTYHAPFJLCop9eH1jMyil6VldYrqOoGhskZgL614bS94zPmSkV086ROaLTYYjDvQbVihIWZas/UXkMbKpVgZ+VQqK4dKYWWHpLJzQ1PY2aEJ6eyrccnplE/472nHbwVfrTjUnrWTAAAAAElFTkSuQmCC) right center no-repeat; padding-right:25px;}
      .rightDbIconSmall {background:transparent url(data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABAAAAAQCAYAAAAf8/9hAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAgY0hSTQAAeiYAAICEAAD6AAAAgOgAAHUwAADqYAAAOpgAABdwnLpRPAAAAAlwSFlzAAALEQAACxEBf2RfkQAAAsRJREFUOE91k0tME0EYx5ftc1taaKHlUSgUEFAQA4oe1AiJGkEPPhL1IF6MXkw8eDGcPWiC0QhGeQSMIjWKCkZQQ+QlCEHRgIAoKQgqobTIFmJJ7Xb7d7rYmqY4yZfJJjO/+f3nm6UABFUlRVFDJXuVr7dl17cYoucbVXK3WUbjnpSUjHY3MCKrOVx8524MpfStDQH052bubzbGufoz4jGRyGAyicGXZAbDRgZ98XK81EvxQCVGrYR2VTOi4hDAU4POOVe0BT93roNtgwY/UhSwEMAoAQ0SYI9BjrY4OZp0MlRLaGcI4KFOg+VjBVjcnQ37phjMpSowZVLgEwF8IIC+BDnaiUlrrBzVYjp4swDQa+E8XwK2OA8LW42YJ4AZAvDHGEhg0EUsnhOLqrUA9zUquGsu4be5Eo6TxbBvNmA2UwNLmhqjqSq8M6nQkahEY4wMN0VhoYB6tQJLV0rBXS3Fr7NHwB7Ihy1Hj6/pkfhoCkcPMfDp10VLccMHqFBL8L+aOFyAb4VZsGWoMWli8IaoP9ZJcY104XqEZPUOfAsDg+cBtxte14oAGN21Ed2pUXiilcIcKUZjlBTNRL0jQYlHxOKW3wAeLuhkr2NRmJ3njoM9tB22fCOms/QYSYtEb6ICz8lms172L4J3xQnvsgNedgFeuxW89bsAYM8cBFuUG4gwRLrQHkf6r5HgMokhdEEALNrB2+YCBvzUZ2Huyk1BS5IWTX8j3I4Qo45sfhbLoIlYVPoN+NkZ8DMW8JZxeMaH4RkZXI1w4RQcRwth25GO6TwDxtZHoy85HC/iGTSQ+ygXhS0JAI9lLHC68D3QLczWE/tg25MTiPCevMJX5BLriX45Tbku0tTpVcDwW3CDveD6O8H1tIHrbBUAtSrGUSeXcLXiMNQQ3SoRxVWIKLaMpjrLKEor/I3+U9cqMkIeWnCB+gOP0TjDKUQRBgAAAABJRU5ErkJggg==) left center no-repeat; padding-left:25px;}
      .keyIcon {background:transparent url(data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABAAAAAQCAYAAAAf8/9hAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAgY0hSTQAAeiYAAICEAAD6AAAAgOgAAHUwAADqYAAAOpgAABdwnLpRPAAAAAlwSFlzAAALEQAACxEBf2RfkQAAAeNJREFUOE+VkktIW0EUhieJlFizsujC92Nh6a64cVUXgqBSN1W68IGCMTVBqwuX3bgSFd1koTE3gl0outGqpWSlUCuKlipNGzB6G6jREFHB58293t+Z60SUK23ywWGGmfP/Z+ZwSAyXlfROfkg//dRXoM715yljnebASCsp4df/ZrzT8tO3VIdraQDHYhsu/zbhPFCNz31pGLYRK097HFrF+ftrI4LL5aB7bcWJQ1tlfzFme56q9Dybp+uZ7smKKlGnJqYwQ82AxYW/FPveZ6Dfc/N0Pd7hlwhvdzx8wVErcNgEhF4jumaG4DBt8XQ984NFOBJvK2rBxJFG4KAWCL6CtJIEt924wdP1fOy2nEphO6T9Fi5uoOI3wF4l4MtAaMEMt42s83Q9oy2kftFVhKvdKkhilfZsBEupOBPyqglysAIu221/WDwKbdJb71C6Gll+DvzKhrJuQfiLGdJmIbankyH/icNksIuksm57Oky7nnajT3AYznyTqfBPWfBDMEIWy/5vch86QC8Eu+FccJCD8ffE/82ZpCp7NYmZ0P7ke5pJ2lQtecKEgblcKIdtiZnEYCIW3ycKVCVivTNJCDpoOaPvDGpgNhOx7/Cr+GHTqu6UQJxJ0SaXH8cPq8qELFw2ghvl51sKtlRwUgAAAABJRU5ErkJggg==) 5px center no-repeat; padding-left:25px;}
      -->
      </style>
</head>
<body> 
    """)

    # 功能号覆盖情况
    runCaseList = {}
    runCasesSql = "select funcid,                              " \
                  "       funcname,                            " \
                  "       count(funcid) count,                 " \
                  "       (select count(funcid)                " \
                  "          from tbl_cases_run_result         " \
                  "         where result = 'pass'              " \
                  "           and funcid = r.funcid) succount, " \
                  "       (select count(funcid)                " \
                  "          from tbl_cases_run_result         " \
                  "         where result = 'fail'              " \
                  "           and funcid = r.funcid) failcount " \
                  "  from tbl_cases_run_result r               " \
                  " group by funcid                            " \
                  " order by funcid                            "
    ds = executeCaseSQL(runCasesSql)

    tIndex = 1
    for rec in ds:
        runCaseList[str(rec["funcid"])] = {}
        runCaseList[str(rec["funcid"])]["funcname"] = str(rec["funcname"])
        runCaseList[str(rec["funcid"])]["totalcount"] = str(rec["count"])
        runCaseList[str(rec["funcid"])]["successcount"] = str(rec["succount"])
        runCaseList[str(rec["funcid"])]["failcount"] = str(rec["failcount"])

    if len(runCaseList) > 0:
        html.write("""
    <p id="casesCoverage"><H1 align="center">功能号已覆盖情况汇总</H1></p>
    <table width="1600px" id="casesCoverageDetail">
        <thead>
            <tr>
                <th>系统支持功能号</th>
                <th>功能号描述</th>
                <th>覆盖情况</th>
                <th>案例个数</th>
                <th>成功个数</th>
                <th>失败个数</th>""")
        if len(runCaseList) > 1:
            html.write("""
                <th>系统支持功能号</th>
                <th>功能号描述</th>
                <th>覆盖情况</th>
                <th>案例个数</th>
                <th>成功个数</th>
                <th>失败个数</th>""")
        html.write("""
            </tr>
        </thead>""")
        for funcid in runCaseList.keys():
            if tIndex == 1 or len(runCaseList) == 1:
                html.write("""
            <tr align="center">
                <td>""" + funcid + """</td>
                <td>""" + runCaseList[funcid]['funcname'] + """</td>
                <td bgcolor="28FF28">覆盖</td>
                <td>""" + runCaseList[funcid]['totalcount'] + """</td>
                <td>""" + runCaseList[funcid]['successcount'] + """</td>""")
                if runCaseList[funcid]['failcount'] != "0":
                    html.write("""
                <td bgcolor="#FF0000">""" + runCaseList[funcid]['failcount'] + """</td>""")
                else:
                    html.write("""
                <td>0</td>""")

                tIndex = tIndex + 1
            else:
                html.write("""
                <td>""" + funcid + """</td>
                <td>""" + runCaseList[funcid]['funcname'] + """</td>
                <td bgcolor="28FF28">覆盖</td>
                <td>""" + runCaseList[funcid]['totalcount'] + """</td>
                <td>""" + runCaseList[funcid]['successcount'] + """</td>""")
                if runCaseList[funcid]['failcount'] != "0":
                    html.write("""
                <td bgcolor="#FF0000">""" + runCaseList[funcid]['failcount'] + """</td>""")
                else:
                    html.write("""
                <td>0</td>
            </tr>""")

                tIndex = 1

        html.write("""
    </table></br>""")

    if len(gl.g_SysCasesList) > 0:
        html.write("""
    <p id="casesNoCoverage"><H1 align="center">功能号未覆盖情况汇总</H1></p>
    <table width="1600px" id="casesNoCoverageDetail">
        <thead>
            <tr>
                <th>系统支持功能号</th>
                <th>功能号描述</th>
                <th>覆盖情况</th>
                <th>案例个数</th>
                <th>成功个数</th>
                <th>失败个数</th>
                <th>系统支持功能号</th>
                <th>功能号描述</th>
                <th>覆盖情况</th>
                <th>案例个数</th>
                <th>成功个数</th>
                <th>失败个数</th>
            </tr>
        </thead>""")
    funcindex = 1
    isCover = ""
    funcname = ""
    totalcount = "0"
    successcount = "0"
    failcount = "0"
    # for index,funcid in enumerate(gl.g_SysCasesList):
    for index, funcid in enumerate(gl.g_SysCasesList.keys()):
        funcname = gl.g_SysCasesList[funcid]["method"]

        if runCaseList.has_key(funcid):
            # isCover = "覆盖"
            # totalcount = runCaseList[funcid]['totalcount']
            # successcount = runCaseList[funcid]['successcount']
            # failcount = runCaseList[funcid]['failcount']
            continue
        else:
            isCover = "未覆盖"
            totalcount = "0"
            successcount = "0"
            failcount = "0"
        if funcindex == 1:
            html.write("""
                <tr align="center">
                    <td>""" + funcid + """</td>
                    <td>""" + funcname + """</td>""")
            if runCaseList.has_key(funcid):
                html.write("""
                    <td bgcolor="28FF28">""" + isCover + """</td>""")
            else:
                html.write("""
                    <td>""" + isCover + """</td>""")
            html.write("""
                    <td>""" + totalcount + """</td>
                    <td>""" + successcount + """</td>""")
            if failcount != "0":
                html.write("""
                    <td bgcolor="#FF0000">""" + failcount + """</td>""")
            else:
                html.write("""
                    <td>""" + failcount + """</td>""")

            funcindex = funcindex + 1
        else:
            html.write("""
                    <td>""" + funcid + """</td>
                    <td>""" + funcname + """</td>""")
            if runCaseList.has_key(funcid):
                html.write("""
                    <td bgcolor="28FF28">""" + isCover + """</td>""")
            else:
                html.write("""
                    <td>""" + isCover + """</td>""")
            html.write("""
                    <td>""" + totalcount + """</td>
                    <td>""" + successcount + """</td>""")
            if failcount != "0":
                html.write("""
                    <td bgcolor="#FF0000">""" + failcount + """</td>
                </tr>""")
            else:
                html.write("""
                    <td>""" + failcount + """</td>
                </tr>""")

            funcindex = 1

    if len(gl.g_SysCasesList) > 0:
        html.write("""
    </table>""")


def runingResultBy_need_log(running_dir):
    """
    通过配置文件中的日志设置来决定是否保留该案例运行日志

    该标志只针对全部成功的案例组下的子案例，有错误运行记录的案例组必须全部保留日志.

    Args:
        running_dir: 运行目录

    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息
        """
    try:
        cwd = os.getcwd()
        xml_total_report_path = os.path.join(cwd, "Autotest_report\\Autotest_Total_report\\%s.xml" % running_dir)

        doc = etree.ElementTree(file=xml_total_report_path)
        root = doc.getroot()
        i = 0
        if root.attrib[u"运行状态"].startswith(u"全部失败"):
            return 0, ""
        for InterfaceItem in root:  # 轮询每一个接口
            if InterfaceItem.attrib[u"运行状态"].startswith(u"全部成功"):

                for groupItem in InterfaceItem:
                    if i == 1:
                        i = 0
                        break
                    if groupItem.attrib[u"运行状态"].startswith(u"全部成功"):
                        for caseItem in groupItem:
                            need_log = getFuncidNeed_log(caseItem.attrib["funcid"])
                            if need_log == 0:
                                if caseItem.attrib["funcid"] in InterfaceItem.attrib[u"业务描述"]:
                                    InterfaceItem.getparent().remove(InterfaceItem)
                                    i = 1
                                    break
                                else:
                                    caseItem.getparent().remove(caseItem)
        xml_error_report_path = os.path.join(cwd, "Autotest_report\\Autotest_Total_report\\%s.xml" % running_dir)
        doc.write(xml_error_report_path, pretty_print=True, xml_declaration=True, encoding='utf-8')
    except:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        # 保存对象待查
        xml_path = os.path.join(cwd, "Autotest_report\\Autotest_Total_report\\%s.xml" % running_dir)
        f = file(xml_path, 'w')
        cPickle.dump(doc, f)
        f.close()
        return -1, exc_info
        # print (time.time() - time1)


def runningErrorResult_to_xml(running_dir):
    """
    从运行报告中导出错误报告

    移除运行报告中全部成功的部分，剩下的就是需要的部分。
    Args:
        running_dir: 运行目录

    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息
    """

    try:
        cwd = os.getcwd()
        xml_total_report_path = os.path.join(cwd, "Autotest_report\\Autotest_Total_report\\%s.xml" % running_dir)

        doc = etree.ElementTree(file=xml_total_report_path)
        root = doc.getroot()

        if root.attrib[u"运行状态"].startswith(u"全部成功"):
            return 0, ""
        for InterfaceItem in root:  # 轮询每一个接口
            if InterfaceItem.attrib[u"运行状态"].startswith(u"全部成功"):
                InterfaceItem.getparent().remove(InterfaceItem)
            else:
                for groupItem in InterfaceItem:
                    if groupItem.attrib[u"运行状态"].startswith(u"全部成功"):
                        groupItem.getparent().remove(groupItem)

        xml_error_report_path = os.path.join(cwd, "Autotest_report\\Autotest_Error_report\\%s_Error.xml" % running_dir)
        doc.write(xml_error_report_path, pretty_print=True, xml_declaration=True, encoding='utf-8')

        # 生成执行html报告
        cover_path = os.path.join(cwd, "Autotest_report\\Autotest_Coverage_html_report\\%s.html" % running_dir)

        html_path = os.path.join(cwd, "Autotest_report\\Autotest_Error_report\\%s_Error.html" % running_dir)
        exportHtml(html_path, 'fail', cover_path)

        return 0, ""
    except:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)

        # 保存对象待查
        xml_path = os.path.join(root, "Autotest_report\\Autotest_Error_report\\%s_Error.pkl" % running_dir)
        f = file(xml_path, 'w')
        cPickle.dump(doc, f)
        f.close()
        return -1, exc_info
        # print (time.time() - time1)


def runningCorrectResult_to_xml(running_dir):
    """
    从运行报告中导出正确报告

    移除运行报告中全部错误和部分错误的部分，剩下的就是需要的部分。
    Args:
        running_dir: 运行目录

    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息
    """

    try:
        cwd = os.getcwd()
        xml_total_report_path = os.path.join(cwd, "Autotest_report\\Autotest_Total_report\\%s.xml" % running_dir)

        doc = etree.ElementTree(file=xml_total_report_path)
        root = doc.getroot()

        if root.attrib[u"运行状态"].startswith(u"全部成功"):
            shutil.copyfile(os.path.join(cwd, "Autotest_report\\Autotest_Total_report\\%s.xml" % running_dir),
                            os.path.join(cwd, "Autotest_report\\Autotest_Correct_report\\%s_Correct.xml" % running_dir))

            # 生成执行html报告
            cover_path = os.path.join(cwd, "Autotest_report\\Autotest_Coverage_html_report\\%s.html" % running_dir)

            html_path = os.path.join(cwd, "Autotest_report\\Autotest_Correct_report\\%s_Correct.html" % running_dir)
            exportHtml(html_path, 'succ', cover_path)

            return 0, ""
        for InterfaceItem in root:  # 轮询每一个接口
            if InterfaceItem.attrib[u"运行状态"].startswith(u"部分成功"):
                InterfaceItem.getparent().remove(InterfaceItem)
            elif InterfaceItem.attrib[u"运行状态"].startswith(u"全部失败"):
                InterfaceItem.getparent().remove(InterfaceItem)
            else:
                continue

        xml_correct_report_path = os.path.join(cwd,
                                               "Autotest_report\\Autotest_Correct_report\\%s_Correct.xml" % running_dir)
        doc.write(xml_correct_report_path, pretty_print=True, xml_declaration=True, encoding='utf-8')

        # 生成执行html报告
        cover_path = os.path.join(cwd, "Autotest_report\\Autotest_Coverage_html_report\\%s.html" % running_dir)

        html_path = os.path.join(cwd, "Autotest_report\\Autotest_Correct_report\\%s_Correct.html" % running_dir)
        exportHtml(html_path, 'succ', cover_path)

        return 0, ""
    except:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)

        # 保存对象待查
        xml_path = os.path.join(root, "Autotest_report\\Autotest_Correct_report\\%s_Correct.pkl" % running_dir)
        f = file(xml_path, 'w')
        cPickle.dump(doc, f)
        f.close()
        return -1, exc_info


def getFuncidTimeout(funcid):
    '''
    根据功能号获取超时设置

    返回系统需要设置的值，如果没有找到对应功能号，或没有具体的超时设置就返回0
    '''
    if gl.g_interfaceTree == None:
        gl.g_interfaceTree = ET.parse(".\\Adapter\\T2\\T2Interfaces.xml")
    root = gl.g_interfaceTree.getroot()
    xpath = ".//Interface[@funcid='%s']" % (funcid)
    InterfaceElem = root.find(xpath)
    if InterfaceElem == None:
        return 0
    if not InterfaceElem.attrib.has_key("timeout") or InterfaceElem.attrib["timeout"] == "":
        return 0
    else:
        return int(InterfaceElem.attrib["timeout"])


def getFuncidNeed_log(funcid):
    '''
    根据功能号获取是否记录日志标签

    返回系统需要设置的值，如果没有找到对应功能号，或没有具体的日志标签就返回0
    '''
    if gl.g_interfaceTree == None:
        gl.g_interfaceTree = ET.parse("T2Interfaces.xml")
    root = gl.g_interfaceTree.getroot()
    xpath = ".//Interface[@funcid='%s']" % (funcid)
    InterfaceElem = root.find(xpath)
    if InterfaceElem == None:
        return 0
    if not InterfaceElem.attrib.has_key("need_log") or InterfaceElem.attrib["need_log"] == "":
        return 0
    else:
        return int(InterfaceElem.attrib["need_log"])


def getInparamListFromCmdstring(funcid, cmdstring):
    """
    将‘|’分割的命令行字符串转化为使用','分割的“字段名：字段值”的字符串，便于参看

    Args:
        funcid: 功能id
        cmdstring: 命令串

    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        inParam : 转化后的字符串
    """
    try:
        # 命令行值列表
        cmdList = cmdstring.split(',')

        # cmd_list = cmdstring.split('|')
        cmd_list = []
        for i in xrange(len(cmdList)):
            cmdStr = cmdList[i].split(":")
            if len(cmdStr) == 2:
                field, value = cmdStr
                # 如果值中有&comma&，予以替换
                value = value.replace('&comma&', ',')
            else:
                field = cmdStr[0]
                value = ""

                for cindex, cmditem in enumerate(cmdStr):
                    if cindex > 0:
                        if cindex == 1:
                            value = cmditem
                        else:
                            value = value + ":" + cmditem
                    # 如果值中有&comma&，予以替换
                    value = value.replace('&comma&', ',')

            cmd_list.append(value)

        # 获取字段列表
        root = gl.g_interfaceDetailTree.getroot()
        column_name = []
        for interfaceElem in root:
            if interfaceElem.attrib["funcid"] == funcid:
                for input in interfaceElem.iter("in"):
                    column_name.append(input.attrib["name"])

        inParam = ",".join(
            "%s:%s" % (column_name[i], cmd_list[i]) for i in xrange(min(len(column_name), len(cmd_list))))
        return 0, inParam

    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error("异常发生：%s" % (exc_info))
        return -1, exc_info


def getExceptionInfo():
    exc_type, exc_value, exc_traceback = sys.exc_info()
    more_info = ""
    if len(exc_value.args) == 2:
        more_info = exc_value.args[1].decode('gbk')
    traceback_details = {
        'filename': exc_traceback.tb_frame.f_code.co_filename,
        'lineno': exc_traceback.tb_lineno,
        'name': exc_traceback.tb_frame.f_code.co_name,
        'type': exc_type.__name__,
        'message': exc_value.message,  # or see traceback._some_str()
        'more_info': more_info
    }
    # logging.error("%s" % (gl.g_traceback_template % traceback_details))
    return "%s" % (gl.g_traceback_template % traceback_details)


def getInParamList(funcid):
    interface_detail_root = gl.g_interfaceDetailTree.getroot()
    xpath = "Interface[@funcid='%s']" % (funcid)
    interface_elem = interface_detail_root.find(xpath)
    if interface_elem == None:
        msg = "funcid:%s not exist in document" % (funcid)
        print msg
        return -1, msg

    xpath = "InputParams//in"
    interface_in_list = interface_elem.findall(xpath)

    cmd_list = []
    for input in interface_in_list:
        cmd_list.append(input.attrib["name"])
    return 0, cmd_list


def addFixParam(in_param_list):
    fix_param_list = ['funcid', 'custid', 'custorgid', 'trdpwd', 'netaddr', 'orgid', 'operway', 'ext']
    return fix_param_list + in_param_list


def addFixParamName(in_param_name_list):
    fix_param_name_list = [u'功能号', u'客户代码', u'客户机构', u'交易密码', u'操作站点', u'操作机构', u'操作方式', u'扩展']
    return fix_param_name_list + in_param_name_list


# def updateXmlout(distlist):
#     x = -1
#     doc = etree.ElementTree(file='T2InterfaceDetails.xml')
#     root = doc.getroot()
#     i = 0
#     for funcname in distlist:
#         list = string[:distlist[funcname].find('\r') - 0].split(',')
#         xpath = "Interface[@funcid='%s']" % funcname
#         interface_elem = root.find(xpath)
#         xpath = "Interface[@funcid='%s']/OutParams" % funcname
#         OutputParams_elem = root.find(xpath)
#         if OutputParams_elem == None:
#             x = 0
#             schemaOut = etree.SubElement(interface_elem, "OutParams")
#             for n in list:
#                 xpath = "Interface[@funcid='%s']/OutParams/out[@name='%s']" % (funcname, n)
#                 elemOut = root.find(xpath)
#                 if elemOut == None:
#                     etree.SubElement(schemaOut, "out", name=n, type="", desc="", disableflag="0")
#                     i += 1
#         else:
#             for i, item in enumerate(OutputParams_elem):
#                 if OutputParams_elem[i].attrib['name'] not in list:
#                     x = 0
#                     OutputParams_elem.remove(OutputParams_elem[i])
#             for n in list:
#                 xpath = "Interface[@funcid='%s']/OutParams/out[@name='%s']" % (funcname, n)
#                 elemOut = root.find(xpath)
#                 if elemOut == None:
#                     x = 0
#                     etree.SubElement(OutputParams_elem, "out", name=n, type="", desc="", disableflag="0")
#                     i += 1
#     if x == 0:
#         info = u"检测到当前组有%d个案例配置文件中出参信息不准确是否需要更新~！" % i
#         dlg = wx.MessageDialog(None, info, u"标题信息", wx.YES_NO | wx.ICON_QUESTION)
#         if dlg.ShowModal() == wx.ID_YES:
# 格式化配置文件
#             doc = etree.ElementTree(root)  # root 为根元素
#             doc.write("T2InterfaceDetails.xml", pretty_print=True, xml_declaration=True, encoding='utf-8')
#             updateT2InterfaceDetails.copyFile()
#         dlg.Destroy()

def action_update(action_list, VARIABLE):
    try:
        cases = []
        actions = []
        all_action = []
        child_action = []
        flag = False
        for action in action_list:
            action = action.strip()
            if action.strip() == "":
                continue
            if action.startswith('"'):
                action = action[1:]
            if action.startswith('$bfh$'):
                action = action[5:]
            if "normal_RT_EXPRESSION" in action:
                action = action.replace("normal_RT_EXPRESSION", "normal_call:ASSERT_EXPRESSION")
            if action.startswith("="):
                a = []
                a.append(action)
                all_action.append(a)
                continue
            if flag:
                flag = False
                child_action = []
            if action.startswith("#") or action.startswith("--"):
                child_action.append(action)
            else:
                child_action.append(action)
                flag = True
                all_action.append(child_action)
        if action_list and not all_action:
            child_action = []
            child_action.append(",".join(action_list).strip())
            all_action.append(child_action)
        for i in all_action:
            action_flag = False
            retrun_json = {"call_when": "", "type": "", "content": "",
                           "desc": "", "PUBLIC_FUNC_FLAG": 0}
            for j in i:
                if 'expression_evaluation' in j:
                    index_ = j.find("=")
                    param1 = j[:index_]
                    param2, param3 = param1.split(":")
                    param2 = param2.strip()
                    param3 = param3.strip()
                    param4 = j[index_ + 1:].replace("expression_evaluation", "").strip()[1:-1]
                    j = param2 + ":expression_evaluation(" + param3 + "," + param4 + ")"
                if "mobile_data_by_normalOfTable_to_dormancyOfTable" in j:
                    j = "normal_call:" + j
                if j.startswith("="):
                    retrun_json["desc"] = ",".join(i)
                    action_flag = True
                    actions.append(retrun_json)
                    continue
                if not j.startswith("#") or not j.startswith("--"):
                    action_flag = True
                    action = j
                    action = action.split("--")[0]
                    # print "huozhi %s" % action
                    return_list_str = ""
                    return_list = []
                    if "oracle_Database" in j:
                        index1 = action.rfind(",")
                        # 获取用户
                        action_1 = action[:index1]
                        user = action[index1:].replace(",", "").replace(")", "")
                        # 判断是否有变量
                        if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
                            return_list_str = action[action.find(":") + 1:action.find("=")].strip()
                            # param_list_str = action[action.find("=") + 1:].strip()
                        else:
                            param_list_str = action[action.find(":") + 1:].strip()
                        # func = param_list_str[:param_list_str.find("(")]
                        # param_info = param_list_str[param_list_str.find("(") + 1:param_list_str.rfind(")")]
                        if return_list_str != "":
                            return_list = return_list_str.split(',')

                        func = str(user.strip()) + ".exec_sql"
                        func = func.replace('"', "")
                        action = action_1.replace("oracle_Database", func) + ")"
                        if return_list:
                            index4 = action.find("=")
                            action = action[:index4] + "&equ&" + action[index4 + 1:]
                            for param in return_list:
                                param = param.strip()
                                VARIABLE[param] = ""
                        if 'SELECT' in action.upper() and 'DELETE' not in action.upper():
                            action_2 = action.upper()
                            field = "field"
                            if "FROM" not in action_2:
                                index9 = action.rfind(")")
                                action = action[:index9] + " as field " + action[index9:]
                                print action
                            else:
                                index3 = action_2.find("FROM")
                                index8 = action_2.find("SELECT")
                                # action_select = action_2[:index3]
                                if action_2.count("SELECT") == 2 and "TOP 1 *" in action_2:
                                    field = action_2.split("SELECT")[2].split("FROM")[0].replace("TOP 1", "")
                                else:
                                    field = action[index8 + 6:index3].strip()
                                if " as " in field:
                                    field = field.split(" as ")[1].strip()
                                if " AS " in field:
                                    field = field.split(" AS ")[1].strip()
                                if "ltrim(rtrim" in field:
                                    field = field.replace("ltrim(rtrim", "").replace("(", "").replace(")", "").strip()
                                if 'TO_CHAR' in field.upper():
                                    index6 = field.find(".")
                                    field_1 = field[index6 + 1:]
                                    index7 = field_1.find(")")
                                    field_2 = field_1[:index7]
                                    field = field_2
                                    action = action[:index8] + action[index8:index3] + "as %s " % field_2 + action[
                                                                                                            index3:]
                                elif "ROUND" in field.upper():  # fixme
                                    # 'round( t.en_last_price,2)'
                                    field = field.strip()
                                    field = field[6:-1]
                                    field = field.split(",")[0]
                                    field = field.split(".")[1]
                                    action = action[:index3] + " as %s " % field + action[index3:]
                                else:
                                    if "," in field:
                                        field = field.split(",")[0].strip()
                                        field = field.split(".")[1]
                                        # print "select many field %s" % field
                                    elif "." in field:
                                        field = field.split(".")[1]
                            field = field.replace("top 1", "").strip()
                            if "*" in field:
                                field = field.split("*")[0]
                            elif "+" in field:
                                field = field.split("+")[0]
                            elif "-" in field:
                                field = field.split("-")[0]
                            elif "\\" in field:
                                field = field.split("\\")[0]

                            if field.count("(") == 1:
                                str_field = field.split("(")[1].replace(")", "").strip()
                                field_lsit = [str(i).strip().lower() for i in str_field.split(" ") if i != ""]
                                if len(field_lsit) == 2 and field_lsit[0] == field_lsit[1]:
                                    field = field_lsit[0]
                                elif len(field_lsit) == 1:
                                    field = field_lsit[0]
                            elif field.count("(") == 2 and field.count("))") == 1:
                                str_field = field.split("(")[2].replace("))", "").strip()
                                field_lsit = [str(i).strip().lower() for i in str_field.split(" ") if i != ""]
                                if len(field_lsit) == 2 and field_lsit[0] == field_lsit[1]:
                                    field = field_lsit[0]
                                elif len(field_lsit) == 1:
                                    field = field_lsit[0]

                            action = action + "&dot&get_value(@DbTable[1][%s]@)" % field
                        index5 = action.find(":")
                        call_when = action[:index5].strip()
                        action = action[index5 + 1:].strip()
                        if call_when == 'call':
                            call_when = "normal_call"
                            # action = action.replace("call", "normal_call")
                        retrun_json["call_when"] = call_when
                        retrun_json["type"] = "DB_Operation"
                        retrun_json["content"] = action
                        retrun_json["PUBLIC_FUNC_FLAG"] = 0
                        if len(i) > 1:
                            retrun_json["desc"] = ",".join(i[:-1])
                    elif "sqlserver_Database" in j:
                        index1 = action.rfind(",")
                        # 获取用户
                        action_1 = action[:index1]
                        user = action[index1:].replace(",", "").replace(")", "")
                        # 判断是否有变量
                        if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
                            return_list_str = action[action.find(":") + 1:action.find("=")].strip()
                            # param_list_str = action[action.find("=") + 1:].strip()
                        else:
                            param_list_str = action[action.find(":") + 1:].strip()
                        # func = param_list_str[:param_list_str.find("(")]
                        # param_info = param_list_str[param_list_str.find("(") + 1:param_list_str.rfind(")")]
                        if return_list_str != "":
                            return_list = return_list_str.split(',')

                        func = str(user.strip()) + ".exec_sql"
                        func = func.replace('"', "")
                        action = action_1.replace("sqlserver_Database", func) + ")"

                        if return_list:
                            index4 = action.find("=")
                            action = action[:index4] + "&equ&" + action[index4 + 1:]
                            for param in return_list:
                                param = param.strip()
                                VARIABLE[param] = ""

                        if 'INSERT' not in action.upper() and 'SELECT' in action.upper() and 'DELETE' not in action.upper():
                            action_2 = action.upper()
                            field = "filed"
                            if "FROM" not in action_2:
                                index9 = action.rfind(")")
                                action = action[:index9] + " as filed " + action[index9:]
                                print action
                            else:
                                index3 = action_2.find("FROM")
                                index8 = action_2.find("SELECT")
                                if action_2.count("SELECT") == 2 and "TOP 1 *" in action_2:
                                    field = action_2.split("SELECT")[2].split("FROM")[0].replace("TOP 1", "")
                                else:
                                    # action_select = action_2[:index3]
                                    field = action[index8 + 6:index3].strip()
                                    # print field
                                    field = field.replace(',*', "")
                                if " as " in field:
                                    field = field.split(" as ")[1].strip()
                                if " AS " in field:
                                    field = field.split(" AS ")[1].strip()
                                # if "ltrim(rtrim" in field:
                                #     field = field.replace("ltrim(rtrim", "").replace("(", "").replace(")", "").strip()
                                if 'TO_CHAR' in field.upper():
                                    index6 = field.find(".")
                                    field_1 = field[index6 + 1:]
                                    index7 = field_1.find(")")
                                    field_2 = field_1[:index7]
                                    field = field_2
                                    action = action[:index8] + action[index8:index3] + "as %s " % field_2 + action[
                                                                                                            index3:]
                                elif "ROUND" in field.upper():  # fixme
                                    # 'round( t.en_last_price,2)'
                                    field = field.strip()
                                    field = field[6:-1]
                                    field = field.split(",")[0]
                                    field = field.split(".")[1]
                                    action = action[:index3] + " as %s " % field + action[index3:]
                                else:
                                    if "," in field:
                                        field = field.split(",")[0].strip()
                                        field = field.split(".")[1]
                                        # print "select many field %s" % field
                                    elif "." in field:
                                        field = field.split(".")[1]
                            field = field.replace("top 1", "").strip()
                            if "*" in field:
                                field = field.split("*")[0]
                            elif "+" in field:
                                field = field.split("+")[0]
                            elif "-" in field:
                                field = field.split("-")[0]
                            elif "\\" in field:
                                field = field.split("\\")[0]

                            if field.count("(") == 1:
                                str_field = field.split("(")[1].replace(")", "").strip()
                                field_lsit = [str(i).strip().lower() for i in str_field.split(" ") if i != ""]
                                if len(field_lsit) == 2 and field_lsit[0] == field_lsit[1]:
                                    field = field_lsit[0]
                                elif len(field_lsit) == 1:
                                    field = field_lsit[0]
                            elif field.count("(") == 2 and field.count("))") == 1:
                                str_field = field.split("(")[2].replace("))", "").strip()
                                field_lsit = [str(i).strip().lower() for i in str_field.split(" ") if i != ""]
                                if len(field_lsit) == 2 and field_lsit[0] == field_lsit[1]:
                                    field = field_lsit[0]
                                elif len(field_lsit) == 1:
                                    field = field_lsit[0]

                            action = action + "&dot&get_value(@DbTable[1][%s]@)" % field
                        # call_when = action.split(":")[0]
                        index5 = action.find(":")
                        call_when = action[:index5].strip()
                        action = action[index5 + 1:].strip()
                        if call_when == 'call':
                            call_when = "normal_call"
                            # action = action.replace("call", "normal_call")
                        retrun_json["call_when"] = call_when
                        retrun_json["type"] = "DB_Operation"
                        retrun_json["content"] = action
                        retrun_json["PUBLIC_FUNC_FLAG"] = 0
                        if len(i) > 1:
                            retrun_json["desc"] = ",".join(i[:-1])
                    elif 'call' in j or "normal_call" in j or "except_call" in j or "final_call" in j:
                        if "normal_call" in j:
                            call_when = "normal_call"
                        elif "except_call" in j:
                            call_when = "except_call"
                        elif "final_call" in j:
                            call_when = "final_call"
                        elif 'call' in j:
                            action = j.replace("call", 'normal_call')
                            call_when = "normal_call"
                        else:
                            pass
                        if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
                            return_list_str = action[action.find(":") + 1:action.find("=")].strip()
                            if return_list_str != "":
                                return_list = return_list_str.split(',')
                                if return_list:
                                    index4 = action.find("=")
                                    action = action[:index4] + "&equ&" + action[index4 + 1:]
                                    for param in return_list:
                                        param = param.strip()
                                        VARIABLE[param] = ""
                        index5 = action.find(":")
                        call_when = action[:index5].strip()
                        action = action[index5 + 1:].strip()
                        retrun_json["call_when"] = call_when
                        retrun_json["type"] = "General_Operation"
                        retrun_json["content"] = action
                        retrun_json["PUBLIC_FUNC_FLAG"] = 0
                        if len(i) > 1:
                            retrun_json["desc"] = ",".join(i[:-1])
                    else:
                        if "save_param" in action:
                            action_new = action.replace("save_param:", "")
                            action_new_list = action_new.split(",")
                            action_param = []
                            key_str = []
                            value_str = []
                            for action_new_param in action_new_list:
                                if "=" in action_new_param:
                                    key = action_new_param.split("=")[0]
                                    value = action_new_param.split("=")[1]
                                    key_str.append(key)
                                    value_str.append(value)
                                else:
                                    index2 = action_new_param.find("[")
                                    key = action_new_param[:index2]
                                    value = action_new_param[index2:]
                                    key_str.append(key)
                                    value_str.append(value)
                            if key_str:
                                for param in key_str:
                                    param = param.strip()
                                    VARIABLE[param] = ""
                                key_str = ",".join(key_str)
                            if value_str:
                                value_str = ",".join(value_str)

                            action = "normal_call:" + key_str + "&equ&save_param(" + value_str + ")"
                            index5 = action.find(":")
                            call_when = action[:index5].strip()
                            action = action[index5 + 1:].strip()
                            retrun_json["call_when"] = "normal_call"
                            retrun_json["type"] = "General_Operation"
                            retrun_json["content"] = action
                            retrun_json["PUBLIC_FUNC_FLAG"] = 0
                            if len(i) > 1:
                                retrun_json["desc"] = ",".join(i[:-1])
                        elif "save_in_param" in action:
                            action_new = action.replace("save_in_param:", "")
                            action_new_list = action_new.split(",")
                            action_param = []
                            key_str = []
                            value_str = []
                            for action_new_param in action_new_list:
                                if "=" in action_new_param:
                                    key = action_new_param.split("=")[0]
                                    value = action_new_param.split("=")[1]
                                    key_str.append(key)
                                    value_str.append(value)
                                else:
                                    index2 = action_new_param.find("[")
                                    key = action_new_param[:index2]
                                    value = action_new_param[index2:]
                                    key_str.append(key)
                                    value_str.append(value)
                            if key_str:
                                for param in key_str:
                                    param = param.strip()
                                    VARIABLE[param] = ""
                                key_str = ",".join(key_str)
                            if value_str:
                                value_str = ",".join(value_str)

                            action = "normal_call:" + key_str + "&equ&save_in_param(" + value_str + ")"
                            index5 = action.find(":")
                            call_when = action[:index5].strip()
                            action = action[index5 + 1:].strip()
                            retrun_json["call_when"] = "normal_call"
                            retrun_json["type"] = "General_Operation"
                            retrun_json["content"] = action
                            retrun_json["PUBLIC_FUNC_FLAG"] = 0
                            if len(i) > 1:
                                retrun_json["desc"] = ",".join(i[:-1])
                        else:
                            print "have no wirte action : %s" % action
                    # print "huozhi new %s" % action
                    actions.append(retrun_json)
                else:
                    print "this is desc %s" % j
            if action_flag == False:
                retrun_json["desc"] = ",".join(i)
                actions.append(retrun_json)
        # f.close()
        return VARIABLE, actions
    except Exception as e:
        print e


# def action_update(action_list, VARIABLE):
#     try:
#         actions = []
#         all_action = []
#         child_action = []
#         flag = False
#         for action in action_list:
#             if action.strip() == "":
#                 continue
#             if flag:
#                 flag = False
#                 child_action = []
#             if "#" in action:
#                 child_action.append(action)
#             else:
#                 child_action.append(action)
#                 flag = True
#                 all_action.append(child_action)
#         for i in all_action:
#             for j in i:
#                 if not j.startswith("#"):
#                     retrun_json = {"call_when": "", "type": "", "content": "",
#                                    "desc": "", "PUBLIC_FUNC_FLAG": 0}
#                     action = j
#                     print "huozhi %s" % action
#                     return_list_str = ""
#                     return_list = []
#                     if "oracle_Database" in j:
#                         index1 = action.rfind(",")
#                         action_1 = action[:index1]
#                         user = action[index1:].replace(",", "").replace(")", "")
#                         if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
#                             return_list_str = action[action.find(":") + 1:action.find("=")].strip()
#                         else:
#                             param_list_str = action[action.find(":") + 1:].strip()
#                         if return_list_str != "":
#                             return_list = return_list_str.split(',')
#                         func = str(user.strip()) + ".exec_sql"
#                         action = action_1.replace("oracle_Database", func) + ")"
#                         if return_list:
#                             index4 = action.find("=")
#                             action = action[:index4] + "&equ&" + action[index4 + 1:]
#                             for param in return_list:
#                                 param = param.strip()
#                                 VARIABLE[param] = ""
#                         if 'SELECT' in action.upper():
#                             action_2 = action.upper()
#                             index3 = action_2.find("FROM")
#                             action_select = action_2[:index3]
#                             field = action_select.split("SELECT")[1].strip()
#                             if "," in field:
#                                 print "select many field %s" % field
#                             elif "." in field:
#                                 field = field.split(".")[1]
#                             action = action + "&comma&get_value(@DbTable[1]['%s']@)" % field
#                         index5 = action.find(":")
#                         call_when = action[:index5]
#                         action = action[index5 + 1:].strip()
#                         if call_when == 'call':
#                             call_when = "normal_call"
#                             action = action.replace("call", "normal_call")
#                         retrun_json["call_when"] = call_when
#                         retrun_json["type"] = "DB_Operation"
#                         retrun_json["content"] = action
#                         retrun_json["PUBLIC_FUNC_FLAG"] = 0
#                         if len(i) > 1:
#                             retrun_json["desc"] = ",".join(i[:-1])
#                     elif "sqlserver_Database" in j:
#                         index1 = action.rfind(",")
#                         action_1 = action[:index1]
#                         user = action[index1:].replace(",", "").replace(")", "")
#                         if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
#                             return_list_str = action[action.find(":") + 1:action.find("=")].strip()
#                         else:
#                             param_list_str = action[action.find(":") + 1:].strip()
#                         if return_list_str != "":
#                             return_list = return_list_str.split(',')
#                         func = str(user.strip()) + ".exec_sql"
#                         action = action_1.replace("sqlserver_Database", func) + ")"
#                         if return_list:
#                             index4 = action.find("=")
#                             action = action[:index4] + "&equ&" + action[index4 + 1:]
#                             for param in return_list:
#                                 param = param.strip()
#                                 VARIABLE[param] = ""
#                         if 'SELECT' in action.upper():
#                             action_2 = action.upper()
#                             index3 = action_2.find("FROM")
#                             action_select = action_2[:index3]
#                             field = action_select.split("SELECT")[1].strip()
#                             if "," in field:
#                                 print "select many field%s" % field
#                             elif "." in field:
#                                 field = field.split(".")[1]
#                             action = action + "&comma&get_value(@DbTable[1]['%s']@)" % field
#                         index5 = action.find(":")
#                         call_when = action[:index5]
#                         action = action[index5 + 1:].strip()
#                         if call_when == 'call':
#                             call_when = "normal_call"
#                             action = action.replace("call", "normal_call")
#                         retrun_json["call_when"] = call_when
#                         retrun_json["type"] = "DB_Operation"
#                         retrun_json["content"] = action
#                         retrun_json["PUBLIC_FUNC_FLAG"] = 0
#                         if len(i) > 1:
#                             retrun_json["desc"] = ",".join(i[:-1])
#                     elif 'call' in j or "normal_call" in j or "except_call" in j or "final_call" in j:
#                         if "normal_call" in j:
#                             call_when = "normal_call"
#                         elif "except_call" in j:
#                             call_when = "except_call"
#                         elif "final_call" in j:
#                             call_when = "final_call"
#                         elif 'call' in j:
#                             action = j.replace("call", 'normal_call')
#                             call_when = "normal_call"
#                         else:
#                             pass
#                         if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
#                             return_list_str = action[action.find(":") + 1:action.find("=")].strip()
#                             if return_list_str != "":
#                                 return_list = return_list_str.split(',')
#                                 if return_list:
#                                     index4 = action.find("=")
#                                     action = action[:index4] + "&equ&" + action[index4 + 1:]
#                                     for param in return_list:
#                                         param = param.strip()
#                                         VARIABLE[param] = ""
#                         index5 = action.find(":")
#                         call_when = action[:index5]
#                         action = action[index5 + 1:].strip()
#                         retrun_json["call_when"] = call_when
#                         retrun_json["type"] = "General_Operation"
#                         retrun_json["content"] = action
#                         retrun_json["PUBLIC_FUNC_FLAG"] = 0
#                         if len(i) > 1:
#                             retrun_json["desc"] = ",".join(i[:-1])
#                     else:
#                         if "save_param" in action:
#                             action_new = action.replace("save_param:", "")
#                             action_new_list = action_new.split(",")
#                             action_param = []
#                             key_str = []
#                             value_str = []
#                             for action_new_param in action_new_list:
#                                 if "=" in action_new_param:
#                                     key = action_new_param.split("=")[0]
#                                     value = action_new_param.split("=")[1]
#                                     key_str.append(key)
#                                     value_str.append(value)
#                                 else:
#                                     index2 = action_new_param.find("[")
#                                     key = action_new_param[:index2]
#                                     value = action_new_param[index2:]
#                                     key_str.append(key)
#                                     value_str.append(value)
#                             if key_str:
#                                 for param in key_str:
#                                     param = param.strip()
#                                     VARIABLE[param] = ""
#                                 key_str = ",".join(key_str)
#                             if value_str:
#                                 value_str = ",".join(value_str)
#                             action = "normal_call:" + key_str + "&equ&save_param(" + value_str + ")"
#                             index5 = action.find(":")
#                             call_when = action[:index5]
#                             action = action[index5 + 1:].strip()
#                             retrun_json["call_when"] = "normal_call"
#                             retrun_json["type"] = "General_Operation"
#                             retrun_json["content"] = action
#                             retrun_json["PUBLIC_FUNC_FLAG"] = 0
#                             if len(i) > 1:
#                                 retrun_json["desc"] = ",".join(i[:-1])
#                         elif "save_in_param" in action:
#                             action_new = action.replace("save_in_param:", "")
#                             action_new_list = action_new.split(",")
#                             action_param = []
#                             key_str = []
#                             value_str = []
#                             for action_new_param in action_new_list:
#                                 if "=" in action_new_param:
#                                     key = action_new_param.split("=")[0]
#                                     value = action_new_param.split("=")[1]
#                                     key_str.append(key)
#                                     value_str.append(value)
#                                 else:
#                                     index2 = action_new_param.find("[")
#                                     key = action_new_param[:index2]
#                                     value = action_new_param[index2:]
#                                     key_str.append(key)
#                                     value_str.append(value)
#                             if key_str:
#                                 for param in key_str:
#                                     param = param.strip()
#                                     VARIABLE[param] = ""
#                                 key_str = ",".join(key_str)
#                             if value_str:
#                                 value_str = ",".join(value_str)
#                             action = "normal_call:" + key_str + "&equ&save_in_param(" + value_str + ")"
#                             index5 = action.find(":")
#                             call_when = action[:index5]
#                             action = action[index5 + 1:].strip()
#                             retrun_json["call_when"] = "normal_call"
#                             retrun_json["type"] = "General_Operation"
#                             retrun_json["content"] = action
#                             retrun_json["PUBLIC_FUNC_FLAG"] = 0
#                             if len(i) > 1:
#                                 retrun_json["desc"] = ",".join(i[:-1])
#                         else:
#                             print "have no wirte action : %s" % action
#                     print "huozhi new %s" % action
#                     actions.append(retrun_json)
#                 else:
#                     print "this is desc %s" % j
#         return VARIABLE, actions
#     except Exception as e:
#         exc_info = getExceptionInfo()
#         Logging.getLog().error(exc_info)
#         return -1, exc_info


def RunInitScriptEx(filename):
    '''
    支持
    '''
    try:
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        elif __file__:
            application_path = os.path.dirname(__file__)

        init_script_file = os.path.join(application_path, filename)
        Logging.getLog().debug(init_script_file)

        # 额外信息
        # extra_dict = {}

        # extra_dict["serverid"] = gl.g_connectSqlServerSetting[servertype]["serverid"]
        # extra_dict["today"] = datetime.datetime.now().strftime('%Y%m%d')

        # 更新全局变量
        # gl.g_accountDict[str(account_group_id)].update(extra_dict)

        group_variable_dict = {}

        # 默认用户组变量为0
        # group_variable_dict['account_group_id'] = 0

        f = open(init_script_file, 'r')
        lines = f.readlines()
        for j, line in enumerate(lines):
            line = line.strip()
            if line.startswith('--') or line.startswith('#'):
                continue
            if len(line) == 0:
                continue
            # line = line.decode('utf-8')
            # account_group_id = int(group_variable_dict['account_group_id'])
            account_group_id = 0
            action_list = []
            action_list.append(line)
            VARIABLE, actions = action_update(action_list, {})
            # actions = json.dumps(actions)
            ret, msg = action_process(actions, group_variable_dict, [], None, None, account_group_id)
            if ret < 0:
                Logging.getLog().critical(u"初始化脚本失败，原因是%s" % msg)
                return ret, msg


    except:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info
    return 0, "OK"


def RunInitScript(filename, account_group_id=0, servertype='p'):
    try:
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        elif __file__:
            application_path = os.path.dirname(__file__)

        init_script_file = os.path.join(application_path, filename)
        print init_script_file

        # 额外信息
        extra_dict = {}
        serverid = cF.get_core(servertype)
        extra_dict["serverid"] = serverid
        extra_dict["today"] = datetime.datetime.now().strftime('%Y%m%d')

        # 更新全局变量
        gl.g_accountDict[str(account_group_id)].update(extra_dict)

        f = open(init_script_file, 'r')
        lines = f.readlines()
        for j, line in enumerate(lines):
            line = line.strip()
            if line.startswith('--'):
                continue
            if len(line) == 0:
                continue
            # line = line.decode('utf-8')

            sql = GlobalParameterReplace(line, str(account_group_id))
            print sql
            executeUpdateSQL(sql, servertype)
    except:
        exc_info = getExceptionInfo()
        print exc_info
        return -1, exc_info
    return 0, "OK"


def getStkcodeByString(cmd):
    flag = False
    lpos = cmd.find("stkcode=")
    if lpos == -1:
        flag = True
        lpos = cmd.find("ofcode=")
        if lpos == -1:
            return None
    rpos = cmd.find("#", lpos)
    if rpos == -1:
        if flag == False:
            return cmd[cmd.find('stkcode') + 7:]
        else:
            return cmd[cmd.find('ofcode') + 8:]
    else:
        if (lpos + 8 == rpos and flag == False) or (lpos + 7 == rpos and flag == True):
            return None
        else:
            if flag == False:
                return cmd[lpos + len("stkcode:"):rpos]
            else:
                return cmd[lpos + len("ofcode:"):rpos]


def getPriceByStkcode(stkcode, servertype='p'):
    sql = "select  openprice from stkprice where stkcode = '%s'" % (stkcode)
    ds = executeRunSQL(sql, servertype)
    if ds == []:
        return 0
    else:
        return ds[0]['openprice']


def getDebugInfo():
    return "%s:%s" % (sys._getframe().f_back.f_code.co_name, sys._getframe().f_back.f_lineno)


def zip_dir(dirname, zipfilename):
    filelist = []
    if os.path.isfile(dirname):
        filelist.append(dirname)
    else:
        for root, dirs, files in os.walk(dirname):
            for name in files:
                filelist.append(os.path.join(root, name))
    zf = zipfile.ZipFile(zipfilename, "w", zipfile.zlib.DEFLATED)
    for tar in filelist:
        arcname = tar[len(dirname):]
        arcname = os.path.basename(dirname) + "\\" + arcname
        # print arcname
        zf.write(tar, arcname)
    zf.close()


def unzip_file(zipfilename, unziptodir):
    if not os.path.exists(unziptodir): os.mkdir(unziptodir)
    zfobj = zipfile.ZipFile(zipfilename)
    for name in zfobj.namelist():
        name = name.replace('\\', '/')

        if name.endswith('/'):
            if not os.path.exists(os.path.join(unziptodir, name)):
                os.mkdir(os.path.join(unziptodir, name))
        else:
            ext_filename = os.path.join(unziptodir, name)
            ext_dir = os.path.dirname(ext_filename)
            if not os.path.exists(ext_dir): os.makedirs(ext_dir)
            outfile = open(ext_filename, 'wb')
            outfile.write(zfobj.read(name))
            outfile.close()


def ModifyRemoteServerDateTime(ip, port, year, month, day, hour, min, sec):
    # ServerTime = "%04d-%02d-%02d %02d:%02d:%02d" %(int(year),int(month),int(day),int(hour),int(min),int(sec))
    # ltime = time.strptime(ServerTime,"%Y-%m-%d %H:%M:%S")
    # ttime=time.localtime(time.mktime( ltime ))
    try:
        # ServerUrl = "http://" + ip + ":" + port
        # s = ServerProxy(ServerUrl)
        # ret, msg = s.ModifyRemoteServerDateTime(int(year), int(month), int(day), int(hour), int(min), int(sec))
        return 0, ''
    except:
        exc = getExceptionInfo()
        Logging.getLog().error(exc)
        return -1, exc


def get_suffix(system_id):
    suffix = "ZDHCS"
    # if system_id == "142891263425277DZ1":
    #     suffix = 'JZJY'
    # elif system_id == "142891090059622AV2":
    #     suffix = 'ZHLC'
    # elif system_id == "1523514095162K1QR6":
    #     suffix = 'WAIHUI'
    # elif system_id == "1481617036190F5IU8":
    #     suffix = 'YHSS'
    # elif system_id == "142890425731984KH9":
    #     suffix = 'OTC'
    # elif system_id == "1428908551835S0NXP":
    #     suffix = 'YHZX'
    # elif system_id == "14774643386617P4VW":
    #     suffix = 'JZYY'
    # elif system_id == "1481879345901LDJNJ":
    #     suffix = 'FY'
    # elif system_id == "1490844074559VS0IY":
    #     suffix = 'HR'
    # elif system_id == "1428910825111C5EQ4":
    #     suffix = 'O32'
    # elif system_id == "1477107781285K6OZC":
    #     suffix = 'GJ'
    # else:
    #     suffix = ""
    return suffix


def read_xml_info(OldSettingPath):
    """
    读取老版本的账号组配置文件
    :param OldSettingPath:
    :return:
    """
    try:
        result = []
        data = []
        child = []
        flag = False
        f = open(OldSettingPath)  # 返回一个文件对象
        line = f.readline()  # 调用文件的 readline()方法
        while line:
            if '</AccountInfo>' in line:
                # child.append(line)
                data.append(child)
                child = []
                flag = False
            if '<AccountInfo' in line:
                flag = True
            if flag:
                line = line.replace("\n", "").replace("\t", "").strip(). \
                           replace("<!--", "").replace('-->', "").replace("<", '||||').replace(">", "||||")[4:]
                child.append(line)
            line = f.readline()
        for i in data:
            id = ""
            for j in i:
                if "AccountInfo" in j:
                    id_index = j.find('id')
                    j = j[id_index:]
                    desc_index = j.find("desc")
                    id = j[:desc_index].replace("id=", "").replace('"', "").replace("'", "").strip()
                    break
            for k in i:
                if k:
                    # print id, k
                    if "AccountInfo" in k:
                        continue
                    index = k.find("||||")
                    replace_data = "||||/" + k[:index]
                    k = k.replace(replace_data, "").split("||||")
                    id = id.replace("|||", '')
                    one_record = dict(GROUP_ID=int(id),
                                      FILELD_NAME=k[0],
                                      FIELD_VALUE=k[1],
                                      FIELD_MEMO=k[2])
                    result.append(one_record)
        f.close()
        return result
    except Exception, ex:
        exc = getExceptionInfo()
        Logging.getLog().error(exc)
        return -1, exc


def read_funcid_param_tools(system_id, funcid, old_param_new_params):
    try:
        Logging.getLog().debug(str(funcid))
        wb = load_workbook("JZYY_EXCEL/%s.xlsx" % (str(funcid[0:3])))
        sheet = wb.get_sheet_by_name(funcid)
        data = []
        count = 0
        col_names = []
        for row in sheet.iter_rows():
            count += 1
            if count == 1:
                col_names = [col.value for col in row]
                continue
            row_data = []
            for i in row:
                values = i.value
                for param in old_param_new_params:
                    if str(param) in str(values):
                        values = old_param_new_params[str(param)]
                row_data.append(values)
            data.append(row_data)
        return 0, data, col_names
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def read_funcid_param(system_id, funcid, row):
    try:
        row = int(row)
        wb = load_workbook("init/%s/%s.xlsx" % (system_id, funcid[0:3]))
        sheet = wb.get_sheet_by_name(funcid)
        row_data = []
        for i in sheet[row + 1]:
            row_data.append(i.value)
        return 0, row_data
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def read_funcid_param_mang_step(system_id, funcid, row, param_data_id):
    try:
        # row = int(row)
        row_data = []
        wb = load_workbook("init/%s/%s.xlsx" % (system_id, funcid[0:3]))
        # sheet = wb.get_sheet_by_name(funcid)
        sheet = ""
        try:
            sheet = wb.get_sheet_by_name(funcid)
        except Exception as e:
            exc_info = getExceptionInfo()
            Logging.getLog().error(exc_info)
            return 0, u"功能号不存在, 读取excel失败"
        for row in sheet:
            flag = False
            count = 0
            for i in row:
                # 获取每行第二个参数， 根据param_data_id 获取行数据
                count += 1
                if count == 2 and str(i.value) == str(param_data_id):
                    flag = True
                    break
            if flag == True:
                for j in row:
                    row_data.append(j.value)
        # row_data = []
        # for i in sheet[row + 1]:
        #     row_data.append(i.value)
        return 0, row_data
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def write_funcid_param_tool(system_id, funcid, row, row_value_list):
    try:
        row = int(row)
        wb = load_workbook("JZYY_EXCEL/%s.xlsx" % (funcid[0:3]))
        sheet = wb.get_sheet_by_name(funcid)
        # print sheet['A2'].value
        for i, col in enumerate(sheet[row + 1]):
            col.number_format = '@'
            if type(row_value_list[i]) == str:
                # TOOLS
                # col.value = (row_value_list[i]).decode('gbk')
                # Testagent
                col.value = (row_value_list[i]).decode('utf8')
            else:
                col.value = row_value_list[i]
        wb.save("JZYY_EXCEL/%s.xlsx" % (funcid[0:3]))
        return 0, ""
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def write_funcid_param(system_id, funcid, row, row_value_list):
    try:
        row = int(row)
        wb = load_workbook("init/%s/%s.xlsx" % (system_id, funcid[0:3]))
        sheet = wb.get_sheet_by_name(funcid)
        # print sheet['A2'].value
        for i, col in enumerate(sheet[row + 1]):
            col.number_format = '@'
            if type(row_value_list[i]) == str:
                # TOOLS
                # col.value = (row_value_list[i]).decode('gbk')
                # Testagent
                col.value = (row_value_list[i]).decode('utf8')
            else:
                col.value = row_value_list[i]
        wb.save("init/%s/%s.xlsx" % (system_id, funcid[0:3]))
        return 0, ""
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def write_funcid_param_many_step(system_id, funcid, row, row_value_list, param_data_id):
    try:
        row = int(row)
        wb = load_workbook("init/%s/%s.xlsx" % (system_id, funcid[0:3]))
        sheet = ""
        try:
            sheet = wb.get_sheet_by_name(funcid)
        except Exception as e:
            exc_info = getExceptionInfo()
            Logging.getLog().error(exc_info)
            return 0, u"功能号不存在, 读取excel失败"
        # print sheet['A2'].value
        for row in sheet:
            flag = False
            count = 0
            for i in row:
                # 获取每行第二个参数， 根据param_data_id 获取行数据
                count += 1
                if count == 2 and str(i.value) == str(param_data_id):
                    flag = True
                    break
            if flag == True:
                for j, col in enumerate(row):
                    # row_data.append(j.value)
                    if type(row_value_list[j]) == str:
                        # TOOLS
                        # col.value = (row_value_list[i]).decode('gbk')
                        # Testagent
                        try:
                            col.value = (row_value_list[j]).decode('utf8').replace("&comma&", ",")
                        except BaseException, ex:
                            col.value = (row_value_list[j]).decode('gbk').replace("&comma&", ",")
                    else:
                        col.value = row_value_list[j]
        # for i, col in enumerate(sheet[row + 1]):
        #     col.number_format = '@'
        #     if type(row_value_list[i]) == str:
        #         #TOOLS
        #         # col.value = (row_value_list[i]).decode('gbk')
        #         #Testagent
        #         col.value = (row_value_list[i]).decode('utf8')
        #     else:
        #         col.value = row_value_list[i]
        wb.save("init/%s/%s.xlsx" % (system_id, funcid[0:3]))
        return 0, ""
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


# def write_funcid_param_tools(system_id, funcid, row_value_list):
#     try:
#         wb = load_workbook("init/%s/%s.xlsx" % (system_id, funcid[0:3]))
#         sheet = wb.get_sheet_by_name(funcid)
#         # print sheet['A2'].value
#         count = 0
#         for rows in row_value_list:
#             for row in sheet.iter_rows():
#                 count += 1
#                 if count ==1:
#                     continue
#                 for i, col in enumerate(row):
#                     col.number_format = '@'
#                     if type(rows[i]) == str:
#                         col.value = (rows[i]).decode('utf8')
#                     else:
#                         col.value = rows[i]
#         wb.save("init/%s/%s.xlsx" % (system_id, funcid[0:3]))
#         return 0, ""
#     except BaseException, ex:
#         exc_info = getExceptionInfo()
#         Logging.getLog().error(exc_info)
#         return -1, exc_info
#
def mobile_data_oracle_sqlserver(client_id, gt_other_type):
    try:
        if client_id:
            select_tables = ['CLIENT', 'CLIENTINFO', 'ORGANINFO',
                             'FUNDACCOUNT', 'BUSINACCOUNT',
                             'SECUBUSINACCT', 'BKEXCHACCOUNT',
                             'STOCKHOLDER', 'OFSTOCKHOLDER'
                             ]
            JZJY_TABLES = ['customer', 'custbaseinfo', 'custcorpinfo',
                           'custotherinfo', 'custlinkinfo', 'fundinfo',
                           'fundasset', 'banktranid', 'secuid', 'oftaacc'
                           ]
            for table in select_tables:
                user = 'HS_ACCT'
                if table == 'STOCKHOLDER':
                    pass
                print 'table>>>>>>>>>>>%s' % table
                # param_info = GlobalParameterReplace(sql, 0)
                if table == 'BKEXCHACCOUNT':
                    user = 'HS_BANK'
                sql = "select COLUMN_NAME from user_tab_columns where TABLE_NAME='%s'" % table
                normal_table = []
                sleep_table = []
                ret, ret_param = eval("bizmod.change_oracle")(sql, user)
                table_sleep = table + "_SLEEP"
                sql = "select COLUMN_NAME from user_tab_columns where TABLE_NAME='%s'" % table_sleep
                ret_sleep, ret_param_sleep = eval("bizmod.change_oracle")(sql, user)
                print ret, ret_param

                if ret == 0:
                    for i in ret_param:
                        if i[0]:
                            normal_table.append(i[0])

                if ret_sleep == 0:
                    for j in ret_param_sleep:
                        if j[0]:
                            sleep_table.append(j[0])
                # 获取休眠表和正常表字段的交集
                ret_list = list(
                    (set(normal_table).union(set(sleep_table))) ^ (set(normal_table) ^ set(sleep_table)))
                print ret_list

                col_name = ",".join(ret_list)
                # 查询休眠表该客户号的数据

                sql = "select " + col_name + " from %s.%s where client_id = '%s' and " \
                                             "gt_other_type='%s'" % (user, table_sleep, client_id, gt_other_type)
                ret, ret_param = eval("bizmod.change_oracle")(sql, user)
                print ret, ret_param
                if ret == 0 and not ret_param:
                    Logging.getLog().debug('休眠表中没有该客户的数据--%s' % client_id)
                    # 查询正常表该客户号的数据
                    sql = "select " + col_name + " from %s.%s where client_id = '%s' " % (user, table, client_id)
                    ret, ret_param = eval("bizmod.change_oracle")(sql, user)
                    print ret, ret_param
                    sufix = ","
                    values = ""
                    if ret_param:
                        for value in ret_param:
                            value = list(value)
                            value.append(str(gt_other_type))
                            value = json.dumps(value, ensure_ascii=False, encoding="gbk")
                            value = value.replace("[", "(").replace("]", ")")
                            value = value + sufix
                            values += value
                    if ret == 0 and not ret_param:
                        Logging.getLog().debug('休眠表和正常表中都没有该客户的数据--%s' % client_id)
                    else:
                        Logging.getLog().debug('正常表数据转移到休眠表并且删除--%s' % client_id)
                        sql_insert = "insert into %s.%s" % (user, table_sleep)
                        col_name = col_name + ",GT_OTHER_TYPE"
                        col_name = "(" + col_name + ")"
                        values = values.replace('"', "'")
                        sql = sql_insert + col_name + ' values ' + values
                        sql = sql[:-1]
                        ret, ret_param = eval("bizmod.change_oracle")(sql, user)
                        if ret == 0:
                            Logging.getLog().debug('正常表数据成功转移到休眠表--%s' % client_id)
                        # 删除正常表的数据
                        sql = "delete from %s.%s where client_id = '%s'" % (user, table, client_id)
                        ret, ret_param = eval("bizmod.change_oracle")(sql, user)
                        if ret == 0:
                            Logging.getLog().debug('正常表数据已删除--%s' % client_id)
                else:
                    Logging.getLog().debug('休眠表中已存在该客户的数据--%s' % client_id)

            ###########################################sqlserver##############################################
            # 不合格账户激活
            Logging.getLog().debug('oracle客户号--%s' % client_id)
            client_id = int(client_id[2:])
            Logging.getLog().debug('sqlserver客户号--%s' % client_id)
            servertype = 'p'
            for table in JZJY_TABLES:
                print 'table>>>>>>>>>>>%s' % table
                # param_info = GlobalParameterReplace(sql, 0)
                sql = "select name from syscolumns where id=object_id(N'%s')" % table
                normal_table = []
                sleep_table = []
                ret, ret_param = eval("bizmod.change_db")(sql, servertype)
                table_sleep = table + "_other"
                print table_sleep
                sql = "select name from syscolumns where id=object_id(N'%s')" % table_sleep
                ret_sleep, ret_param_sleep = eval("bizmod.change_db")(sql, servertype)
                print ret, ret_param

                if ret == 0:
                    for i in ret_param:
                        if i[0]:
                            normal_table.append(i[0])

                if ret_sleep == 0:
                    for j in ret_param_sleep:
                        if j[0]:
                            sleep_table.append(j[0])
                # 获取休眠表和正常表字段的交集
                ret_list = list(
                    (set(normal_table).union(set(sleep_table))) ^ (set(normal_table) ^ set(sleep_table)))
                print ret_list

                col_name = ",".join(ret_list)
                # 查询休眠表该客户号的数据

                sql = "select " + col_name + " from %s where custid = '%s'" % (table_sleep, client_id)
                ret, ret_param = eval("bizmod.change_db")(sql, servertype)
                print ret, ret_param
                if ret == -1:
                    Logging.getLog().debug('休眠表中没有该客户的数据--%s' % client_id)
                    # 查询正常表该客户号的数据
                    sql = "select " + col_name + " from %s where custid = '%s'" % (table, client_id)
                    ret, ret_param = eval("bizmod.change_db")(sql, servertype)
                    print ret, ret_param
                    sufix = ","
                    values = ""
                    if ret_param:
                        for value in ret_param:
                            new_values = []
                            for i in value:
                                if type(i) == str:
                                    i = i.decode('gbk').strip()
                                new_values.append(i)
                            new_values.append(str(gt_other_type))
                            new_values = json.dumps(new_values, ensure_ascii=False, encoding="gbk", default=default)
                            _value = new_values.replace("[", "(").replace("]", ")")
                            _value = _value + sufix
                            values += _value
                    if ret == -1:
                        Logging.getLog().debug('休眠表和正常表中都没有该客户的数据--%s' % client_id)
                    else:
                        Logging.getLog().debug('正常表数据转移到休眠表并且删除--%s' % client_id)
                        sql_insert = "insert into %s" % (table_sleep)
                        col_name = col_name + ",othertype"
                        col_name = "(" + col_name + ")"
                        values = values.replace('"', "'")
                        sql = sql_insert + col_name + ' values ' + values
                        sql = sql[:-1]
                        ret, ret_param = eval("bizmod.change_db")(sql, servertype)
                        if ret == 0:
                            Logging.getLog().debug('正常表数据成功转移到休眠表--%s' % client_id)
                            # if table_sleep == 'customer_other':
                            #     sql = "update customer_other set othertype=1 where custid='%s'" % client_id
                            #     ret, ret_param = eval("bizmod.change_db")(sql, servertype)
                        # 删除正常表的数据
                        sql = "delete from %s where custid = '%s'" % (table, client_id)
                        ret, ret_param = eval("bizmod.change_db")(sql, servertype)
                        if ret == 0:
                            Logging.getLog().debug('正常表数据已删除--%s' % client_id)
                else:
                    Logging.getLog().debug('休眠表中已存在该客户的数据--%s' % client_id)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def getUserToken(sdk, account_group_id):
    try:
        # cmdstring = "operator_no=@operator_no@#password=@password@#mac_address=FF-FF-FF-FF-FF-FF#ip_address=127.0.0.1#op_station=HuaBaoPB#authorization_id=279959C0257235196EC3CE"
        cmdstring = "operator_no=@operator_no@#password=@password@#mac_address=FF-FF-FF-FF-FF-FF#ip_address=127.0.0.1#op_station=HuaBaoPB#authorization_id=279959C0257235196EC3CE#hd_volserial=#login_time=#verification_code="
        cmdstring = GlobalParameterReplace(cmdstring, str(account_group_id))
        print "ExecuteOneTestCase:%s" % cmdstring
        ret, msg = ExecuteOneTestCase(sdk, '10001', cmdstring)
        dataset_list = []
        if ret < 0:
            return ret, msg
        try:
            dataset_list = GetReturnRows(sdk)
        except Exception as e:
            # 报错会造成agent异常退出
            return -1, u"未获取返回信息"
        user_token = ""
        if len(dataset_list) > 0:
            user_token = dataset_list[1][1][0]
            return 0, user_token
    except:
        exc_info = getExceptionInfo()
        print exc_info
        return -1, exc_info
    return -1, "nothing found"


def mobile_data(client_id, gt_other_type):
    """
        不合格账户激活：
        执行前用例前综合理财表
        client_sleep,clientinfo_sleep,organinfo_sleep,fundaccount_sleep，businaccount_sleep,secubusinacct_sleep,                       bkexchaccount_sleep, stockholder_sleep,ofstockholder_sleep 含有相应数据
        执行后以上表相应数据迁入以下综合理财表
        client,clientinfo,organinfo,fundaccount,businaccount,secubusinacct,bkexchaccount,stockholder,ofstockholder

        执行前集中交易表
        customer_other,custbaseinfo_other,custcorpinfo_other,custotherinfo_other
        custlinkinfo_other,fundinfo_other,fundasset_other,banktranid_other,secuid_other,oftaacc_other含有相应数据

        执行用例后以上数据迁入集中交易表
        customer,custbaseinfo,custcorpinfo,custotherinfo，custlinkinfo,fundinfo,fundasset,banktranid,secuid,oftaacc
        小额休眠账户激活综合理财表

        #1、寻找小额休眠的客户号
        call:login_id_no134 = oracle_Database(select client_id from hs_acct.client_sleep s where s.branch_no='@branch_nono@'           and s.gt_other_type='1'  and  rownum = '1' and s.client_id not in ( select distinct client_id from stockholder_sleep          where branch_no='@branch_nono@'),HS_ACCT)

        #1、寻找不合格账户的客户号
        call:login_id_no133 = oracle_Database(select client_id from hs_acct.client_sleep s where s.branch_no='@branch_nono@'
                  and s.gt_other_type='2'and rownum = '1',HS_ACCT)

        #1、寻找小额休眠含证券账号的客户号
        call:login_id_no134 = oracle_Database(select client_id from hs_acct.client_sleep s where s.branch_no='@branch_nono@'
                  and s.gt_other_type='1'  and  rownum = '1' and s.client_id in
                  ( select distinct client_id from stockholder_sleep where   branch_no='@branch_nono@'),HS_ACCT)
    """
    try:
        # client_ids = []
        # if gt_other_type == '2':
        #     # 1、寻找小额休眠的客户号
        #     sql = """select client_id from hs_acct.client_sleep s where s.branch_no='@branch_nono@'
        #               and s.gt_other_type='2'and rownum = '1'"""
        #     sql = GlobalParameterReplace(sql, 0)
        #     ret_sleep, ret_param_sleep = eval("bizmod.change_oracle")(sql, user)
        #     if ret_sleep == 0:
        #         client_ids.append(str(ret_param_sleep[0][0]))
        #
        # if gt_other_type == '1':
        #     # 1、寻找小额休眠的客户号
        #     sql = "select client_id from hs_acct.client_sleep s where s.branch_no='@branch_nono@' and s.gt_other_type='1' " \
        #           "and  rownum = '1' and s.client_id in (select distinct client_id from stockholder_sleep where " \
        #           "branch_no='@branch_nono@') and s.client_id not in (select distinct client_id from " \
        #           "hs_bank.bkexchaccount_sleep where branch_no='@branch_nono@')"
        #     sql = GlobalParameterReplace(sql, 0)
        #     ret_sleep, ret_param_sleep = eval("bizmod.change_oracle")(sql, user)
        #     if ret_sleep == 0:
        #         client_ids.append(str(ret_param_sleep[0][0]))v
        #
        #     # 1、寻找小额休眠的客户号n
        #     sql = "select client_id from hs_acct.client_sleep s where s.branch_no='@branch_nono@'  and s.gt_other_type='1'  " \
        #           "and  rownum = '1' and s.client_id not in ( select distinct client_id from hs_acct.stockholder_sleep where " \
        #           "branch_no='@branch_nono@') and s.client_id not in (select distinct client_id from hs_bank.bkexchaccount_sleep" \
        #           " where branch_no='@branch_nono@')"
        #     sql = GlobalParameterReplace(sql, 0)
        #     ret_sleep, ret_param_sleep = eval("bizmod.change_oracle")(sql, user)
        #     if ret_sleep == 0:
        # client_ids.append(client_id)
        # gl.g_clients = client_ids
        mobile_data_oracle_sqlserver(client_id, gt_other_type)
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def default(obj):
    if isinstance(obj, Decimal):
        return str(obj)
    raise TypeError("Object of type '%s' is not JSON serializable" % type(obj).__name__)


def exeRunSQLOracle(sqlstr, login_info):
    try:
        Logging.getLog().debug("sqlstr:%s,login_info:%s" % (sqlstr, str(login_info)))
        conn = cx_Oracle.connect(login_info['username'], login_info['password'], login_info['db_alias'])
        crs = conn.cursor()
        crs.execute(sqlstr)
        ds = []
        if sqlstr.lower().__contains__("select"):
            try:
                rows = crs.fetchall()
                for row in rows:
                    rec = {}
                    for i, desc in enumerate(crs.description):
                        rec[desc[0]] = row[i]
                    ds.append(rec)
            except BaseException, ex2:
                pass
        conn.commit()
        crs.close()
        conn.close()
        return 0, ds
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def executeRunSQL_oracle(sqlstr, servertype):
    try:
        Logging.getLog().debug("sqlstr:%s,servertype:%s" % (sqlstr, servertype))
        core = ''
        for server in gl.g_connectOracleSetting.keys():
            # Logging.getLog().debug("servertype:%s,gl.g_connectOracleSetting[server][servertype]:%s" % (servertype, gl.g_connectOracleSetting[server]["servertype"]))
            if servertype.upper().strip() == gl.g_connectOracleSetting[server]["servertype"].upper():
                core = server
                break
        Logging.getLog().debug("core:%s" % core)
        if core == '':
            return -1, "sql传入用户名%s,在配置文件内不存在" % (servertype)
        conn = ConnectOracleDB(core)
        crs = conn.cursor()
        crs.execute(sqlstr)
        ds = []
        if sqlstr.lower().__contains__("select"):
            try:
                rows = crs.fetchall()
                for row in rows:
                    rec = {}
                    for i, desc in enumerate(crs.description):
                        if isinstance(row[i], decimal.Decimal):
                            rec[desc[0]] = str(row[i])
                        elif isinstance(row[i], str):
                            rec[desc[0]] = row[i].strip().decode("gbk")
                        else:
                            rec[desc[0]] = row[i]
                    ds.append(rec)
            except BaseException, ex2:
                pass
        conn.commit()
        crs.close()
        conn.close()
        return 0, ds
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def fault_tolerant(func_id='', ret='', msg_info='', log_info='', dataset_list='', return_json_string='',
                   data_change_info_json_string=''):  # 容错处理操作
    try:
        Logging.getLog().info(msg_info)
        Logging.getLog().info(func_id)
        # Logging.getLog().info(gl.g_allows_error_dict)
        allow_errors_dict = gl.g_allows_error_dict  # 容错字典
        error_code = msg_info.replace("\n", "").replace("\n\r", "")
        if func_id in allow_errors_dict.keys():
            key_list = allow_errors_dict[func_id]
            for i in key_list:
                i = i.strip()
                if i in error_code:
                    ret = 0
                    break
        return ret, u'error_tolerant'
    except BaseException, ex:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info


def wait_resultinfo():
    if 0 in gl.all_core.values():
        time.sleep(1)
        wait_resultinfo()
    else:
        return


def run_test_http(type, url, cmdstring):
    try:
        ret = 0
        return_data = []
        data = ""
        if type.lower() == "get":
            # url = "http://10.176.65.17:8080/fyauto/cases/queryDetail?CASES_ID=1"
            req = urllib2.Request(url)
            res_data = urllib2.urlopen(req)
            status = res_data.code
            if status == 200:
                ret = 0
            else:
                ret = -1
            data = res_data.read()
        if type.lower() == "post":
            # test_data = {'ServiceCode': 'aaaa', 'b': 'bbbbb'}
            # test_data_urlencode = urllib.urlencode(test_data)
            # url = "http://10.176.65.17:8080/fyauto/cases/getTagName"
            req = urllib2.Request(url=url)
            print req
            res_data = urllib2.urlopen(req)
            status = res_data.code
            if status == 200:
                ret = 0
            else:
                ret = -1
            data = res_data.read()

        try:
            response_msg = json.loads(data)
            if isinstance(response_msg, list):
                flag = True
                for c in response_msg:
                    if not isinstance(c, json):
                        flag = False
                if flag:
                    return_data.append(response_msg[0].keys())
                    for c in response_msg:
                        return_data.append(c.values())
                else:
                    return_data = [["response_msg"], [json.dumps(response_msg)]]
            elif isinstance(response_msg, json):
                return_data.append(response_msg.keys())
                return_data.append(response_msg.values())
            else:
                return_data = [["response_msg"], [json.dumps(response_msg)]]
        except:
            return_data = [["response_msg"], [data]]
        return ret, return_data
    except Exception as e:
        Logging.getLog().error(e)
        return -1, e


def while_test():
    count = 0
    while True:
        count += 1
        print count
        t2sdk, msg = connectT2("10.189.49.200:19001", "license.dat")
        t2sdk.Close()  # fixme


def trans_web_interface(task_id='', pt_tag=''):
    if not task_id:
        task_id = 22251
    if not pt_tag:
        pt_tag = 'ZDH'
    url = 'http://127.0.0.1:8080/fyauto/services/csglService?wsdl'
    client = suds.client.Client(url)
    service = client.service

    rec_result = service.returnRunningState(task_id, pt_tag)
    print 'java接口调用成功', rec_result
    return 0, rec_result

def get_task_sx_interface_info(CASES_DETAIL_IDS_DICT):
    try:
        data = {}
        PARAM_ID_STR = ",".join(CASES_DETAIL_IDS_DICT.values())
        sql = "select INTERFACE_ID, PARAM_NAME, PARAM_TYPE from SX_INTERFACE_DETAIL_INFO where INTERFACE_ID in ("+PARAM_ID_STR+") "
        rs = executeCaseSQL(sql)
        for i in rs:
            INTERFACE_ID = i["INTERFACE_ID"]
            PARAM_NAME = i["PARAM_NAME"]
            PARAM_TYPE = i["PARAM_TYPE"]
            if PARAM_TYPE in [None, 'null', ""]:
                continue
            PARAM_TYPE = i["PARAM_TYPE"]
            if PARAM_TYPE in [None, 'null', ""]:
                continue
            name_type = {PARAM_NAME:PARAM_TYPE}
            if data:
                if str(INTERFACE_ID) in data.keys():
                    data[str(INTERFACE_ID)].update(name_type)
                else:
                    data[str(INTERFACE_ID)] = name_type
            else:
                data[str(INTERFACE_ID)] = name_type
        return 0, data
    except:
        return -1, {}
def read_config(fileName):
    try:
        f = open(fileName, 'r')
        f_data = json.load(fp=f, encoding='gbk')
        return f_data
    except Exception as e:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info
def deal_core_data(all_config_dict):
    try:
        core_first_list = all_config_dict['core'].split(',')
        core_list = []
        for i in core_first_list:
            i = i.strip()
            core_list.append(i)
        return 0, core_list
    except Exception as e:
        exc_info = getExceptionInfo()
        Logging.getLog().error(exc_info)
        return -1, exc_info
def get_conn(ip, db, user, passwd):
    dbstring = u"DRIVER={SQL Server};SERVER=%s,1433;" \
               u"DATABASE=%s;UID=%s;PWD=%s;autocommit=True"%(ip, db, user, passwd)
    try:
        conn = pyodbc.connect(dbstring)
        cursor = conn.cursor()
        return cursor
    except Exception as e:
        Logging.getLog().error(e.args[1].decode('gb2312'))
def auto_exec_sql(cursor, sql, core):
    """
    if PRE_ACTION.__contains__("\r\n"):
                action_list = PRE_ACTION.strip().split('\r\n')
            elif PRE_ACTION.__contains__("\r"):
                action_list = PRE_ACTION.strip().split('\r')
            elif PRE_ACTION.__contains__("\n"):
                action_list = PRE_ACTION.strip().split('\n')
            else:
                action_list = PRE_ACTION.strip().split('\\n')
    :param cursor:
    :param sql:
    :param core:
    :return:
    """
    try:
        if sql.upper().__contains__('EXEC') and sql.upper().__contains__('END'):
            try:
                print sql
                cursor.execute(sql)
                cursor.commit()
            except Exception as e:
                print e
        else:
            try:
                stringi = ''
                sql_list = ret_sql_list(sql, core)
                for sql_str in sql_list:
                    if sql_str.startswith('--') or sql_str == '' or sql_str == None:
                        pass
                    else:
                        try:
                            sql_str_new = ''
                            if sql_str.__contains__('stkname'):
                                chinese = re.findall('\'(.*?)\'', sql_str)[5]
                                flag = chardet.detect(chinese)['encoding']
                                chinese_rep = chinese.decode(encoding=flag)
                                sql_str = sql_str.replace(chinese, '%s')
                                sql_str_new = sql_str%(chinese_rep)
                            if sql_str_new:
                                print u'sql_new_str -- 正在执行sql ...'
                                print sql_str_new, '\n'
                                try:
                                    cursor.execute(sql_str_new)
                                    cursor.commit()
                                except Exception as e:
                                    Logging.getLog().error(sql_str_new)
                                    Logging.getLog().error(e.args[1].decode('gb2312'))
                            else:
                                print u'正在执行sql ...'
                                print sql_str, '\n'
                                try:
                                    cursor.execute(sql_str)
                                    cursor.commit()
                                except Exception as e:
                                    Logging.getLog().error(sql_str)
                                    Logging.getLog().error(e.args[1].decode('gb2312'))
                        except Exception as e:
                            Logging.getLog().error(e.args[1].decode('gb2312'))
            except Exception as e:
                print e
    except Exception as e:
        Logging.getLog().error(e.args[1].decode('gb2312'))
def ret_sql_list(sql, core=''):
    try:
        sql_list = []
        if core:
            if sql.lower().__contains__('@serverid'):
                sql = sql.replace('@serverid', "'%s'" % str(core))
        if sql.__contains__("\r\n"):
            sql_list = sql.strip().split('\r\n')
        elif sql.__contains__("\r"):
            sql_list = sql.strip().split('\r')
        elif sql.__contains__("\n"):
            sql_list = sql.strip().split('\n')
        else:
            sql_list = sql.strip().split('\\n')
        return sql_list
    except Exception as e:
        Logging.getLog().error(e.args[1].decode('gb2312'))

def test():
    import requests
    # c  = """httpRequestMethod=post#httpurl=http://10.187.160.112:8080/StrategyOperationFrontEnd/strategymanager/H12207001#OP_CODE=@OP_CODE@#OP_BRANCH_NO=@OP_BRANCH_NO@#OP_SITE=@OP_SITE@#OP_WAY=@OP_WAY@#OP_PROGRAM=@OP_PROGRAM@#USER_CODE=@USER_CODE@#BRANCH_NO=@BRANCH_NO@#FUNC_ID=@FUNC_ID@#PASSWORD=@PASSWORD@#USER_TOKEN=@USER_TOKEN@#TOKEN_SERIAL_NO=@TOKEN_SERIAL_NO@#AUTH_BIND_STATION=@AUTH_BIND_STATION@#ACTION_TYPE=A#STRATEGY_ID=192#STRATEGY_NAME=199#TARGET_TYPE=@TARGET_TYPE@#TARGET_DEFINITION={\"TAG_ID\":\"814\"}#BEGIN_TRIGGER_DATE=20190712#END_TRIGGER_DATE=20190712#BEGIN_TRIGGER_TIME=53000#END_TRIGGER_TIME=203000#MAX_TRIGGER_TIMES=30#TRIGGER_TIME_GAP=1#MSG_PLATEFORM=app#MSG_FORMAT=json#MSG_TITLE=计划70#MSG_BRIEF=55#MSG_CONTENT=1#CONDITION_DEFINITION={\"&bfb&1\":{\"EVENT_TYPE\":\"GZ\",\"OPERATE\":\">\",\"INDEX\":\"RISE\",\"VALUE\":\"-200\",\"STOCK_CODE\":\"sh000001\",\"VALUE_TYPE\":\"2\"},\"EXP\":\"&bfb&1 \"}#REMARK=#VALIDATION=1#TRIGGER_DATE_PARAM=0#TRIGGER_DATE_TYPE=0#TRIGGER_TIME=-1#MSG_ACCOUNT_TYPE=tel#MSG_EXTRA_DEFINITION={\"TOPIC_ID\":\"144\",\"MSG_TYPE\":\"A\"}"""
    c  = """httpRequestMethod=post#httpurl=http://10.187.160.112:8080/StrategyOperationFrontEnd/strategymanager/H12207001#OP_CODE=@OP_CODE@#OP_BRANCH_NO=@OP_BRANCH_NO@#OP_SITE=@OP_SITE@#OP_WAY=@OP_WAY@#OP_PROGRAM=@OP_PROGRAM@#USER_CODE=@USER_CODE@#BRANCH_NO=@BRANCH_NO@#FUNC_ID=@FUNC_ID@#PASSWORD=@PASSWORD@#USER_TOKEN=@USER_TOKEN@#TOKEN_SERIAL_NO=@TOKEN_SERIAL_NO@#AUTH_BIND_STATION=@AUTH_BIND_STATION@#ACTION_TYPE=A#STRATEGY_ID=192#STRATEGY_NAME=199#TARGET_TYPE=@TARGET_TYPE@#TARGET_DEFINITION={\"TAG_ID\":\"814\"}#BEGIN_TRIGGER_DATE=20190712#END_TRIGGER_DATE=20190712#BEGIN_TRIGGER_TIME=53000#END_TRIGGER_TIME=203000#MAX_TRIGGER_TIMES=30#TRIGGER_TIME_GAP=1#MSG_PLATEFORM=app#MSG_FORMAT=json#MSG_TITLE=计划70#MSG_BRIEF=55#MSG_CONTENT=1#CONDITION_DEFINITION={\"&bfb&1\":{\"EVENT_TYPE\":\"GZ\",\"OPERATE\":\">\",\"INDEX\":\"RISE\",\"VALUE\":\"-200\",\"STOCK_CODE\":\"sh000001\",\"VALUE_TYPE\":\"2\"},\"EXP\":\"&bfb&1 \"}#REMARK=#VALIDATION=1#TRIGGER_DATE_PARAM=0#TRIGGER_DATE_TYPE=0#TRIGGER_TIME=-1#MSG_ACCOUNT_TYPE=tel#MSG_EXTRA_DEFINITION={\"TOPIC_ID\":\"144\",\"MSG_TYPE\":\"A\"}"""
    cmd_list = c.split("#")
    json_field = {}
    send_data = {}
    # cmd_list = ['TARGET_DEFINITION={"TAG_ID":"814"}']
    interface_info = {"httpRequestMethod":"C255","httpurl":"C255","OP_CODE":"C100","OP_BRANCH_NO":"C100","OP_SITE":"C100","OP_WAY":"C100","OP_PROGRAM":"C100","USER_CODE":"C20","BRANCH_NO":"C100","FUNC_ID":"HsFunctionNo","PASSWORD":"C100","USER_TOKEN":"C100","TOKEN_SERIAL_NO":"C100","AUTH_BIND_STATION":"C254","ACTION_TYPE":"C32","STRATEGY_ID":"C24","STRATEGY_NAME":"C32","TARGET_TYPE":"C32","TARGET_DEFINITION":"HttpJson","BEGIN_TRIGGER_DATE":"C8","END_TRIGGER_DATE":"C8","BEGIN_TRIGGER_TIME":"C8","END_TRIGGER_TIME":"C8","MAX_TRIGGER_TIMES":"C8","TRIGGER_TIME_GAP":"C8","MSG_PLATEFORM":"c16","MSG_FORMAT":"C16","MSG_TITLE":"C255","MSG_BRIEF":"C255","MSG_CONTENT":"C4096","CONDITION_DEFINITION":"HttpJson","REMARK":"C1024","VALIDATION":"C11","TRIGGER_DATE_PARAM":"C11","TRIGGER_DATE_TYPE":"C16","TRIGGER_TIME":"C11","MSG_ACCOUNT_TYPE":"C16","MSG_EXTRA_DEFINITION":"HttpJson"}
    for cmd in cmd_list:
        field_type = ""
        field = str(cmd.split("=")[0])
        value = cmd.split("=")[1].replace("&bfb&", "#")
        print field, value
        if field in interface_info.keys():
            try:
                field_type = interface_info[str(field)]
            except:
                pass
        if field == "httpurl":
            url = value
        if field == "httpRequestMethod":
            type = value
        if field not in ["httpurl", "httpRequestMethod"]:
            if field_type == 'HttpList':
                try:
                    send_data[field] = eval(value)
                except:
                    pass
            elif field_type == 'HttpJson':
                print "HttpJson"
                try:
                    loads_value = eval(value)
                    if "CONDITION_DEFINITION" !=field:
                        json_field[str(loads_value)] = value.replace('"', '\\"')
                    send_data[field] = loads_value
                except:
                    pass
            elif field_type == 'HttpInt':
                print "HttpInt"
                try:
                    send_data[field] = int(value)
                except:
                    pass
            elif field_type == 'HttpFloat':
                print "HttpFloat"
                try:
                    send_data[field] = float(value)
                except:pass
            else:
                send_data[str(field)] = value
    ret = 0
    dataset_list = []
    data = ""
    try:
        print send_data, "ur:%s" % url
        if type.lower() == "get":
            r = ""
            if send_data:
                r = requests.get(url, params=send_data)
                data = r.text
            else:
                r = requests.get(url)
                data = r.text
            status = r.status_code
            if status == 200:
                ret = 0
            else:
                ret = -1
        if type.lower() == "post":
            r = ""
            if send_data:
                data = str(send_data)
                for j in json_field:
                    data = data.replace(j, str(json_field[j]))
                print "data: %s" % data
                r = requests.post(url, data.encode('gbk'))
                data = r.text
            else:
                r = requests.post(url)
                data = r.text
            status = r.status_code
            if status == 200:
                ret = 0
            else:
                ret = -1
    except Exception as e:
        pass
if __name__ == '__main__':
    test()
    # while_test()
    # readJzjyAccountInfo("14774643386617P4VW")
    # 测试http接口
    # run_test_http("post", "http://10.176.173.72:5000/http_test", "")
    # gl.g_server_id = '7'
    # readSystemConfigFile('142891263425277DZ1')


    # getUserToken(t2sdk, 170)
    # executeRunSQL("select SERIAL_NO from USER_NOTES_924  WHERE USER_CODE = '44924'", '833')
    # getCardID()
    # unzip_file(r"D:\PYTHON_CODE\localsvn\TestAgent\init\142891090059622AV2\test.zip", r"D:\PYTHON_CODE\localsvn\TestAgent\init\142891090059622AV2")
    # readConfigFile()
    # ExcelProcessParam("14774643386617P4VW", "0030070010", 0, {}, 0, 240332)
    # read_funcid_param("14774643386617P4VW", '0030070010', 0, 240323)
    # readSystemConfigFile('1428910825111C5EQ4')
    # readSystemConfigFile('14774643386617P4VW')
    readSystemConfigFile('142891090059622AV2')
    # get_task_sx_interface_info()
    # connectUFX(gl.g_ZHLCSystemSetting["T2Server"], "license_o32.dat")
    # add_tnsnames_info()
    # RunInitScriptEx("init\\1481879345901LDJNJ\\initialize.ini")
    # executeRunSQL_oracle("select t.CLIENT_ID AS CBS,t.BRANCH_NO from client t where t.cif_account='860003862032'",
    #                      servertype="HS_ACCT")
    # updateZHLCAutoTestSettingXml("142891090059622AV2","", 52)
    # # sdk, msg = connectT2(gl.g_ZHLCSystemSetting["T2Server"])
    #
    # add_tnsnames_info()

    # action_process_selenium
    # action_process_selenium("""""")
    action_process("""[{u'content': u'idkind,branchno,idno,fullname&equ&HS_ACCT.exec_sql(select id_no,id_kind,full_name,branch_no from gtja_regapplyclient t where rownum<=1 order by curr_date )&dot&get_value(@DbTable[1][ID_KIND]@,@DbTable[1][BRANCH_NO]@,@DbTable[1][ID_NO]@,@DbTable[1][FULL_NAME]@)', u'PUBLIC_FUNC_FLAG': 0, u'call_when': u'normal_call', u'type': u'DB_Operation', u'desc': u''}]
""")
    # action_process("""[{u'content': u'v2.exec_sql(select * into run..vip_kcbp_orgids_1 from run..vip_kcbp_orgids)&dot&get_value(@DbTable[1][]@)', u'PUBLIC_FUNC_FLAG': 0, u'call_when': u'normal_call', u'type': u'DB_Operation', u'desc': u''}]""")
    # cmdstring = "OP_CODE:,OP_BRANCH_NO:,OP_SITE:,OP_WAY:,OP_PROGRAM:,USER_CODE:0,BRANCH_NO:,FUNC_ID:12000308,PASSWORD:,USER_TOKEN:8911afba02f95954086b0db4f547b673,TOKEN_SERIAL_NO:2016000000000011003000040020000000000000000000000000000000000000001212126868000000000000000000000000,ACTIVITY_CODE:AUTOTESTACTIVITY,ACTIVITY_NAME:自动化测试活动配置,ACTIVITY_BUDGET:1000000000,BEGIN_DATE:20161124,END_DATE:20181230,ACTION_FLAG:1"
    # cmdstring = cmdstring.replace(':','=').replace(',',"#")
    # print cmdstring
    # cmdstring = "OP_CODE=#OP_BRANCH_NO=#OP_SITE=#OP_WAY=#OP_PROGRAM=#USER_CODE=#BRANCH_NO=#FUNC_ID=12000001#PASSWORD=#USER_TOKEN=#TOKEN_SERIAL_NO=#USER_TYPE=#USER_NAME=#USER_FNAME=#ID_TYPE=#ID_NO=#ACCT_TYPE=6#USER_SOURCE=#MOBILE=15021863640#EMAIL=#QQ=#WECHAT=#GENERATOR_NO=#NICKNAME=自动化测试#SEX=#ACCT_CODE=#CHANNEL_SOURCE="
    # ret, msg = ExecuteOneTestCase(sdk, '12000001', cmdstring)
    # ret, msg = ExecuteOneTestCase(sdk, '12007207',
    #                               'OP_CODE=#OP_BRANCH_NO=#OP_SITE=#OP_WAY=#OP_PROGRAM=#USER_CODE=0#BRANCH_NO=#FUNC_ID=12007207#PASSWORD=#USER_TOKEN=b4b147bc522828731f1a016bfa72c073#TOKEN_SERIAL_NO=2016000000000011003000040020000000000000000000000000000000000000001212126868000000000000000000000000#PROD_ID=51422#ISSUE_ID=0000028533#PAGE_SIZE=6#PAGE_NO=1#BEGIN_DATE=20150705#END_DATE=20151016')
    # sendT2Biz()
    # mobile_data('2')
    # ConnectCaseDB('DB2')
    # sql = "SELECT PRE_ACTION FROM SX_CASES_DETAIL WHERE cases_id =227"
    # ds = executeCaseSQL_DB2(sql)
    # group_dict= {}
    # sdk, msg = connectT2("10.187.57.45:19001")
    # if sdk != None:
    #     ret,msg=ExecuteOneTestCase(sdk,'12001905','OP_CODE=#OP_BRANCH_NO=#OP_SITE=#OP_WAY=#OP_PROGRAM=#USER_CODE=0#BRANCH_NO=#FUNC_ID=#PASSWORD=#USER_TOKEN=b4b147bc522828731f1a016bfa72c073#TOKEN_SERIAL_NO=2016000000000011003000040020000000000000000000000000000000000000001212126868000000000000000000000000#CUST_CODE=610015#PROD_CODE_LIST=#TERMINAL_INFO=')
    # ret,msg = RunInitScript(r'C:\01_projects\TestAgent\init\fuyi\initialize-p.sql',0,'p')
    # if ret<0:
    #    print msg
    # ret,msg = RunInitScript(r'C:\01_projects\TestAgent\init\fuyi\initialize-r.sql',0,'r')
    # if ret < 0:
    #    print msg
    #     action_process("""#1.综合理财表进行销户根据一户通号及营业部将客户号全部销户
    #     call:oracle_Database(update client set client_status = '3' where cif_account = '@cif_account25@' and branch_no = '3688',HS_ACCT)
    #     #2.
    #     call:oracle_Database(update aasclient set client_status = '3' where cif_account = '@cif_account25@' and branch_no = '3688'，HS_AAS)
    #     #3.
    #     call:oracle_Database(update hs_user.clientauth set client_status = '3' where cif_account = '@cif_account25@' and branch_no = '3688'，HS_AAS)
    # """)
    #     action_process("call:oracle_Database(update cif.tywqq set clzt = '8' where YHTH='@client_id3@',CIF)")
    # action_process_selenium("")
    # GlobalParameterReplace("user=22983#pass=6401011#stkcode=000001#price=@openprice@#qty=10000#seat=000000#agreementid=01#linkman=xxxx#linkway=13333333333")
    # print getDebugInfo()
    # ConnectCaseDB('DB2')
    # executeCaseSQL_DB2('SELECT * FROM SX_CASES_COLLECTION_RUN_TASK')
    # print ret,msg
    # pass
