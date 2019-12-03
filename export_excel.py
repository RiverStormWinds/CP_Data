# coding=utf-8
"""
workbook JZYY 导出excel
"""
import time
import os
import collections
from ftpClient import XFer
import pyodbc
from openpyxl import Workbook
import commFuncs as cF
from gl import GlobalLogging as Logging
import gl

def upload_interface_excel(file_path, system_id):  # fixme
    """文件上传到FTP"""
    try:
        homedir = os.getcwd()
        _type = '.xlsx'
        for root, dirs, files in os.walk(file_path):
            for filename in files:
                if _type in filename:
                    f = XFer(gl.gl_ftp_info['FTPServer']['IP'], gl.gl_ftp_info['FTPServer']['USER'], gl.gl_ftp_info['FTPServer']['PASSWORD'], '.', int(gl.gl_ftp_info['FTPServer']['PORT']))
                    f.login()
                    f.upload_file_ex('%s/%s' % (homedir, filename), './init/%s' % system_id)
        print 'ftp ok'
        return 0, ""
    except Exception as e:
        raise e


def get_param_data_info():
    """
    获取param_data的数据
    :return:
    """
    type = 'DB2'
    conn = None
    if type == 'DB2':
        # dbstring = "driver={IBM DB2 ODBC DRIVER};database=CSZDH;hostname=10.189.96.6;port=50000;protocol=TCPIP;uid=TIMESVC;pwd=timesvc;"
        # dbstring = "driver={IBM DB2 ODBC DRIVER};database=ZDH;hostname=10.187.96.46;port=50000;protocol=TCPIP;uid=timesvc;pwd=csgl@2017;LONGDATACOMPAT=1;LOBMAXCOLUMNSIZE=10485875;"
        conn = cF.ConnectCaseDB(type)
    cursor = conn.cursor()
    try:
        Logging.getLog().debug(u"ready to get_param_data_info--zzp")
        data = []
        CASES_DETAIL_IDs = []
        # sql = "SELECT CASES_DETAIL_ID, CASES_ID, PARAM_ID, STEP FROM SX_CASES_DETAIL WHERE CASES_ID IN (SELECT " \
        #       "CASES_ID FROM SX_CASES WHERE SYSTEM_ID='14774643386617P4VW' AND IS_DELETE = '0' AND (%s))"%str_like_remarks
        sql = "SELECT FUNCID, CASES_DETAIL_ID, CASES_ID, PARAM_ID, STEP FROM SX_CASES_DETAIL WHERE CASES_ID IN (SELECT " \
              "CASES_ID FROM SX_CASES WHERE SYSTEM_ID='14774643386617P4VW' AND IS_DELETE = '0')"
        cursor.execute(sql)
        Logging.getLog().debug(sql)
        for row in cursor.fetchall():
            # print row
            flag = True
            FUNCID, CASES_DETAIL_ID, CASES_ID, PARAM_ID, STEP = row
            CASES_DETAIL_IDs.append(str(CASES_DETAIL_ID))
            if data:
                for child_data in data:
                    if child_data['CASES_ID'] == CASES_ID and child_data['PARAM_ID'] == PARAM_ID:
                        if str(CASES_DETAIL_ID) not in child_data["CASES_DETAIL_ID"].keys():
                            child_data["CASES_DETAIL_ID"][str(CASES_DETAIL_ID)] = {str(CASES_DETAIL_ID):
                                                                                       {'STEP': STEP, "PARAM_DATA": ""}}
                    else:
                        flag = False
            if not data or flag == False:
                data.append({'FUNCID': FUNCID, 'PARAM_ID': PARAM_ID, 'CASES_ID': CASES_ID,
                             "CASES_DETAIL_ID": {str(CASES_DETAIL_ID): {'STEP': STEP, "PARAM_DATA": ""}}})
        # print data
        CASES_DETAIL_ID_str = ",".join(CASES_DETAIL_IDs)
        sql = "SELECT CASES_DETAIL_ID, PARAM_DATA,PARAM_DATA_ID from SX_CASES_DETAIL_PARAM_DATA where CASES_DETAIL_ID in (" + CASES_DETAIL_ID_str + ")"
        cursor.execute(sql)
        # Logging.getLog().debug(sql)
        for row in cursor.fetchall():
            # print row
            CASES_DETAIL_ID, PARAM_DATA, PARAM_DATA_ID = row
            PARAM_DATA = PARAM_DATA.decode('gbk', 'ignore').encode('utf-8')
            for i in data:
                # 判断CASES_DETAIL_ID是否存在，添加PARAM_DATA
                keys = i['CASES_DETAIL_ID'].keys()
                if str(CASES_DETAIL_ID) in keys:
                    i["CASES_DETAIL_ID"][str(CASES_DETAIL_ID)]["PARAM_DATA"] = PARAM_DATA
                    i["CASES_DETAIL_ID"][str(CASES_DETAIL_ID)]["PARAM_DATA_ID"] = PARAM_DATA_ID

        new_data = []
        for d in data:
            key = d["CASES_DETAIL_ID"].keys()[0]
            if d["CASES_DETAIL_ID"][key].has_key('PARAM_DATA_ID'):
                new_data.append(d)
        return new_data
    except Exception, ex:
        print ex
    finally:
        cursor.close()
        conn.close()


def get_interface_info():
    type = 'DB2'
    conn = None
    if type == 'DB2':
        # dbstring = "driver={IBM DB2 ODBC DRIVER};database=CSZDH;hostname=10.189.96.6;port=50000;protocol=TCPIP;uid=TIMESVC;pwd=timesvc;"
        # dbstring = "driver={IBM DB2 ODBC DRIVER};database=ZDH;hostname=10.187.96.46;port=50000;protocol=TCPIP;uid=timesvc;pwd=csgl@2017;LONGDATACOMPAT=1;LOBMAXCOLUMNSIZE=10485875;"
        conn = conn = cF.ConnectCaseDB(type)
    cursor = conn.cursor()
    ID_FUNCID = collections.OrderedDict()
    try:
        # sql = """select i.ID as ID, i.FUNCID as FUNCID, d.PARAM_NAME as PARAM_NAME, d.PARAM_INDEX as
        # PARAM_INDEX from SX_INTERFACE_INFO i
        # left join SX_INTERFACE_DETAIL_INFO d on i.ID = d.INTERFACE_ID where i.SYSTEM_ID = '14774643386617P4VW' and (%s)
        # order by i.FUNCID, d.PARAM_INDEX"""%str_like_funcid
        sql = """select i.ID as ID, i.FUNCID as FUNCID, d.PARAM_NAME as PARAM_NAME, d.PARAM_INDEX as
                PARAM_INDEX from SX_INTERFACE_INFO i
                left join SX_INTERFACE_DETAIL_INFO d on i.ID = d.INTERFACE_ID where i.SYSTEM_ID = '14774643386617P4VW'
                order by i.FUNCID, d.PARAM_INDEX"""
        cursor.execute(sql)
        all_data = collections.OrderedDict()
        all_file_name = []
        # 每个FUNCID的对应的数据以{'FUNCID':[PARAM_NAME,PARAM_NAME,PARAM_NAME]}的形式保存
        for row in cursor.fetchall():
            ID, FUNCID, PARAM_NAME, PARAM_INDEX = row
            ID_FUNCID[str(FUNCID)] = ID
            if str(FUNCID) in all_data.keys():
                all_data[str(FUNCID)].append(PARAM_NAME.decode('gbk', 'ignore').encode('utf-8'))
            else:
                if PARAM_NAME:
                    all_data[str(FUNCID)] = []
                    all_data[str(FUNCID)].append(PARAM_NAME.decode('gbk', 'ignore').encode('utf-8'))
                else:
                    all_data[str(FUNCID)] = []
            if FUNCID:
                if str(FUNCID)[0:3] == '000':
                    continue
                else:
                    all_file_name.append(str(FUNCID)[0:3])
        all_file_name = set(all_file_name)
        return all_file_name, all_data, ID_FUNCID
    except Exception as e:
        raise e
    finally:
        cursor.close()
        conn.close()


def get_sx_cases_info_export_excel(rs, system_id):
    # func_ids_prefix = ['001', '008']
    # if not func_ids_prefix:
    #     return 0, ""
    # func_ids_prefix_list = ["'" + i + "'" for i in func_ids_prefix]
    # func_ids_prefix_str = ",".join(func_ids_prefix_list)
    try:
        # str_like_remarks = ""
        # for func_id_str in func_ids_prefix:
        #     func_id_str = func_id_str + '%'
        #     func_id_str = " or detail.FUNCID like '%s'" % func_id_str
        #     str_like_remarks += func_id_str
        # str_like_remarks = str_like_remarks[3:]
        # if not str_like_remarks:
        #     return 0, ""
        # Logging.getLog().debug(u"查询数据库时间%s" % str(time.time()))
        # sql = "select sx.CASE_NAME ,detail.CASES_ID,detail.CASES_DETAIL_ID,detail.PARAM_ID, detail.FUNCID,DETAIL_PARAM_DATA.PARAM_DATA_ID " \
        #       ",DETAIL_PARAM_DATA.PARAM_DATA from sx_cases as sx, sx_cases_detail as detail, SX_CASES_DETAIL_PARAM_DATA as DETAIL_PARAM_DATA " \
        #       "where sx.CASES_ID = detail.CASES_ID and detail.CASES_DETAIL_ID =DETAIL_PARAM_DATA.CASES_DETAIL_ID and detail.cases_id in (select cases_id " \
        #       "from sx_cases WHERE  SYSTEM_ID='14774643386617P4VW') AND detail.IS_DELETE = '0'  AND detail.FUNCID IN (%s) order by detail.FUNCID"%func_ids_prefix_str
        # rs = cF.executeCaseSQL_DB2(sql)
        # Logging.getLog().debug(u"%s 数据量：%s" % (str(sql), len(rs)))
        # Logging.getLog().debug(u"查询数据库时间戳%s" % str(time.time()))
        # 获取excel的名称, 根据excel_name分类
        FUNCIDS = []
        func_ids_prefix = []
        for line in rs:
            FUNCID = line['FUNCID']
            if FUNCIDS or FUNCIDS != 'null':
                FUNCIDS.append(FUNCID)
                func_ids_prefix.append(FUNCID[0:3])
        FUNCIDS = [line['FUNCID'] for line in rs]
        func_ids_prefix = [line['FUNCID'][0:3] for line in rs]
        excel_names = {}
        for excel_name in func_ids_prefix:
            excel_names[str(excel_name)] = [FUNCID for FUNCID in FUNCIDS if FUNCID.startswith(str(excel_name))]
        Logging.getLog().debug(u"分类时间戳%s" % str(time.time()))
        for name in excel_names:
            wb = Workbook()
            funcids = excel_names[name]
            funcidlist = list(set(funcids))
            funcidlist.sort()
            for funcid in funcidlist:
                col_name = ['CASES_ID', 'PARAM_DATA_ID']
                # ws = wb.active
                ws = wb.create_sheet(str(funcid))
                # 创建sheet
                for ch in xrange(0x41, 0x5A):
                    # 设置列宽
                    ws.column_dimensions[str(unichr(ch))].width = 25
                # 判断列名是否添加
                flag = False
                # 获取功能号对应的数据
                for line in rs:
                    col_data = []
                    FUNCID = line['FUNCID']
                    if str(funcid) == str(FUNCID):
                        PARAM_DATA = line['PARAM_DATA']
                        CASES_ID = line['CASES_ID']
                        col_data.append(CASES_ID)
                        PARAM_DATA_ID = line['PARAM_DATA_ID']
                        col_data.append(PARAM_DATA_ID)
                        # 获取全部列名
                        # Logging.getLog().debug(u"%s" % str(PARAM_DATA))
                        if not PARAM_DATA or PARAM_DATA == "null":
                            continue
                        params = PARAM_DATA.split("#")
                        for param in params:
                            colname = param.split('=')[0]
                            col_name.append(colname)
                            coldata = param.split('=')[1].replace("&bfb&", "#").replace("&bft&", "=")
                            col_data.append(coldata)
                        # 设置列名
                        if flag == False:
                            ws.append(col_name)
                            flag = True
                        # 添加多行数据数据,实现多步骤
                        ws.append(col_data)
                        # 创建excel时， 有默认的sheet没有用到，删除
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
            wb.save(".//init//%s//%s.xlsx" % (system_id, name))
            Logging.getLog().debug(u"%s" % str(time.time()))
            Logging.getLog().debug(u">>>生成：%s.xlsx" % name)
    except BaseException, ex:
        exc_info = cF.getExceptionInfo()
        Logging.getLog().error(exc_info)


# def export_interface_info():
#     """导出excel"""
#     try:
#         error_func_id = ""
#         Logging.getLog().debug(u"开始获取接口数据")
#         data = get_param_data_info()
#         all_file_name, all_data, ID_FUNCID = get_interface_info()
#         Logging.getLog().debug(u"获取接口数据完成")
#         for i in all_file_name:
#             # all_file_name文件名，都是funcid的前三个字符组成
#             wb = Workbook()
#             # ws = wb.active
#             ws = wb.create_sheet(str(j))
#             # 创建sheet
#             for ch in xrange(0x41, 0x5A):
#                 # 设置列宽
#                 ws.column_dimensions[str(unichr(ch))].width = 25
#             for j in all_data:
#                 col_name = ['CASES_ID', 'PARAM_ID']
#                 # j是FUNCID
#                 # all_data是字典类型， key值是sheet的名称， values是对应的行数据
#                 ws = wb.create_sheet(str(j))
#                 # 创建sheet
#                 for ch in xrange(0x41, 0x5A):
#                     # 设置列宽
#                     ws.column_dimensions[str(unichr(ch))].width = 25
#                 if j.startswith(i):
#                     # 判断sheet的名称是不是以文件名开头的数据，如果是是同一个文件的数据，保存在同一个文件，否则反正
#                     error_func_id = str(j)
#                     # 数据列名可以直接分配到单元格中
#                     if all_data[str(j)]:
#                         # 添加数据
#                         col_name.extend(all_data[str(j)])
#                         ws.append(col_name)
#                     if str(j) in ID_FUNCID.keys():
#                         ID = ID_FUNCID[str(j)]
#                     sheet_data = {}
#                     for param_data in data:
#                         flag = True
#                         if int(ID) == int(param_data['PARAM_ID']):
#                             if flag == False:
#                                 sheet_data['CASES_DETAIL_ID'].update(param_data['CASES_DETAIL_ID'])
#                             else:
#                                 sheet_data.update(param_data)
#                             flag = False
#                     if sheet_data:
#                         CASES_DETAIL_IDS = sheet_data['CASES_DETAIL_ID'].keys()
#                         CASES_DETAIL_IDS.sort()
#                         for CASES_DETAIL_ID in CASES_DETAIL_IDS:
#                             paramdata = []
#                             one_record = {}
#                             paramdata.append(sheet_data['CASES_ID'])
#                             PARAM_DATA_ID = sheet_data['CASES_DETAIL_ID'][str(CASES_DETAIL_ID)]['PARAM_DATA_ID']
#                             paramdata.append(PARAM_DATA_ID)
#                             PARAM_DATA = sheet_data['CASES_DETAIL_ID'][str(CASES_DETAIL_ID)]['PARAM_DATA']
#                             if PARAM_DATA:
#                                 params = PARAM_DATA.split("#")
#                                 for param in params:
#                                     param_list = param.split('=')
#                                     paramList = [param.replace("&bfb&", "#").replace("&bft&", "=") for param in param_list]
#                                     if paramList:
#                                         one_record[str(paramList[0])] = paramList[1]
#                                 for colname in col_name:
#                                     if str(colname) in ('CASES_ID', 'PARAM_ID'):
#                                         continue
#                                     if str(colname) in one_record.keys():
#                                         paramdata.append(one_record[str(colname)])
#                                     else:
#                                         paramdata.append("")
#                                 ws.append(paramdata)
#             # 创建excel时， 有默认的sheet没有用到，删除
#             wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
#             wb.save("%s.xlsx" % i)
#             Logging.getLog().debug(u">>>生成：%s.xlsx" % i)
#     except BaseException, ex:
#         exc_info = cF.getExceptionInfo()
#         Logging.getLog().error(exc_info)
#         Logging.getLog().error(error_func_id)
def export_interface_info():
    """导出excel"""
    try:
        error_func_id = ""
        Logging.getLog().debug(u"开始获取接口数据")
        data = get_param_data_info()
        all_file_name, all_data, ID_FUNCID = get_interface_info()
        Logging.getLog().debug(u"获取接口数据完成")
        for i in all_file_name:
            # all_file_name文件名，都是funcid的前三个字符组成
            wb = Workbook()
            # ws = wb.active
            for j in all_data:
                col_name = ['CASES_ID', 'PARAM_ID']
                # j是FUNCID
                # all_data是字典类型， key值是sheet的名称， values是对应的行数据
                if j.startswith(i):
                    # 判断sheet的名称是不是以文件名开头的数据，如果是是同一个文件的数据，保存在同一个文件，否则反正
                    error_func_id = str(j)
                    ws = wb.create_sheet(str(j))
                    # 创建sheet
                    for ch in xrange(0x41, 0x5A):
                        # 设置列宽
                        ws.column_dimensions[str(unichr(ch))].width = 25
                    # 数据可以直接分配到单元格中
                    if all_data[str(j)]:
                        # 添加数据
                        col_name.extend(all_data[str(j)])
                        ws.append(col_name)
                    if str(j) in ID_FUNCID.keys():
                        ID = ID_FUNCID[str(j)]
                    sheet_data = {}
                    for param_data in data:
                        flag = True
                        if int(ID) == int(param_data['PARAM_ID']):
                            if flag == False:
                                sheet_data['CASES_DETAIL_ID'].update(param_data['CASES_DETAIL_ID'])
                            else:
                                sheet_data.update(param_data)
                            flag = False
                    if sheet_data:
                        CASES_DETAIL_IDS = sheet_data['CASES_DETAIL_ID'].keys()
                        CASES_DETAIL_IDS.sort()
                        for CASES_DETAIL_ID in CASES_DETAIL_IDS:
                            paramdata = []
                            one_record = {}
                            paramdata.append(sheet_data['CASES_ID'])
                            PARAM_DATA_ID = sheet_data['CASES_DETAIL_ID'][str(CASES_DETAIL_ID)]['PARAM_DATA_ID']
                            paramdata.append(PARAM_DATA_ID)
                            PARAM_DATA = sheet_data['CASES_DETAIL_ID'][str(CASES_DETAIL_ID)]['PARAM_DATA']
                            if PARAM_DATA:
                                params = PARAM_DATA.split("#")
                                for param in params:
                                    param_list = param.split('=')
                                    paramList = [param.replace("&bfb&", "#").replace("&bft&", "=") for param in
                                                 param_list]
                                    if paramList:
                                        one_record[str(paramList[0])] = paramList[1]
                                for colname in col_name:
                                    if str(colname) in ('CASES_ID', 'PARAM_ID'):
                                        continue
                                    if str(colname) in one_record.keys():
                                        paramdata.append(one_record[str(colname)])
                                    else:
                                        paramdata.append("")
                                ws.append(paramdata)
            # 创建excel时， 有默认的sheet没有用到，删除
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
            wb.save("%s.xlsx" % i)
            Logging.getLog().debug(u">>>生成：%s.xlsx" % i)
    except BaseException, ex:
        exc_info = cF.getExceptionInfo()
        Logging.getLog().error(exc_info)
        Logging.getLog().error(error_func_id)


if __name__ == '__main__':
    cF.readConfigFile()
    # 导出excel
    get_sx_cases_info_export_excel('')
    # 上传到FTP file_path 保存文件的路径
    # file_path = r"D:/"
    # upload_interface_excel(r'./', '14774643386617P4VW')
