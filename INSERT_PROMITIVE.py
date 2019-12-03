# coding=GBK
"""
导入原语
"""
import json
import os
import pyodbc
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import OrderedDict
import MySQLdb
# def getConnected_zdh():
#     try:
#         type = 'DB2'
#         conn = None
#         if type == 'DB2':
#             # dbstring = "driver={IBM DB2 ODBC DRIVER};database=TEST;hostname=10.189.111.9;port=50000;protocol=TCPIP;uid=AUTOTEST;pwd=gtja.8888;"
#             dbstring = "driver={IBM DB2 ODBC DRIVER};database=ZDH;hostname=10.189.111.9;port=50000;protocol=TCPIP;uid=AUTOTEST;pwd=gtja.8888;"
#             conn = pyodbc.connect(dbstring)
#         cursor = conn.cursor()
#         return conn, cursor
#     except Exception as e:
#         raise e

def getConnected1():
    try:
        type = 'DB2'
        conn = None
        if type == 'DB2':
            dbstring = "driver={IBM DB2 ODBC DRIVER};database=TEST;hostname=10.189.111.9;port=50000;protocol=TCPIP;uid=AUTOTEST;pwd=gtja.8888;"
            # dbstring = "driver={IBM DB2 ODBC DRIVER};database=ZDH;hostname=10.189.111.9;port=50000;protocol=TCPIP;uid=AUTOTEST;pwd=gtja.8888;"
            conn = pyodbc.connect(dbstring)
        cursor = conn.cursor()
        return conn, cursor
    except Exception as e:
        raise e

def getConnected():
    try:
        type = 'DB2'
        conn = None
        if type == 'DB2':
            dbstring = "driver={IBM DB2 ODBC DRIVER};database=CSGLGJ;hostname=5.5.6.66;port=50000;protocol=TCPIP;uid=ZDH;pwd=P@ssw0rd;"

            # dbstring = "driver={IBM DB2 ODBC DRIVER};database=TEST;hostname=10.189.111.9;port=50000;protocol=TCPIP;uid=AUTOTEST;pwd=gtja.8888;"
            # dbstring = "driver={IBM DB2 ODBC DRIVER};database=AUTOTEST;hostname=10.181.101.151;port=50000;protocol=TCPIP;uid=autotest;pwd=1GaTjest;"
            conn = pyodbc.connect(dbstring)
        cursor = conn.cursor()
        return conn, cursor
    except Exception as e:
        raise e
def getConnected_onling():
    try:
        type = 'DB2'
        conn = None
        if type == 'DB2':
            # dbstring = "driver={IBM DB2 ODBC DRIVER};database=ZDH;hostname=10.189.111.9;port=50000;protocol=TCPIP;uid=AUTOTEST;pwd=gtja.8888;"
            # dbstring = "driver={IBM DB2 ODBC DRIVER};database=TEST;hostname=10.189.111.9;port=50000;protocol=TCPIP;uid=AUTOTEST;pwd=gtja.8888;"
            dbstring = "driver={IBM DB2 ODBC DRIVER};database=AUTOTEST;hostname=10.181.101.151;port=50000;protocol=TCPIP;uid=autotest;pwd=1GaTjest;"
            #conn = pyodbc.connect(dbstring)
        cursor = conn.cursor()
        return conn, cursor
    except Exception as e:
        raise e


def INSERT_PROMITIVE(path):
    try:
        conn, cursor = getConnected()
        data = []
        wb = load_workbook(path)
        sheets = wb.get_sheet_names()
        # 获取所有的sheet
        if sheets:
            for i in sheets:
                # 读取每一个sheet的内容，只取第一行数据作为入参数据
                sheet = wb.get_sheet_by_name(i)
                for row in sheet.iter_rows():
                    child = []
                    for cell in row:
                        value = cell.value
                        if not value:
                            value = ""
                        child.append(value)
                    data.append(child)
        col_names = data.pop(0)
        dict_data = []
        for i in data:
            dict_data.append(dict(zip(col_names, i)))
        for child_dict_data in dict_data:
            SYSTEM_IDS = child_dict_data['VISIBLE_RANGE']
            FUNC_DESC = child_dict_data['THE_PRIMITIVE_DETAILS']
            OUT_PARAM = child_dict_data['OUT_PARAM']
            """
            # {
            #     “序号1”：{“IN_PARAM
            #     入参名称”：NAME，“IN_PARAM_DETAIL入参描述”：DEMO}，
            #     “序号2”：{“IN_PARAM
            #     入参名称”：NAME，“IN_PARAM_DETAIL入参描述”：DEMO}
            #     }
            # 
            #     OUT_PARAM
            #         出参
            #     {
            #     “序号1”：{“OUT_PARAM
            #     出参名称”：NAME，“OUT_PARAM_DETAIL出参描述”：DEMO}，
            #     “序号2”：{“OUT_PARAM
            #     出参名称”：NAME，“OUT_PARAM_DETAIL出参描述”：DEMO}
            #     }
            
            {
                "in_param":{
                    "in_param_1_name":"in_param_1_desc",
                    "in_param_2_name":"in_param_2_desc"
                },
                "out_param":{
                    "out_param_1_name":"out_param_1_desc",
                    "out_param_2_name":"out_param_2_desc"
                }
            }
                
                """
            PARAMMETER = {
                "in_param": OrderedDict(),
                "out_param": OrderedDict()
            }
            if OUT_PARAM:
                OUT_PARAMS = OUT_PARAM.split(",")
                for PARAM in OUT_PARAMS:
                    key = PARAM.split(":")[0].replace("\n", "").strip()
                    value = PARAM.split(":")[1].strip()
                    PARAMMETER['out_param'][key] = value

            IN_PARAM = child_dict_data['IN_PARAM']
            if IN_PARAM:
                IN_PARAM = IN_PARAM.split(",")
                for PARAM in IN_PARAM:
                    key = PARAM.split(":")[0].replace("\n", "").strip()
                    value = PARAM.split(":")[1].strip()
                    PARAMMETER['in_param'][key] = value
            PARAMMETER = json.dumps(PARAMMETER)
            FUNCNAME = child_dict_data['THE_FUNCTION_NAME']
            sql = "insert into SX_PUBLIC_FUNC_INFO (FUNCNAME,PARAMMETER,FUNC_DESC, SYSTEM_IDS,PUBLIC_FUNC_FLAG)" \
                  " VALUES ('%s','%s','%s','%s',0)" % (
                      FUNCNAME, PARAMMETER, FUNC_DESC, SYSTEM_IDS)
            # print sql
            cursor.execute(sql)
            # conn.commit()
            # print sql
    except Exception as e:
        print e


def select():
    try:
        conn, cursor = getConnected()

        sql = "select PRE_ACTION, PRO_ACTION from SX_CASES_DETAIL where cases_id in (select cases_id from sx_cases WHERE  SYSTEM_ID='1428910825111C5EQ4')"
        # sql = "select CASES_DETAIL_ID, PRE_ACTION, PRO_ACTION from SX_CASES_DETAIL"
        cursor.execute(sql)
        for row in cursor.fetchall():
            PRE_ACTION, PRO_ACTION = row
            action_list = []
            if PRE_ACTION.__contains__("\r\n"):
                action_list = PRE_ACTION.strip().split('\r\n')
            elif PRE_ACTION.__contains__("\r"):
                action_list = PRE_ACTION.strip().split('\r')
            elif PRE_ACTION.__contains__("\n"):
                action_list = PRE_ACTION.strip().split('\n')
            else:
                action_list = PRE_ACTION.strip().split('\\n')
            for i in action_list:
                # print i
                if 'SELECT' in i or 'select' in i:
                    print i

            action_list = []
            if PRO_ACTION.__contains__("\r\n"):
                action_list = PRO_ACTION.strip().split('\r\n')
            elif PRO_ACTION.__contains__("\r"):
                action_list = PRO_ACTION.strip().split('\r')
            elif PRO_ACTION.__contains__("\n"):
                action_list = PRO_ACTION.strip().split('\n')
            else:
                action_list = PRO_ACTION.strip().split('\\n')
            for i in action_list:
                # print i
                if 'SELECT' in i or 'select' in i:
                    print i
    except Exception as e:
        print e


def change_data_type():
    try:
        f = open("error_log.txt", 'a+')
        conn, cursor = getConnected()
        sql = "select CASES_DETAIL_ID,CASES_ID, PRE_ACTION, PRO_ACTION from SX_CASES_DETAIL where IS_DELETE='0'"
        # sql = "select CASES_DETAIL_ID,CASES_ID, PRE_ACTION, PRO_ACTION from SX_CASES_DETAIL where cases_id in (select cases_id from sx_cases WHERE  SYSTEM_ID='142891090059622AV2')"
        # sql = "select CASES_DETAIL_ID,CASES_ID, PRE_ACTION, PRO_ACTION from TIMESVC.SX_CASES_DETAIL where IS_DELETE='0' and  cases_id in (select cases_id from sx_cases where IS_DELETE='0' and DISABLEFLAG='0')"
        # sql = "select CASES_DETAIL_ID,CASES_ID, PRE_ACTION, PRO_ACTION from SX_CASES_DETAIL where IS_DELETE='0' and  cases_id in (select cases_id from sx_cases where IS_DELETE='0' and DISABLEFLAG='0' and SYSTEM_ID='1428910825111C5EQ4')"

        cursor.execute(sql)
        count = 0
        cases = []
        for row in cursor.fetchall():
            CASES_DETAIL_ID, CASES_ID, PRE_ACTION, PRO_ACTION = row
            # if "263770" == str(CASES_DETAIL_ID):
            #     pass
            # print 'cbs', CASES_DETAIL_ID, CASES_ID, PRE_ACTION, PRO_ACTION
            print 'cbs', CASES_DETAIL_ID, CASES_ID
            # f.write("CASES_DETAIL_ID:%s, CASES_ID=%s"%(CASES_DETAIL_ID, CASES_ID))
            count += 1
            print "count = %s"%count
            pre_action_dict, pro_action_dict, actions_pro, actions_pre = "", "", "", ""
            VARIABLE = {}
            if PRE_ACTION :
                action_list = []
                if PRE_ACTION.__contains__("\r\n"):
                    action_list = PRE_ACTION.strip().split('\r\n')
                elif PRE_ACTION.__contains__("\r"):
                    action_list = PRE_ACTION.strip().split('\r')
                elif PRE_ACTION.__contains__("\n"):
                    action_list = PRE_ACTION.strip().split('\n')
                else:
                    action_list = PRE_ACTION.strip().split('\\n')

                # action_list = ['call:n = sqlserver_Database(select convert(varchar(10),dateadd(day, -1, getdate()),112),m)']
                VARIABLE, actions_pre, cases_list = action_update(action_list, VARIABLE, CASES_ID)
                cases.extend(cases_list)
                # pre_action_dict = json.dumps(actions_pre, encoding="gbk")
                pre_action_dict = json.dumps(actions_pre, indent=4, ensure_ascii=False, encoding="gbk")
                pre_action_dict = pre_action_dict.replace("'", "''")
            # pre_action_dict = MySQLdb.escape_string(pre_action_dict)
            if PRO_ACTION :
                action_list = []
                if PRO_ACTION.__contains__("\r\n"):
                    action_list = PRO_ACTION.strip().split('\r\n')
                elif PRO_ACTION.__contains__("\r"):
                    action_list = PRO_ACTION.strip().split('\r')
                elif PRO_ACTION.__contains__("\n"):
                    action_list = PRO_ACTION.strip().split('\n')
                else:
                    action_list = PRO_ACTION.strip().split('\\n')

                VARIABLE, actions_pro, cases_list = action_update(action_list, VARIABLE, CASES_ID)
                cases.extend(cases_list)
                pro_action_dict = json.dumps(actions_pro, indent=4, ensure_ascii=False, encoding="gbk")
                pro_action_dict = pro_action_dict.replace("'", "''")
                # pro_action_dict = MySQLdb.escape_string(pro_action_dict)
            # print 'start'
            # sql = "update SX_CASES_DETAIL set PRE_ACTION='%s' , PRO_ACTION = '%s' where CASES_DETAIL_ID=%s" % (
            #     pre_action_dict, pro_action_dict, CASES_DETAIL_ID)
            # if not actions_pro and actions_pre:
            #     sql = "update SX_CASES_DETAIL set PRE_ACTION='%s' where CASES_DETAIL_ID=%s" % (
            #         pre_action_dict, CASES_DETAIL_ID)
            #     # print sql
            #     cursor.execute(sql)
            #     # conn.commit()
            if actions_pro:
                try:
                    sql = "update SX_CASES_DETAIL set PRO_ACTION = '%s' where CASES_DETAIL_ID=%s" % (
                        pro_action_dict, CASES_DETAIL_ID)
                    print sql
                    cursor.execute(sql)
                    # conn.commit()
                except Exception as e:
                    f.write("CASES_DETAIL_ID:%s, CASES_ID=%s" % (CASES_DETAIL_ID, CASES_ID))
                    f.write("\n")
                    f.write("PRO_ACTION  ERROR")
                    f.write("\n")
                    sql = "update SX_CASES_DETAIL set PRO_ACTION = '' where CASES_DETAIL_ID=%s" % CASES_DETAIL_ID
                    # print sql
                    cursor.execute(sql)
            if actions_pre:
                # print sql
                try:
                    sql = "update SX_CASES_DETAIL set PRE_ACTION = '%s' where CASES_DETAIL_ID=%s" % (
                        pre_action_dict, CASES_DETAIL_ID)
                    # print sql
                    cursor.execute(sql)
                except Exception as e:
                    f.write("CASES_DETAIL_ID:%s, CASES_ID=%s" % (CASES_DETAIL_ID, CASES_ID))
                    f.write("\n")
                    f.write("PRE_ACTION  ERROR")
                    f.write("\n")
                    sql = "update SX_CASES_DETAIL set PRE_ACTION = '' where CASES_DETAIL_ID=%s" % CASES_DETAIL_ID
                    # print sql
                    cursor.execute(sql)
            if VARIABLE:
                VARIABLE = json.dumps(VARIABLE, indent=4, ensure_ascii=False, encoding="gbk")
                VARIABLE = VARIABLE.replace("'", "''")
                sql = "update SX_CASES set VARIABLE='%s' where CASES_ID=%s" % (VARIABLE, CASES_ID)
                # print sql_1
                cursor.execute(sql)
        f.write("SUCCESS")
        f.close()
        # conn.commit()
        # print "cases :%s"%str(cases)
        cursor.close()
        conn.close()
    except Exception as e:
        print e


def action_update(action_list, VARIABLE, CASES_ID):
    try:
        # f = open("pre_pro_log.txt", "a+")
        cases = []
        actions = []
        all_action = []
        child_action = []
        flag = False
        for action in action_list:
            action = action.strip()
            if action.strip() == "":
                continue
            if action.startswith('"'):
                action = action[1:]
            if action.startswith('$bfh$'):
                action = action[5:]
            if "normal_RT_EXPRESSION" in action:
                action = action.replace("normal_RT_EXPRESSION", "normal_call:ASSERT_EXPRESSION")
            if action.startswith("="):
                a = []
                a.append(action)
                all_action.append(a)
                continue
            if flag:
                flag = False
                child_action = []
            if "#" in action:
                child_action.append(action)
            else:
                child_action.append(action)
                flag = True
                all_action.append(child_action)
        if action_list and not all_action:
            child_action = []
            child_action.append(",".join(action_list).strip())
            all_action.append(child_action)
        for i in all_action:
            action_flag = False
            retrun_json = {"call_when": "", "type": "", "content": "",
                           "desc": "", "PUBLIC_FUNC_FLAG": 0}
            for j in i:
                if "mobile_data_by_normalOfTable_to_dormancyOfTable" in j:
                    j = "normal_call:"+j
                if j.startswith("="):
                    retrun_json["desc"] = ",".join(i)
                    action_flag = True
                    actions.append(retrun_json)
                    continue
                if not j.startswith("#"):
                    action_flag = True
                    action = j
                    # print "huozhi %s" % action
                    return_list_str = ""
                    return_list = []
                    if "oracle_Database" in j:
                        index1 = action.rfind(",")
                        # 获取用户
                        action_1 = action[:index1]
                        user = action[index1:].replace(",", "").replace(")", "")
                        # 判断是否有变量
                        if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
                            return_list_str = action[action.find(":") + 1:action.find("=")].strip()
                            # param_list_str = action[action.find("=") + 1:].strip()
                        else:
                            param_list_str = action[action.find(":") + 1:].strip()
                        # func = param_list_str[:param_list_str.find("(")]
                        # param_info = param_list_str[param_list_str.find("(") + 1:param_list_str.rfind(")")]
                        if return_list_str != "":
                            return_list = return_list_str.split(',')

                        func = str(user.strip()) + ".exec_sql"
                        func = func.replace('"',"")
                        action = action_1.replace("oracle_Database", func) + ")"
                        if return_list:
                            index4 = action.find("=")
                            action = action[:index4] + "&equ&" +action[index4+1:]
                            for param in return_list:
                                param = param.strip()
                                VARIABLE[param] = ""
                        if 'SELECT' in action.upper() and 'DELETE' not in action.upper():
                            action_2 = action.upper()
                            field = "field"
                            if "FROM" not in action_2:
                                index9 = action.rfind(")")
                                action = action[:index9] + " as field " + action[index9:]
                                print action
                            else:
                                index3 = action_2.find("FROM")
                                index8 = action_2.find("SELECT")
                                # action_select = action_2[:index3]
                                field = action[index8 + 6:index3].strip()
                                # print field
                                if 'TO_CHAR' in field.upper():
                                    index6 = field.find(".")
                                    field_1 = field[index6 + 1:]
                                    index7 = field_1.find(")")
                                    field_2 = field_1[:index7]
                                    field = field_2
                                    action = action[:index8] + action[index8:index3] + "as %s " % field_2 + action[
                                                                                                            index3:]
                                elif "ROUND" in field.upper():#fixme
                                    # 'round( t.en_last_price,2)'
                                    field = field.strip()
                                    field = field[6:-1]
                                    field = field.split(",")[0]
                                    field = field.split(".")[1]
                                    action = action[:index3] + " as %s " % field + action[index3:]
                                else:
                                    if "," in field:
                                        field = field.split(",")[0].strip()
                                        field = field.split(".")[1]
                                        # print "select many field %s" % field
                                    elif "." in field:
                                        field = field.split(".")[1]
                            field = field.replace("top 1", "").strip()
                            action = action + "&dot&get_value(@DbTable[1][%s]@)" % field
                        index5 = action.find(":")
                        call_when = action[:index5]
                        action = action[index5 + 1:].strip()
                        if call_when == 'call':
                            call_when = "normal_call"
                            action = action.replace("call", "normal_call")
                        retrun_json["call_when"] = call_when
                        retrun_json["type"] = "DB_Operation"
                        retrun_json["content"] = action
                        retrun_json["PUBLIC_FUNC_FLAG"] = 0
                        if len(i) > 1:
                            retrun_json["desc"] = ",".join(i[:-1])
                    elif "exec_sql_database" in j:
                                index1 = action.rfind(",")
                                # 获取用户
                                action_1 = action[:index1]
                                user = action[index1:].replace(",", "").replace(")", "")
                                # 判断是否有变量
                                if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
                                    return_list_str = action[action.find(":") + 1:action.find("=")].strip()
                                    # param_list_str = action[action.find("=") + 1:].strip()
                                else:
                                    return_list_str=""
                                    param_list_str = action[action.find(":") + 1:].strip()
                                # func = param_list_str[:param_list_str.find("(")]
                                # param_info = param_list_str[param_list_str.find("(") + 1:param_list_str.rfind(")")]
                                if return_list_str != "":
                                    return_list = return_list_str.split(',')

                                if return_list:
                                    index4 = action.find("=")
                                    action = action[:index4] + "&equ&" + action[index4 + 1:]
                                    for param in return_list:
                                        param = param.strip()
                                        VARIABLE[param] = ""

                                index5 = action.find(":")
                                call_when = action[:index5]
                                action = action[index5 + 1:].strip()
                                if call_when == 'call':
                                    call_when = "normal_call"
                                    action = action.replace("call", "normal_call")
                                retrun_json["call_when"] = call_when
                                retrun_json["type"] = "DB_Operation"
                                retrun_json["content"] = action
                                retrun_json["PUBLIC_FUNC_FLAG"] = 0
                                if len(i) > 1:
                                    retrun_json["desc"] = ",".join(i[:-1])

                    elif "sqlserver_Database" in j:
                        index1 = action.rfind(",")
                        # 获取用户
                        action_1 = action[:index1]
                        user = action[index1:].replace(",", "").replace(")", "")
                        # 判断是否有变量
                        if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
                            return_list_str = action[action.find(":") + 1:action.find("=")].strip()
                            # param_list_str = action[action.find("=") + 1:].strip()
                        else:
                            param_list_str = action[action.find(":") + 1:].strip()
                        # func = param_list_str[:param_list_str.find("(")]
                        # param_info = param_list_str[param_list_str.find("(") + 1:param_list_str.rfind(")")]
                        if return_list_str != "":
                            return_list = return_list_str.split(',')

                        func = str(user.strip()) + ".exec_sql"
                        func = func.replace('"', "")
                        action = action_1.replace("sqlserver_Database", func) + ")"

                        if return_list:
                            index4 = action.find("=")
                            action = action[:index4] + "&equ&" + action[index4 + 1:]
                            for param in return_list:
                                param = param.strip()
                                VARIABLE[param] = ""

                        if 'SELECT' in action.upper() and 'DELETE' not in action.upper():
                            action_2 = action.upper()
                            field = "filed"
                            if "FROM" not in action_2:
                                index9 = action.rfind(")")
                                action = action[:index9] + " as filed " + action[index9:]
                                print action
                            else:
                                index3 = action_2.find("FROM")
                                index8 = action_2.find("SELECT")
                                # action_select = action_2[:index3]
                                field = action[index8+6:index3].strip()
                                # print field
                                if 'TO_CHAR' in field.upper():
                                    index6 = field.find(".")
                                    field_1 = field[index6+1:]
                                    index7 = field_1.find(")")
                                    field_2 = field_1[:index7]
                                    field = field_2
                                    action = action[:index8] + action[index8:index3]+"as %s "%field_2 +action[index3:]
                                elif "ROUND" in field.upper():#fixme
                                    # 'round( t.en_last_price,2)'
                                    field = field.strip()
                                    field = field[6:-1]
                                    field = field.split(",")[0]
                                    field = field.split(".")[1]
                                    action = action[:index3] + " as %s " % field + action[index3:]
                                else:
                                    if "," in field:
                                        field = field.split(",")[0].strip()
                                        field = field.split(".")[1]
                                        # print "select many field %s" % field
                                    elif "." in field:
                                        field = field.split(".")[1]
                            field = field.replace("top 1", "").strip()
                            action = action + "&dot&get_value(@DbTable[1][%s]@)" % field
                        # call_when = action.split(":")[0]
                        index5 = action.find(":")
                        call_when = action[:index5]
                        action = action[index5 + 1:].strip()
                        if call_when == 'call':
                            call_when = "normal_call"
                            action = action.replace("call", "normal_call")
                        retrun_json["call_when"] = call_when
                        retrun_json["type"] = "DB_Operation"
                        retrun_json["content"] = action
                        retrun_json["PUBLIC_FUNC_FLAG"] = 0
                        if len(i) > 1:
                            retrun_json["desc"] = ",".join(i[:-1])
                    elif 'call' in j or "normal_call" in j or "except_call" in j or "final_call" in j:
                        if "normal_call" in j:
                            call_when = "normal_call"
                        elif "except_call" in j:
                            call_when = "except_call"
                        elif "final_call" in j:
                            call_when = "final_call"
                        elif 'call' in j:
                            action = j.replace("call", 'normal_call')
                            call_when = "normal_call"
                        else:
                            pass
                        if action[:action.find("(")].find("=") >= 0:  # 有返回值模式
                            return_list_str = action[action.find(":") + 1:action.find("=")].strip()
                            if return_list_str != "":
                                return_list = return_list_str.split(',')
                                if return_list:
                                    index4 = action.find("=")
                                    action = action[:index4] + "&equ&" + action[index4 + 1:]
                                    for param in return_list:
                                        param = param.strip()
                                        VARIABLE[param] = ""
                        index5 = action.find(":")
                        call_when = action[:index5]
                        action = action[index5 + 1:].strip()
                        retrun_json["call_when"] = call_when
                        retrun_json["type"] = "General_Operation"
                        retrun_json["content"] = action
                        retrun_json["PUBLIC_FUNC_FLAG"] = 0
                        if len(i) > 1:
                            retrun_json["desc"] = ",".join(i[:-1])
                    else:
                        if "save_param" in action:
                            action_new = action.replace("save_param:", "")
                            action_new_list = action_new.split(",")
                            action_param = []
                            key_str = []
                            value_str = []
                            for action_new_param in action_new_list:
                                if "=" in action_new_param:
                                    key = action_new_param.split("=")[0]
                                    value = action_new_param.split("=")[1]
                                    key_str.append(key)
                                    value_str.append(value)
                                else:
                                    index2 = action_new_param.find("[")
                                    key = action_new_param[:index2]
                                    value = action_new_param[index2:]
                                    key_str.append(key)
                                    value_str.append(value)
                            if key_str:
                                for param in key_str:
                                    param = param.strip()
                                    VARIABLE[param] = ""
                                key_str = ",".join(key_str)
                            if value_str:
                                value_str = ",".join(value_str)

                            action = "normal_call:" + key_str + "&equ&save_param(" + value_str + ")"
                            index5 = action.find(":")
                            call_when = action[:index5]
                            action = action[index5 + 1:].strip()
                            retrun_json["call_when"] = "normal_call"
                            retrun_json["type"] = "General_Operation"
                            retrun_json["content"] = action
                            retrun_json["PUBLIC_FUNC_FLAG"] = 0
                            if len(i) > 1:
                                retrun_json["desc"] = ",".join(i[:-1])
                        elif "save_in_param" in action:
                            action_new = action.replace("save_in_param:", "")
                            action_new_list = action_new.split(",")
                            action_param = []
                            key_str = []
                            value_str = []
                            for action_new_param in action_new_list:
                                if "=" in action_new_param:
                                    key = action_new_param.split("=")[0]
                                    value = action_new_param.split("=")[1]
                                    key_str.append(key)
                                    value_str.append(value)
                                else:
                                    index2 = action_new_param.find("[")
                                    key = action_new_param[:index2]
                                    value = action_new_param[index2:]
                                    key_str.append(key)
                                    value_str.append(value)
                            if key_str:
                                for param in key_str:
                                    param = param.strip()
                                    VARIABLE[param] = ""
                                key_str = ",".join(key_str)
                            if value_str:
                                value_str = ",".join(value_str)

                            action = "normal_call:" + key_str + "&equ&save_in_param(" + value_str + ")"
                            index5 = action.find(":")
                            call_when = action[:index5]
                            action = action[index5 + 1:].strip()
                            retrun_json["call_when"] = "normal_call"
                            retrun_json["type"] = "General_Operation"
                            retrun_json["content"] = action
                            retrun_json["PUBLIC_FUNC_FLAG"] = 0
                            if len(i) > 1:
                                retrun_json["desc"] = ",".join(i[:-1])
                        else:
                            print "have no wirte action : %s"%action
                            # f.write("have no wirte action : %s"%action)
                            cases.append(CASES_ID)
                    # print "huozhi new %s" % action
                    actions.append(retrun_json)
                else:
                    print "this is desc %s"%j
            if action_flag == False:
                retrun_json["desc"] = ",".join(i)
                actions.append(retrun_json)
        # f.close()
        return VARIABLE, actions, cases
    except Exception as e:
        print e

def update_yuanyu():
    conn, cursor = getConnected1()
    sql = "select FUNCNAME,SYSTEM_IDS from TIMESVC.SX_PUBLIC_FUNC_INFO"
    cursor.execute(sql)
    data = {}
    for row in cursor.fetchall():
        FUNCNAME, SYSTEM_IDS = row
        data[FUNCNAME] = SYSTEM_IDS
    # conn.commit()
    # print "cases :%s" % str(cases)
    cursor.close()
    conn.close()

    conn, cursor = getConnected()
    for i in data:
        sql = "update SX_PUBLIC_FUNC_INFO set SYSTEM_IDS='%s' where FUNCNAME='%s'"%(data[i], i)
        cursor.execute(sql)
    # conn.commit()
    # print "cases :%s" % str(cases)
    cursor.close()
    conn.close()


def test():
    conn, cursor = getConnected()
    sql = "select cast(PRE_ACTION as VARCHAR(32672)) as PRE_ACTION ,cast(PRO_ACTION as VARCHAR(32672)) as PRO_ACTION from SX_CASES_DETAIL_CBS_TEST"
    cursor.execute(sql)
    ds = cursor.fetchall()
    print ds
    for row in ds:
        PRE_ACTION, PRO_ACTION = row
        print PRE_ACTION, PRO_ACTION


def get_duoyu_content():
    try:
        # conn, cursor = getConnected_onling()
        conn, cursor = getConnected()
        sql = "select c.CASE_NAME, d.CASES_DETAIL_ID, d.CASES_ID, cast(d.PRE_ACTION as VARCHAR(32672)) as PRE_ACTION ,cast(d.PRO_ACTION as VARCHAR(32672)) as PRO_ACTION, c.SYSTEM_ID from SX_CASES_DETAIL d, sx_cases c where c.cases_id = d.cases_id and d.IS_DELETE='0' and  c.IS_DELETE='0' and c.DISABLEFLAG='0'"
        cursor.execute(sql)
        ds = cursor.fetchall()
        for row in ds:
            flag = False
            CASE_NAME, CASES_DETAIL_ID, CASES_ID, PRE_ACTION, PRO_ACTION, SYSTEM_ID = row
            # print CASE_NAME, CASES_DETAIL_ID, CASES_ID
            # pro
            if not PRE_ACTION or PRE_ACTION == 'null':
                pass
            else:
                PRE_ACTION_LIST = json.loads(PRE_ACTION, encoding='gbk')
                PRE_ACTION_LIST_NEW = []
                for i in PRE_ACTION_LIST:
                    # {
                    #     "content": "p--设置回购状态bondreg为0，不需要了：--.exec_sql(update run..secuid set regflag = 'B' WHERE secuid ='@stock_account1@')",
                    #     "PUBLIC_FUNC_FLAG": 0,
                    #     "call_when": "normal_call",
                    #     "type": "DB_Operation",
                    #     "desc": ""
                    # },
                    if "--" in i["content"]:
                        flag = True
                        content = i["content"]
                        print content
                        print "PRE_ACTION"
                        index0 = content.find("--")
                        index1 = content.find(".exec_sql")
                        content1 = content[:index0]
                        content2 = content[index1:]
                        content = content1 + content2
                        i["content"] = content
                        PRE_ACTION_LIST_NEW.append(i)
                    else:
                        PRE_ACTION_LIST_NEW.append(i)
                pre_action_dict = json.dumps(PRE_ACTION_LIST_NEW, indent=4, ensure_ascii=False, encoding="gbk")
                pre_action_dict = pre_action_dict.replace("'", "''")
                sql = "update SX_CASES_DETAIL set PRE_ACTION = '%s' where CASES_DETAIL_ID=%s" % (
                    pre_action_dict, CASES_DETAIL_ID)
                cursor.execute(sql)

            #pro
            if not PRO_ACTION or PRO_ACTION == 'null':
                pass
            else:
                PRO_ACTION_LIST = json.loads(PRO_ACTION, encoding='gbk')
                PRO_ACTION_LIST_NEW = []
                for j in PRO_ACTION_LIST:
                    # {
                    #     "content": "p--设置回购状态bondreg为0，不需要了：--.exec_sql(update run..secuid set regflag = 'B' WHERE secuid ='@stock_account1@')",
                    #     "PUBLIC_FUNC_FLAG": 0,
                    #     "call_when": "normal_call",
                    #     "type": "DB_Operation",
                    #     "desc": ""
                    # },
                    if "--" in j["content"]:
                        flag = True
                        content = j["content"]
                        print content
                        print PRO_ACTION
                        index0 = content.find("--")
                        index1 = content.find(".exec_sql")
                        content1 = content[:index0]
                        content2 = content[index1:]
                        content = content1 + content2
                        j["content"] = content
                        PRO_ACTION_LIST_NEW.append(j)
                    else:
                        PRO_ACTION_LIST_NEW.append(j)
                pro_action_dict = json.dumps(PRO_ACTION_LIST_NEW, indent=4, ensure_ascii=False, encoding="gbk")
                pro_action_dict = pro_action_dict.replace("'", "''")
                sql = "update SX_CASES_DETAIL set PRO_ACTION = '%s' where CASES_DETAIL_ID=%s" % (
                    pro_action_dict, CASES_DETAIL_ID)
                cursor.execute(sql)
            if flag == True:
                print CASE_NAME, CASES_DETAIL_ID, CASES_ID
        print "ok"
        # conn.commit()
        cursor.close()
        conn.close()

    except Exception as e:
        print e


def select_get_duoyu_content():
    try:
        print "starting"
        conn, cursor = getConnected_onling()
        # conn, cursor = getConnected()
        # sql = "select c.CASE_NAME, d.CASES_DETAIL_ID, d.CASES_ID, cast(d.PRE_ACTION as VARCHAR(32672)) as PRE_ACTION ,cast(d.PRO_ACTION as VARCHAR(32672)) as PRO_ACTION, c.SYSTEM_ID from SX_CASES_DETAIL d, sx_cases c where c.cases_id = d.cases_id and d.IS_DELETE='0' and  c.IS_DELETE='0' and c.DISABLEFLAG='0'"
        sql = "select c.CASE_NAME, d.CASES_DETAIL_ID, d.CASES_ID, d.PRE_ACTION ,d.PRO_ACTION, c.SYSTEM_ID from SX_CASES_DETAIL d, sx_cases c where c.cases_id = d.cases_id and d.IS_DELETE='0' and  c.IS_DELETE='0' and c.DISABLEFLAG='0'"
        cursor.execute(sql)
        ds = cursor.fetchall()
        print "ds"
        for row in ds:
            # count +=1
            # print count
            flag = False
            CASE_NAME, CASES_DETAIL_ID, CASES_ID, PRE_ACTION, PRO_ACTION, SYSTEM_ID = row
            # print CASE_NAME, CASES_DETAIL_ID, CASES_ID
            # pro
            if not PRE_ACTION or PRE_ACTION == 'null':
                pass
            else:
                PRE_ACTION_LIST = json.loads(PRE_ACTION, encoding='gbk')
                PRE_ACTION_LIST_NEW = []
                for i in PRE_ACTION_LIST:
                    # {
                    #     "content": "p--设置回购状态bondreg为0，不需要了：--.exec_sql(update run..secuid set regflag = 'B' WHERE secuid ='@stock_account1@')",
                    #     "PUBLIC_FUNC_FLAG": 0,
                    #     "call_when": "normal_call",
                    #     "type": "DB_Operation",
                    #     "desc": ""
                    # },
                    if "--" in i["content"]:
                        flag = True
                        content = i["content"]
                        print content
                        print "PRE_ACTION"
                        index0 = content.find("--")
                        index1 = content.find(".exec_sql")
                        content1 = content[:index0]
                        content2 = content[index1:]
                        content = content1 + content2
                        i["content"] = content
                        PRE_ACTION_LIST_NEW.append(i)
                    else:
                        PRE_ACTION_LIST_NEW.append(i)
                # pre_action_dict = json.dumps(PRE_ACTION_LIST_NEW, indent=4, ensure_ascii=False, encoding="gbk")
                # pre_action_dict = pre_action_dict.replace("'", "''")
                # sql = "update SX_CASES_DETAIL set PRE_ACTION = '%s' where CASES_DETAIL_ID=%s" % (
                #     pre_action_dict, CASES_DETAIL_ID)
                # cursor.execute(sql)

            #pro
            if not PRO_ACTION or PRO_ACTION == 'null':
                pass
            else:
                PRO_ACTION_LIST = json.loads(PRO_ACTION, encoding='gbk')
                PRO_ACTION_LIST_NEW = []
                for j in PRO_ACTION_LIST:
                    # {
                    #     "content": "p--设置回购状态bondreg为0，不需要了：--.exec_sql(update run..secuid set regflag = 'B' WHERE secuid ='@stock_account1@')",
                    #     "PUBLIC_FUNC_FLAG": 0,
                    #     "call_when": "normal_call",
                    #     "type": "DB_Operation",
                    #     "desc": ""
                    # },
                    if "--" in j["content"]:
                        flag = True
                        content = j["content"]
                        print content
                        print PRO_ACTION
                        index0 = content.find("--")
                        index1 = content.find(".exec_sql")
                        content1 = content[:index0]
                        content2 = content[index1:]
                        content = content1 + content2
                        j["content"] = content
                        PRO_ACTION_LIST_NEW.append(j)
                    else:
                        PRO_ACTION_LIST_NEW.append(j)
                # pro_action_dict = json.dumps(PRO_ACTION_LIST_NEW, indent=4, ensure_ascii=False, encoding="gbk")
                # pro_action_dict = pro_action_dict.replace("'", "''")
                # sql = "update SX_CASES_DETAIL set PRO_ACTION = '%s' where CASES_DETAIL_ID=%s" % (
                #     pro_action_dict, CASES_DETAIL_ID)
                # cursor.execute(sql)
            if flag == True:
                print CASE_NAME, CASES_DETAIL_ID, CASES_ID
        print "ok"
    except Exception as e:
        print e
if __name__ == '__main__':
    change_data_type()