# -*- coding:gbk -*-
import time, Image
import os, win32gui, win32ui, win32con, win32api
import zlib
import base64
import os
from collections import OrderedDict
import uuid
import json
import commFuncs as cF
import csv
from lxml import etree
import Adapter
from Adapter.T2 import T2Adapter
import time
import os
# import anytree
import logging
import os.path
import zipfile
from ftpClient import XFer
import shutil
import sys
import MySQLdb
import datetime
import subprocess
import gl
from PIL import ImageGrab
from openpyxl import load_workbook, Workbook

try:
    import xml.etree.cElementTree as ET
except ImportError:
    import xml.etree.ElementTree as ET

error_photo_path = ""


def log_record(funcid):
    try:
        global error_photo_path
        timestr = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        timestr = str(timestr) + "_" + str(funcid)
        home_dir = os.getcwd()
        file_dir = home_dir + "\\tools_log\\%s\\" % timestr
        filedir = os.path.split(file_dir)[0]
        if not os.path.isdir(filedir):
            os.makedirs(filedir)
            error_photo_path = filedir

        file_path = file_dir + 'logs.txt'
        if not os.path.exists(file_path):
            os.system(r"%s" % file_path)

        logging.basicConfig(filename=file_path, format='[%(asctime)s-%(filename)s-%(levelname)s:%(message)s]',
                            level=logging.DEBUG, filemode='a', datefmt='%Y-%m-%d%I:%M:%S %p')
    except BaseException, ex:
        exc_info = cF.getExceptionInfo()
        print u"异常发生，请检查日志文件,提示信息为:%s" % (exc_info)
        logging.error(u"异常发生，请检查日志文件,提示信息为:%s" % (exc_info))
        return -1, u"异常发生，请检查日志文件,提示信息为:%s" % (exc_info), "", ""


def encode_a_file(path):
    f = open(path, 'r')
    str = base64.b64encode(zlib.compress(f.read()))
    print len(str)
    print str
    return str


def decode_to_file(str, path):
    f = open(path, 'w')
    f.write(zlib.decompress(base64.b64decode(str)))
    f.close()
    return 0


def testzlib():
    s = "hello word, 00000000000000000000000000000000"
    print len(s)
    c = zlib.compress(s)
    print len(c)

    base = base64.b64encode(c)

    print len(base)

    print len(base64.b64encode(s))

    t = base64.b64decode(base)

    d = zlib.decompress(t)
    print d


def scanScriptToJson(path):
    '''
    遍历脚本目录，产生相应的配置文件
    '''
    interfaceDetail = {}
    detail_list = []

    for parent, dirnames, filenames in os.walk(path):
        for dirname in dirnames:
            if dirname.find('jy') < 0:
                continue
            detail = OrderedDict()
            detail["funcid"] = str(uuid.uuid1()).replace('-', '')
            detail["funcname"] = ""
            detail["scriptname"] = dirname
            input_param = OrderedDict()
            input_param["name"] = ""
            input_param["desc"] = ""
            detail["InputParams"] = []
            detail["InputParams"].append(input_param)
            detail["InputParams"].append(input_param)
            detail["OutputParams"] = []

            detail_list.append(detail)
        break
    interfaceDetail["intrfacedetails"] = detail_list
    # dump to interface details json
    open('interfacedetail.json', 'w').write(json.dumps(interfaceDetail, ensure_ascii=False, encoding="utf8"))


def importCasesFromDetailJson(path):
    detail_interface = json.loads(open(path, 'r').read(), object_pairs_hook=OrderedDict)
    details = detail_interface["interfacedetails"]
    cases_auto_id = getAutoIncrement('sx_cases')
    case_detail_auto_id = getAutoIncrement('sx_cases_detail')
    param_data_auto_id = getAutoIncrement('sx_cases_detail_param_data')
    for i, func in enumerate(details):
        case_id = i + cases_auto_id
        case_code = 'FY' + str(i + 1).zfill(7)
        case_name = func["funcname"]

        sql = "insert into sx_cases (CASES_ID,CASES_CODE,CASE_NAME,TREE_DIRECTORY_ID,TEST_SYSTEM_ID,CASE_LEVEL,IS_DELETE) values(%d,'%s','%s',2,'TEST_SYSTEM_FY','CASE_LEVEL_HIGH','0')" % (
            case_id, case_code, case_name)
        print sql
        cF.executeCaseSQL(sql)

        # case_detail_id = generateUUID()#str(uuid.uuid1()).replace('-','')
        case_detail_id = case_detail_auto_id + i
        step = '1'
        adapter_type = 'SPLX_SIKULI'
        input_string_list = []
        input_string_list.append("path=%s" % func["scriptname"])
        input_string_list.append("user=")
        input_string_list.append("pass=")
        for input_param in func["InputParams"]:
            input_string_list.append("%s=" % input_param["name"])

        input_string = '#'.join(input_string_list)
        sql = "insert into sx_cases_detail (CASES_DETAIL_ID,CASES_ID,STEP,ADAPTER_TYPE,INPUT_STRING) values(%d,%d,'%s','%s','%s')" % (
            case_detail_id, case_id, step, adapter_type, input_string)
        print sql
        cF.executeCaseSQL(sql)

        param_data_id = param_data_auto_id + i
        param_data = ""

        input_string_list.pop(0)
        if input_string_list != []:
            param_data = "#".join(input_string_list)
        sql = "insert into sx_cases_detail_param_data (PARAM_DATA_ID,CASES_DETAIL_ID,PARAM_DATA,EXPECTED_VALUE) values(%d,%d,'%s','0')" % (
            param_data_id, case_detail_id, param_data)
        print sql
        cF.executeCaseSQL(sql)


def calculateDeeper(d, p):
    level = 0
    if p == '' or d[p] == '':
        return level
    while d[p] != '':
        level += 1
        p = d[p]
    return level


def exportFuncTree(path):
    '''
    '''
    sql = "select * from sx_func_tree_directory"
    ds = cF.executeCaseSQL(sql)
    dataset = []
    field_list = ['no', 'name', 'parent', 'parent_info', 'IS_LEAF']
    dataset.append(field_list)
    for rec in ds:
        value_list = [rec['TREE_DIRECTORY_ID'], rec['DIRECTORY_NAME'], rec['PARENT_ID'], rec['PARAM_INFO'],
                      rec['IS_LEAF']]
        dataset.append(value_list)

    writer = csv.writer(file('menu1.csv', 'wb'), quoting=csv.QUOTE_NONE)
    writer.writerows(dataset)


def importFuncTree(path):
    '''
    需要确保TREE_DIRECTORY_ID 从1 开始编号
    '''
    tree_parent_no_dict = OrderedDict()
    tree_parent_no_dict[0] = ''

    csvfile = file(path, 'r')
    reader = csv.reader((line.replace('\0', '') for line in csvfile), delimiter=',')
    for i, line in enumerate(reader):
        if i == 0:
            continue
        tree_parent_no_dict[int(line[0])] = int(line[2])
        directory_name = line[1].decode('gbk')
        directory_level = 'DIRECTORY_LEVEL_%d' % (calculateDeeper(tree_parent_no_dict, int(line[2])))
        sql = "insert into sx_func_tree_directory (TREE_DIRECTORY_ID,DIRECTORY_NAME,DIRECTORY_TYPE,PARENT_ID,PARAM_INFO,IS_LEAF) values(%d,'%s','%s',%d,'%s',%d)" % (
            int(line[0]), directory_name, directory_level, int(line[2]), line[3], int(line[4]))
        print sql
        cF.executeCaseSQL(sql)


'''    
def importFuncTree(path):
    csvfile = file(path,'r')
    reader = csv.reader((line.replace('\0','') for line in csvfile),delimiter=',')

    tree_parent_no_dict = OrderedDict()
    tree_parent_no_dict['0']= ""
    tree_id_dict = OrderedDict()
    tree_id_dict['0']= 1
    
    for i,line in enumerate(reader):
        if i == 0:
            continue
        # tree_id = str(uuid.uuid1()).replace('-','')
        tree_parent_no_dict[str(line[0])] = str(line[2])
        # tree_id_dict[str(line[0])] = tree_id
        directory_name = line[1].decode('gbk')
        if str(line[2]) == '0': #文档中的根节点
            directory_level = 'DIRECTORY_LEVEL_%d'%(calculateDeeper(tree_parent_no_dict,str(line[2])))
            sql = "insert into sx_func_tree_directory (DIRECTORY_NAME,DIRECTORY_TYPE,PARENT_ID) values('%s','%s',0)" %(directory_name,directory_level)
            print sql
            # 执行
            cF.executeCaseSQL(sql)

            # 获取TREE_DIRECTORY_ID 号
            sql = "select tree_directory_id from sx_func_tree_directory where directory_name = '%s'" %directory_name
            print sql
            ds = cF.executeCaseSQL(sql)
            tree_id_dict[str(line[0])] = ds[0]["tree_directory_id"]
        else:
            parent_id = tree_id_dict[line[2]]
            directory_level = 'DIRECTORY_LEVEL_%d'%calculateDeeper(tree_parent_no_dict,str(line[2]))
            sql = "insert into sx_func_tree_directory (DIRECTORY_NAME,DIRECTORY_TYPE,PARENT_ID) values('%s','%s',%d)" %(directory_name,directory_level,parent_id)
            print sql
            cF.executeCaseSQL(sql)
            # 获取TREE_DIRECTORY_ID 号
            sql = "select tree_directory_id from sx_func_tree_directory where directory_name = '%s'" %directory_name
            print sql
            ds = cF.executeCaseSQL(sql)
            if len(ds)!= 1:
                print "%s has two more info" %directory_name
            tree_id_dict[str(line[0])] = ds[0]["tree_directory_id"]
'''


def addScriptname():
    sql = 'select cases_detail_id,param_data_id from sx_cases_detail_param_data'
    ds = cF.executeCaseSQL(sql)
    for rec in ds:
        sql = "select input_string from sx_cases_detail where cases_detail_id = '%s'" % rec['CASES_DETAIL_ID']
        ds1 = cF.executeCaseSQL(sql)
        input_string = ds1[0]['input_string']
        path = input_string.split('#')[0]
        scriptname = path[5:]
        sql = "update sx_cases_detail_param_data set script_name='%s' where param_data_id='%s'" % (
            scriptname, rec['PARAM_DATA_ID'])
        print sql
        cF.executeCaseSQL(sql)


def addCollectionDetail():
    sql = 'select param_data_id from sx_cases_detail_param_data'
    ds = cF.executeCaseSQL(sql)

    # ds=[{"param_data_id":"8078b4b07ff811e69ebcb8763f280014"},{"param_data_id":"807b25b07ff811e69cffb8763f280014"},{"param_data_id":"808d75307ff811e697e4b8763f280014"},{"param_data_id":"80905b617ff811e6bbbdb8763f280014"}]
    ds = [{"param_data_id": "807e32f07ff811e687e0b8763f280014"}, {"param_data_id": "8081b5617ff811e6bd0eb8763f280014"},
          {"param_data_id": "8084e9b07ff811e6a3a7b8763f280014"}, {"param_data_id": "8087cfe17ff811e6a57bb8763f280014"}]

    for rec in ds:
        collection_detail_id = str(uuid.uuid1()).replace('-', '')
        sql = "insert into sx_cases_collection_detail(collection_detail_id,collection_id,param_data_id) values('%s','63b7d82184a211e6b69ab8763f280014','%s')" % (
            collection_detail_id, rec['PARAM_DATA_ID'])
        print sql
        cF.executeCaseSQL(sql)


def generateUUID():
    return str(uuid.uuid1()).replace('-', '')


def convertT2CasesToJson(path):
    '''
    将导出的测试用例，转化为json案例
    '''
    cases_list = []  # 用来产生json
    case_dict = OrderedDict()
    doc = etree.ElementTree(file=path)
    root = doc.getroot()
    xpath = './Test/sub'
    subs = root.findall(xpath)
    case_dict["caseid"] = generateUUID()
    step_list = []

    for stepno, sub in enumerate(subs):
        step_dict = OrderedDict()
        step_dict["stepid"] = generateUUID()
        step_dict["step_no"] = str(stepno + 1)
        step_dict["step_note"] = sub.attrib["note"]
        step_dict["adapter_type"] = "ADAPTER_T2"
        step_dict["addtional_info"] = "funcid=%s#livetime=%s#pack_ver=%s#accountInfo_id=0#note=%s" % (
            sub.attrib["id"], sub.attrib["livetime"], sub.attrib["pack_ver"], sub.attrib["note"])

        cmd_list = []

        test_data_list = []
        test_data_dict = OrderedDict()
        test_data_dict["param_data_id"] = ""
        test_data_dict["path"] = ""
        test_data_dict["param"] = ""
        test_data_dict["expectation"] = "0"
        test_data_dict["pres"] = sub[2].attrib["test"]
        test_data_dict["pros"] = sub[3].attrib["test"]
        for item in sub[1]:
            cmd_list.append("%s=%s" % (item.attrib["name"], item.attrib["value"]))
        test_data_dict["param"] = "#".join(cmd_list)
        test_data_list.append(test_data_dict)
        step_dict["test_data"] = test_data_list
        step_list.append(step_dict)
    case_dict["caseSteps"] = step_list
    open("init.json", "w").write(json.dumps(case_dict, ensure_ascii=False, encoding="utf8"))


def importParamDetailintoParamData(path):
    '''
    将脚本中参数.txt的相关信息导入到sx_cases_detail_param_data中
    '''
    lines = open(path, 'r').readlines()
    for line in lines:
        if line.startswith('runsikulix.cmd') == False:
            # print line
            continue
        # print line

        r = line.find('.sikuli') + 7
        l = line.rfind('\\', 0, r)
        print line[l + 1:r]

        script_name = line[l + 1:r]
        print script_name
        param = line[line.find('-- ') + 3:].strip('\r\n').strip(' ')
        param_list = param.split(' ')
        param_list.pop(0)  # 将第一个参数剔除

        sql = "select cases_detail_id,input_string from sx_cases_detail where input_string like '%%%s%%'" % script_name

        print sql
        ds = cF.executeCaseSQL(sql)
        if len(ds) == 0:
            print u"没找到案例细节"
            continue
        elif len(ds) > 1:
            print u"找到多个案例，展示不支持，只取第一个"

        cases_detail_id = ds[0]['cases_detail_id']  # 展示只考虑一个的情况
        input_string = ds[0]['input_string']
        value_data = input_string[input_string.find('#') + 1:]

        sql = "select param_data from sx_cases_detail_param_data where cases_detail_id = %d" % cases_detail_id
        print sql
        ds = cF.executeCaseSQL(sql)
        if len(ds) == 0:
            print u"没找到对应案例数据"
            continue
        rec = ds[0]  # 目前只考虑每个案例对应一个接口的情况
        # value_data = rec["param_data"]
        value_data_list = value_data.split('#')
        if len(value_data_list) != len(param_list):
            print u"接口参数个数不匹配 %d vs %d" % (len(value_data_list), len(param_list))
            continue
        for i in xrange(len(value_data_list)):
            value_data_list[i] = value_data_list[i] + param_list[i]
            print value_data_list[i]
        print '#'.join(value_data_list)
        sql = "update sx_cases_detail_param_data set param_data='%s' where cases_detail_id = %d" % (
            '#'.join(value_data_list), cases_detail_id)
        print sql
        cF.executeCaseSQL(sql)


def createACaseCollection():
    '''
    将当前所有案例，创建为一个用例集
    '''
    sql = "select b.cases_id,b.cases_detail_id,b.step,c.param_data_id from sx_cases_detail b, sx_cases_detail_param_data c where c.cases_detail_id = b.cases_detail_id and b.adapter_type='SPLX_T2' and b.cases_id=5933"
    ds = cF.executeCaseSQL(sql)
    if len(ds) <= 0:
        return
    c_id = ""
    c_auto_id = getAutoIncrement('sx_cases_collection')
    d_auto_id = getAutoIncrement('sx_cases_collection_detail')
    for i, rec in enumerate(ds):

        if i == 0:
            c_id = c_auto_id
            sql = "insert into sx_cases_collection(collection_id,collection_name,status) values (%d,'%s','0')" % (
                c_id, u"综合理财开户案例%s" % time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime()))
            print sql
            cF.executeCaseSQL(sql)
        d_id = d_auto_id + i
        sql = "insert into sx_cases_collection_detail(collection_id,collection_detail_id,cases_id,param_data_id) values(%d,%d,%d,%d)" % (
            c_id, d_id, int(rec["CASES_ID"]), int(rec["PARAM_DATA_ID"]))
        cF.executeCaseSQL(sql)
        print sql


def getAutoIncrement(tablename):
    sql = "show table status where name = '%s'" % tablename
    ds = cF.executeCaseSQL(sql)
    return ds[0]['Auto_increment']
    pass


def initCasesParamData():
    '''
    根据案例表更新案例数据表
    '''
    # sql = "select cases_id from sx_cases"
    # ds = cF.executeCaseSQL(sql)
    # for rec in ds:
    #    sql = "select case_detail_id,step from sx_cases_detail where cases_id = %d order by step" %rec['cases_id']
    #    ds1 = cF.executeCaseSQL(sql)

    # pass
    for i in xrange(133):
        sql = "insert into sx_cases_param_data (cases_param_data_id,cases_id,param_data_id_str) values(%d,%d,%d)" % (
            i + 1, i + 1, i + 1)
        print sql
        cF.executeCaseSQL(sql)


def scanTestAndCreateDirForCompare():
    root = os.getcwd()
    sql = "select input_string from sx_cases_detail order by cases_detail_id"
    ds = cF.executeCaseSQL(sql)
    for rec in ds:
        if not rec['INPUT_STRING'].startswith('path='):
            continue
        r = rec['INPUT_STRING'].find('#')
        script_name = rec['INPUT_STRING'][5:r]
        dir = os.path.join(root, 'verify_script\\%s' % script_name)
        print dir
        if not os.path.exists(dir):
            os.makedirs(dir)
        else:
            print "%s exists" % dir


# class FuncTree(anytree.NodeMixin):
#     def __init__(self,id,name,parent):
#         super(FuncTree,self).__init__()
#         self.id = id
#         self.name = name
#         self.parent = parent

def createANewFuncTree(path):
    tree_list = []
    csvfile = file(path, 'r')
    reader = csv.reader((line.replace('\0', '') for line in csvfile), delimiter=',')
    for i, line in enumerate(reader):
        if i == 0:
            continue
        if line[2] == "0":
            tree_list.append(FuncTree(line[0], line[1].decode('gbk'), None))
        else:
            tree_list.append(FuncTree(line[0], line[1].decode('gbk'), tree_list[int(line[2]) - 1]))

    for pre, _, node in anytree.RenderTree(tree_list[0]):
        treestr = u"%s %s %s" % (pre, node.id, node.name)
        print(treestr.ljust(8))


def importTestCases(path):
    tree_list = []
    csvfile = file(path, 'r')
    reader = csv.reader((line.replace('\0', '') for line in csvfile), delimiter=',')
    for i, line in enumerate(reader):
        if i == 0:
            continue
        detail_id = line[0]
        tree_id = line[1]
        sql = "update sx_cases_detail set TREE_DIRECTORY_ID = '%s' WHERE CASES_DETAIL_ID = %d" % (
            tree_id, int(detail_id))
        print sql
        cF.executeCaseSQL(sql)


def processLeafNode(path):
    row_list = []
    new_row_list = []
    csvfile = file(path, 'r')
    reader = csv.reader((line.replace('\0', '') for line in csvfile), delimiter=',')
    for i, line in enumerate(reader):
        row = [line[0], line[1], line[2], line[3]]
        row_list.append(row)

    # 判断是否是叶子节点
    for i, row in enumerate(row_list):
        if i == 0:
            row.append('IS_LEAF')
            new_row_list.append(row)
            continue
        has_children = False
        id = row[0]
        for j, row1 in enumerate(row_list):
            if j == 0:
                continue
            if row1[2] == id:
                has_children = True
                break
        if has_children == True:
            row.append(0)
        else:
            row.append(1)
        new_row_list.append(row)

    writer = csv.writer(file(r'C:\01_projects\TestAgent\menu1.csv', 'w'))
    writer.writerows(new_row_list)


def createANewCollectionByTaskID(id):
    '''
    根据一个任务的运行情况，重新创建一个用例集合
    '''
    sql = "select CASES_ID, CASES_PARAM_DATA_ID from sx_run_record WHERE RESULT !='STATE_PASS' and task_id = %d" % id
    ds = cF.executeCaseSQL(sql)
    c_id = ""
    c_auto_id = getAutoIncrement('sx_cases_collection')
    d_auto_id = getAutoIncrement('sx_cases_collection_detail')
    for i, rec in enumerate(ds):

        if i == 0:
            c_id = c_auto_id
            sql = "insert into sx_cases_collection(collection_id,collection_name,status) values (%d,'%s','0')" % (
                c_id, u"任务%d未通过案例集%s" % (id, time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime())))
            print sql
            cF.executeCaseSQL(sql)
        d_id = d_auto_id + i
        sql = "insert into sx_cases_collection_detail(collection_id,collection_detail_id,cases_id,param_data_id) values(%d,%d,%d,%d)" % (
            c_id, d_id, int(rec["CASES_ID"]), int(rec["CASES_PARAM_DATA_ID"]))
        cF.executeCaseSQL(sql)
        print sql


def zip_dir(dirname, zipfilename):
    filelist = []
    if os.path.isfile(dirname):
        filelist.append(dirname)
    else:
        for root, dirs, files in os.walk(dirname):
            for name in files:
                filelist.append(os.path.join(root, name))
    zf = zipfile.ZipFile(zipfilename, "w", zipfile.zlib.DEFLATED)

    # 希望保留一层目录

    for tar in filelist:
        arcname = tar[len(dirname):]
        arcname = os.path.basename(dirname) + "\\" + arcname
        # print arcname
        zf.write(tar, arcname)
    zf.close()


def unzip_file(zipfilename, unziptodir):
    if not os.path.exists(unziptodir): os.mkdir(unziptodir)
    zfobj = zipfile.ZipFile(zipfilename)
    for name in zfobj.namelist():
        name = name.replace('\\', '/')

        if name.endswith('/'):
            os.mkdir(os.path.join(unziptodir, name))
        else:
            ext_filename = os.path.join(unziptodir, name)
            ext_dir = os.path.dirname(ext_filename)
            if not os.path.exists(ext_dir): os.mkdir(ext_dir)
            outfile = open(ext_filename, 'wb')
            outfile.write(zfobj.read(name))
            outfile.close()


def zipScript(path):
    for root, dirnames, filenames in os.walk(path):
        for dir in dirnames:
            # if dir == 'bizmod.sikuli' or dir == 'common':
            #    continue
            zip_dir(os.path.join(root, dir), dir + ".zip")

        break


def UnzipScript(path):
    for root, dirnames, filenames in os.walk(path):
        for file in filenames:
            unzip_file(os.path.join(path, file), os.path.join('.\\Script', file[:-4]))

        break


def testUpdateFTPScript():
    f = XFer("gl.gl_ftp_info['FTPServer']['IP']", gl.gl_ftp_info['FTPServer']['USER'], gl.gl_ftp_info['FTPServer']['PASSWORD'], '.', gl.gl_ftp_info['FTPServer']['PORT'])
    f.login()
    f.download_files(r'.\ScriptRcv', r'.\Script')
    updated_file_list = f.get_updated_file_list()
    for file in updated_file_list:  # 只处理一层目录 fixme xiaorz
        filename = file[file.rfind('\\') + 1:-4]
        if os.path.exists(os.path.join(r'.\Script', filename)):
            shutil.rmtree(os.path.join(r'.\Script', filename))
            os.mkdir(os.path.join(r'.\Script', filename))
        unzip_file(file, os.path.join(r'.\Script', filename))
        print filename


def generateNewCollectionFromFail():
    pass


def get_cur_info():
    print sys._getframe().f_code.co_filename  # 当前文件名，可以通过__file__获得
    print sys._getframe().f_code.co_name  # 当前函数名
    print sys._getframe().f_lineno  # 当前行号


def scan_appium_inteface_csv():
    inteface_dict = OrderedDict()
    with open(r'e:\a.csv', 'r') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            inteface_dict[row[0].decode('gbk').encode('utf8')] = row[1].decode('gbk').encode('utf8')
    f = open(r"e:\setting.json", "wb")
    json.dump(inteface_dict, f, indent=4, ensure_ascii=False, encoding="utf8")


def transfer_func_param_info_to_sx_interface_info():
    # sql = "select * from sx_func_param_info"
    # ds = cF.executeCaseSQL(sql)

    conn = cF.ConnectCaseDB()
    crs = conn.cursor(MySQLdb.cursors.DictCursor)

    interface_id = 0
    interface_detail_id = 0

    # 计算合适的table_id
    sql = "SELECT MAX(ID) AS max_id FROM sx_interface_info"
    crs.execute(sql)
    ds = list(crs.fetchall())
    if len(ds) == 0 or ds[0]["max_id"] == "" or ds[0]["max_id"] == None:
        interface_id = 1
    else:
        interface_id = int(ds[0]["max_id"]) + 1

    # 计算合适的column_id
    sql = "SELECT MAX(ID) AS max_id FROM sx_interface_detail_info"
    crs.execute(sql)
    ds = list(crs.fetchall())
    if len(ds) == 0 or ds[0]["max_id"] == "" or ds[0]["max_id"] == None:
        interface_detail_id = 1
    else:
        interface_detail_id = int(ds[0]["max_id"]) + 1

    sql = "select * from sx_func_param_info"
    crs.execute(sql)
    ds = list(crs.fetchall())

    interface_param_list = []
    interface_detail_param_list = []

    # 需要找出原来的接口ID和现有ID的对照关系，并逐一修改
    case_detail_param_id_dict = OrderedDict()
    for rec in ds:
        # 一条记录对应sx_interface_info 一条记录
        # param_info的个数对应sx_interface_detail_info 的个数

        case_detail_param_id_dict[rec["PARAM_ID"]] = str(interface_id)
        interface_param = (interface_id, rec["PARAM_NAME"], rec["TREE_DIRECTORY_ID"], rec["SYSTEM_ID"])
        input_param = rec["PARAM_INFO"].replace("=", "")
        input_describe = rec["PARAM_DESCRIBE"]
        for i, param in enumerate(input_param.split('#')):
            interface_detail_param = (
                interface_detail_id, interface_id, "IN", param, input_describe.split("&seac&")[i], i + 1)
            interface_detail_param_list.append(interface_detail_param)
            interface_detail_id = interface_detail_id + 1
        interface_param_list.append(interface_param)
        interface_id = interface_id + 1

    sql = "insert into sx_interface_info (ID,interface_name,tree_directory_id,system_id) values (%s,%s,%s,%s)"
    crs.executemany(sql, interface_param_list)

    sql = "insert into sx_interface_detail_info (ID,interface_id,input_type,param_name,param_desc,param_index) values (%s,%s,%s,%s,%s,%s)"
    crs.executemany(sql, interface_detail_param_list)

    for id in case_detail_param_id_dict.keys():
        sql = "update sx_cases_detail set param_id = %s where param_id = %s" % (case_detail_param_id_dict[id], id)
        crs.execute(sql)
    crs.close()
    conn.commit()
    conn.close()


def start_appium_server():
    '''
    启动appium server 需要使用线程启动，以避免阻塞
    :return:
    '''
    # if self._appium_path == "":
    #     Logging.getLog().debug(u"未找到appium_path")
    #     return 0
    _appium_is_started = False
    today = datetime.datetime.now().strftime('%Y%m%d')
    f = open('.\\logs\\appium_server_%s.log' % today, "a")
    print("begin to start appium server")

    # 移除日志文件
    if os.path.exists("appium.log"):
        print "remove appium.log"
        os.remove("appium.log")
    cmd_str = 'node "C:\\Program Files (x86)\\Appium\\node_modules\\appium\\bin\\appium.js" >>appium.log 2>&1'
    si = subprocess.STARTUPINFO
    # si.dwFlags = subprocess.CREATE_NEW_CONSOLE | subprocess.STARTF_USESHOWWINDOW
    # si.wShowWindow = subprocess.SW_HIDE
    # proc = subprocess.Popen(cmd_str, shell=False, startupinfo=si, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, stdin=subprocess.PIPE,creationflags=subprocess.CREATE_NEW_CONSOLE)
    proc = subprocess.Popen(cmd_str, shell=True, creationflags=subprocess.CREATE_NEW_CONSOLE)

    while proc.poll() is None:  # 检测后台进程是否终止
        if _appium_is_started == False:
            if os.path.exists("appium.log"):
                f = open("appium.log", "r")
                buf = f.readlines()
                for line in buf:
                    if line.__contains__("Appium REST http interface listener started"):
                        print line
                        _appium_is_started = True
                f.close()

                # line = proc.stdout.readline()
                # f.write(line)
                # #print line
                # if _appium_is_started == False:
                #     if line.__contains__("Appium REST http interface listener started"):
                #         print line
                #         _appium_is_started = True
    # f.close()
    _appium_is_started = False
    exit_code = proc.returncode


def testDB2Clob():
    sql = "UPDATE sx_run_record set data_change_info ='asdf' where run_record_id = 1643"
    ds = cF.executeCaseSQL_DB2(sql)
    print ds


def readConfigFileForTools():
    '''读取XML系统配置文件 '''
    try:
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        elif __file__:
            application_path = os.path.dirname(__file__)
        else:
            print(u"无法获取application_path")
            return -1
        print("application_path = %s" % application_path)

        print('try to switch cwd to %s' % application_path)
        os.chdir(application_path)

        tree = ET.parse(".\\init\\14774643386617P4VW\\ZHLCAutoTestSetting.xml")  # fixme 不能使用这个固定的路径，以后更改
        root = tree.getroot()
    except BaseException, ex:
        print ex
        return -1
    try:
        # 解析帐号信息
        # 帐号信息格式为以类型为键值的字典类型 {type1:value1,type2:value2 ……}

        account_group_id = tree.findall('.//AccountInfo')
        dist = {}
        for elem in account_group_id:
            for xx in elem.getchildren():
                tx = xx.text.replace(',', '&comma&')
                tx = tx.replace(':', '&colon&')
                dist[xx.tag] = tx
            gl.g_accountDict[elem.attrib["id"]] = dist  # fixme 账号信息以后尽量不从这个文件中读取
            dist = {}
        #
        # 解析案例数据库信息

        # 将用例数据库配置信息
        # testcaseDBElem = root.find("TestCaseDataBase")
        # for elem in testcaseDBElem:
        #     gl.g_testcaseDBDict[elem.tag] = elem.text

        if os.path.exists("settings.json"):
            _setting = json.loads(open("settings.json", "r").read())
            gl.g_testcaseDBDict["DBType"] = _setting["casedb"]["DBType"]
            gl.g_testcaseDBDict["casedb"] = _setting["casedb"][gl.g_testcaseDBDict["DBType"]]

        # 解析BeyondCompare程序路径
        # ByondCmpElem = root.find("BeyondComparePath/path")
        # gl.g_bcmppath = ByondCmpElem.text

        # 解析集中交易数据库信息
        # SqlServerElem = root.find("ConnectSqlServerSetting")
        # for serverElem in SqlServerElem:
        #     gl.g_connectSqlServerSetting[serverElem.attrib["core"]] ={}
        #     for elem in serverElem:
        #         gl.g_connectSqlServerSetting[serverElem.attrib["core"]][elem.tag]=elem.text
        #
        #
        # 解析综合理财数据库信息
        OracleElem = root.find("ConnectOracleSetting")
        for serverElem in OracleElem:
            gl.g_connectOracleSetting[serverElem.attrib["core"]] = {}
            for elem in serverElem:
                gl.g_connectOracleSetting[serverElem.attrib["core"]][elem.tag] = elem.text

        # 解析综合理财系统信息
        ZHLCSystemSettingElem = root.find("ZHLCSystemSetting")
        for elem in ZHLCSystemSettingElem:
            gl.g_ZHLCSystemSetting[elem.tag] = elem.text

            # #解析需要修改运行时间的服务器列表
            # TimeSyncServerSettingElem = root.find("TimeSyncServerSetting")
            # if TimeSyncServerSettingElem != None:
            #     for elem in TimeSyncServerSettingElem:
            #         gl.g_timeSyncServerSettingList.append(elem.attrib["ip"])
            # loadSYSCases()
    except BaseException, ex:
        print ex
        return -1
    print(u"解析通用配置文件ZHLCAutoTestSetting.xml 完毕")
    return 0


def RunSeleniumScript(system_id, funcid, param_data_id, runtaskid, run_record_id):
    """
    运行Selenium脚本，Selenium使用了统一的一个脚本，通过参数的传递来调用相关功能。
    Args:
        system_id: 系统ID
        funcid: 已经经过参数拼接的命令串
        param_data_id: 参数ID
        runtaskid:运行任务ID
        run_record_id: 运行记录ID

    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息
        log_info : 日志信息
    """

    log_info = ""

    cmd_str = 'java -jar .\\JZYYAutoTest.jar %s %s' % (funcid, param_data_id)

    try:
        si = subprocess.STARTUPINFO
        si.wShowWindow = subprocess.SW_HIDE
        # Logging.getLog().debug("ready to execute")
        proc = subprocess.Popen(cmd_str, shell=True, startupinfo=si, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                stdin=subprocess.PIPE)
        # proc.wait()
        while proc.poll() is None:
            line = proc.stdout.readline()
            line = line.strip()
            if line:
                print "%s" % line
                logging.info("%s" % line)

        # 退出监控线程
        # Logging.getLog().debug("quit subprocess")
        print "quit subprocess with returncode = %s" % str(proc.returncode)

        # output_info = proc.stdout.readlines()
        # log_info = "".join(output_info)
        if proc.returncode == 0:
            return 0, "", "", ""
        else:
            return -1, "", "", ""
            # 启动监控线程
            #
            # print cmd_str
            # f = os.popen(cmd_str.encode('gb2312'))
            # output_info = f.read()
            # exit_code = f.close()
            # if exit_code == None:
            #     exit_code = 0
            #
            # # 退出监控线程
            # print output_info
            # print exit_code
            # output_info = proc.stdout.readlines()
            # print (output_info)
            # exit_code = proc.returncode
            # print "return code is %s" % str(exit_code)

    except BaseException, ex:
        exc_info = cF.getExceptionInfo()
        print u"异常发生，请检查日志文件,提示信息为:%s" % (exc_info)
        return -1, u"异常发生，请检查日志文件,提示信息为:%s" % (exc_info), "", ""


def deal_with_pre_pro_info(colnames, FLAG, path, old_param_new_params, funcid, ACCOUNT_NUM_INFO):
    try:
        ACCOUNT_GROUP_ID = ""
        result = []
        wb = load_workbook(path)
        sheets = wb.get_sheet_names()
        pre_pro = ""
        count = 0
        # f = open("funcid_account_info", 'a+')
        for i in sheets:
            sheet = wb.get_sheet_by_name(i)
            for row in sheet.iter_rows():
                count += 1
                if count == 1:
                    continue
                ACCOUNT = ""
                line_data = []
                CASE_NAME = row[0].value
                ACCOUNTGROUPID = row[1].value
                func_id = row[2].value
                PRE_PRO = row[3].value
                if not ACCOUNTGROUPID or ACCOUNTGROUPID == 'None':
                    ACCOUNTGROUPID = ""
                ACCOUNT = ACCOUNTGROUPID
                if func_id and str(funcid) == str(func_id) and PRE_PRO:
                    # pre_pro = PRE_PRO_ACTION
                    for accout_group in old_param_new_params:
                        if str(accout_group) in PRE_PRO:
                            if ACCOUNTGROUPID and ACCOUNTGROUPID != 'None':
                                ACCOUNT = ACCOUNTGROUPID
                            else:
                                ACCOUNT = ACCOUNT_NUM_INFO[str(accout_group)]
                            funcid_account = str(func_id) + "=" + str(ACCOUNT) + "\n"
                            # f.write(funcid_account)
                            PRE_PRO = PRE_PRO.replace(str(accout_group), old_param_new_params[str(accout_group)])
                    if ACCOUNT == "":
                        ACCOUNT_GROUP_ID = 0
                    else:
                        ACCOUNT_GROUP_ID = int(ACCOUNT)
                    pre_pro = PRE_PRO
                    # break
                line_data.append(CASE_NAME)
                line_data.append(str(ACCOUNT))
                line_data.append(func_id)
                line_data.append(PRE_PRO)
                result.append(line_data)
        # f.close()
        cF.export_data(colnames, FLAG, result)
        return pre_pro, ACCOUNT_GROUP_ID
    except BaseException, ex:
        exc_info = cF.getExceptionInfo()
        print u"异常发生，请检查日志文件,提示信息为:%s" % (exc_info)
        return -1, u"异常发生，请检查日志文件,提示信息为:%s" % (exc_info), "", ""


def deal_account_info(path, old_param_new_param):
    ACCOUNT_NUM_INFO = {}
    wb = load_workbook(path)
    sheets = wb.get_sheet_names()
    for i in sheets:
        sheet = wb.get_sheet_by_name(i)
        count = 0
        for row in sheet.iter_rows():
            count += 1
            if count == 1:
                continue
            old_account = row[0].value
            ACCOUNT_NUM = row[1].value
            new_account = row[2].value

            key = "@" + str(old_account) + "@"
            value = "@" + str(new_account) + "@"
            old_param_new_param[key] = value
            ACCOUNT_NUM_INFO[key] = ACCOUNT_NUM

    return old_param_new_param, ACCOUNT_NUM_INFO


def testSeleniumScript(funcid):
    '''
        自动化测试工具 便于本地测试案例
        replace_account_group_info文件当中的信息，例如A=B。 A是旧的案例账户组信息，B是新的案例账户组信息
        PRE_ACTION.xlsx  生产上存储的前置信息
        PRO_ACTION.xlsx  生产上存储的后置信息
    '''
    try:
        funcid = funcid.strip()
        flag = 0
        old_param_new_param = {}
        homedir = os.getcwd()
        # path_txt = homedir + "/replace_account_group_info.txt"
        # f = open(path_txt, "r")
        # text = f.read()
        # f.close()
        account_group_path = homedir + "/replace_account_group_info.xlsx"
        old_param_new_param, ACCOUNT_NUM_INFO = deal_account_info(account_group_path, old_param_new_param)
        datetime = time.strftime('%Y%m%d', time.localtime())
        log_path = homedir + "/tools_log/" + str(datetime) + ".txt"
        fw = open(log_path, "a")
        fw.write("[" + str(funcid) + "]" + str(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())) + "\n")
        # old_param_new_param = text.split("\n")
        # for param in old_param_new_param:
        #     if param:
        #         param = param.split("#")[0]
        #         key = "@"+str(param.split("=")[0])+"@"
        #         value = "@"+str(param.split("=")[1])+"@"
        #         old_param_new_params[key] = value
        log_record(funcid)
        group_variable_dict = {}
        pre_path = homedir + "\\PRE_ACTION.xlsx"
        pro_path = homedir + "\\PRO_ACTION.xlsx"
        # 前置
        pres_name = ['CASE_NAME', 'ACCOUNT_GROUP_ID', 'REMARKS', 'PRE_ACTION']
        # 默认获取0组的信息
        account_group_id = 0
        pres, account_group_id = deal_with_pre_pro_info(pres_name, "PRE_ACTION", pre_path, old_param_new_param, funcid,
                                                        ACCOUNT_NUM_INFO)
        # 后置
        pros_name = ['CASE_NAME', 'ACCOUNT_GROUP_ID', 'REMARKS', 'PRO_ACTION']
        pros, account_group_id2 = deal_with_pre_pro_info(pros_name, "PRO_ACTION", pro_path, old_param_new_param, funcid,
                                                         ACCOUNT_NUM_INFO)
        if account_group_id == '':
            account_group_id = 0
        print "账户组信息:%s" % account_group_id
        pres = cF.GlobalParameterReplace(pres, account_group_id)
        ret, msg_info = cF.action_process(pres, group_variable_dict, dataset=[], funcid=None, cmdstring=None,
                                          account_group_id=account_group_id)
        if ret != 0:
            flag = 1
            print u"前置失败，原因%s" % msg_info
            fw.write(u"前置失败，原因%s\n" % msg_info)
            fw.write("运行失败\n")
            return
        else:
            fw.write("前置success\n")
            logging.info("前置success")

        PARAM_DATA_IDS = cF.ExcelProcessParamTools('14774643386617P4VW', funcid, group_variable_dict, account_group_id,
                                                   old_param_new_param)
        if PARAM_DATA_IDS:
            PARAM_DATA_IDS = list(set(PARAM_DATA_IDS))
            for PARAM_DATA_ID in PARAM_DATA_IDS:
                ret, msg_info, log_info, data_change_info_json_string = RunSeleniumScript('14774643386617P4VW', funcid,
                                                                                          PARAM_DATA_ID, 1, 1)
                if ret < 0:
                    flag = 1
                    fw.write("jzyy 脚本运行 fail\n")
                    logging.info("jzyy 脚本运行 fail")
                    # CaptureScreen(funcid)
                    window_capture(funcid)
                else:
                    fw.write("jzyy 脚本运行 success\n")
                    logging.info("jzyy 脚本运行 success")
        if ret < 0:
            pros = cF.GlobalParameterReplace(pros, account_group_id)
            pros = cF.GroupParameterReplace(pros, group_variable_dict)
            pros_ret, pros_msg_info = cF.action_process_selenium(pros, group_variable_dict, -1, dataset=[], funcid=None,
                                                                 cmdstring=None, account_group_id=account_group_id)
            if pros_ret < 0:
                print u"run Selenium fail，原因:msg_info=%s,log_info=%s；后置失败，原因：pros_msg_info=%s" % (
                    msg_info, log_info, pros_msg_info)
            else:
                print u"run Selenium fail，原因:msg_info=%s,log_info=%s" % (msg_info, log_info)
                # return ret

        pros = cF.GlobalParameterReplace(pros, account_group_id)
        pros = cF.GroupParameterReplace(pros, group_variable_dict)
        ret, msg_info = cF.action_process_selenium(pros, group_variable_dict, 0, dataset=[], funcid=None,
                                                   cmdstring=None, account_group_id=account_group_id)
        if ret != 0:
            flag = 1
            fw.write(u"后置失败，原因%s\n" % msg_info)
            print u"后置失败，原因%s" % msg_info
        else:
            fw.write("后置success\n")
            logging.info("后置success")
        if flag == 0:
            fw.write("运行成功\n")
        else:
            fw.write("运行失败\n")
        fw.write("\n")
        fw.close()
    except BaseException, ex:
        exc_info = cF.getExceptionInfo()
        print u"异常发生，请检查日志文件,提示信息为:%s" % (exc_info)
        logging.error(u"异常发生，请检查日志文件,提示信息为:%s" % (exc_info))
        return -1, u"异常发生，请检查日志文件,提示信息为:%s" % (exc_info), "", ""


def window_capture(funcid):
    '''''
  截屏函数,调用方法window_capture('d:\\') ,参数为指定保存的目录
  返回图片文件名,文件名格式:日期.jpg 如:2009328224853.jpg
    '''
    try:
        global error_photo_path
        if not os.path.exists(error_photo_path):
            os.makedirs(error_photo_path)
        hwnd = 0
        hwndDC = win32gui.GetWindowDC(hwnd)
        mfcDC = win32ui.CreateDCFromHandle(hwndDC)
        saveDC = mfcDC.CreateCompatibleDC()
        saveBitMap = win32ui.CreateBitmap()
        MoniterDev = win32api.EnumDisplayMonitors(None, None)
        w = MoniterDev[0][2][2]
        h = MoniterDev[0][2][3]
        # print w,h　　　#图片大小
        saveBitMap.CreateCompatibleBitmap(mfcDC, w, h)
        saveDC.SelectObject(saveBitMap)
        saveDC.BitBlt((0, 0), (w, h), mfcDC, (0, 0), win32con.SRCCOPY)
        # cc=time.gmtime()
        # bmpname=str(cc[0])+str(cc[1])+str(cc[2])+str(cc[3]+8)+str(cc[4])+str(cc[5])+'.bmp'
        bmpname = "%s.bmp" % (str(funcid))
        saveBitMap.SaveBitmapFile(saveDC, bmpname)
        Image.open(bmpname).save(bmpname[:-4] + ".jpeg")
        os.remove(bmpname)
        jpgname = bmpname[:-4] + '.jpeg'
        djpgname = error_photo_path + "\\" + jpgname
        copy_command = "move %s %s" % (jpgname, djpgname)
        os.popen(copy_command)
        return 0, ''
    except:
        exc_info = cF.getExceptionInfo()
        logging.error().error(exc_info)
        return -1, exc_info


def CaptureScreen(funcid):
    """
    截屏，在Agent本地Log中保存一份，再上传FTP一份
    Args:
        runtaskid: 任务id
        filename: 数据ID
    Returns:
        一个元组（ret，msg）：
        ret : 0 表示 成功，-1 表示 失败
        msg : 对应信息
    """
    try:
        print "CaptureScreen"
        im = ImageGrab.grab()
        global error_photo_path
        if not os.path.exists(error_photo_path):
            os.makedirs(error_photo_path)

        # if filename.startswith('common\\'):
        #    filename = filename[7:]
        # 保存错误截图
        jpg_file = "%s\\%s.jpeg" % (error_photo_path, str(funcid))
        print "jpgfile = %s" % jpg_file
        im.save(jpg_file, 'jpeg')
    except:
        exc_info = cF.getExceptionInfo()
        logging.error().error(exc_info)
        return -1, exc_info


def testSeleniumBysubprocess():
    si = subprocess.STARTUPINFO
    si.wShowWindow = subprocess.SW_HIDE
    # Logging.getLog().debug("ready to execute")
    proc = subprocess.Popen(r"java -jar .\Script\JZYYAutotest.jar 0010010001 140489", shell=True, startupinfo=si,
                            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                            stdin=subprocess.PIPE)
    # proc.wait()
    while proc.poll() is None:
        line = proc.stdout.readline()
        line = line.strip()
        if line:
            print "output %s" % line

    # 退出监控线程
    # Logging.getLog().debug("quit subprocess")
    print "quit subprocess"

    # output_info = proc.stdout.readlines()
    # log_info = "".join(output_info)
    print proc.returncode
    # print log_info

def test_hradapter():
    import hradapter
    for i in range(1000):
        print u"第%s次"%i
        hr_quant_handle = hradapter.HRAdapter()
        ret,msg = hr_quant_handle.ha_init('10.187.112.160',9056)
        print ret
        if ret != 0:
            print msg
        hr_quant_handle.ha_exit()
        # req = open("req.json", "r").read()
        # print req
        # ret,info = hr_quant_handle.send_msg(830020,req)
        # ret, info = hr_quant_handle.send_msg(830012, req)
        # print ret
        # print info
        #
        # req = open("req3.json","r").read()
        # print req
        # ret,info = hr_quant_handle.send_msg(830014,req)
        # # ret,info = hr_quant_handle.send_msg(830150,req)
        # print ret
        # print info

if __name__ == '__main__':
    # test_hradapter()
    # test_hradapter()
    # if len(sys.argv) != 2:
    #     print u"参数个数不对，参数为pre_path,pro_path,funcid,param_index"
    #     sys.exit(-1)

    # cF.readConfigFileForTools()
    # # # testSeleniumBysubprocess()
    # # # testDB2Clob()
    # # # testSeleniumScript(sys.argv[1],sys.argv[2],sys.argv[3],sys.argv[4])
    # # funcid = sys.argv[1]
    # testSeleniumScript("0010010001")
    # testDB2Clob()
    # start_appium_server()
    # transfer_func_param_info_to_sx_interface_info()
    # createACaseCollection()
    # scan_appium_inteface_csv()
    cF.readConfigFile()
    # createANewCollectionByTaskID(2)
    # zipScript(r'C:\01_projects\FUYIAutoTest\Script')
    # UnzipScript(r'C:\01_projects\TestAgent\ScriptRcv')
    # testUpdateFTPScript()
    # exportFuncTree(None)
    # importFuncTree(r'C:\01_projects\TestAgent\menu.csv')
    # zip_dir(r"c:\test-zip","test-zip.zip")
    # unzip_file(r"C:\01_projects\TestAgent\test-zip.zip",".")
    # str_a = "asdf'a'sdf'"
    # sql = "update sx_cases_detail set pro_action = ? where cases_detail_id=287" %str_a
    result_list = []
    result_list.append(('RUN_TYPE_IMMI', 1010))
    result_list.append(('RUN_TYPE_IMMI', 1011))
    # cF.executeCaseSQL(sql)
    conn = cF.ConnectCaseDB('DB2')
    crs = conn.cursor()
    crs.executemany("update sx_run_record set RUN_TYPE=? where RUN_RECORD_ID=?", result_list)
    # crs.execute("update sx_cases_detail set pro_action = ?,pre_action = ? where cases_detail_id=287",(str_a,str_a))
    crs.close()
    conn.commit()
    conn.close()
    # get_cur_info()
    pass
