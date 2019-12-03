# coding=GBK
"""
workbook JZYY
"""
import os
import pyodbc
from openpyxl import Workbook
from openpyxl import load_workbook

def read_funcid_param():
    try:
        INPUT_TYPE = "IN"
        tree_id = 21178
        system_id = 'JZYY'
        path = r"D:\db2_data\ZHLC\work_book_info\JZYY_DATA"
        file_path = ""
        for root, dirs, files in os.walk(path):
            print files
            for i in files:
                file_path = path + '\\' + i
                wb = load_workbook(file_path)
                sheets = wb.get_sheet_names()
                for i in sheets:
                    print 'insert_intface_info'
                    INTERFACE_ID = insert_intface_info(i, i, i, tree_id, system_id)
                    sheet = wb.get_sheet_by_name(i)
                    for row in sheet.iter_rows():
                        PARAM_INDEX = 0
                        for cell in row:
                            col_name = cell.value
                            if col_name == 'None':
                                continue
                            if col_name:
                                PARAM_INDEX += 1
                                insert_intfacedetail_info(INTERFACE_ID, INPUT_TYPE, col_name, col_name, PARAM_INDEX)
                        break
    except Exception, e:
        raise e

def insert_intface_info(INTERFACE_NAME, FUNCID, FUNCID_DESC, TREE_DIRECTORY_ID, system_id):
    type = 'DB2'
    conn = None
    if type == 'DB2':
        # dbstring = "driver={IBM DB2 ODBC DRIVER};database=CSZDH;hostname=10.189.96.6;port=50000;protocol=TCPIP;uid=TIMESVC;pwd=timesvc;"
        dbstring = "driver={IBM DB2 ODBC DRIVER};database=TIMESVC;hostname=10.187.96.46;port=50000;protocol=TCPIP;uid=timesvc;pwd=csgl@2017;"
        conn = pyodbc.connect(dbstring)
    cursor = conn.cursor()
    try:
        args = (INTERFACE_NAME, FUNCID, FUNCID_DESC, int(TREE_DIRECTORY_ID), system_id)
        sql = "insert into TIMESVC.SX_INTERFACE_INFO(INTERFACE_NAME, FUNCID, FUNCID_DESC, TREE_DIRECTORY_ID, SYSTEM_ID) " \
              "VALUES ('%s', '%s', '%s', %s, '%s')" % args
        cursor.execute(sql)
        conn.commit()
        max_id = get_row_id()
        return max_id
    except Exception as e:
        raise e
    finally:
        cursor.close()
        conn.close()

def insert_intfacedetail_info(INTERFACE_ID, INPUT_TYPE, PARAM_NAME, PARAM_DESC, PARAM_INDEX):
    type = 'DB2'
    conn = None
    if type == 'DB2':
        # dbstring = "driver={IBM DB2 ODBC DRIVER};database=CSZDH;hostname=10.189.96.6;port=50000;protocol=TCPIP;uid=TIMESVC;pwd=timesvc;"
        dbstring = "driver={IBM DB2 ODBC DRIVER};database=TIMESVC;hostname=10.187.96.46;port=50000;protocol=TCPIP;uid=timesvc;pwd=csgl@2017;"
        conn = pyodbc.connect(dbstring)
    cursor = conn.cursor()
    try:
        args = (int(INTERFACE_ID), INPUT_TYPE, PARAM_NAME, PARAM_DESC, PARAM_INDEX)
        sql = "insert into TIMESVC.SX_INTERFACE_DETAIL_INFO(INTERFACE_ID, INPUT_TYPE, PARAM_NAME, " \
              "PARAM_DESC, PARAM_INDEX) VALUES(%s,'%s','%s','%s',%s)" % args
        cursor.execute(sql)
    except Exception as e:
        raise e
    finally:
        conn.commit()
        cursor.close()
        conn.close()

# def write_funcid_param(system_id, funcid, row, row_value_list):
#     try:
#         wb = load_workbook("init/%s/%s.xlsx" % (system_id, funcid[0:3]))
#         sheet = wb.get_sheet_by_name(funcid)
#         # print sheet['A2'].value
#         for i, col in enumerate(sheet[row + 1]):
#             col.number_format = '@'
#             col.value = row_value_list[i]
#         wb.save("init/%s/%s.xlsx" % (system_id, funcid[0:3]))
#     except Exception, e:
#         raise e


def get_row_id():
    # 获取旧版本的综合理财自动化测试系统的测试用库获tbl_cases
    type = 'DB2'
    conn = None
    if type == 'DB2':
        # dbstring = "driver={IBM DB2 ODBC DRIVER};database=CSZDH;hostname=10.189.96.6;port=50000;protocol=TCPIP;uid=TIMESVC;pwd=timesvc;"
        dbstring = "driver={IBM DB2 ODBC DRIVER};database=TIMESVC;hostname=10.187.96.46;port=50000;protocol=TCPIP;uid=timesvc;pwd=csgl@2017;"
        conn = pyodbc.connect(dbstring)
    cursor = conn.cursor()
    try:
        sql = "select max(ID) FROM TIMESVC.SX_INTERFACE_INFO"
        cursor.execute(sql)
        result = cursor.fetchall()
        max_id = 0
        if result:
            max_id = result[0][0]
        return max_id
    except Exception as e:
        raise e
    finally:
        cursor.close()
        conn.close()

if __name__ == '__main__':
    read_funcid_param()